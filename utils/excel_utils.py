from openpyxl import load_workbook


def iter_rows_from_xlsx(file_stream):
    """
    Reads an uploaded .xlsx file using openpyxl and returns:
    - headers as a list of strings
    - data rows as list of dicts (header:value pairs)
    Uses the exact logic currently used in the Lifts upload module.
    """
    wb = load_workbook(file_stream, data_only=True)
    try:
        sheet = wb.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], []

        headers = [str(h).strip() if h else "" for h in rows[0]]
        data_rows = []

        for row in rows[1:]:
            row_dict = {}
            for idx, cell_value in enumerate(row):
                key = headers[idx]
                row_dict[key] = cell_value
            data_rows.append(row_dict)

        return headers, data_rows
    finally:
        wb.close()

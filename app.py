from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    jsonify,
    abort,
    session,
    send_file,
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager,
    login_user,
    login_required,
    logout_user,
    UserMixin,
    current_user,
)
from werkzeug.utils import secure_filename
import os, json, datetime, sqlite3, threading, re, uuid, random, string, copy, calendar, base64, shutil
import importlib.util
import csv
from io import BytesIO, StringIO
from collections import OrderedDict, Counter, defaultdict
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

from sqlalchemy import case, inspect, func, or_, and_
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import joinedload, subqueryload

OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None

if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
else:
    Workbook = load_workbook = Alignment = Font = None  # type: ignore[assignment]

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__, instance_relative_config=True)
app.config["SECRET_KEY"] = "dev-eleva-secret"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(BASE_DIR, "instance", "eleva.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = os.path.join("static", "uploads")
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200MB uploads

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "instance"), exist_ok=True)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"


class MissingDependencyError(RuntimeError):
    """Raised when an optional dependency required for a feature is unavailable."""


OPENPYXL_MISSING_MESSAGE = (
    "Excel support requires the openpyxl package. Install dependencies with "
    "`pip install -r requirements.txt` and restart the application."
)


def _ensure_openpyxl():
    if not OPENPYXL_AVAILABLE:
        raise MissingDependencyError(OPENPYXL_MISSING_MESSAGE)


def _extract_tabular_upload(upload, *, sheet_name=None):
    filename = (upload.filename or "").lower()
    if filename.endswith(".xlsx"):
        _ensure_openpyxl()
        upload.stream.seek(0)
        workbook = load_workbook(upload, data_only=True)
        worksheet = (
            workbook[sheet_name]
            if sheet_name and sheet_name in workbook.sheetnames
            else workbook.active
        )
        header = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            [],
        )

        def row_iter():
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                yield row

        return header, row_iter()

    if filename.endswith(".csv"):
        upload.stream.seek(0)
        raw_bytes = upload.read()
        upload.stream.seek(0)
        try:
            text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_bytes.decode("latin-1")
        reader = csv.reader(StringIO(text))
        rows = list(reader)
        header = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        return header, data_rows

    raise ValueError("Unsupported file type")


PENDING_UPLOAD_SUBDIR = "pending"


@dataclass
class UploadOutcome:
    header_map: Dict[str, int]
    created_count: int
    updated_count: int
    processed_rows: int
    created_items: List[Dict[str, Any]] = field(default_factory=list)
    updated_items: List[Dict[str, Any]] = field(default_factory=list)
    row_errors: List[str] = field(default_factory=list)


def save_pending_upload_file(upload):
    upload_root = app.config["UPLOAD_FOLDER"]
    pending_root = os.path.join(upload_root, PENDING_UPLOAD_SUBDIR)
    os.makedirs(pending_root, exist_ok=True)
    original_name = upload.filename or "upload"
    extension = os.path.splitext(original_name)[1].lower()
    token = uuid.uuid4().hex
    dest_path = os.path.join(pending_root, f"{token}{extension}")
    upload.stream.seek(0)
    upload.save(dest_path)
    upload.stream.seek(0)
    return dest_path


def _clear_pending_upload(pending_token, *, remove_file=False):
    pending_uploads = session.get("pending_uploads", {})
    pending = pending_uploads.pop(pending_token, None)
    session["pending_uploads"] = pending_uploads
    session.modified = True
    if remove_file:
        file_path = (pending or {}).get("path")
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                pass
    return pending


def _extract_tabular_upload_from_path(file_path, *, sheet_name=None):
    filename = file_path.lower()
    if filename.endswith(".xlsx"):
        _ensure_openpyxl()
        workbook = load_workbook(file_path, data_only=True)
        worksheet = (
            workbook[sheet_name]
            if sheet_name and sheet_name in workbook.sheetnames
            else workbook.active
        )
        header = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            [],
        )
        data_rows = [row for row in worksheet.iter_rows(min_row=2, values_only=True)]
        workbook.close()
        return header, data_rows

    if filename.endswith(".csv"):
        with open(file_path, "rb") as fh:
            raw_bytes = fh.read()
        try:
            text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_bytes.decode("latin-1")
        reader = csv.reader(StringIO(text))
        rows = list(reader)
        header = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        return header, data_rows

    raise ValueError("Unsupported file type")


def _customer_identifier(customer, *, fallback):
    if isinstance(customer, Customer) and getattr(customer, "id", None) is not None:
        return f"existing:{customer.id}"
    return fallback


def process_customer_upload_file(file_path, *, apply_changes):
    header_cells, data_rows = _extract_tabular_upload_from_path(
        file_path, sheet_name=CUSTOMER_UPLOAD_TEMPLATE_SHEET_NAME
    )
    header_cells = header_cells or []
    header_map: Dict[str, int] = {}
    for idx, header in enumerate(header_cells or []):
        label = stringify_cell(header)
        if label:
            header_map[label] = idx

    outcome = UploadOutcome(
        header_map=header_map,
        created_count=0,
        updated_count=0,
        processed_rows=0,
    )

    customers = Customer.query.all()
    existing_by_code: Dict[str, Any] = {
        customer.customer_code.lower(): customer
        for customer in customers
        if customer.customer_code
    }
    existing_by_external: Dict[str, Any] = {
        customer.external_customer_id.lower(): customer
        for customer in customers
        if customer.external_customer_id
    }

    service_routes = ServiceRoute.query.all()
    route_lookup: Dict[str, Any] = {}
    for route in service_routes:
        options = {route.state.lower()}
        display_name = clean_str(route.display_name)
        if display_name:
            options.add(display_name.lower())
        if route.branch:
            options.add(route.branch.lower())
            options.add(f"{route.state.lower()} · {route.branch.lower()}")
            options.add(f"{route.state.lower()}-{route.branch.lower()}")
        for option in options:
            route_lookup[option] = route

    processed_codes: Dict[str, str] = {}
    processed_external_ids: Dict[str, str] = {}
    generated_codes: set[str] = set()

    for row_index, row_values in enumerate(data_rows, start=2):
        if not row_values:
            continue
        row_data: Dict[str, Any] = {}
        for header, position in header_map.items():
            value = row_values[position] if position < len(row_values) else None
            row_data[header] = value

        key_fields = [
            row_data.get("External Customer ID"),
            row_data.get("Customer Code"),
            row_data.get("Company Name"),
            row_data.get("Email"),
            row_data.get("Phone"),
        ]
        if not any(clean_str(stringify_cell(value)) for value in key_fields):
            continue

        outcome.processed_rows += 1

        external_id_value = clean_str(stringify_cell(row_data.get("External Customer ID")))
        customer_code_value = clean_str(stringify_cell(row_data.get("Customer Code")))
        company_name_value = clean_str(stringify_cell(row_data.get("Company Name")))
        contact_person_value = clean_str(stringify_cell(row_data.get("Contact Person")))
        phone_value = clean_str(stringify_cell(row_data.get("Phone")))
        mobile_value = clean_str(stringify_cell(row_data.get("Mobile")))
        email_value = clean_str(stringify_cell(row_data.get("Email")))
        gst_value = clean_str(stringify_cell(row_data.get("GST Number")))
        billing_line1 = clean_str(stringify_cell(row_data.get("Billing Address Line 1")))
        billing_line2 = clean_str(stringify_cell(row_data.get("Billing Address Line 2")))
        city_value = clean_str(stringify_cell(row_data.get("City")))
        state_value = clean_str(stringify_cell(row_data.get("State")))
        pincode_value = clean_str(stringify_cell(row_data.get("Pincode")))
        country_value = clean_str(stringify_cell(row_data.get("Country")))
        route_value_raw = clean_str(stringify_cell(row_data.get("Route")))
        sector_value = clean_str(stringify_cell(row_data.get("Sector")))
        notes_value = clean_str(stringify_cell(row_data.get("Notes")))
        office_line1 = clean_str(stringify_cell(row_data.get("Office Address Line 1")))
        office_line2 = clean_str(stringify_cell(row_data.get("Office Address Line 2")))
        office_city_value = clean_str(stringify_cell(row_data.get("Office City")))
        office_state_value = clean_str(stringify_cell(row_data.get("Office State")))
        office_pincode_value = clean_str(stringify_cell(row_data.get("Office Pincode")))

        branch_value = None
        branch_error = None
        if "Branch" in header_map:
            branch_value, branch_error = validate_branch(
                row_data.get("Branch"), label="Branch", required=False
            )
            if branch_error:
                outcome.row_errors.append(f"Row {row_index}: {branch_error}")
                continue

        route_value = None
        if route_value_raw:
            lookup_key = route_value_raw.lower()
            route = route_lookup.get(lookup_key)
            if not route:
                outcome.row_errors.append(
                    f"Row {row_index}: Route '{route_value_raw}' does not match an active service route."
                )
                continue
            route_value = route.state

        existing_customer = None
        if customer_code_value:
            lookup_code = customer_code_value.lower()
            existing_customer = existing_by_code.get(lookup_code)

        if external_id_value:
            lookup_external = external_id_value.lower()
            external_match = existing_by_external.get(lookup_external)
            if isinstance(existing_customer, Customer) and isinstance(external_match, Customer):
                if existing_customer.id != external_match.id:
                    outcome.row_errors.append(
                        f"Row {row_index}: Customer code '{customer_code_value}' does not match the external customer ID '{external_id_value}'."
                    )
                    continue
            elif external_match and existing_customer and external_match is not existing_customer:
                outcome.row_errors.append(
                    f"Row {row_index}: Customer code '{customer_code_value}' does not match the external customer ID '{external_id_value}'."
                )
                continue
            if isinstance(external_match, Customer):
                existing_customer = external_match

        updates = []
        if company_name_value:
            updates.append(("company_name", company_name_value, "Company Name"))
        if contact_person_value is not None:
            updates.append(("contact_person", contact_person_value, "Contact Person"))
        if phone_value is not None:
            updates.append(("phone", phone_value, "Phone"))
        if mobile_value is not None:
            updates.append(("mobile", mobile_value, "Mobile"))
        if email_value is not None:
            updates.append(("email", email_value, "Email"))
        if gst_value is not None:
            updates.append(("gst_no", gst_value, "GST Number"))
        if billing_line1 is not None:
            updates.append(("billing_address_line1", billing_line1, "Billing Address Line 1"))
        if billing_line2 is not None:
            updates.append(("billing_address_line2", billing_line2, "Billing Address Line 2"))
        if city_value is not None:
            updates.append(("city", city_value, "City"))
        if state_value is not None:
            updates.append(("state", state_value, "State"))
        if pincode_value is not None:
            updates.append(("pincode", pincode_value, "Pincode"))
        if country_value is not None:
            updates.append(("country", country_value, "Country"))
        if route_value is not None:
            updates.append(("route", route_value, "Route"))
        if sector_value is not None:
            updates.append(("sector", sector_value, "Sector"))
        if branch_value is not None:
            updates.append(("branch", branch_value, "Branch"))
        if notes_value is not None:
            updates.append(("notes", notes_value, "Notes"))
        if office_line1 is not None:
            updates.append(("office_address_line1", office_line1, "Office Address Line 1"))
        if office_line2 is not None:
            updates.append(("office_address_line2", office_line2, "Office Address Line 2"))
        if office_city_value is not None:
            updates.append(("office_city", office_city_value, "Office City"))
        if office_state_value is not None:
            updates.append(("office_state", office_state_value, "Office State"))
        if office_pincode_value is not None:
            updates.append(("office_pincode", office_pincode_value, "Office Pincode"))

        if existing_customer:
            customer = existing_customer
            if not customer.customer_code:
                outcome.row_errors.append(
                    f"Row {row_index}: Customer record is missing a customer code and cannot be updated."
                )
                continue
            code_key = customer.customer_code.lower()
            identifier = _customer_identifier(customer, fallback=f"existing:{id(customer)}")
            owner = processed_codes.get(code_key)
            if owner and owner != identifier:
                outcome.row_errors.append(
                    f"Row {row_index}: Customer code '{customer.customer_code}' is duplicated in the upload."
                )
                continue
            processed_codes[code_key] = identifier
            outcome.updated_count += 1

            changes = []
            for attr, value, label in updates:
                if value is None:
                    continue
                current = getattr(customer, attr)
                if (current or None) != value:
                    changes.append({
                        "field": label,
                        "from": current,
                        "to": value,
                    })
                    if apply_changes:
                        setattr(customer, attr, value)

            if apply_changes:
                if country_value is None and not customer.country:
                    customer.country = "India"
                if customer.office_country is None:
                    customer.office_country = "India"

            if changes and len(outcome.updated_items) < 20:
                outcome.updated_items.append(
                    {
                        "customer_code": customer.customer_code,
                        "company_name": customer.company_name,
                        "changes": changes,
                    }
                )

        else:
            if not company_name_value:
                outcome.row_errors.append(
                    f"Row {row_index}: Company name is required for new customers."
                )
                continue

            if customer_code_value:
                normalized_code = customer_code_value.lower()
                if normalized_code in existing_by_code:
                    outcome.row_errors.append(
                        f"Row {row_index}: Customer code '{customer_code_value}' already exists."
                    )
                    continue
                if normalized_code in processed_codes:
                    outcome.row_errors.append(
                        f"Row {row_index}: Customer code '{customer_code_value}' is duplicated in the upload."
                    )
                    continue
                customer_code = customer_code_value
            else:
                customer_code = generate_next_customer_code()
                while (
                    customer_code.lower() in existing_by_code
                    or customer_code.lower() in processed_codes
                    or customer_code.lower() in generated_codes
                ):
                    customer_code = generate_next_customer_code()
                generated_codes.add(customer_code.lower())

            identifier = f"new:{customer_code.lower()}"
            processed_codes[customer_code.lower()] = identifier

            if apply_changes:
                customer = Customer(
                    customer_code=customer_code,
                    company_name=company_name_value,
                )
                db.session.add(customer)
                existing_by_code[customer_code.lower()] = customer
            else:
                customer = Customer(
                    customer_code=customer_code,
                    company_name=company_name_value,
                )
                existing_by_code[customer_code.lower()] = customer

            outcome.created_count += 1
            if len(outcome.created_items) < 20:
                outcome.created_items.append(
                    {
                        "customer_code": customer_code,
                        "company_name": company_name_value,
                        "route": route_value,
                        "branch": branch_value,
                    }
                )

            if apply_changes:
                for attr, value, _ in updates:
                    if value is not None:
                        setattr(customer, attr, value)
                if not customer.country:
                    customer.country = "India"
                if not customer.office_country:
                    customer.office_country = "India"

        target_identifier = None
        if existing_customer:
            customer_ref = existing_customer
        else:
            customer_ref = customer
        if isinstance(customer_ref, Customer):
            if getattr(customer_ref, "customer_code", None):
                target_identifier = processed_codes.get(customer_ref.customer_code.lower())
            else:
                target_identifier = identifier or f"new:{id(customer_ref)}"

        if external_id_value:
            normalized_external = external_id_value.lower()
            conflict = existing_by_external.get(normalized_external)
            conflict_identifier = None
            if isinstance(conflict, Customer):
                conflict_identifier = _customer_identifier(conflict, fallback=f"existing:{id(conflict)}")
            elif isinstance(conflict, str):
                conflict_identifier = conflict
            if conflict_identifier and target_identifier and conflict_identifier != target_identifier:
                outcome.row_errors.append(
                    f"Row {row_index}: External customer ID '{external_id_value}' is already linked to another customer."
                )
                continue
            previous = processed_external_ids.get(normalized_external)
            if previous and target_identifier and previous != target_identifier:
                outcome.row_errors.append(
                    f"Row {row_index}: External customer ID '{external_id_value}' is duplicated in the upload."
                )
                continue
            processed_external_ids[normalized_external] = target_identifier or conflict_identifier or identifier or normalized_external
            if apply_changes and isinstance(customer_ref, Customer):
                customer_ref.external_customer_id = external_id_value
                existing_by_external[normalized_external] = customer_ref
            elif not apply_changes:
                existing_by_external[normalized_external] = target_identifier or identifier or normalized_external
        elif isinstance(customer_ref, Customer) and customer_ref.external_customer_id:
            processed_external_ids[customer_ref.external_customer_id.lower()] = target_identifier or identifier or f"existing:{id(customer_ref)}"

    if apply_changes and (outcome.created_count or outcome.updated_count):
        db.session.commit()

    return outcome


def _normalize_lookup_key(value):
    normalized = clean_str(stringify_cell(value))
    if isinstance(normalized, str) and normalized:
        return normalized.lower()
    return None


def process_lift_upload_file(file_path, *, apply_changes):
    header_cells, data_rows = _extract_tabular_upload_from_path(
        file_path, sheet_name=AMC_LIFT_TEMPLATE_SHEET_NAME
    )
    header_cells = header_cells or []
    header_map: Dict[str, int] = {}
    for idx, header in enumerate(header_cells or []):
        label = stringify_cell(header)
        if label:
            header_map[label] = idx

    outcome = UploadOutcome(
        header_map=header_map,
        created_count=0,
        updated_count=0,
        processed_rows=0,
    )

    customers = Customer.query.all()
    customer_by_code = {}
    customer_by_external = {}
    customer_by_name = {}
    for customer in customers:
        code_key = _normalize_lookup_key(customer.customer_code)
        if code_key:
            customer_by_code[code_key] = customer
        external_key = _normalize_lookup_key(customer.external_customer_id)
        if external_key:
            customer_by_external[external_key] = customer
        name_key = _normalize_lookup_key(customer.company_name)
        if name_key:
            customer_by_name[name_key] = customer

    service_routes = ServiceRoute.query.all()
    route_lookup: Dict[str, Any] = {}
    for route in service_routes:
        options = {route.state.lower()}
        display_name = clean_str(route.display_name)
        if display_name:
            options.add(display_name.lower())
        if route.branch:
            options.add(route.branch.lower())
            options.add(f"{route.state.lower()} · {route.branch.lower()}")
            options.add(f"{route.state.lower()}-{route.branch.lower()}")
        for option in options:
            route_lookup[option] = route

    existing_by_code: Dict[str, Optional[Lift]] = {}
    existing_by_external: Dict[str, Optional[Lift]] = {}
    processed_codes: set[str] = set()
    processed_external_ids: set[str] = set()
    generated_codes: set[str] = set()

    lift_brand_present = "Lift Brand" in header_map

    for row_index, row_values in enumerate(data_rows, start=2):
        if not row_values:
            continue
        row_data: Dict[str, Any] = {}
        for header, position in header_map.items():
            value = row_values[position] if position < len(row_values) else None
            row_data[header] = value

        key_fields = [
            row_data.get("Customer External ID"),
            row_data.get("Customer Code"),
            row_data.get("Customer Name"),
            row_data.get("External Lift ID"),
            row_data.get("Lift Code"),
            row_data.get("AMC Status"),
        ]
        if not any(clean_str(stringify_cell(value)) for value in key_fields):
            continue

        outcome.processed_rows += 1

        customer_external_id_value = clean_str(
            stringify_cell(row_data.get("Customer External ID"))
        )
        customer_code_value = clean_str(stringify_cell(row_data.get("Customer Code")))
        customer_name_value = clean_str(stringify_cell(row_data.get("Customer Name")))
        customer_external_id_key = (
            customer_external_id_value.lower() if customer_external_id_value else None
        )
        customer_code_key = customer_code_value.lower() if customer_code_value else None
        customer_name_key = customer_name_value.lower() if customer_name_value else None

        customer = None
        if customer_external_id_key and customer_external_id_key in customer_by_external:
            customer = customer_by_external[customer_external_id_key]
        if customer_code_key and customer_code_key in customer_by_code:
            customer = customer_by_code[customer_code_key]
        if (
            customer_external_id_key
            and customer_code_key
            and customer_external_id_key in customer_by_external
            and customer_code_key in customer_by_code
            and customer_by_external[customer_external_id_key].id
            != customer_by_code[customer_code_key].id
        ):
            outcome.row_errors.append(
                f"Row {row_index}: Customer code '{customer_code_value}' does not match external ID '{customer_external_id_value}'."
            )
            continue
        if not customer and customer_name_key and customer_name_key in customer_by_name:
            customer = customer_by_name[customer_name_key]
        if not customer:
            missing_reference = (
                customer_external_id_value
                or customer_code_value
                or customer_name_value
                or "—"
            )
            outcome.row_errors.append(
                f"Row {row_index}: Customer '{missing_reference}' was not found. Upload customers first or use customer external ID."
            )
            continue

        route_value_raw = clean_str(stringify_cell(row_data.get("Route")))
        route_value = None
        if route_value_raw:
            lookup_key = route_value_raw.lower()
            route = route_lookup.get(lookup_key)
            if not route:
                outcome.row_errors.append(
                    f"Row {row_index}: Route '{route_value_raw}' does not match an active service route."
                )
                continue
            route_value = route.state

        existing_lift = None
        provided_code = clean_str(stringify_cell(row_data.get("Lift Code")))
        provided_external = clean_str(stringify_cell(row_data.get("External Lift ID")))

        if provided_code:
            lookup_code = provided_code.lower()
            if lookup_code in existing_by_code:
                existing_lift = existing_by_code[lookup_code]
            else:
                existing_lift = (
                    Lift.query.filter(func.lower(Lift.lift_code) == lookup_code).first()
                )
                existing_by_code[lookup_code] = existing_lift
        if not existing_lift and provided_external:
            lookup_external = provided_external.lower()
            if lookup_external in existing_by_external:
                existing_lift = existing_by_external[lookup_external]
            else:
                existing_lift = (
                    Lift.query.filter(func.lower(Lift.external_lift_id) == lookup_external).first()
                )
                existing_by_external[lookup_external] = existing_lift

        if existing_lift and existing_lift.lift_code:
            existing_by_code[existing_lift.lift_code.lower()] = existing_lift
            if existing_lift.external_lift_id:
                existing_by_external[existing_lift.external_lift_id.lower()] = existing_lift

        if provided_external:
            normalized_external = provided_external.lower()
            if normalized_external in processed_external_ids:
                outcome.row_errors.append(
                    f"Row {row_index}: External lift ID '{provided_external}' is duplicated in the upload."
                )
                continue
            processed_external_ids.add(normalized_external)

        code_key = None
        if existing_lift and existing_lift.lift_code:
            code_key = existing_lift.lift_code.lower()
        elif provided_code:
            code_key = provided_code.lower()
        if code_key and code_key in processed_codes:
            display_code = provided_code or (existing_lift.lift_code if existing_lift else None)
            outcome.row_errors.append(
                f"Row {row_index}: Lift code '{display_code}' is duplicated in the upload."
            )
            continue

        amc_status_value, status_error = normalize_amc_status(row_data.get("AMC Status"))
        if status_error:
            outcome.row_errors.append(f"Row {row_index}: {status_error}")
            continue
        if not amc_status_value:
            outcome.row_errors.append(f"Row {row_index}: AMC status is required.")
            continue

        duration_key, duration_error = normalize_amc_duration(row_data.get("AMC Duration"))
        if duration_error:
            outcome.row_errors.append(f"Row {row_index}: {duration_error}")
            continue
        if not duration_key:
            outcome.row_errors.append(f"Row {row_index}: AMC duration is required.")
            continue

        amc_start, error = parse_date_field(
            row_data.get("AMC Start (YYYY-MM-DD)"),
            "AMC start date",
        )
        if error:
            outcome.row_errors.append(f"Row {row_index}: {error}")
            continue
        if not amc_start:
            outcome.row_errors.append(f"Row {row_index}: AMC start date is required.")
            continue

        amc_end, error = parse_date_field(
            row_data.get("AMC End (YYYY-MM-DD)"),
            "AMC end date",
        )
        if error:
            outcome.row_errors.append(f"Row {row_index}: {error}")
            continue
        if not amc_end:
            amc_end = calculate_amc_end_date(amc_start, duration_key)

        preferred_days_source = row_data.get("Preferred Service Days")
        preferred_days_display = (
            clean_str(stringify_cell(preferred_days_source))
            if preferred_days_source is not None
            else None
        )
        preferred_days = preferred_days_display

        preferred_date_source = row_data.get("Preferred Service Date")
        preferred_date = None
        if preferred_date_source not in (None, ""):
            preferred_date, error = parse_date_field(
                preferred_date_source,
                "Preferred service date",
            )
            if error:
                outcome.row_errors.append(f"Row {row_index}: {error}")
                continue

        preferred_time_source = row_data.get("Preferred Service Time")
        preferred_time = None
        if preferred_time_source not in (None, ""):
            preferred_time, error = parse_time_field(
                preferred_time_source,
                "Preferred service time",
            )
            if error:
                outcome.row_errors.append(f"Row {row_index}: {error}")
                continue

        next_service_due_source = row_data.get("Next Service Due")
        next_service_due = None
        if next_service_due_source not in (None, ""):
            next_service_due, error = parse_date_field(
                next_service_due_source,
                "Next service due",
            )
            if error:
                outcome.row_errors.append(f"Row {row_index}: {error}")
                continue

        capacity_persons, error = parse_int_field(
            row_data.get("Capacity (persons)"),
            "Capacity (persons)",
        )
        if error:
            outcome.row_errors.append(f"Row {row_index}: {error}")
            continue

        capacity_kg, error = parse_int_field(
            row_data.get("Capacity (kg)"),
            "Capacity (kg)",
        )
        if error:
            outcome.row_errors.append(f"Row {row_index}: {error}")
            continue

        speed_mps, error = parse_float_field(
            row_data.get("Speed (m/s)"),
            "Speed (m/s)",
        )
        if error:
            outcome.row_errors.append(f"Row {row_index}: {error}")
            continue

        lift_type_value = clean_str(stringify_cell(row_data.get("Lift Type")))
        lift_brand_value = clean_str(stringify_cell(row_data.get("Lift Brand")))
        site_address_line1 = clean_str(stringify_cell(row_data.get("Site Address Line 1")))
        site_address_line2 = clean_str(stringify_cell(row_data.get("Site Address Line 2")))
        building_villa_number = clean_str(stringify_cell(row_data.get("Building / Villa No.")))
        city_value = clean_str(stringify_cell(row_data.get("City")))
        state_value = clean_str(stringify_cell(row_data.get("State")))
        pincode_value = clean_str(stringify_cell(row_data.get("Pincode")))
        notes_value = clean_str(stringify_cell(row_data.get("Notes")))

        if existing_lift:
            lift = existing_lift
            if not lift.lift_code:
                outcome.row_errors.append(
                    f"Row {row_index}: Lift record is missing a lift code and cannot be updated."
                )
                continue
            outcome.updated_count += 1
        else:
            if provided_code:
                lift_code = provided_code
            else:
                lift_code = generate_next_lift_code()
                while (
                    lift_code.lower() in generated_codes
                    or lift_code.lower() in processed_codes
                ):
                    lift_code = generate_next_lift_code()
            if apply_changes:
                lift = Lift(lift_code=lift_code)
                db.session.add(lift)
            else:
                lift = Lift(lift_code=lift_code)
            generated_codes.add(lift.lift_code.lower())
            outcome.created_count += 1
            if len(outcome.created_items) < 20:
                outcome.created_items.append(
                    {
                        "lift_code": lift.lift_code,
                        "customer_code": customer.customer_code,
                        "amc_status": amc_status_value,
                        "amc_start": amc_start.isoformat(),
                        "amc_end": amc_end.isoformat() if amc_end else None,
                    }
                )

        processed_codes.add((lift.lift_code or "").lower())

        if len(outcome.updated_items) < 20 and existing_lift:
            outcome.updated_items.append(
                {
                    "lift_code": lift.lift_code,
                    "customer_code": customer.customer_code,
                    "amc_status": amc_status_value,
                    "amc_start": amc_start.isoformat(),
                    "amc_end": amc_end.isoformat() if amc_end else None,
                }
            )

        if apply_changes:
            if provided_external:
                lift.external_lift_id = provided_external
            lift.customer_code = customer.customer_code
            lift.customer = customer
            if building_villa_number is not None:
                lift.building_villa_number = building_villa_number
            if site_address_line1 is not None:
                lift.site_address_line1 = site_address_line1
            if site_address_line2 is not None:
                lift.site_address_line2 = site_address_line2
            if city_value is not None:
                lift.city = city_value
            elif not existing_lift and not lift.city and customer.city:
                lift.city = customer.city
            if state_value is not None:
                lift.state = state_value
            elif not existing_lift and not lift.state and customer.state:
                lift.state = customer.state
            if pincode_value is not None:
                lift.pincode = pincode_value
            elif not existing_lift and not lift.pincode and customer.pincode:
                lift.pincode = customer.pincode
            if route_value:
                lift.route = route_value
            elif not existing_lift and not lift.route and customer.route:
                lift.route = customer.route
            if lift_type_value:
                lift.lift_type = lift_type_value
            if lift_brand_present:
                lift.lift_brand = lift_brand_value
            if capacity_persons is not None:
                lift.capacity_persons = capacity_persons
            if capacity_kg is not None:
                lift.capacity_kg = capacity_kg
            if speed_mps is not None:
                lift.speed_mps = speed_mps
            lift.amc_status = amc_status_value
            lift.amc_start = amc_start
            lift.amc_duration_key = duration_key
            lift.amc_end = amc_end
            if preferred_days_display is not None:
                lift.preferred_service_days = preferred_days
            elif not existing_lift:
                lift.preferred_service_days = preferred_days
            if preferred_date_source not in (None, ""):
                lift.preferred_service_date = preferred_date
            if preferred_time_source not in (None, ""):
                lift.preferred_service_time = preferred_time
            if next_service_due is not None:
                lift.next_service_due = next_service_due
            if notes_value is not None:
                lift.notes = notes_value
            lift.last_updated_by = current_user.id
            lift.set_capacity_display()

        existing_by_code[lift.lift_code.lower()] = lift
        if provided_external:
            existing_by_external[provided_external.lower()] = lift
        elif lift.external_lift_id:
            existing_by_external[lift.external_lift_id.lower()] = lift

    if apply_changes and (outcome.created_count or outcome.updated_count):
        db.session.commit()

    return outcome


def _random_digits(length=10):
    return "".join(random.choice(string.digits) for _ in range(length))


def generate_random_phone(country_code="+91"):
    return f"{country_code}-{_random_digits(10)}"


def generate_random_email(domains=None):
    domain_pool = domains or [
        "example.com",
        "maildrop.cc",
        "inbound.test",
        "demo.local",
    ]
    local_part = "".join(random.choice(string.ascii_lowercase + string.digits) for _ in range(8))
    return f"{local_part}@{random.choice(domain_pool)}"


def generate_linked_task_id():
    """Generate a short, human-friendly identifier for linked tasks."""
    return f"TASK-{uuid.uuid4().hex[:6].upper()}"


def parse_optional_date(value):
    if not value:
        return None
    if isinstance(value, datetime.date):
        return value
    try:
        return datetime.datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def clean_str(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def parse_int_field(value, label):
    value = clean_str(value)
    if value is None:
        return None, None
    try:
        return int(value), None
    except (TypeError, ValueError):
        return None, f"{label} must be a whole number."


def parse_float_field(value, label):
    value = clean_str(value)
    if value is None:
        return None, None
    try:
        return float(value), None
    except (TypeError, ValueError):
        return None, f"{label} must be a number."


def parse_date_field(value, label):
    value = clean_str(value)
    if not value:
        return None, None
    if isinstance(value, str) and value.lower() == "none":
        return None, None
    parsed = parse_optional_date(value)
    if not parsed:
        return None, f"{label} must be in YYYY-MM-DD format."
    return parsed, None


def parse_time_field(value, label):
    value = clean_str(value)
    if not value:
        return None, None
    try:
        parsed = datetime.datetime.strptime(value, "%H:%M").time()
    except (TypeError, ValueError):
        return None, f"{label} must be in 24-hour HH:MM format."
    return parsed, None


PREFERRED_SERVICE_DATE_BASE_YEAR = 2000
PREFERRED_SERVICE_DATE_BASE_MONTH = 1


def parse_preferred_service_date(value):
    value = clean_str(value)
    if not value:
        return None, None
    if re.fullmatch(r"0?[1-9]|[12][0-9]|30", value):
        day = int(value)
        return (
            datetime.date(
                PREFERRED_SERVICE_DATE_BASE_YEAR,
                PREFERRED_SERVICE_DATE_BASE_MONTH,
                day,
            ),
            None,
        )
    return None, "Preferred service date must be a day between 01 and 30."


def parse_preferred_service_days(values):
    if not values:
        return [], None
    if isinstance(values, str):
        values = [values]
    selected = []
    for value in values:
        cleaned = clean_str(value)
        if not cleaned:
            continue
        lowered = cleaned.lower()
        if lowered not in SERVICE_PREFERRED_DAY_LABELS:
            return [], "Select valid preferred service days."
        if lowered == "any":
            selected = ["any"]
            break
        if lowered not in selected:
            selected.append(lowered)
    return selected, None


def is_monthly_preference_date(value):
    if isinstance(value, datetime.datetime):
        value = value.date()
    return (
        isinstance(value, datetime.date)
        and value.year == PREFERRED_SERVICE_DATE_BASE_YEAR
        and value.month == PREFERRED_SERVICE_DATE_BASE_MONTH
    )


def add_months(date_obj, months):
    if not isinstance(date_obj, datetime.date):
        return None
    month = date_obj.month - 1 + int(months)
    year = date_obj.year + month // 12
    month = month % 12 + 1
    day = min(date_obj.day, calendar.monthrange(year, month)[1])
    return datetime.date(year, month, day)


def calculate_amc_end_date(start_date, duration_key):
    if not isinstance(start_date, datetime.date):
        return None
    if not duration_key:
        return None
    duration_key = duration_key.strip().lower()
    months = AMC_DURATION_MONTHS.get(duration_key)
    if not months:
        return None
    target = add_months(start_date, months)
    if not target:
        return None
    return target - datetime.timedelta(days=1)


def normalize_amc_status(value):
    value = clean_str(value)
    if not value:
        return None, None
    lowered = value.lower()
    if lowered in AMC_STATUS_NORMALIZED:
        return AMC_STATUS_NORMALIZED[lowered], None
    valid_options = ", ".join(sorted(AMC_STATUS_OPTIONS))
    return None, f"AMC status must be one of: {valid_options}."


def normalize_amc_duration(value):
    value = clean_str(value)
    if not value:
        return None, None
    lowered = value.lower().strip()
    normalized = re.sub(r"[^a-z0-9]+", "_", lowered).strip("_")
    if normalized in AMC_DURATION_MONTHS:
        return normalized, None
    for key, label in AMC_DURATION_CHOICES:
        if not key or not label:
            continue
        if label.lower() == lowered:
            return key, None
    try:
        months = int(float(value))
    except (TypeError, ValueError):
        months = None
    if months:
        for key, duration_months in AMC_DURATION_MONTHS.items():
            if duration_months == months:
                return key, None
    valid_options = ", ".join(sorted(AMC_DURATION_LABELS.values()))
    return None, f"AMC duration must match one of: {valid_options}."


def parse_preferred_service_days_from_string(value):
    if not value:
        return [], None
    if isinstance(value, str):
        raw_values = [item.strip() for item in re.split(r"[,;/]", value) if item.strip()]
    elif isinstance(value, (list, tuple, set)):
        raw_values = [clean_str(item) for item in value if clean_str(item)]
    else:
        raw_values = [clean_str(value)] if clean_str(value) else []
    if not raw_values:
        return [], None
    parsed, error = parse_preferred_service_days(raw_values)
    if error:
        return [], error
    return parsed, None


def stringify_cell(value):
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        value = value.date()
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _customer_upload_row(customer):
    return [
        customer.external_customer_id or "",
        customer.customer_code or "",
        customer.company_name or "",
        customer.contact_person or "",
        customer.phone or "",
        customer.mobile or "",
        customer.email or "",
        customer.gst_no or "",
        customer.billing_address_line1 or "",
        customer.billing_address_line2 or "",
        customer.city or "",
        customer.state or "",
        customer.pincode or "",
        customer.country or "",
        customer.route or "",
        customer.sector or "",
        customer.branch or "",
        customer.notes or "",
        customer.office_address_line1 or "",
        customer.office_address_line2 or "",
        customer.office_city or "",
        customer.office_state or "",
        customer.office_pincode or "",
    ]


def _build_csv_output(headers, rows):
    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    output = BytesIO(csv_buffer.getvalue().encode("utf-8-sig"))
    output.seek(0)
    return output


def build_customer_upload_workbook():
    _ensure_openpyxl()
    workbook = Workbook()
    instructions_sheet = workbook.active
    instructions_sheet.title = "Instructions"
    instructions_sheet["A1"] = "Customer upload template"
    instructions_sheet["A1"].font = Font(bold=True, size=14)
    instructions_sheet["A3"] = "Fill the Customers sheet with one customer per row."
    instructions_sheet["A4"] = (
        "Provide an External Customer ID to map lifts during AMC uploads."
    )
    instructions_sheet["A5"] = (
        "Customer codes are auto-generated if left blank. Keep them unique if provided."
    )
    instructions_sheet["A6"] = (
        "Route should match an active service route. Branch must be one of "
        + ", ".join(SERVICE_BRANCH_OPTIONS)
        + "."
    )

    data_sheet = workbook.create_sheet(CUSTOMER_UPLOAD_TEMPLATE_SHEET_NAME)
    data_sheet.append(CUSTOMER_UPLOAD_TEMPLATE_HEADERS)
    example_row = ["" for _ in CUSTOMER_UPLOAD_TEMPLATE_HEADERS]
    data_sheet.append(example_row)

    for idx, header in enumerate(CUSTOMER_UPLOAD_TEMPLATE_HEADERS, start=1):
        cell = data_sheet.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        column_letter = data_sheet.cell(row=1, column=idx).column_letter
        data_sheet.column_dimensions[column_letter].width = max(18, len(header) + 2)

    return workbook


def build_customer_export_workbook(customers):
    _ensure_openpyxl()
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = CUSTOMER_UPLOAD_TEMPLATE_SHEET_NAME
    data_sheet.append(CUSTOMER_UPLOAD_TEMPLATE_HEADERS)

    for customer in customers:
        data_sheet.append(_customer_upload_row(customer))

    for idx, header in enumerate(CUSTOMER_UPLOAD_TEMPLATE_HEADERS, start=1):
        cell = data_sheet.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        column_letter = data_sheet.cell(row=1, column=idx).column_letter
        data_sheet.column_dimensions[column_letter].width = max(18, len(header) + 2)

    return workbook


def _sales_client_upload_row(client):
    owner_email = ""
    if client.owner and getattr(client.owner, "email", None):
        owner_email = client.owner.email or ""
    return [
        client.display_name or "",
        client.company_name or "",
        client.email or "",
        client.phone or "",
        client.tag or "",
        client.category or "",
        owner_email,
        client.lifecycle_stage or "",
        client.description or "",
    ]


def build_sales_client_upload_workbook():
    _ensure_openpyxl()
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions["A1"] = "Sales clients upload template"
    instructions["A1"].font = Font(bold=True, size=14)
    instructions["A3"] = "Add or update client records in bulk. One client per row."
    instructions["A4"] = "Provide the owner email to assign the client. Leave blank to assign yourself."
    if SALES_CLIENT_LIFECYCLE_STAGES:
        instructions["A5"] = (
            "Lifecycle Stage must be one of: "
            + ", ".join(SALES_CLIENT_LIFECYCLE_STAGES)
            + "."
        )

    data_sheet = workbook.create_sheet(SALES_CLIENT_TEMPLATE_SHEET_NAME)
    data_sheet.append(SALES_CLIENT_UPLOAD_HEADERS)
    data_sheet.append(["" for _ in SALES_CLIENT_UPLOAD_HEADERS])

    for idx, header in enumerate(SALES_CLIENT_UPLOAD_HEADERS, start=1):
        cell = data_sheet.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        column_letter = data_sheet.cell(row=1, column=idx).column_letter
        data_sheet.column_dimensions[column_letter].width = max(18, len(header) + 2)

    return workbook


def build_sales_client_export_workbook(clients):
    _ensure_openpyxl()
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = SALES_CLIENT_TEMPLATE_SHEET_NAME
    data_sheet.append(SALES_CLIENT_UPLOAD_HEADERS)

    for client in clients:
        data_sheet.append(_sales_client_upload_row(client))

    for idx, header in enumerate(SALES_CLIENT_UPLOAD_HEADERS, start=1):
        cell = data_sheet.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        column_letter = data_sheet.cell(row=1, column=idx).column_letter
        data_sheet.column_dimensions[column_letter].width = max(18, len(header) + 2)

    return workbook


def _sales_opportunity_upload_row(opportunity):
    owner_email = ""
    if opportunity.owner and getattr(opportunity.owner, "email", None):
        owner_email = opportunity.owner.email or ""
    client_name = opportunity.client.display_name if opportunity.client else ""
    expected_close = (
        opportunity.expected_close_date.isoformat()
        if opportunity.expected_close_date
        else ""
    )
    return [
        opportunity.title or "",
        opportunity.pipeline or "",
        opportunity.stage or "",
        opportunity.status or "",
        (opportunity.temperature or "") if opportunity.temperature else "",
        opportunity.amount if opportunity.amount is not None else "",
        opportunity.currency or "₹",
        expected_close,
        opportunity.probability if opportunity.probability is not None else "",
        owner_email,
        client_name,
        opportunity.related_project or "",
        opportunity.description or "",
    ]


def build_sales_opportunity_upload_workbook(pipeline_key):
    _ensure_openpyxl()
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions["A1"] = "Sales opportunities upload template"
    instructions["A1"].font = Font(bold=True, size=14)
    instructions["A3"] = "Populate the Opportunities sheet with one deal per row."
    instructions["A4"] = (
        "Pipeline accepts the keys: "
        + ", ".join(f"{key} ({cfg['label']})" for key, cfg in SALES_PIPELINES.items())
        + "."
    )
    instructions["A5"] = (
        "Leave Pipeline blank to default to the selected pipeline in the app."
    )

    row_index = 7
    for key, cfg in SALES_PIPELINES.items():
        instructions[f"A{row_index}"] = f"{cfg['label']} pipeline stages ({key}):"
        instructions[f"A{row_index + 1}"] = ", ".join(cfg.get("stages", []))
        row_index += 2

    data_sheet = workbook.create_sheet(SALES_OPPORTUNITY_TEMPLATE_SHEET_NAME)
    data_sheet.append(SALES_OPPORTUNITY_UPLOAD_HEADERS)
    config = get_pipeline_config(pipeline_key)
    sample_stage = config["stages"][0] if config.get("stages") else ""
    data_sheet.append(
        [
            "Sample Opportunity",
            pipeline_key,
            sample_stage,
            "Open",
            "warm",
            "500000",
            "₹",
            datetime.date.today().isoformat(),
            "40",
            "sales@example.com",
            "Sample Client",
            "Reference Project",
            "Notes about the deal",
        ]
    )

    for idx, header in enumerate(SALES_OPPORTUNITY_UPLOAD_HEADERS, start=1):
        cell = data_sheet.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        column_letter = data_sheet.cell(row=1, column=idx).column_letter
        data_sheet.column_dimensions[column_letter].width = max(18, len(header) + 2)

    return workbook


def build_sales_opportunity_export_workbook(opportunities):
    _ensure_openpyxl()
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = SALES_OPPORTUNITY_TEMPLATE_SHEET_NAME
    data_sheet.append(SALES_OPPORTUNITY_UPLOAD_HEADERS)

    for opportunity in opportunities:
        data_sheet.append(_sales_opportunity_upload_row(opportunity))

    for idx, header in enumerate(SALES_OPPORTUNITY_UPLOAD_HEADERS, start=1):
        cell = data_sheet.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        column_letter = data_sheet.cell(row=1, column=idx).column_letter
        data_sheet.column_dimensions[column_letter].width = max(18, len(header) + 2)

    return workbook


def _customer_query_for_export(search_query):
    query = Customer.query
    if search_query:
        like = f"%{search_query.lower()}%"
        query = query.filter(
            or_(
                func.lower(Customer.customer_code).like(like),
                func.lower(Customer.company_name).like(like),
                func.lower(Customer.contact_person).like(like),
                func.lower(Customer.city).like(like),
                func.lower(Customer.state).like(like),
                func.lower(Customer.route).like(like),
                func.lower(Customer.branch).like(like),
                func.lower(Customer.sector).like(like),
                func.lower(Customer.notes).like(like),
            )
        )
    return query.order_by(func.lower(Customer.company_name))


def build_amc_lift_upload_workbook():
    _ensure_openpyxl()
    workbook = Workbook()
    instructions_sheet = workbook.active
    instructions_sheet.title = "Instructions"
    instructions_sheet["A1"] = "AMC lift upload template"
    instructions_sheet["A1"].font = Font(bold=True, size=14)
    instructions_sheet["A3"] = "Fill the AMC Lifts sheet with one lift per row."
    instructions_sheet["A4"] = (
        "Provide the customer external ID from the customer upload or the customer code/name."
    )
    instructions_sheet["A5"] = "Valid AMC statuses: " + ", ".join(AMC_STATUS_OPTIONS)
    instructions_sheet["A6"] = "Valid AMC durations: " + ", ".join(AMC_DURATION_LABELS.values())
    instructions_sheet["A7"] = "Dates must be in YYYY-MM-DD format. Preferred service date accepts 01-30."
    instructions_sheet["A8"] = "Preferred service days can be comma separated (e.g. Monday,Wednesday)."

    data_sheet = workbook.create_sheet(AMC_LIFT_TEMPLATE_SHEET_NAME)
    data_sheet.append(AMC_LIFT_TEMPLATE_HEADERS)
    example_row = ["" for _ in AMC_LIFT_TEMPLATE_HEADERS]
    data_sheet.append(example_row)

    for idx, header in enumerate(AMC_LIFT_TEMPLATE_HEADERS, start=1):
        cell = data_sheet.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        column_letter = data_sheet.cell(row=1, column=idx).column_letter
        data_sheet.column_dimensions[column_letter].width = max(18, len(header) + 2)

    return workbook


def format_preferred_service_date(value):
    if not value:
        return "—"
    if isinstance(value, datetime.datetime):
        value = value.date()
    if is_monthly_preference_date(value):
        return f"Day {value.day:02d}"
    return value.strftime("%d %b %Y")


def preferred_service_date_matches(preferred_value, visit_date):
    if not preferred_value or not visit_date:
        return False
    if isinstance(preferred_value, datetime.datetime):
        preferred_value = preferred_value.date()
    if isinstance(visit_date, datetime.datetime):
        visit_date = visit_date.date()
    if not isinstance(preferred_value, datetime.date) or not isinstance(visit_date, datetime.date):
        return False
    if is_monthly_preference_date(preferred_value):
        return preferred_value.day == visit_date.day
    return preferred_value == visit_date


def validate_branch(value, *, label="Branch", required=False):
    branch_value = clean_str(value)
    if not branch_value:
        if required:
            return None, f"{label} is required."
        return None, None

    lowered = branch_value.lower()
    if lowered not in SERVICE_BRANCH_OPTION_SET:
        allowed = ", ".join(SERVICE_BRANCH_OPTIONS)
        return None, f"{label} must be one of {allowed}."

    for option in SERVICE_BRANCH_OPTIONS:
        if option.lower() == lowered:
            return option, None

    return branch_value, None


def _next_sequential_code(model, column_attr, *, prefix, width):
    column = getattr(model, column_attr)
    max_value = 0
    for (code,) in db.session.query(column).filter(column.isnot(None)).all():
        match = re.search(r"(\d+)$", code or "")
        if not match:
            continue
        max_value = max(max_value, int(match.group(1)))
    next_value = max_value + 1
    return f"{prefix}{next_value:0{width}d}"


def generate_next_customer_code():
    return _next_sequential_code(Customer, "customer_code", prefix="CUS", width=4)


def generate_next_lift_code():
    return _next_sequential_code(Lift, "lift_code", prefix="LFT", width=4)


def get_service_contract_by_id(contract_id):
    if not contract_id:
        return None
    for contract in SERVICE_CONTRACTS:
        contract_code = str(contract.get("id") or "").strip()
        if contract_code and contract_code.lower() == str(contract_id).strip().lower():
            return contract
    return None


# ---------------------- QC Profile choices (visible in UI) ----------------------
STAGES = [
    "Template QC", "Stage 1", "Stage 2", "Stage 3",
    "Completion", "Completion QC", "Structure", "Cladding", "Service", "Repair", "Material"
]
LIFT_TYPE_OPTIONS = ["Hydraulic", "MR", "MRL", "Goods", "Dumbwaiter"]
LIFT_TYPES = list(LIFT_TYPE_OPTIONS)

LIFT_CAPACITY_PERSON_OPTIONS = [
    ("", "Select capacity"),
    ("0", "0"),
    ("4", "4 Pass"),
    ("6", "6 Pass"),
    ("8", "8 Pass"),
    ("10", "10 Pass"),
    ("13", "13 Pass"),
    ("15", "15 Pass"),
    ("20", "20 Pass"),
    ("25", "25 Pass"),
]

MACHINE_TYPE_OPTIONS = ["Geared", "Gearless", "Hydraulic", "Drum", "Stiltz"]
DOOR_TYPE_OPTIONS = ["ATO-LH", "ATO-RH", "ACO", "Swing", "Collapsible", "Gate", "IMP"]
DOOR_FINISH_OPTIONS = ["SS H/L", "SS Mirror", "MS"]
POWER_SUPPLY_OPTIONS = ["1 Phase", "3 Phase"]
AMC_STATUS_OPTIONS = ["Active", "Expired", "Renewal Pending", "Call Basis"]
AMC_STATUS_NORMALIZED = {option.lower(): option for option in AMC_STATUS_OPTIONS}

SERVICE_VISIT_STATUS_OPTIONS = [
    ("scheduled", "Scheduled"),
    ("completed", "Completed"),
    ("overdue", "Overdue"),
]
SERVICE_VISIT_STATUS_LABELS = {
    value: label for value, label in SERVICE_VISIT_STATUS_OPTIONS
}
LIFT_STATUS_OPTIONS = ["On", "Off", "Protected", "Decommissioned"]
SERVICE_BRANCH_OPTIONS = ["Goa", "Mumbai"]
SERVICE_BRANCH_OPTION_SET = {option.lower() for option in SERVICE_BRANCH_OPTIONS}

SERVICE_PREFERRED_DAY_OPTIONS = [
    ("", "No preference"),
    ("any", "Any day"),
    ("monday", "Monday"),
    ("tuesday", "Tuesday"),
    ("wednesday", "Wednesday"),
    ("thursday", "Thursday"),
    ("friday", "Friday"),
    ("saturday", "Saturday"),
    ("sunday", "Sunday"),
]
SERVICE_PREFERRED_DAY_LABELS = {
    key: label for key, label in SERVICE_PREFERRED_DAY_OPTIONS if key
}

AMC_DURATION_CHOICES = [
    ("", "Select AMC duration"),
    ("1_year", "1 year"),
    ("2_years", "2 years"),
    ("3_years", "3 years"),
    ("4_years", "4 years"),
    ("5_years", "5 years"),
    ("1_year_bimonthly", "1 year bimonthly"),
    ("1_year_quarterly", "1 year quarterly"),
    ("6_months", "6 months"),
]
AMC_DURATION_LABELS = {key: label for key, label in AMC_DURATION_CHOICES if key}
AMC_DURATION_MONTHS = {
    "1_year": 12,
    "2_years": 24,
    "3_years": 36,
    "4_years": 48,
    "5_years": 60,
    "1_year_bimonthly": 12,
    "1_year_quarterly": 12,
    "6_months": 6,
}
AMC_LIFT_TEMPLATE_SHEET_NAME = "AMC Lifts"
AMC_LIFT_TEMPLATE_HEADERS = [
    "Customer External ID",
    "Lift Code",
    "External Lift ID",
    "Customer Code",
    "Customer Name",
    "Building / Villa No.",
    "Site Address Line 1",
    "Site Address Line 2",
    "City",
    "State",
    "Pincode",
    "Route",
    "Lift Type",
    "Lift Brand",
    "Capacity (persons)",
    "Capacity (kg)",
    "Speed (m/s)",
    "AMC Status",
    "AMC Start (YYYY-MM-DD)",
    "AMC Duration",
    "AMC End (YYYY-MM-DD)",
    "Preferred Service Days",
    "Preferred Service Date (DD)",
    "Preferred Service Time (HH:MM)",
    "Next Service Due (YYYY-MM-DD)",
    "Notes",
]

CUSTOMER_UPLOAD_TEMPLATE_SHEET_NAME = "Customers"
CUSTOMER_UPLOAD_TEMPLATE_FIELDS = [
    ("External Customer ID", "external_customer_id"),
    ("Customer Code", "customer_code"),
    ("Company Name", "company_name"),
    ("Contact Person", "contact_person"),
    ("Phone", "phone"),
    ("Mobile", "mobile"),
    ("Email", "email"),
    ("GST No", "gst_no"),
    ("Billing Address Line 1", "billing_address_line1"),
    ("Billing Address Line 2", "billing_address_line2"),
    ("City", "city"),
    ("State", "state"),
    ("Pincode", "pincode"),
    ("Country", "country"),
    ("Route", "route"),
    ("Sector", "sector"),
    ("Branch", "branch"),
    ("Notes", "notes"),
    ("Office Address Line 1", "office_address_line1"),
    ("Office Address Line 2", "office_address_line2"),
    ("Office City", "office_city"),
    ("Office State", "office_state"),
    ("Office Pincode", "office_pincode"),
]
CUSTOMER_UPLOAD_TEMPLATE_HEADERS = [label for label, _ in CUSTOMER_UPLOAD_TEMPLATE_FIELDS]

DROPDOWN_FIELD_DEFINITIONS = {
    "lift_type": {
        "label": "Lift Type",
        "value_editable": False,
        "default_options": [
            {"value": option, "label": option}
            for option in ["Hydraulic", "MR", "MRL", "Goods", "Dumbwaiter", "Passenger"]
        ],
    },
    "door_type": {
        "label": "Door Type",
        "value_editable": False,
        "default_options": [
            {"value": option, "label": option}
            for option in [
                "ATO-LH",
                "ATO-RH",
                "ACO",
                "Swing",
                "Collapsible",
                "Gate",
                "IMP",
                "Manual",
            ]
        ],
    },
    "door_finish": {
        "label": "Door Finish",
        "value_editable": False,
        "default_options": [
            {"value": option, "label": option}
            for option in ["SS H/L", "SS Mirror", "MS", "Powder Coated"]
        ],
    },
    "power_supply": {
        "label": "Power Supply",
        "value_editable": False,
        "default_options": [
            {"value": option, "label": option}
            for option in ["1 Phase", "3 Phase"]
        ],
    },
    "machine_type": {
        "label": "Machine Type",
        "value_editable": False,
        "default_options": [
            {"value": option, "label": option}
            for option in ["Geared", "Gearless", "Hydraulic", "Drum", "Stiltz"]
        ],
    },
    "passenger_capacity": {
        "label": "Passenger Capacity",
        "value_editable": True,
        "default_options": [
            {"value": value, "label": label}
            for value, label in [
                ("0", "0"),
                ("4", "4 Pass"),
                ("6", "6 Pass"),
                ("8", "8 Pass"),
                ("10", "10 Pass"),
                ("13", "13 Pass"),
                ("15", "15 Pass"),
                ("20", "20 Pass"),
                ("25", "25 Pass"),
            ]
        ],
    },
    "load_capacity": {
        "label": "Load Capacity (Kg)",
        "value_editable": True,
        "default_options": [
            {"value": value, "label": f"{value} Kg"}
            for value in ["170", "272", "408", "544", "680", "800", "1000", "1500"]
        ],
    },
}


LEGACY_DEMO_CUSTOMERS = {
    "CUS0001": "St. Marys Convent",
    "CUS0002": "Kilowott Agency Pvt. Ltd.",
    "CUS0003": "Satguru Apartments Society",
    "CUS0004": "Jonathan Fernandes",
    "CUS0005": "Mr. Anirudh",
}

LEGACY_DEMO_LIFTS = {
    "G192": ("CUS0001", LEGACY_DEMO_CUSTOMERS["CUS0001"]),
    "G208": ("CUS0005", LEGACY_DEMO_CUSTOMERS["CUS0005"]),
    "G167": ("CUS0004", LEGACY_DEMO_CUSTOMERS["CUS0004"]),
    "G084": ("CUS0002", LEGACY_DEMO_CUSTOMERS["CUS0002"]),
    "G044": ("CUS0003", LEGACY_DEMO_CUSTOMERS["CUS0003"]),
}


def purge_legacy_demo_records():
    """Remove legacy demo customers and lifts shipped in earlier seeds."""

    removed_lifts = []
    removed_customers = []

    if "Lift" in globals():
        legacy_lift_codes = list(LEGACY_DEMO_LIFTS.keys())
        if legacy_lift_codes:
            lifts = (
                Lift.query.filter(Lift.lift_code.in_(legacy_lift_codes)).all()
            )
            for lift in lifts:
                expected_customer_code, expected_customer_name = LEGACY_DEMO_LIFTS.get(
                    lift.lift_code, (None, None)
                )
                customer_code_matches = (
                    (lift.customer_code or "").strip().upper() == expected_customer_code
                )
                customer_name_matches = True
                related_customer = getattr(lift, "customer", None)
                if related_customer and expected_customer_name:
                    customer_name_matches = (
                        (related_customer.company_name or "").strip()
                        == expected_customer_name
                    )
                if customer_code_matches and customer_name_matches:
                    delete_lift_record(lift)
                    removed_lifts.append(lift.lift_code)

    if removed_lifts:
        db.session.flush()

    if "Customer" in globals():
        legacy_customer_codes = list(LEGACY_DEMO_CUSTOMERS.keys())
        if legacy_customer_codes:
            customers = (
                Customer.query.filter(Customer.customer_code.in_(legacy_customer_codes)).all()
            )
            for customer in customers:
                expected_name = LEGACY_DEMO_CUSTOMERS.get(customer.customer_code)
                if (customer.company_name or "").strip() == expected_name:
                    db.session.delete(customer)
                    removed_customers.append(customer.customer_code)

    if removed_lifts:
        print(
            "♻️ Removed legacy demo lifts: " + ", ".join(sorted(removed_lifts))
        )
    if removed_customers:
        print(
            "♻️ Removed legacy demo customers: "
            + ", ".join(sorted(removed_customers))
        )


def ensure_dropdown_options_seed():
    for field_key, definition in DROPDOWN_FIELD_DEFINITIONS.items():
        if DropdownOption.query.filter_by(field_key=field_key).count() > 0:
            continue
        default_options = definition.get("default_options") or []
        for index, option in enumerate(default_options):
            db.session.add(
                DropdownOption(
                    field_key=field_key,
                    value=option.get("value"),
                    label=option.get("label"),
                    order_index=index,
                )
            )
    db.session.commit()


def get_dropdown_choices(field_key):
    definition = DROPDOWN_FIELD_DEFINITIONS.get(field_key)
    if not definition:
        return []
    options = (
        DropdownOption.query.filter_by(field_key=field_key)
        .order_by(DropdownOption.order_index.asc(), DropdownOption.id.asc())
        .all()
    )
    if not options:
        return [option.copy() for option in definition.get("default_options", [])]
    return [option.as_choice() for option in options]


def get_dropdown_options_map():
    return {
        field_key: get_dropdown_choices(field_key)
        for field_key in DROPDOWN_FIELD_DEFINITIONS.keys()
    }
DEFAULT_TASK_FORM_NAME = "Generic Task Tracker"
TASK_MILESTONES = [
    "Order Milestone",
    "Design Milestone",
    "Production Milestone",
    "Installation Stage 1",
    "Installation Stage 2",
    "Commissioning",
]

PROJECT_PRIORITIES = ["Immediate", "Urgent", "Normal"]
PROJECT_OPENING_TYPES = ["Single", "Adjacent", "Opposite Opening"]
PROJECT_LOCATIONS = ["Internal", "External"]
PROJECT_STRUCTURE_TYPES = ["NA", "RCC", "MS", "GI"]
PROJECT_CLADDING_TYPES = ["ACP", "Glass", "Hybrid", "Clients Scope", "Other"]
PROJECT_CABIN_FINISHES = ["SS", "MS", "Glass", "SS+Glass", "Designer", "Cage", "Half Cabin", "Other"]
PROJECT_DOOR_OPERATION_TYPES = ["Manual", "Auto"]
PROJECT_DOOR_FINISHES = ["SS", "MS", "Collapsible", "BiParting", "Gate"]
DEPARTMENT_BRANCHES = ["Goa", "Maharashtra"]


WORKSPACE_MODULES = [
    {
        "key": "customer_support",
        "label": "Customer Support",
        "description": "Support overview, ticket triage and linked tasks.",
        "visibility_label": "Show Customer Support workspace",
        "assignment_label": "Allow Customer Support ownership",
    },
    {
        "key": "service",
        "label": "Service",
        "description": "Post-installation service delivery, tasks and maintenance orchestration.",
        "visibility_label": "Show Service workspace",
        "assignment_label": "Allow Service task assignment",
    },
    {
        "key": "sales",
        "label": "Sales",
        "description": "Pipeline, clients and revenue forecasting dashboards.",
        "visibility_label": "Show Sales workspace",
        "assignment_label": "Allow Sales ownership",
    },
    {
        "key": "operations",
        "label": "Operations",
        "description": "Project delivery tools inside the New Installation area.",
        "visibility_label": "Show Operations workspace",
        "assignment_label": "Allow Operations task assignment",
    },
    {
        "key": "srt",
        "label": "SRT",
        "description": "Service Response Team dashboards and templates.",
        "visibility_label": "Show SRT workspace",
        "assignment_label": "Allow SRT ownership",
    },
    {
        "key": "qc",
        "label": "Quality Control",
        "description": "QC boards, task tracking and submission reviews.",
        "visibility_label": "Show QC workspace",
        "assignment_label": "Allow QC task assignment",
    },
]

WORKSPACE_MODULE_MAP = {module["key"]: module for module in WORKSPACE_MODULES}


def slugify(value):
    value = (value or "").lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or _random_digits(6)


SRT_SAMPLE_TASKS = []


SRT_TASK_ACTIVITY = {}


SERVICE_TASKS = []


SERVICE_COMPLAINTS = []


SERVICE_CONTRACTS = []


SERVICE_PARTS_LEDGER = {
    "stock_alerts": [],
    "consumption": [],
    "returns": [],
}


SERVICE_AUTOMATIONS = {
    "flows": [
        {
            "from": "Sarv",
            "to": "Customer Support",
            "summary": "Incoming calls auto-create tickets with recording",
        },
        {
            "from": "Customer Support",
            "to": "Service (AMC)",
            "summary": "Marking as AMC support converts to AMC service task",
        },
        {
            "from": "Customer Support",
            "to": "Service (Non-AMC)",
            "summary": "Creates chargeable task + optional Sales alert",
        },
        {
            "from": "Service Task",
            "to": "Inventory",
            "summary": "Parts usage deducts stock / raises PR",
        },
        {
            "from": "Service Task",
            "to": "Overview",
            "summary": "Live metrics refresh once task updates",
        },
    ],
    "roles": {
        "Admin": ["Full access", "Configure automations"],
        "Service Manager": ["Planner, contracts, reports", "Approve conversions"],
        "Technician": ["My tasks", "Upload media", "View assigned lifts"],
    },
    "config": [
        "Complaint categories & priorities",
        "SLA presets and escalation timers",
        "PM checklist templates per lift type",
        "Task statuses and close codes",
        "Dashboard widget toggles + alert thresholds",
    ],
}


DEFAULT_LIFT_INSIGHT = {
    "lifetime_value": {
        "total_breakdowns_this_year": 0,
        "breakdowns_completed_this_year": 0,
        "breakdowns_pending_this_year": 0,
        "total_amc_value": 0,
        "total_repair_revenue": 0,
        "repair_revenue_this_year": 0,
        "total_cost": 0,
        "total_cost_this_year": 0,
        "total_revenue_till_date": 0,
        "net_lifetime_profitability": 0,
    },
    "amc": {
        "status": "Not captured",
        "type": "Not captured",
        "start": None,
        "end": None,
        "contract_value": 0,
        "payment_terms": "—",
        "services_per_year": 0,
        "pending_services_count": 0,
        "service_owner": "—",
        "service_contact": "—",
        "renewal_history": [],
        "attachments": [],
        "service_schedule": [],
    },
    "machine": {
        "make": "—",
        "model": "—",
        "serial": "—",
    },
    "drive_type": "—",
    "controller_type": "—",
    "door_configuration": None,
    "commissioned_date": None,
    "floors_served": None,
    "breakdowns": [],
    "breakdown_summary": [],
    "uploads": {
        "documents": [],
        "media": [],
        "other": [],
    },
    "timeline": [],
    "site_name": None,
}


LIFT_INSIGHT_LIBRARY = {}


CUSTOMER_SUPPORT_CATEGORIES = [
    {
        "id": "sales-ni",
        "label": "Sales – NI",
        "description": "Pre-sales and new installation related follow-ups.",
        "default_first_response_hours": 4,
        "default_resolution_hours": 24,
    },
    {
        "id": "sales-amc",
        "label": "Sales – AMC",
        "description": "Annual maintenance contract conversations and renewals.",
        "default_first_response_hours": 6,
        "default_resolution_hours": 48,
    },
    {
        "id": "support-amc",
        "label": "Support – AMC",
        "description": "Breakdown, callbacks and reactive maintenance tickets.",
        "default_first_response_hours": 2,
        "default_resolution_hours": 18,
    },
    {
        "id": "other-dept",
        "label": "Other Department",
        "description": "Requests that need to be routed to internal departments.",
        "default_first_response_hours": 8,
        "default_resolution_hours": 72,
    },
    {
        "id": "other-query",
        "label": "Other Query",
        "description": "General enquiries that do not fit the above categories.",
        "default_first_response_hours": 12,
        "default_resolution_hours": 120,
    },
]

CUSTOMER_SUPPORT_CHANNELS = [
    {"id": "phone", "label": "Phone", "icon": "📞"},
    {"id": "email", "label": "Email", "icon": "✉️"},
    {"id": "web", "label": "Web", "icon": "🌐"},
    {"id": "walk-in", "label": "Walk-in", "icon": "🚪"},
    {"id": "whatsapp", "label": "WhatsApp", "icon": "💬"},
]

CUSTOMER_SUPPORT_AMC_SITES = []

CUSTOMER_SUPPORT_SLA_PRESETS = [
    {
        "id": "standard",
        "label": "Standard",
        "first_response_hours": 6,
        "resolution_hours": 48,
    },
    {
        "id": "priority",
        "label": "Priority",
        "first_response_hours": 2,
        "resolution_hours": 18,
    },
    {
        "id": "critical",
        "label": "Critical",
        "first_response_hours": 1,
        "resolution_hours": 8,
    },
]


CUSTOMER_SUPPORT_DEFAULT_TEAM = {
    "Service Desk",
    "Field Team 3",
    "Customer Care Desk",
}

CUSTOMER_SUPPORT_TICKETS = []

CUSTOMER_SUPPORT_CALL_LOGS = []


def _get_srt_task(task_id):
    return next((task for task in SRT_SAMPLE_TASKS if task["id"] == task_id), None)


def _log_srt_activity(task_id, **payload):
    if not task_id:
        return

    event = apply_actor_context(payload)
    event.setdefault("timestamp", datetime.datetime.utcnow())
    SRT_TASK_ACTIVITY.setdefault(task_id, []).append(event)


def _default_srt_item():
    return {
        "label": "New Checklist Item",
        "type": "select",
        "options": ["OK", "Not OK", "Need Client Input"],
        "required": True,
        "allow_photo": True,
        "allow_remark": True,
        "photo_required_if_ng": False,
        "display_image": "",
    }


def _default_srt_schema():
    return [
        {
            "section": "General",
            "display_image": "",
            "items": [_default_srt_item()],
        }
    ]


def _get_customer_support_ticket(ticket_id):
    if not ticket_id:
        return None

    return next((ticket for ticket in CUSTOMER_SUPPORT_TICKETS if ticket["id"] == ticket_id), None)


def _ticket_has_open_linked_tasks(ticket):
    if not ticket:
        return False

    closing_statuses = {"closed", "resolved", "completed", "done", "cancelled"}
    for task in ticket.get("linked_tasks", []) or []:
        status = (task.get("status") or "").strip().lower()
        if status and status in closing_statuses:
            continue
        if status:
            return True
        if not status:
            return True
    return False


def _resolve_customer_support_channel_label(channel_value):
    if not channel_value:
        return None

    lowered = channel_value.lower()
    for channel in CUSTOMER_SUPPORT_CHANNELS:
        channel_id = (channel.get("id") or "").lower()
        channel_label = (channel.get("label") or "").lower()
        if lowered in {channel_id, channel_label}:
            return channel.get("label")
    return None


def _format_customer_support_amc_site_from_lift(lift):
    if not lift:
        return None

    location_parts = [
        lift.site_address_line1,
        lift.site_address_line2,
        lift.city,
        lift.state,
    ]
    location = ", ".join(part for part in location_parts if part)

    base_label = lift.lift_code or f"Lift {lift.id}"
    label = f"{base_label} · {location}" if location else base_label

    customer_name = None
    customer_code = None
    if lift.customer:
        customer_name = lift.customer.company_name or None
        customer_code = lift.customer.customer_code or None
    else:
        customer_code = lift.customer_code or None

    return {
        "id": str(lift.id),
        "label": label,
        "client": customer_name or customer_code,
        "customer_name": customer_name,
        "customer_code": customer_code,
        "amc_status": lift.amc_status or None,
    }


def _customer_support_amc_site_options():
    lifts = (
        Lift.query.options(joinedload(Lift.customer))
        .order_by(func.lower(Lift.lift_code))
        .all()
    )

    options = []
    for lift in lifts:
        formatted = _format_customer_support_amc_site_from_lift(lift)
        if formatted:
            options.append(formatted)

    return sorted(options, key=lambda item: (item.get("label") or "").lower())


def _resolve_customer_support_amc_site(site_id):
    if not site_id:
        return None

    try:
        lift_id = int(site_id)
    except (TypeError, ValueError):
        lift_id = None

    if lift_id is not None:
        lift = db.session.get(Lift, lift_id)
        formatted = _format_customer_support_amc_site_from_lift(lift)
        if formatted:
            return formatted

    lowered = site_id.lower()
    for site in CUSTOMER_SUPPORT_AMC_SITES:
        if lowered == (site.get("id") or "").lower():
            return site
    return None


def _generate_customer_support_ticket_id():
    existing_numbers = []
    for ticket in CUSTOMER_SUPPORT_TICKETS:
        match = re.match(r"CS-(\d+)$", str(ticket.get("id") or ""))
        if match:
            try:
                existing_numbers.append(int(match.group(1)))
            except ValueError:
                continue

    next_number = (max(existing_numbers) + 1) if existing_numbers else 1001
    while any(ticket.get("id") == f"CS-{next_number}" for ticket in CUSTOMER_SUPPORT_TICKETS):
        next_number += 1
    return f"CS-{next_number}"


def _customer_support_summary():
    summary = {
        "Open": 0,
        "In Progress": 0,
        "Resolved": 0,
        "Closed": 0,
    }

    for ticket in CUSTOMER_SUPPORT_TICKETS:
        summary.setdefault(ticket["status"], 0)
        summary[ticket["status"]] += 1

    total = sum(summary.values())
    return {
        "counts": summary,
        "total": total,
        "recent": sorted(
            CUSTOMER_SUPPORT_TICKETS,
            key=lambda item: item.get("updated_at") or item.get("created_at"),
            reverse=True,
        ),
    }


def _calculate_ticket_sla_due(ticket):
    if not isinstance(ticket, dict):
        return None

    if ticket.get("due_at"):
        return ticket.get("due_at")

    sla = ticket.get("sla") or {}
    created_at = ticket.get("created_at")
    resolution_hours = sla.get("resolution_hours")
    if created_at and resolution_hours:
        try:
            hours = float(resolution_hours)
        except (TypeError, ValueError):
            return None
        return created_at + datetime.timedelta(hours=hours)
    return None


def _customer_support_team_members():
    members = set(CUSTOMER_SUPPORT_DEFAULT_TEAM)
    for user in get_assignable_users_for_module("customer_support"):
        if user.is_active:
            members.add(user.display_name)
    for ticket in CUSTOMER_SUPPORT_TICKETS:
        assignee = ticket.get("assignee")
        if assignee:
            members.add(assignee)
        for event in ticket.get("timeline", []):
            actor = event.get("actor")
            if actor:
                members.add(actor)
    if current_user.is_authenticated:
        members.add(current_user.display_name)
    members.add("Unassigned")
    return sorted(member for member in members if member)


def _resolve_ticket_assignee_user(ticket, module_key="customer_support"):
    if not isinstance(ticket, dict):
        return None

    assignee_user_id = ticket.get("assignee_user_id")
    if assignee_user_id:
        try:
            user_id = int(assignee_user_id)
        except (TypeError, ValueError):
            user_id = None
        if user_id is not None:
            user = User.query.get(user_id)
            if user and user.is_active:
                if not module_key or user.can_be_assigned_module(module_key):
                    return user

    assignee_name = (ticket.get("assignee") or "").strip()
    if not assignee_name or assignee_name.lower() == "unassigned":
        return None

    lowered_name = assignee_name.lower()
    potential_users = get_assignable_users_for_module(module_key) if module_key else User.query.all()
    for user in potential_users:
        if not user.is_active:
            continue
        if user.display_name.strip().lower() == lowered_name or user.username.strip().lower() == lowered_name:
            ticket["assignee_user_id"] = user.id
            ticket["assignee"] = user.display_name
            return user

    user = User.query.filter(func.lower(User.username) == lowered_name).first()
    if user and user.is_active:
        if not module_key or user.can_be_assigned_module(module_key):
            ticket["assignee_user_id"] = user.id
            ticket["assignee"] = user.display_name
            return user

    return None


def _user_is_service_team_member(user, service_user_ids):
    if not user or not user.is_active:
        return False
    if user.id in service_user_ids:
        return True
    department = (getattr(user, "department", "") or "").strip().lower()
    return department == "service"


def _service_complaint_tasks_from_support():
    service_users = get_assignable_users_for_module("service")
    service_user_ids = {user.id for user in service_users if user.is_active}

    complaint_tasks = []
    for ticket in CUSTOMER_SUPPORT_TICKETS:
        status_value = (ticket.get("status") or "").strip().lower()
        if status_value in {"resolved", "closed"}:
            continue

        assigned_user = _resolve_ticket_assignee_user(ticket)
        if not _user_is_service_team_member(assigned_user, service_user_ids):
            continue

        site_label = ticket.get("location") or ticket.get("customer") or "Site not specified"
        client_label = ticket.get("customer") or ticket.get("contact_name") or "Customer pending"
        lift_label = (ticket.get("amc_site") or {}).get("label") or "—"

        due_at = ticket.get("due_at") or _calculate_ticket_sla_due(ticket)
        if isinstance(due_at, datetime.datetime):
            schedule_window = due_at.strftime("%d %b · %H:%M")
        else:
            schedule_window = "Awaiting scheduling"

        sla_info = ticket.get("sla") or {}
        resolution_hours = sla_info.get("resolution_hours")
        if resolution_hours:
            sla_label = f"Resolve within {resolution_hours}h"
        else:
            sla_label = "No SLA defined"

        worklog_entries = []
        created_at = ticket.get("created_at")
        if isinstance(created_at, datetime.datetime):
            worklog_entries.append(
                {
                    "label": ticket.get("subject") or "Ticket logged",
                    "time": created_at.strftime("%d %b %H:%M"),
                }
            )
        updated_at = ticket.get("updated_at")
        if (
            isinstance(updated_at, datetime.datetime)
            and updated_at != created_at
            and ticket.get("status")
        ):
            worklog_entries.append(
                {
                    "label": f"Status: {ticket.get('status')}",
                    "time": updated_at.strftime("%d %b %H:%M"),
                }
            )

        complaint_tasks.append(
            {
                "id": ticket.get("id"),
                "site": site_label,
                "client": client_label,
                "lift_id": lift_label,
                "call_type": f"Complaint · {ticket.get('category')}" if ticket.get("category") else "Complaint",
                "priority": ticket.get("priority") or "Medium",
                "technicians": [assigned_user.display_name] if assigned_user else [],
                "schedule_window": schedule_window,
                "sla": sla_label,
                "status": ticket.get("status") or "Open",
                "worklog": worklog_entries,
                "requires_media": False,
                "parts_used": [],
                "origin_ticket_id": ticket.get("id"),
            }
        )

    return complaint_tasks


def _infer_attachment_type(filename, mimetype=None):
    ext = (os.path.splitext(filename)[1] or "").lower()
    if mimetype:
        major = mimetype.split("/", 1)[0].lower()
        if major == "image":
            return "image"
        if major == "video":
            return "video"
    if ext in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}:
        return "image"
    if ext in {".mp4", ".mov", ".avi", ".mkv", ".webm"}:
        return "video"
    return "file"


def _save_customer_support_attachments(files):
    upload_root = app.config["UPLOAD_FOLDER"]
    os.makedirs(upload_root, exist_ok=True)

    saved = []
    timestamp_prefix = datetime.datetime.utcnow().strftime("%Y%m%d%H%M%S")

    for index, file in enumerate(files or []):
        if not file or not file.filename:
            continue

        original_name = secure_filename(file.filename)
        if not original_name:
            continue

        dest_name = f"{timestamp_prefix}_{index}_{original_name}"
        dest_path = os.path.join(upload_root, dest_name)

        try:
            file.save(dest_path)
        except Exception:
            flash(f"Could not save attachment '{original_name}'.", "error")
            continue

        attachment_type = _infer_attachment_type(original_name, getattr(file, "mimetype", None))
        saved.append(
            {
                "label": original_name,
                "type": attachment_type,
                "url": url_for("static", filename=f"uploads/{dest_name}"),
            }
        )

    return saved


def _customer_support_filter_calls(category=None, status=None, search=None):
    records = CUSTOMER_SUPPORT_CALL_LOGS

    if category:
        category = category.lower()
        records = [
            record
            for record in records
            if (record.get("category") or "").lower() == category
        ]

    if status:
        status = status.lower()
        records = [
            record
            for record in records
            if (record.get("status") or "").lower() == status
        ]

    if search:
        term = search.lower()
        records = [
            record
            for record in records
            if term in (record.get("subject") or "").lower()
            or term in (record.get("caller") or "").lower()
            or term in (record.get("ticket_id") or "").lower()
        ]

    return sorted(records, key=lambda item: item.get("logged_at"), reverse=True)


def _handle_customer_support_ticket_creation():
    form_name = (request.form.get("form_name") or "").strip().lower()
    if form_name != "create_ticket":
        return None

    customer = (request.form.get("customer") or "").strip()
    contact_name = (request.form.get("contact_name") or "").strip()
    contact_phone = (request.form.get("contact_phone") or "").strip()
    contact_email = (request.form.get("contact_email") or "").strip()
    location = (request.form.get("location") or "").strip()
    subject = (request.form.get("subject") or "").strip()
    category_id = (request.form.get("category") or "").strip()
    amc_site_id = (request.form.get("amc_site") or "").strip()
    channel_value = (request.form.get("channel") or "").strip()
    sla_priority_id = (request.form.get("sla_priority") or "").strip()
    assignee_value = (request.form.get("assignee") or "").strip()
    due_raw = (request.form.get("due_datetime") or "").strip()
    remarks = (request.form.get("remarks") or "").strip()
    uploaded_files = request.files.getlist("attachments") or []
    linked_sales_client_id_raw = (request.form.get("linked_sales_client_id") or "").strip()
    linked_project_id_raw = (request.form.get("linked_project_id") or "").strip()
    linked_customer_id_raw = (request.form.get("linked_customer_id") or "").strip()
    linked_lift_id_raw = (request.form.get("linked_lift_id") or "").strip()

    errors = []
    assignee_user = None
    assignee_user_id = None
    if not subject:
        errors.append("Provide a summary of the customer issue.")
    if not category_id:
        errors.append("Select a ticket category.")
    if not channel_value:
        errors.append("Select the intake channel for the ticket.")

    category_label = None
    category_requires_customer = True
    if category_id:
        category_label = next(
            (item.get("label") for item in CUSTOMER_SUPPORT_CATEGORIES if item.get("id") == category_id),
            None,
        )
        if not category_label:
            errors.append("Choose a valid ticket category.")
        lowered_category = category_id.lower()
        if lowered_category in {"other-dept", "other-query"}:
            category_requires_customer = False

    amc_site_record = None
    if category_id.lower() == "support-amc".lower():
        if not amc_site_id:
            errors.append("Select the AMC site for support AMC tickets.")
        else:
            amc_site_record = _resolve_customer_support_amc_site(amc_site_id)
            if not amc_site_record:
                errors.append("Choose a valid AMC site from the list.")

    if amc_site_record:
        linked_customer_name = (
            amc_site_record.get("customer_name")
            or amc_site_record.get("client")
            or amc_site_record.get("customer")
        )
        if linked_customer_name and not customer:
            customer = linked_customer_name

    linked_sales_client = None
    if linked_sales_client_id_raw:
        try:
            linked_sales_client_id = int(linked_sales_client_id_raw)
        except (TypeError, ValueError):
            errors.append("Select a valid sales client to link.")
        else:
            linked_sales_client = db.session.get(SalesClient, linked_sales_client_id)
            if not linked_sales_client:
                errors.append("Select a valid sales client to link.")
            elif not customer:
                customer = (
                    linked_sales_client.display_name
                    or linked_sales_client.company_name
                    or linked_sales_client.description
                    or f"Client {linked_sales_client.id}"
                )

    linked_project = None
    if linked_project_id_raw:
        try:
            linked_project_id = int(linked_project_id_raw)
        except (TypeError, ValueError):
            errors.append("Select a valid installation project to link.")
        else:
            linked_project = db.session.get(Project, linked_project_id)
            if not linked_project:
                errors.append("Select a valid installation project to link.")
            elif not customer:
                customer = linked_project.customer_name or linked_project.name

    linked_customer = None
    if linked_customer_id_raw:
        try:
            linked_customer_id = int(linked_customer_id_raw)
        except (TypeError, ValueError):
            errors.append("Select a valid AMC customer to link.")
        else:
            linked_customer = db.session.get(Customer, linked_customer_id)
            if not linked_customer:
                errors.append("Select a valid AMC customer to link.")
            elif not customer:
                customer = linked_customer.company_name or linked_customer.customer_code

    linked_lift = None
    if linked_lift_id_raw:
        try:
            linked_lift_id = int(linked_lift_id_raw)
        except (TypeError, ValueError):
            errors.append("Select a valid lift to link.")
        else:
            linked_lift = db.session.get(Lift, linked_lift_id)
            if not linked_lift:
                errors.append("Select a valid lift to link.")
            elif not customer:
                if linked_lift.customer and linked_lift.customer.company_name:
                    customer = linked_lift.customer.company_name
                elif linked_lift.customer_code:
                    customer = linked_lift.customer_code

    if category_requires_customer and not customer:
        errors.append("Enter the customer name for the ticket.")

    channel_label = _resolve_customer_support_channel_label(channel_value)
    if channel_value and not channel_label:
        errors.append("Choose a valid ticket channel.")

    sla_preset = None
    if sla_priority_id:
        sla_preset = next(
            (preset for preset in CUSTOMER_SUPPORT_SLA_PRESETS if preset.get("id") == sla_priority_id),
            None,
        )
    if not sla_preset and CUSTOMER_SUPPORT_SLA_PRESETS:
        sla_preset = CUSTOMER_SUPPORT_SLA_PRESETS[0]

    due_at = None
    if due_raw:
        try:
            due_at = datetime.datetime.strptime(due_raw, "%Y-%m-%dT%H:%M")
        except ValueError:
            errors.append("Enter the due date in YYYY-MM-DD HH:MM format.")

    if assignee_value:
        try:
            assignee_user_id = int(assignee_value)
        except (TypeError, ValueError):
            assignee_user_id = None
        if assignee_user_id is None:
            errors.append("Select a valid assignee from the ERP user list.")
        else:
            assignee_user = User.query.get(assignee_user_id)
            if (
                not assignee_user
                or not assignee_user.is_active
                or not assignee_user.can_be_assigned_module("customer_support")
            ):
                errors.append("Select a valid assignee from the ERP user list.")

    if errors:
        for message in errors:
            flash(message, "error")
        return redirect(url_for("customer_support_tasks"))

    ticket_id = _generate_customer_support_ticket_id()
    created_at = datetime.datetime.utcnow()
    actor_info = timeline_actor_context()

    timeline_detail_parts = []
    if channel_label:
        timeline_detail_parts.append(f"Channel: {channel_label}")
    if location:
        timeline_detail_parts.append(f"Location: {location}")
    if amc_site_record:
        timeline_detail_parts.append(f"AMC site: {amc_site_record.get('label')}")
    if subject:
        timeline_detail_parts.append(subject)
    if remarks:
        timeline_detail_parts.append(remarks)
    timeline_detail = " · ".join(part for part in timeline_detail_parts if part)

    attachments_added = _save_customer_support_attachments(uploaded_files)

    ticket_summary_parts = [ticket_id]
    if subject:
        ticket_summary_parts.append(subject)
    ticket_summary = " · ".join(part for part in ticket_summary_parts if part)
    ticket_url = url_for("customer_support_tasks", ticket=ticket_id, _external=True)
    comment_message = (
        f"Support ticket {ticket_summary} linked from customer support.\n{ticket_url}"
        if ticket_summary
        else f"Support ticket linked from customer support.\n{ticket_url}"
    )

    linked_entities = []
    if linked_sales_client:
        name = (
            linked_sales_client.display_name
            or linked_sales_client.company_name
            or linked_sales_client.description
            or f"Client {linked_sales_client.id}"
        )
        entity = {
            "type": "sales_client",
            "label": "Sales client",
            "name": name,
            "url": url_for("sales_client_detail", client_id=linked_sales_client.id),
        }
        if linked_sales_client.company_name and linked_sales_client.company_name != name:
            entity["description"] = linked_sales_client.company_name
        linked_entities.append(entity)

    if linked_project:
        project_name = linked_project.name or f"Project {linked_project.id}"
        entity = {
            "type": "installation_project",
            "label": "Installation project",
            "name": project_name,
            "url": url_for("project_detail", project_id=linked_project.id),
        }
        if linked_project.customer_name:
            entity["description"] = linked_project.customer_name
        linked_entities.append(entity)

    if linked_customer:
        customer_name = (
            linked_customer.company_name
            or linked_customer.customer_code
            or f"Customer {linked_customer.id}"
        )
        entity = {
            "type": "amc_customer",
            "label": "AMC customer",
            "name": customer_name,
            "url": url_for("service_customer_detail", customer_id=linked_customer.id),
        }
        if linked_customer.customer_code:
            entity["description"] = linked_customer.customer_code
        linked_entities.append(entity)

    if linked_lift:
        lift_name = linked_lift.lift_code or f"Lift {linked_lift.id}"
        entity = {
            "type": "lift",
            "label": "Lift",
            "name": lift_name,
            "url": url_for("service_lift_detail", lift_id=linked_lift.id),
        }
        lift_details = []
        if linked_lift.customer and linked_lift.customer.company_name:
            lift_details.append(linked_lift.customer.company_name)
        elif linked_lift.customer_code:
            lift_details.append(linked_lift.customer_code)
        if linked_lift.city:
            lift_details.append(linked_lift.city)
        if lift_details:
            entity["description"] = " · ".join(lift_details)
        linked_entities.append(entity)

    ticket_record = {
        "id": ticket_id,
        "subject": subject,
        "customer": customer,
        "contact_name": contact_name or "",
        "contact_phone": contact_phone or "",
        "contact_email": contact_email or "",
        "category": category_label or category_id,
        "channel": channel_label or channel_value,
        "priority": "Medium",
        "status": "Open",
        "assignee": assignee_user.display_name if assignee_user else "Unassigned",
        "assignee_user_id": assignee_user.id if assignee_user else None,
        "created_at": created_at,
        "updated_at": created_at,
        "sla": {
            "first_response_hours": sla_preset.get("first_response_hours", 0) if sla_preset else 0,
            "resolution_hours": sla_preset.get("resolution_hours", 0) if sla_preset else 0,
        },
        "attachments": attachments_added,
        "timeline": [
            {
                "timestamp": created_at,
                "type": "status",
                "label": "Ticket logged",
                "visibility": "external",
                "detail": timeline_detail or "Ticket created manually.",
                **actor_info,
            }
        ],
        "linked_tasks": [],
    }

    if location:
        ticket_record["location"] = location
    if due_at:
        ticket_record["due_at"] = due_at
    if remarks:
        ticket_record["remarks"] = remarks
    if amc_site_record:
        ticket_record["amc_site"] = {
            "id": amc_site_record.get("id"),
            "label": amc_site_record.get("label"),
            "client": amc_site_record.get("client"),
        }
        if amc_site_record.get("customer_name"):
            ticket_record["amc_site"]["customer_name"] = amc_site_record.get("customer_name")
        if amc_site_record.get("customer_code"):
            ticket_record["amc_site"]["customer_code"] = amc_site_record.get("customer_code")
        if amc_site_record.get("amc_status"):
            ticket_record["amc_site"]["amc_status"] = amc_site_record.get("amc_status")

    if linked_entities:
        ticket_record["linked_entities"] = linked_entities
        for entity in linked_entities:
            ticket_record[f"linked_{entity['type']}"] = entity

    comments_added = False
    if linked_sales_client:
        log_sales_activity(
            "client",
            linked_sales_client.id,
            f"Support ticket {ticket_id} linked",
            notes=comment_message,
            actor=current_user if current_user.is_authenticated else None,
        )
        comments_added = True

    if linked_project:
        db.session.add(
            ProjectComment(
                project=linked_project,
                body=comment_message,
                author=current_user if current_user.is_authenticated else None,
            )
        )
        comments_added = True

    if linked_customer:
        db.session.add(
            CustomerComment(
                customer=linked_customer,
                body=comment_message,
                author=current_user if current_user.is_authenticated else None,
            )
        )
        comments_added = True

    if linked_lift:
        db.session.add(
            LiftComment(
                lift=linked_lift,
                body=comment_message,
                author=current_user if current_user.is_authenticated else None,
            )
        )
        comments_added = True

    if comments_added:
        db.session.commit()

    CUSTOMER_SUPPORT_TICKETS.append(ticket_record)
    flash(f"Ticket {ticket_id} created successfully.", "success")
    return redirect(url_for("customer_support_tasks", ticket=ticket_id))


def _normalise_srt_schema(raw_schema):
    if not isinstance(raw_schema, list):
        return _default_srt_schema()

    normalised_sections = []
    for raw_section in raw_schema:
        if not isinstance(raw_section, dict):
            continue

        section_name = str(raw_section.get("section", "") or "")
        section_image = str(raw_section.get("display_image", "") or "")
        raw_items = raw_section.get("items")
        normalised_items = []

        if isinstance(raw_items, list):
            for raw_item in raw_items:
                if not isinstance(raw_item, dict):
                    continue

                item_type = str(raw_item.get("type", "select") or "select").lower()
                if item_type == "table":
                    rows = [
                        str(value).strip()
                        for value in raw_item.get("rows", [])
                        if str(value).strip()
                    ]
                    columns = [
                        str(value).strip()
                        for value in raw_item.get("columns", [])
                        if str(value).strip()
                    ]
                    normalised_items.append(
                        {
                            "label": str(raw_item.get("label", "") or ""),
                            "type": "table",
                            "required": bool(raw_item.get("required", False)),
                            "rows": rows or ["Row 1", "Row 2"],
                            "columns": columns or ["Column 1", "Column 2"],
                            "display_image": str(raw_item.get("display_image", "") or ""),
                        }
                    )
                    continue

                allowed_types = {"select", "text", "textarea"}
                if item_type not in allowed_types:
                    item_type = "select"

                if item_type == "select":
                    options = [
                        str(value).strip()
                        for value in raw_item.get("options", [])
                        if str(value).strip()
                    ] or ["OK", "Not OK", "Need Client Input"]
                else:
                    options = []

                allow_photo = bool(raw_item.get("allow_photo", item_type == "select"))
                photo_required = bool(raw_item.get("photo_required_if_ng", False))
                if not allow_photo or item_type != "select":
                    photo_required = False

                normalised_items.append(
                    {
                        "label": str(raw_item.get("label", "") or ""),
                        "type": item_type,
                        "options": options,
                        "required": bool(raw_item.get("required", item_type == "select")),
                        "allow_photo": allow_photo,
                        "allow_remark": bool(raw_item.get("allow_remark", item_type != "text")),
                        "photo_required_if_ng": photo_required,
                        "display_image": str(raw_item.get("display_image", "") or ""),
                    }
                )

        if not normalised_items:
            normalised_items = [_default_srt_item()]

        normalised_sections.append(
            {
                "section": section_name,
                "display_image": section_image,
                "items": normalised_items,
            }
        )

    if not normalised_sections:
        return _default_srt_schema()

    return normalised_sections


_SRT_FORM_TEMPLATES_SEED = [
    {
        "id": "srt-emergency-brake-audit",
        "name": "SRT - Emergency Brake Audit",
        "category": "Safety",
        "last_updated": datetime.date(2024, 4, 28),
        "usage_count": 14,
        "description": "Checklist capturing emergency brake checks, load test confirmation and evidence uploads.",
        "schema": [
            {
                "section": "Emergency Brake Assembly",
                "display_image": "/static/uploads/1761394043.501005_SAVE_20230822_183825.jpg",
                "items": [
                    {
                        "label": "Brake calipers inspected for wear",
                        "type": "select",
                        "options": ["OK", "Not OK", "Need Client Input"],
                        "required": True,
                        "allow_photo": True,
                        "allow_remark": True,
                        "photo_required_if_ng": True,
                        "display_image": "/static/uploads/1761394043.510924_SAVE_20230822_1837491.jpg",
                    },
                    {
                        "label": "Counterweight gap measurement (mm)",
                        "type": "text",
                        "options": [],
                        "required": True,
                        "allow_photo": False,
                        "allow_remark": False,
                        "photo_required_if_ng": False,
                        "display_image": "",
                    },
                    {
                        "label": "Load test observation notes",
                        "type": "textarea",
                        "options": [],
                        "required": False,
                        "allow_photo": True,
                        "allow_remark": True,
                        "photo_required_if_ng": False,
                        "display_image": "",
                    },
                ],
            },
            {
                "section": "Test Documentation",
                "display_image": "",
                "items": [
                    {
                        "label": "Brake torque verification table",
                        "type": "table",
                        "required": True,
                        "rows": ["Test 1", "Test 2", "Test 3"],
                        "columns": ["Recorded", "Expected", "Variance"],
                        "display_image": "/static/uploads/1761394043.506364_SAVE_20230822_183832.jpg",
                    }
                ],
            },
        ],
    },
    {
        "id": "srt-door-operation-review",
        "name": "SRT - Door Operation Review",
        "category": "Doors",
        "last_updated": datetime.date(2024, 5, 9),
        "usage_count": 9,
        "description": "Structured walk-through for door alignment, interlocks and threshold compliance.",
        "schema": [
            {
                "section": "Door Movement",
                "display_image": "",
                "items": [
                    {
                        "label": "Door closing speed within spec",
                        "type": "select",
                        "options": ["OK", "Not OK", "Need Client Input"],
                        "required": True,
                        "allow_photo": True,
                        "allow_remark": True,
                        "photo_required_if_ng": True,
                        "display_image": "",
                    },
                    {
                        "label": "Sill alignment reference",
                        "type": "textarea",
                        "options": [],
                        "required": False,
                        "allow_photo": True,
                        "allow_remark": True,
                        "photo_required_if_ng": False,
                        "display_image": "/static/uploads/1761394043.501005_SAVE_20230822_183825.jpg",
                    },
                ],
            },
            {
                "section": "Interlock Compliance",
                "display_image": "",
                "items": [
                    {
                        "label": "Landing door interlocks",
                        "type": "select",
                        "options": ["OK", "Not OK", "Need Client Input"],
                        "required": True,
                        "allow_photo": True,
                        "allow_remark": True,
                        "photo_required_if_ng": True,
                        "display_image": "",
                    },
                    {
                        "label": "Interlock wiring continuity",
                        "type": "text",
                        "options": [],
                        "required": False,
                        "allow_photo": False,
                        "allow_remark": False,
                        "photo_required_if_ng": False,
                        "display_image": "",
                    },
                ],
            },
        ],
    },
    {
        "id": "srt-post-service-summary",
        "name": "SRT - Post Service Summary",
        "category": "Reporting",
        "last_updated": datetime.date(2024, 3, 19),
        "usage_count": 22,
        "description": "Captures punch-list closure status, photos and pending parts for handover.",
        "schema": [
            {
                "section": "Punch List",
                "display_image": "",
                "items": [
                    {
                        "label": "Outstanding issues",
                        "type": "textarea",
                        "options": [],
                        "required": False,
                        "allow_photo": True,
                        "allow_remark": True,
                        "photo_required_if_ng": False,
                        "display_image": "",
                    },
                    {
                        "label": "Pending parts arrival date",
                        "type": "text",
                        "options": [],
                        "required": False,
                        "allow_photo": False,
                        "allow_remark": False,
                        "photo_required_if_ng": False,
                        "display_image": "",
                    },
                ],
            },
            {
                "section": "Hand-over Evidence",
                "display_image": "/static/uploads/1761394043.506364_SAVE_20230822_183832.jpg",
                "items": [
                    {
                        "label": "Client sign-off table",
                        "type": "table",
                        "required": True,
                        "rows": ["Client", "Technician", "Supervisor"],
                        "columns": ["Name", "Signature", "Date"],
                        "display_image": "",
                    }
                ],
            },
        ],
    },
]


SRT_FORM_TEMPLATES_FILE = os.path.join(BASE_DIR, "instance", "srt_form_templates.json")


def _seed_default_srt_form_templates():
    return copy.deepcopy(_SRT_FORM_TEMPLATES_SEED)


def _coerce_positive_int(value, default=0):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed >= 0 else default


def _parse_srt_template_date(value, default=None):
    if isinstance(value, datetime.date):
        return value

    if isinstance(value, str):
        value = value.strip()
        if not value:
            return default
        for fmt in ("%Y-%m-%d", "%d %b %Y", "%d %B %Y"):
            try:
                return datetime.datetime.strptime(value, fmt).date()
            except ValueError:
                continue

    return default


def _load_srt_form_templates():
    if os.path.exists(SRT_FORM_TEMPLATES_FILE):
        try:
            with open(SRT_FORM_TEMPLATES_FILE, "r", encoding="utf-8") as handle:
                raw_templates = json.load(handle)
        except (OSError, json.JSONDecodeError, TypeError) as exc:
            app.logger.warning("Failed to load SRT templates from disk: %s", exc)
            raw_templates = []
        templates_loaded = []
        for raw in raw_templates:
            if not isinstance(raw, dict):
                continue

            name = str(raw.get("name", "") or "").strip()
            template_id = str(raw.get("id", "") or "").strip()
            if not template_id:
                template_id = slugify(name or f"srt-template-{_random_digits(4)}")

            last_updated = _parse_srt_template_date(
                raw.get("last_updated"), default=datetime.date.today()
            )

            schema_payload = raw.get("schema")
            schema = copy.deepcopy(_normalise_srt_schema(schema_payload))

            templates_loaded.append(
                {
                    "id": template_id,
                    "name": name or "Untitled template",
                    "category": str(raw.get("category", "") or "General").strip() or "General",
                    "description": str(raw.get("description", "") or ""),
                    "usage_count": _coerce_positive_int(raw.get("usage_count"), 0),
                    "last_updated": last_updated or datetime.date.today(),
                    "schema": schema,
                }
            )

        if templates_loaded:
            templates_loaded.sort(key=lambda item: item["name"].lower())
            return templates_loaded

    return _seed_default_srt_form_templates()


def _persist_srt_form_templates():
    payload = []
    for template in SRT_FORM_TEMPLATES:
        record = {
            "id": template.get("id") or slugify(template.get("name") or "srt-template"),
            "name": template.get("name", ""),
            "category": template.get("category", "General"),
            "description": template.get("description", ""),
            "usage_count": _coerce_positive_int(template.get("usage_count"), 0),
            "last_updated": "",
            "schema": copy.deepcopy(template.get("schema") or _default_srt_schema()),
        }

        last_updated = template.get("last_updated")
        if isinstance(last_updated, datetime.date):
            record["last_updated"] = last_updated.isoformat()
        elif isinstance(last_updated, str):
            record["last_updated"] = last_updated
        else:
            record["last_updated"] = datetime.date.today().isoformat()

        payload.append(record)

    try:
        with open(SRT_FORM_TEMPLATES_FILE, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
    except OSError as exc:
        app.logger.error("Failed to persist SRT templates: %s", exc)


SRT_FORM_TEMPLATES = _load_srt_form_templates()

SRT_TEAM_MEMBERS = [
    "Ravi Kumar",
    "Priya Nair",
    "Amol Patil",
    "Sneha Kulkarni",
    "Anita D'silva",
    "Rahul Menezes",
]


SRT_SITES = []

SALES_PIPELINES = {
    "lift": {
        "label": "Lift",
        "stages": [
            "New Enquiry",
            "Site Visit",
            "Quote Submission",
            "Negotiation",
            "Closed Won",
            "Closed Lost",
        ],
    },
    "amc": {
        "label": "AMC",
        "stages": [
            "New AMC Enquiry",
            "Technician Visit",
            "Quote Submission",
            "Negotiation",
            "Closed Won",
            "Closed Lost",
        ],
    },
    "parking": {
        "label": "Parking System",
        "stages": [
            "New Parking Enquiry",
            "Site Assessment",
            "Proposal Shared",
            "Negotiation",
            "Closed Won",
            "Closed Lost",
        ],
    },
}

SALES_TEMPERATURES = [
    ("cold", "Cold"),
    ("warm", "Warm"),
    ("hot", "Hot"),
]

SALES_CLIENT_LIFECYCLE_STAGES = [
    "Prospect",
    "Qualification",
    "Negotiation",
    "Customer",
    "Post-Sales",
]

SALES_CLIENT_TEMPLATE_SHEET_NAME = "Clients"
SALES_CLIENT_UPLOAD_HEADERS = [
    "Display Name",
    "Company Name",
    "Email",
    "Phone",
    "Tag",
    "Category",
    "Owner Email",
    "Lifecycle Stage",
    "Description",
]

SALES_OPPORTUNITY_TEMPLATE_SHEET_NAME = "Opportunities"
SALES_OPPORTUNITY_UPLOAD_HEADERS = [
    "Title",
    "Pipeline",
    "Stage",
    "Status",
    "Temperature",
    "Amount",
    "Currency",
    "Expected Close Date",
    "Probability",
    "Owner Email",
    "Client Name",
    "Related Project",
    "Description",
]

OPPORTUNITY_REMINDER_OPTIONS = [
    ("", "No reminder"),
    ("1h", "1 Hr before due"),
    ("2h", "2 Hr before due"),
    ("3h", "3 Hr before due"),
    ("1d", "1 Day before due"),
]

REMINDER_OPTION_LABELS = {value: label for value, label in OPPORTUNITY_REMINDER_OPTIONS}

OPPORTUNITY_ACTIVITY_LABELS = {
    "meeting": "Meeting",
    "call": "Call",
    "email": "Email",
}


def format_file_size(num_bytes):
    if num_bytes is None:
        return "0 B"

    step = 1024.0
    units = ["B", "KB", "MB", "GB", "TB"]
    size = float(max(num_bytes, 0))
    for unit in units:
        if size < step or unit == units[-1]:
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} {unit}"
        size /= step


def get_pipeline_config(pipeline_key):
    key = (pipeline_key or "lift").lower()
    return SALES_PIPELINES.get(key, SALES_PIPELINES["lift"])


def get_pipeline_stages(pipeline_key):
    return get_pipeline_config(pipeline_key)["stages"]


def format_currency(amount, currency="₹"):
    if amount is None:
        return "—"
    return f"{currency or '₹'}{amount:,.2f}"


def _cleanup_empty_directories(start_path, stop_path):
    current = os.path.dirname(start_path)
    stop_path = os.path.abspath(stop_path)
    while True:
        current = os.path.abspath(current)
        if not current.startswith(stop_path) or current == stop_path:
            break
        try:
            os.rmdir(current)
        except OSError:
            break
        current = os.path.dirname(current)


def remove_static_file(relative_path):
    if not relative_path:
        return False

    static_root = os.path.join(BASE_DIR, "static")
    target_path = os.path.normpath(os.path.join(static_root, relative_path))
    if not target_path.startswith(os.path.abspath(static_root)):
        return False

    if not os.path.isfile(target_path):
        return False

    try:
        os.remove(target_path)
    except OSError:
        return False

    _cleanup_empty_directories(target_path, static_root)
    return True


def remove_static_directory(path):
    if not path:
        return False

    static_root = os.path.join(BASE_DIR, "static")
    target_path = os.path.abspath(path)
    if not target_path.startswith(os.path.abspath(static_root)):
        return False

    if not os.path.isdir(target_path):
        return False

    try:
        shutil.rmtree(target_path)
    except OSError:
        return False

    _cleanup_empty_directories(target_path, static_root)
    return True


def normalize_lifecycle_stage(value):
    value = (value or "").strip()
    if not value:
        return None
    if value not in SALES_CLIENT_LIFECYCLE_STAGES:
        return SALES_CLIENT_LIFECYCLE_STAGES[0]
    return value


def delete_lift_record(lift, *, remove_from_session=True):
    if not lift:
        return

    for attachment in list(getattr(lift, "attachments", [])):
        remove_static_file(attachment.stored_path)

    remove_static_directory(
        os.path.join(BASE_DIR, app.config["UPLOAD_FOLDER"], "lifts", str(lift.id))
    )
    remove_static_directory(
        os.path.join(BASE_DIR, app.config["UPLOAD_FOLDER"], "service_slips", str(lift.id))
    )

    if remove_from_session:
        db.session.delete(lift)


def delete_sales_opportunity_record(opportunity, *, remove_from_session=True):
    if not opportunity:
        return

    for file_record in list(getattr(opportunity, "files", [])):
        remove_static_file(file_record.stored_path)

    remove_static_directory(
        os.path.join(BASE_DIR, app.config["UPLOAD_FOLDER"], "opportunities", str(opportunity.id))
    )

    if remove_from_session:
        db.session.delete(opportunity)


def format_service_date(value):
    if not value:
        return "—"
    if isinstance(value, datetime.datetime):
        value = value.date()
    return value.strftime("%d %b %Y")


def reset_workspace_data():
    summary = {}
    static_root = os.path.join(BASE_DIR, "static")

    def _cleanup_submission_files(raw_json):
        if not raw_json:
            return
        try:
            entries = json.loads(raw_json)
        except (TypeError, ValueError):
            entries = []
        if not isinstance(entries, list):
            return
        for stored_path in entries:
            if not stored_path or not isinstance(stored_path, str):
                continue
            normalized = stored_path.replace("\\", "/")
            if normalized.startswith("static/"):
                normalized = normalized[len("static/"):]
            removed = remove_static_file(normalized)
            if removed:
                continue
            candidate = normalized
            if not os.path.isabs(candidate):
                candidate = os.path.join(BASE_DIR, candidate)
            candidate = os.path.abspath(candidate)
            if not candidate.startswith(static_root):
                continue
            if not os.path.isfile(candidate):
                continue
            try:
                os.remove(candidate)
            except OSError:
                continue
            _cleanup_empty_directories(candidate, static_root)

    def _cleanup_comment_attachments(raw_json):
        if not raw_json:
            return
        try:
            entries = json.loads(raw_json)
        except (TypeError, ValueError):
            entries = []
        if not isinstance(entries, list):
            return
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            path = entry.get("path") or entry.get("stored_path") or entry.get("web_path")
            if not path:
                continue
            normalized = str(path).replace("\\", "/")
            if normalized.startswith("static/"):
                normalized = normalized[len("static/"):]
            removed = remove_static_file(normalized)
            if removed:
                continue
            candidate = normalized
            if not os.path.isabs(candidate):
                candidate = os.path.join(BASE_DIR, candidate)
            candidate = os.path.abspath(candidate)
            if not candidate.startswith(static_root):
                continue
            if not os.path.isfile(candidate):
                continue
            try:
                os.remove(candidate)
            except OSError:
                continue
            _cleanup_empty_directories(candidate, static_root)

    lifts = (
        Lift.query.options(subqueryload(Lift.attachments)).all()
        if "Lift" in globals()
        else []
    )
    summary["lifts"] = len(lifts)
    for lift in lifts:
        delete_lift_record(lift, remove_from_session=False)

    if "LiftFile" in globals():
        db.session.query(LiftFile).delete(synchronize_session=False)
    if "LiftComment" in globals():
        db.session.query(LiftComment).delete(synchronize_session=False)
    if "Lift" in globals():
        db.session.query(Lift).delete(synchronize_session=False)

    customer_count = Customer.query.count() if "Customer" in globals() else 0
    summary["customers"] = customer_count
    if "CustomerComment" in globals():
        db.session.query(CustomerComment).delete(synchronize_session=False)
    if "Customer" in globals():
        db.session.query(Customer).delete(synchronize_session=False)

    submissions = Submission.query.all() if "Submission" in globals() else []
    summary["submissions"] = len(submissions)
    for submission in submissions:
        _cleanup_submission_files(getattr(submission, "photos_json", "[]"))
        _cleanup_submission_files(getattr(submission, "videos_json", "[]"))
    if summary["submissions"]:
        db.session.query(Submission).delete(synchronize_session=False)

    qc_task_count = 0
    if "QCWork" in globals():
        qc_task_count = QCWork.query.count()
        if "QCWorkComment" in globals():
            comments = QCWorkComment.query.all()
            for comment in comments:
                _cleanup_comment_attachments(getattr(comment, "attachments_json", "[]"))
            db.session.query(QCWorkComment).delete(synchronize_session=False)
        if "QCWorkLog" in globals():
            db.session.query(QCWorkLog).delete(synchronize_session=False)
        if "QCWorkDependency" in globals():
            db.session.query(QCWorkDependency).delete(synchronize_session=False)
        db.session.query(QCWork).delete(synchronize_session=False)
    summary["qc_tasks"] = qc_task_count

    project_count = Project.query.count() if "Project" in globals() else 0
    summary["projects"] = project_count
    if "ProjectComment" in globals():
        db.session.query(ProjectComment).delete(synchronize_session=False)
    if "Project" in globals():
        db.session.query(Project).delete(synchronize_session=False)

    opportunities = (
        SalesOpportunity.query.options(subqueryload(SalesOpportunity.files)).all()
        if "SalesOpportunity" in globals()
        else []
    )
    summary["opportunities"] = len(opportunities)
    for opportunity in opportunities:
        delete_sales_opportunity_record(opportunity, remove_from_session=False)

    if summary["opportunities"]:
        if "SalesOpportunityItem" in globals():
            db.session.query(SalesOpportunityItem).delete(synchronize_session=False)
        if "SalesOpportunityEngagement" in globals():
            db.session.query(SalesOpportunityEngagement).delete(synchronize_session=False)
        if "SalesOpportunityFile" in globals():
            db.session.query(SalesOpportunityFile).delete(synchronize_session=False)
        if "SalesOpportunityComment" in globals():
            db.session.query(SalesOpportunityComment).delete(synchronize_session=False)
        db.session.query(SalesOpportunity).delete(synchronize_session=False)

    client_count = SalesClient.query.count() if "SalesClient" in globals() else 0
    summary["sales_clients"] = client_count
    if "SalesClient" in globals():
        db.session.query(SalesClient).delete(synchronize_session=False)

    activities_deleted = (
        db.session.query(SalesActivity).delete(synchronize_session=False)
        if "SalesActivity" in globals()
        else 0
    )
    summary["sales_activities"] = activities_deleted or 0

    global SERVICE_COMPLAINTS, SERVICE_CONTRACTS, SERVICE_PARTS_LEDGER, SRT_SITES

    summary["service_complaints"] = len(SERVICE_COMPLAINTS)
    SERVICE_COMPLAINTS = []

    summary["service_contracts"] = len(SERVICE_CONTRACTS)
    SERVICE_CONTRACTS = []

    parts_removed = sum(len(SERVICE_PARTS_LEDGER.get(key, [])) for key in ("stock_alerts", "consumption", "returns"))
    summary["service_parts_entries"] = parts_removed
    SERVICE_PARTS_LEDGER = {"stock_alerts": [], "consumption": [], "returns": []}

    summary["srt_sites"] = len(SRT_SITES)
    SRT_SITES = []

    return summary


def format_duration_hours(value):
    if value is None:
        return "—"
    try:
        hours = float(value)
    except (TypeError, ValueError):
        return str(value)
    if hours.is_integer():
        return f"{int(hours)} hrs"
    return f"{hours:.1f} hrs"


def timeline_actor_context(actor_name=None, actor_role=None):
    name = (actor_name or "").strip()
    role = (actor_role or "").strip().lower()

    if not name:
        if current_user.is_authenticated:
            name = (
                current_user.display_name
                or current_user.email
                or current_user.username
                or "User"
            )
            role = role or ("admin" if current_user.is_admin else "user")
        else:
            name = "System"

    if not role:
        role = "system" if name.strip().lower() == "system" else "user"

    if role not in {"system", "admin", "user"}:
        role = "user"

    if role == "system":
        label = "System"
        normalized_name = "System"
    else:
        normalized_name = name or ("Admin" if role == "admin" else "User")
        role_label = "Admin" if role == "admin" else "User"
        label = f"{role_label} · {normalized_name}" if normalized_name else role_label

    return {
        "actor": normalized_name,
        "actor_role": role,
        "actor_label": label,
    }


def apply_actor_context(entry, actor_name=None, actor_role=None):
    payload = dict(entry or {})
    actor_fields = timeline_actor_context(
        actor_name or payload.get("actor"),
        actor_role or payload.get("actor_role"),
    )
    payload.update(actor_fields)
    return payload


def is_lift_open(lift):
    status = (lift.status or "").strip().lower()
    return not status or status not in {"inactive", "scrapped", "decommissioned"}


def build_lift_payload(lift):
    insight_config = copy.deepcopy(DEFAULT_LIFT_INSIGHT)

    customer = lift.customer
    route_display = "—"
    if lift.route:
        route_value = lift.route.strip()
        route_record = None
        if route_value:
            route_record = ServiceRoute.query.filter(
                func.lower(ServiceRoute.state) == route_value.lower()
            ).first()
        if route_record:
            route_display = route_record.display_name
        elif route_value:
            route_display = route_value
    site_lines = []
    if insight_config.get("site_name"):
        site_lines.append(insight_config["site_name"])
    elif lift.site_address_line1:
        site_lines.append(lift.site_address_line1)
    elif customer and customer.company_name:
        site_lines.append(customer.company_name)

    if lift.site_address_line2:
        site_lines.append(lift.site_address_line2)

    if lift.building_villa_number:
        site_lines.insert(0, lift.building_villa_number)

    location_parts = [part for part in [lift.city, lift.state, lift.pincode] if part]
    if location_parts:
        site_lines.append(", ".join(location_parts))
    if lift.country:
        if not location_parts:
            site_lines.append(lift.country)
        else:
            last_line = site_lines[-1] if site_lines else ""
            if lift.country not in last_line:
                site_lines.append(lift.country)

    customer_lines = []
    if customer and customer.company_name:
        customer_lines.append(customer.company_name)
    contact_bits = []
    if customer and customer.contact_person:
        contact_bits.append(customer.contact_person)
    if customer and customer.mobile:
        contact_bits.append(customer.mobile)
    elif customer and customer.phone:
        contact_bits.append(customer.phone)
    if contact_bits:
        customer_lines.append(" · ".join(contact_bits))
    if customer and customer.email:
        customer_lines.append(customer.email)

    machine_details = insight_config.get("machine", {}) or {}
    machine_make = machine_details.get("make") or (lift.machine_brand or "—")
    machine_model = machine_details.get("model") or "—"
    machine_serial = machine_details.get("serial") or "—"

    lifetime_value = insight_config.get("lifetime_value", {}) or {}
    amc_config = insight_config.get("amc", {}) or {}
    total_breakdowns = lifetime_value.get("total_breakdowns_this_year", 0)
    average_response_value = (
        lifetime_value.get("average_call_response_time_hours")
        or lifetime_value.get("avg_call_response_time_hours")
        or lifetime_value.get("avg_response_time_hours")
    )
    if average_response_value is not None:
        average_response_display = format_duration_hours(average_response_value)
    else:
        average_response_raw = (
            lifetime_value.get("average_call_response_time")
            or lifetime_value.get("avg_call_response_time")
            or lifetime_value.get("average_response_time")
        )
        average_response_display = (
            format_duration_hours(average_response_raw)
            if average_response_raw not in (None, "")
            else "—"
        )

    average_close_value = (
        lifetime_value.get("average_call_close_time_hours")
        or lifetime_value.get("avg_call_close_time_hours")
        or lifetime_value.get("avg_close_time_hours")
    )
    if average_close_value is not None:
        average_close_display = format_duration_hours(average_close_value)
    else:
        average_close_raw = (
            lifetime_value.get("average_call_close_time")
            or lifetime_value.get("avg_call_close_time")
            or lifetime_value.get("average_close_time")
        )
        average_close_display = (
            format_duration_hours(average_close_raw)
            if average_close_raw not in (None, "")
            else "—"
        )

    repair_revenue_this_year = lifetime_value.get("repair_revenue_this_year")
    if repair_revenue_this_year is None:
        repair_revenue_this_year = lifetime_value.get("total_repair_revenue", 0)
    total_cost_this_year = lifetime_value.get("total_cost_this_year")
    if total_cost_this_year is None:
        total_cost_this_year = lifetime_value.get("total_cost", 0)

    months_to_renewal_value = (
        lifetime_value.get("months_to_renewal")
        if lifetime_value.get("months_to_renewal") not in (None, "")
        else lifetime_value.get("months_until_renewal")
    )

    amc_end_source = amc_config.get("end") or lift.amc_end
    amc_end_date = None
    if isinstance(amc_end_source, datetime.datetime):
        amc_end_date = amc_end_source.date()
    elif isinstance(amc_end_source, datetime.date):
        amc_end_date = amc_end_source
    elif isinstance(amc_end_source, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%d %b %Y", "%d %B %Y"):
            try:
                amc_end_date = datetime.datetime.strptime(amc_end_source, fmt).date()
                break
            except ValueError:
                continue

    if months_to_renewal_value in (None, "") and amc_end_date:
        today = datetime.date.today()
        delta_months = (amc_end_date.year - today.year) * 12 + (amc_end_date.month - today.month)
        if amc_end_date.day < today.day:
            delta_months -= 1
        months_to_renewal_value = max(0, delta_months)

    if months_to_renewal_value in (None, ""):
        months_to_renewal_display = "—"
    else:
        try:
            months_to_renewal_display = str(int(float(months_to_renewal_value)))
        except (TypeError, ValueError):
            months_to_renewal_display = str(months_to_renewal_value)

    lifetime_metrics = [
        {
            "label": "Total breakdowns this year",
            "display": str(total_breakdowns),
        },
        {
            "label": "Average call response time",
            "display": average_response_display,
        },
        {
            "label": "Average call close time",
            "display": average_close_display,
        },
        {
            "label": "Repair revenue this year",
            "display": format_currency(repair_revenue_this_year or 0),
        },
        {
            "label": "Total cost this year",
            "display": format_currency(total_cost_this_year or 0),
        },
        {
            "label": "Months to next renewal",
            "display": months_to_renewal_display,
        },
    ]

    stored_metrics = lift.lifetime_metrics
    if stored_metrics:
        lifetime_metrics = [
            {
                "label": item.get("label", "Metric"),
                "display": item.get("display", "—"),
            }
            for item in stored_metrics
        ]

    amc_start = amc_config.get("start") or lift.amc_start
    amc_end = amc_config.get("end") or lift.amc_end
    linked_contract = get_service_contract_by_id(lift.amc_contract_id)

    preferred_day_keys = lift.preferred_service_days
    preferred_day_labels = []
    for day_key in preferred_day_keys:
        label = SERVICE_PREFERRED_DAY_LABELS.get(day_key)
        if not label:
            label = day_key.title()
        preferred_day_labels.append(label)
    preferred_day_display = ", ".join(preferred_day_labels) if preferred_day_labels else None
    preferred_date_display = format_preferred_service_date(lift.preferred_service_date)
    preferred_time_display = (
        lift.preferred_service_time.strftime("%H:%M")
        if isinstance(lift.preferred_service_time, datetime.time)
        else "—"
    )
    preference_bits = []
    if preferred_day_display:
        preference_bits.append(preferred_day_display)
    if lift.preferred_service_date:
        preference_bits.append(format_preferred_service_date(lift.preferred_service_date))
    elif lift.preferred_service_time and not preferred_day_display:
        preference_bits.append("Any date")
    if lift.preferred_service_time:
        preference_bits.append(lift.preferred_service_time.strftime("%H:%M"))
    preferred_summary = " · ".join(preference_bits)

    amc_payload = {
        "status": (lift.amc_status or amc_config.get("status") or "—"),
        "type": amc_config.get("type") or "—",
        "start_display": format_service_date(amc_start),
        "end_display": format_service_date(amc_end),
        "duration_display": AMC_DURATION_LABELS.get(lift.amc_duration_key) or "—",
        "contract_value_display": format_currency(amc_config.get("contract_value", 0)),
        "payment_terms": amc_config.get("payment_terms") or "—",
        "services_per_year": amc_config.get("services_per_year", 0),
        "pending_services_count": amc_config.get("pending_services_count", 0),
        "service_owner": amc_config.get("service_owner") or "—",
        "service_contact": amc_config.get("service_contact") or "—",
        "renewal_history": [
            {
                "period": item.get("period", "—"),
                "value_display": format_currency(item.get("value", 0)),
                "renewed_on_display": format_service_date(item.get("renewed_on")),
            }
            for item in amc_config.get("renewal_history", [])
        ],
        "attachments": [
            {
                "label": item.get("label", "Document"),
                "filename": item.get("filename"),
                "url": item.get("url") or "#",
            }
            for item in amc_config.get("attachments", [])
        ],
    }

    route_technician_label = (
        route_display if route_display and route_display != "—" else "Route technician"
    )

    schedule_source = lift.service_schedule or []
    service_schedule = []
    for item in schedule_source:
        if not isinstance(item, dict):
            continue
        raw_date = item.get("date")
        visit_date = None
        if isinstance(raw_date, datetime.datetime):
            visit_date = raw_date.date()
        elif isinstance(raw_date, datetime.date):
            visit_date = raw_date
        elif isinstance(raw_date, str):
            try:
                visit_date = datetime.datetime.strptime(raw_date, "%Y-%m-%d").date()
            except ValueError:
                visit_date = None
        date_display = format_service_date(visit_date)
        date_iso = visit_date.isoformat() if isinstance(visit_date, datetime.date) else ""
        technician_raw = clean_str(item.get("technician"))
        technician_display = technician_raw or route_technician_label
        status_raw = clean_str(item.get("status"))
        status_key = (
            status_raw.lower()
            if status_raw and status_raw.lower() in SERVICE_VISIT_STATUS_LABELS
            else "scheduled"
        )
        status_display = SERVICE_VISIT_STATUS_LABELS.get(status_key, "Scheduled")
        slip_raw = clean_str(
            item.get("slip_url") or item.get("slip") or item.get("url") or item.get("href")
        )
        slip_label = clean_str(item.get("slip_label") or item.get("label"))
        slip_href = None
        if slip_raw:
            if slip_raw.lower().startswith(("http://", "https://")):
                slip_href = slip_raw
            elif slip_raw.startswith("/"):
                slip_href = slip_raw
            else:
                normalized = slip_raw.lstrip("/")
                if normalized.startswith("static/"):
                    normalized = normalized.split("static/", 1)[1]
                if normalized:
                    slip_href = url_for("static", filename=normalized)
        slip_display_label = (
            slip_label
            or (os.path.basename(slip_raw) if slip_raw else None)
            or "Service slip"
        )
        service_schedule.append(
            {
                "date": visit_date,
                "date_display": date_display,
                "date_iso": date_iso,
                "technician_display": technician_display,
                "technician_value": technician_raw or "",
                "status_key": status_key,
                "status_display": status_display,
                "slip_url": slip_href,
                "slip_label": slip_label,
                "slip_display_label": slip_display_label,
                "slip_stored": slip_raw,
                "has_slip": bool(slip_href),
                "allow_overdue": True,
            }
        )

    service_schedule.sort(
        key=lambda entry: entry["date"] or datetime.date.max
    )

    contacts_source = lift.amc_contacts or amc_config.get("contacts", []) or []
    amc_contacts = []
    for contact in contacts_source:
        if not isinstance(contact, dict):
            continue
        amc_contacts.append(
            {
                "name": contact.get("name") or "—",
                "designation": contact.get("designation") or "—",
                "phone": contact.get("phone") or "—",
                "email": contact.get("email") or "—",
            }
        )
    if not amc_contacts and amc_payload["service_owner"] != "—":
        amc_contacts.append(
            {
                "name": amc_payload["service_owner"],
                "designation": "Service Owner",
                "phone": amc_payload["service_contact"],
                "email": "—",
            }
        )
    amc_payload["contacts"] = amc_contacts

    if linked_contract and linked_contract.get("id"):
        amc_payload["contract"] = {
            "id": linked_contract.get("id"),
            "type": linked_contract.get("type"),
            "coverage": linked_contract.get("coverage"),
            "url": url_for(
                "service_contracts",
                _anchor=f"contract-{linked_contract.get('id')}"
            ),
        }
    else:
        amc_payload["contract"] = None

    uploads_config = insight_config.get("uploads", {}) or {}
    documents = [
        {
            "label": item.get("label", "Document"),
            "filename": item.get("filename"),
            "description": item.get("description"),
            "updated_display": format_service_date(item.get("updated")),
            "url": item.get("url") or "#",
        }
        for item in uploads_config.get("documents", [])
    ]

    additional_uploads = []
    for bucket in ("media", "other"):
        for item in uploads_config.get(bucket, []) or []:
            additional_uploads.append(
                {
                    "label": item.get("label", "Attachment"),
                    "filename": item.get("filename"),
                    "description": item.get("description"),
                    "updated_display": format_service_date(item.get("updated")),
                    "url": item.get("url") or "#",
                }
            )

    stored_documents = []
    stored_other_uploads = []
    sorted_attachments = sorted(
        lift.attachments,
        key=lambda record: record.created_at or datetime.datetime.min,
        reverse=True,
    )
    for record in sorted_attachments:
        entry = {
            "label": record.display_label,
            "filename": record.original_filename,
            "description": record.description,
            "updated_display": record.uploaded_display,
            "url": url_for("static", filename=record.stored_path) if record.stored_path else "#",
            "uploaded_by": record.uploaded_by.display_name if record.uploaded_by else None,
            "size_display": record.display_size,
        }
        category = (record.category or "other").strip().lower()
        if category == "document":
            stored_documents.append(entry)
        else:
            stored_other_uploads.append(entry)

    def dedupe_uploads(items):
        seen = set()
        unique = []
        for item in items:
            key = (
                (item.get("label") or "").strip().lower(),
                (item.get("filename") or "").strip().lower(),
                item.get("url") or "",
            )
            if key in seen:
                continue
            unique.append(item)
            seen.add(key)
        return unique

    documents = dedupe_uploads(stored_documents + documents)
    additional_uploads = dedupe_uploads(stored_other_uploads + additional_uploads)

    breakdowns = [
        {
            "issue": item.get("issue", "—"),
            "technician": item.get("technician", "—"),
            "response": format_duration_hours(item.get("response_hours")),
            "resolution": format_duration_hours(item.get("resolution_hours")),
            "fault_type": item.get("fault_type", "—"),
            "spares": ", ".join(item.get("spares", []) or []) or "—",
            "status": item.get("status", "—"),
            "media": ", ".join(item.get("media", []) or []) or "—",
            "call_reference": item.get("call_reference", "—"),
        }
        for item in insight_config.get("breakdowns", [])
    ]

    summary_source = insight_config.get("breakdown_summary") or []
    breakdown_summary = []
    for item in summary_source:
        if not isinstance(item, dict):
            continue
        raw_date = item.get("date")
        summary_date = None
        if isinstance(raw_date, datetime.datetime):
            summary_date = raw_date.date()
        elif isinstance(raw_date, datetime.date):
            summary_date = raw_date
        elif isinstance(raw_date, str):
            try:
                summary_date = datetime.datetime.strptime(raw_date, "%Y-%m-%d").date()
            except ValueError:
                summary_date = None
        status_raw = clean_str(item.get("status"))
        status_key = clean_str(item.get("status_key")) or (
            status_raw.lower().replace(" ", "_") if status_raw else ""
        )
        breakdown_summary.append(
            {
                "date": summary_date,
                "date_display": format_service_date(summary_date),
                "date_iso": summary_date.isoformat() if isinstance(summary_date, datetime.date) else "",
                "type": item.get("type") or item.get("fault_type") or "—",
                "status": status_raw or "—",
                "status_key": status_key or "",
                "description": item.get("description")
                or item.get("detail")
                or "No additional details provided.",
                "call_reference": item.get("call_reference")
                or item.get("reference")
                or "—",
                "reported_by": item.get("reported_by")
                or item.get("reported_by_name")
                or item.get("reported_by_contact"),
            }
        )

    if breakdown_summary:
        breakdown_summary.sort(
            key=lambda entry: (
                entry.get("date") is None,
                entry.get("date") or datetime.date.min,
            ),
            reverse=True,
        )

    timeline_entries = []
    stored_timeline = lift.timeline_entries
    if stored_timeline:
        for item in stored_timeline:
            actor_info = apply_actor_context(item)
            timeline_entries.append(
                {
                    "date_display": format_service_date(item.get("date")),
                    "title": item.get("title", "—"),
                    "detail": item.get("detail", ""),
                    "category": item.get("category", "Update"),
                    "actor_label": actor_info.get("actor_label"),
                    "actor_role": actor_info.get("actor_role"),
                    "actor_name": actor_info.get("actor"),
                }
            )
    for item in insight_config.get("timeline", []) or []:
        actor_info = apply_actor_context(item)
        timeline_entries.append(
            {
                "date_display": format_service_date(item.get("date")),
                "title": item.get("title", "—"),
                "detail": item.get("detail", ""),
                "category": item.get("category", "Update"),
                "actor_label": actor_info.get("actor_label"),
                "actor_role": actor_info.get("actor_role"),
                "actor_name": actor_info.get("actor"),
            }
        )

    sorted_comments = sorted(
        lift.comments,
        key=lambda record: record.created_at or datetime.datetime.min,
        reverse=True,
    )

    machine_type_display = (
        lift.machine_type or insight_config.get("drive_type") or "—"
    )
    door_finish_display = (
        lift.door_finish or insight_config.get("door_finish") or "—"
    )
    cabin_finish_display = (
        lift.cabin_finish or insight_config.get("cabin_finish") or "—"
    )
    power_supply_display = (
        lift.power_supply or insight_config.get("power_supply") or "—"
    )
    machine_make_value = machine_make if machine_make != "—" else None
    machine_model_value = machine_model if machine_model != "—" else None
    machine_serial_value = machine_serial if machine_serial != "—" else None
    machine_details_parts = []
    if machine_make_value:
        machine_details_parts.append(machine_make_value)
    if machine_model_value:
        machine_details_parts.append(machine_model_value)
    if machine_serial_value:
        machine_details_parts.append(f"Serial {machine_serial_value}")
    machine_details_display = " · ".join(machine_details_parts) if machine_details_parts else "—"
    controller_type_value = insight_config.get("controller_type")
    controller_brand_value = lift.controller_brand
    controller_parts = []
    if controller_type_value:
        controller_parts.append(controller_type_value)
    if controller_brand_value:
        controller_parts.append(controller_brand_value)
    controller_display = " · ".join(controller_parts) if controller_parts else "—"

    payload = {
        "id": lift.id,
        "lift_code": lift.lift_code,
        "external_lift_id": lift.external_lift_id or "—",
        "customer_code": lift.customer_code or "—",
        "status": lift.status or "—",
        "site_lines": site_lines or ["—"],
        "customer_lines": customer_lines or ["—"],
        "site_summary": " · ".join(site_lines) if site_lines else "—",
        "customer_summary": " · ".join(customer_lines) if customer_lines else "—",
        "lift_type": lift.lift_type or "—",
        "lift_brand": lift.lift_brand or "—",
        "drive_type": machine_type_display,
        "machine_type": machine_type_display,
        "controller_type": controller_type_value or "—",
        "controller_brand": controller_brand_value or "—",
        "controller_display": controller_display,
        "door_type": lift.door_type or "—",
        "door_configuration": insight_config.get("door_configuration") or (lift.door_type or "—"),
        "floors_served": insight_config.get("floors_served") or (lift.building_floors or "—"),
        "route_display": route_display,
        "door_finish": door_finish_display,
        "cabin_finish": cabin_finish_display,
        "power_supply": power_supply_display,
        "site_address_line1": lift.site_address_line1 or "—",
        "site_address_line2": lift.site_address_line2 or "—",
        "city": lift.city or "—",
        "state": lift.state or "—",
        "pincode": lift.pincode or "—",
        "country": lift.country or "—",
        "capacity_display": lift.capacity_display or (f"{lift.capacity_persons} persons / {lift.capacity_kg} kg" if lift.capacity_persons and lift.capacity_kg else "—"),
        "speed_display": f"{lift.speed_mps:.2f} m/s" if lift.speed_mps is not None else "—",
        "machine_details_display": machine_details_display,
        "machine_make": machine_make,
        "machine_model": machine_model,
        "machine_serial": machine_serial,
        "machine_brand": lift.machine_brand or "—",
        "controller_brand": controller_brand_value or "—",
        "install_date_display": format_service_date(lift.install_date),
        "warranty_expiry_display": format_service_date(lift.warranty_expiry),
        "last_service_date_display": format_service_date(lift.last_service_date),
        "commissioned_date_display": format_service_date(insight_config.get("commissioned_date") or lift.install_date),
        "documents": documents,
        "other_uploads": additional_uploads,
        "amc": amc_payload,
        "breakdowns": breakdowns,
        "breakdown_summary": breakdown_summary,
        "timeline": timeline_entries,
        "lifetime_metrics": lifetime_metrics,
        "service_schedule": service_schedule,
        "remarks": lift.remarks or "—",
        "building_villa_number": lift.building_villa_number or "—",
        "preferred_service_day_display": preferred_day_display or "—",
        "preferred_service_date_display": preferred_date_display,
        "preferred_service_time_display": preferred_time_display,
        "preferred_service_summary": preferred_summary,
        "location_display": ", ".join(
            [part for part in [lift.city, lift.state, lift.pincode, lift.country] if part]
        )
        or "—",
        "geo_location": lift.geo_location or "—",
        "qr_code_url": lift.qr_code_url or None,
        "service_notes": lift.notes or "—",
        "comments": [
            {
                "body": comment.body,
                "author": comment.author_name,
                "created_display": comment.created_display,
            }
            for comment in sorted_comments
        ],
    }

    return payload


def log_sales_activity(parent_type, parent_id, title, notes=None, actor=None):
    entry = SalesActivity(
        parent_type=parent_type,
        parent_id=parent_id,
        actor=actor or (current_user if current_user.is_authenticated else None),
        title=title,
        notes=notes,
    )
    db.session.add(entry)
    return entry


def normalize_floor_label(raw_value):
    value = (raw_value or "").strip()
    if not value:
        return None

    compact = re.sub(r"\s+", "", value.upper())
    if compact in {"G", "G+"}:
        return None

    # If the value is purely numeric, prefix it with G+ for consistency.
    if re.fullmatch(r"\d+", compact):
        return f"G+{compact}"

    # Normalize leading G without a plus sign (e.g., G10 -> G+10).
    if compact.startswith("G") and not compact.startswith("G+"):
        compact = f"G+{compact[1:]}"

    # Ensure there is only a single + immediately after G when present.
    if compact.startswith("G+"):
        suffix = compact[2:]
        suffix = suffix.lstrip("+")
        compact = f"G+{suffix}" if suffix else "G+"

    if compact in {"", "G+"}:
        return None

    return compact


def _extract_task_timing(form):
    start_mode = (form.get("start_mode") or "immediate").strip().lower()
    if start_mode not in {"immediate", "scheduled", "after_previous"}:
        start_mode = "immediate"

    start_date_value = None
    start_date_raw = (form.get("start_date") or "").strip()
    if start_mode == "scheduled":
        if not start_date_raw:
            return None, None, None, None, "Provide a start date for scheduled tasks."
        try:
            start_date_value = datetime.datetime.strptime(start_date_raw, "%Y-%m-%d").date()
        except ValueError:
            return None, None, None, None, "Start date must be a valid YYYY-MM-DD date."
    elif start_mode == "after_previous":
        # Ignore any manually provided start date when using sequential scheduling.
        start_date_value = None
    elif start_date_raw:
        # Allow optionally overriding the start date even if immediate was chosen.
        try:
            start_date_value = datetime.datetime.strptime(start_date_raw, "%Y-%m-%d").date()
        except ValueError:
            return None, None, None, None, "Start date must be a valid YYYY-MM-DD date."

    duration_raw = (form.get("duration_days") or "").strip()
    duration_value = None
    if duration_raw:
        try:
            duration_value = int(duration_raw)
        except ValueError:
            return None, None, None, None, "Duration must be a whole number of days."
        if duration_value < 0:
            return None, None, None, None, "Duration must be zero or a positive number of days."

    milestone_value = (form.get("milestone") or "").strip() or None

    # If mode is immediate we ignore scheduled date (unless user typed one to override).
    if start_mode == "immediate" and start_date_raw == "":
        start_date_value = None

    return start_mode, start_date_value, duration_value, milestone_value, None


def normalize_template_task_order(template_id):
    tasks = ProjectTemplateTask.query.filter_by(template_id=template_id).order_by(
        ProjectTemplateTask.order_index.asc(),
        ProjectTemplateTask.id.asc()
    ).all()
    for idx, task in enumerate(tasks, start=1):
        task.order_index = idx
    return tasks


def set_template_task_dependencies(task, dependency_ids):
    if not task:
        return
    cleaned = []
    for dep_id in dependency_ids or []:
        if not dep_id:
            continue
        if dep_id == task.id:
            continue
        if dep_id not in cleaned:
            cleaned.append(dep_id)
    task.depends_on_id = cleaned[0] if cleaned else None
    existing = {link.depends_on_id: link for link in getattr(task, "dependency_links", [])}
    for dep_id, link in list(existing.items()):
        if dep_id not in cleaned:
            db.session.delete(link)
    for dep_id in cleaned:
        if dep_id not in existing:
            db.session.add(ProjectTemplateTaskDependency(task_id=task.id, depends_on_id=dep_id))


def set_qc_work_dependencies(work, dependency_ids):
    if not work:
        return
    cleaned = []
    for dep_id in dependency_ids or []:
        if not dep_id:
            continue
        if dep_id == work.id:
            continue
        if dep_id not in cleaned:
            cleaned.append(dep_id)
    work.depends_on_id = cleaned[0] if cleaned else None
    existing = {link.depends_on_id: link for link in getattr(work, "dependency_links", [])}
    for dep_id, link in list(existing.items()):
        if dep_id not in cleaned:
            db.session.delete(link)
    for dep_id in cleaned:
        if dep_id not in existing:
            db.session.add(QCWorkDependency(task_id=work.id, depends_on_id=dep_id))


def synchronize_dependency_links():
    try:
        template_tasks = ProjectTemplateTask.query.filter(ProjectTemplateTask.depends_on_id.isnot(None)).all()
        for task in template_tasks:
            existing = {link.depends_on_id for link in getattr(task, "dependency_links", [])}
            if task.depends_on_id and task.depends_on_id not in existing:
                db.session.add(ProjectTemplateTaskDependency(task_id=task.id, depends_on_id=task.depends_on_id))

        qc_tasks = QCWork.query.filter(QCWork.depends_on_id.isnot(None)).all()
        for work in qc_tasks:
            existing = {link.depends_on_id for link in getattr(work, "dependency_links", [])}
            if work.depends_on_id and work.depends_on_id not in existing:
                db.session.add(QCWorkDependency(task_id=work.id, depends_on_id=work.depends_on_id))

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        print(f"⚠️ Unable to synchronize dependency links automatically: {exc}")


def build_task_template_blueprint(template):
    tasks = sorted(template.tasks, key=lambda t: ((t.order_index or 0), t.id))
    id_to_index = {task.id: idx for idx, task in enumerate(tasks)}
    blueprint = []
    for idx, task in enumerate(tasks):
        dependency_indexes = []
        for dep in task.dependencies:
            dep_index = id_to_index.get(dep.id)
            if dep_index is not None:
                dependency_indexes.append(dep_index)
        blueprint.append({
            "name": task.name,
            "description": task.description,
            "order_index": task.order_index or (idx + 1),
            "default_assignee_id": task.default_assignee_id,
            "form_template_id": task.form_template_id,
            "start_mode": task.start_mode or "immediate",
            "planned_start_date": task.planned_start_date.isoformat() if task.planned_start_date else None,
            "duration_days": task.duration_days,
            "milestone": task.milestone,
            "dependency_indexes": dependency_indexes,
        })
    return blueprint


def apply_blueprint_to_template(template, blueprint):
    if not isinstance(blueprint, list):
        return []
    created = []
    for idx, entry in enumerate(blueprint):
        if not isinstance(entry, dict):
            continue
        name = (entry.get("name") or f"Task {idx + 1}").strip()
        description = (entry.get("description") or None)
        order_index = entry.get("order_index") or (idx + 1)
        default_assignee_id = entry.get("default_assignee_id")
        form_template_id = entry.get("form_template_id")
        start_mode = (entry.get("start_mode") or "immediate").lower()
        if start_mode not in {"immediate", "scheduled", "after_previous"}:
            start_mode = "immediate"
        planned_start = entry.get("planned_start_date")
        planned_start_date = None
        if planned_start:
            try:
                planned_start_date = datetime.datetime.strptime(planned_start, "%Y-%m-%d").date()
            except ValueError:
                planned_start_date = None
        duration_days = entry.get("duration_days")
        if isinstance(duration_days, str) and duration_days.isdigit():
            duration_days = int(duration_days)
        elif not isinstance(duration_days, int):
            duration_days = None
        milestone = entry.get("milestone") or None

        task = ProjectTemplateTask(
            template_id=template.id,
            name=name,
            description=description,
            order_index=order_index,
            default_assignee_id=default_assignee_id,
            form_template_id=form_template_id,
            start_mode=start_mode,
            planned_start_date=planned_start_date,
            duration_days=duration_days,
            milestone=milestone
        )
        db.session.add(task)
        db.session.flush()
        created.append((task, entry))

    for task, entry in created:
        dependency_indexes = entry.get("dependency_indexes") or []
        dependency_ids = []
        for dep_idx in dependency_indexes:
            if isinstance(dep_idx, int) and 0 <= dep_idx < len(created):
                dependency_ids.append(created[dep_idx][0].id)
        set_template_task_dependencies(task, dependency_ids)

    normalize_template_task_order(template.id)
    return [task for task, _ in created]
# -------------------------------------------------------------------------------


def _normalize_form_schema(schema_raw):
    """Return (sections, is_sectioned) for a stored form schema."""
    if not isinstance(schema_raw, list):
        return [], False

    def _normalize_item(item, idx):
        if not isinstance(item, dict):
            item = {}
        normalized = dict(item)
        normalized["label"] = str(item.get("label") or f"Item {idx + 1}")
        ftype = (item.get("type") or "select").lower()
        if ftype not in {"text", "textarea", "select", "table"}:
            ftype = "select"
        normalized["type"] = ftype
        normalized["required"] = bool(item.get("required", False))
        if ftype == "select":
            opts = item.get("options") or ["OK", "Not OK", "Need Client Input"]
            normalized["options"] = [str(opt) for opt in opts if str(opt).strip()]
            normalized["photo_required_if_ng"] = bool(item.get("photo_required_if_ng", False))
            normalized["allow_photo"] = bool(item.get("allow_photo", normalized["photo_required_if_ng"]))
            normalized["allow_remark"] = bool(item.get("allow_remark", False))
            normalized["reference_image"] = None
            normalized["rows"] = []
            normalized["columns"] = []
        elif ftype in {"text", "textarea"}:
            normalized["options"] = []
            normalized["photo_required_if_ng"] = False
            normalized["allow_photo"] = bool(item.get("allow_photo", False))
            normalized["allow_remark"] = bool(item.get("allow_remark", False))
            normalized["reference_image"] = None
            normalized["rows"] = []
            normalized["columns"] = []
        else:  # table
            rows = item.get("rows") or []
            cols = item.get("columns") or []
            if not isinstance(rows, list):
                rows = []
            if not isinstance(cols, list):
                cols = []
            normalized["rows"] = [str(r) for r in rows if str(r).strip()]
            normalized["columns"] = [str(c) for c in cols if str(c).strip()]
            if not normalized["rows"]:
                normalized["rows"] = ["Row 1", "Row 2"]
            if not normalized["columns"]:
                normalized["columns"] = ["Column 1", "Column 2"]
            ref_img = item.get("reference_image")
            normalized["reference_image"] = str(ref_img) if ref_img is not None else ""
            normalized["options"] = []
            normalized["photo_required_if_ng"] = False
            normalized["allow_photo"] = False
            normalized["allow_remark"] = False
        return normalized

    if schema_raw and isinstance(schema_raw[0], dict) and "section" in schema_raw[0]:
        sections = []
        for s_idx, section in enumerate(schema_raw):
            if not isinstance(section, dict):
                continue
            items = section.get("items") or []
            normalized_items = [_normalize_item(it, idx) for idx, it in enumerate(items)]
            sections.append({
                "section": section.get("section") or f"Section {s_idx + 1}",
                "items": normalized_items
            })
        return sections, True

    normalized_items = [_normalize_item(it, idx) for idx, it in enumerate(schema_raw)]
    return [{"section": "", "items": normalized_items}], False


class Department(db.Model):
    __tablename__ = "department"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    branch = db.Column(db.String(50), nullable=False, default=DEPARTMENT_BRANCHES[0])
    description = db.Column(db.Text, nullable=True)
    parent_id = db.Column(db.Integer, db.ForeignKey("department.id"), nullable=True)
    active = db.Column(db.Boolean, default=True)

    children = db.relationship(
        "Department",
        backref=db.backref("parent", remote_side=[id]),
    )
    positions = db.relationship(
        "Position",
        back_populates="department",
    )

    @property
    def full_name(self):
        parts = [self.name]
        parent = getattr(self, "parent", None)
        while parent:
            parts.append(parent.name)
            parent = getattr(parent, "parent", None)
        return " / ".join(reversed([p for p in parts if p]))


class Position(db.Model):
    __tablename__ = "position"

    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(120), nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey("department.id"), nullable=True)
    reports_to_id = db.Column(db.Integer, db.ForeignKey("position.id"), nullable=True)
    active = db.Column(db.Boolean, default=True)

    department = db.relationship("Department", back_populates="positions")
    reports_to = db.relationship(
        "Position",
        remote_side=[id],
        backref=db.backref("direct_reports"),
    )
    users = db.relationship("User", back_populates="position")

    @property
    def hierarchy_label(self):
        parts = [self.title]
        parent = getattr(self, "reports_to", None)
        visited = {self.id}
        while parent and getattr(parent, "id", None) not in visited:
            parts.append(parent.title)
            visited.add(parent.id)
            parent = getattr(parent, "reports_to", None)
        return " / ".join(reversed([p for p in parts if p]))

    @property
    def display_label(self):
        dept_label = self.department.full_name if self.department else None
        hierarchy = self.hierarchy_label
        if dept_label:
            return f"{dept_label} · {hierarchy}"
        return hierarchy


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    first_name = db.Column(db.String(80), nullable=True)
    last_name = db.Column(db.String(80), nullable=True)
    department = db.Column(db.String(120), nullable=True)
    role = db.Column(db.String(120), nullable=True)
    mobile_number = db.Column(db.String(40), nullable=True)
    email = db.Column(db.String(120), nullable=True)
    display_picture = db.Column(db.String(255), nullable=True)
    active = db.Column(db.Boolean, default=True)
    session_token = db.Column(
        db.String(36),
        nullable=True,
        default=lambda: str(uuid.uuid4()),
    )
    position_id = db.Column(db.Integer, db.ForeignKey("position.id"), nullable=True)
    module_permissions_json = db.Column(db.Text, nullable=False, default="{}")

    position = db.relationship("Position", back_populates="users")

    @property
    def display_name(self):
        parts = [p for p in [self.first_name, self.last_name] if p]
        return " ".join(parts) if parts else self.username

    @property
    def is_active(self):
        return bool(self.active)

    @property
    def is_admin(self):
        role = (self.role or "").strip().lower()
        return role == "admin" or self.username.lower() == "admin"

    def _module_permissions_cache(self):
        cache = getattr(self, "_module_permissions_data", None)
        if cache is None:
            raw = self.module_permissions_json or "{}"
            try:
                loaded = json.loads(raw)
            except (TypeError, ValueError):
                loaded = {}
            if not isinstance(loaded, dict):
                loaded = {}
            normalised = {}
            for key, value in loaded.items():
                if not isinstance(value, dict):
                    continue
                module_key = (key or "").strip().lower()
                if not module_key:
                    continue
                normalised[module_key] = {
                    "visibility": bool(value.get("visibility", True)),
                    "assignment": bool(value.get("assignment", True)),
                }
            self._module_permissions_data = normalised
            cache = normalised
        return cache

    def get_module_permission_settings(self, module_key):
        module_key = (module_key or "").strip().lower()
        data = self._module_permissions_cache().get(module_key, {})
        return {
            "visibility": bool(data.get("visibility", True)),
            "assignment": bool(data.get("assignment", True)),
        }

    def set_module_permissions(self, permissions):
        cleaned = {}
        for key, value in (permissions or {}).items():
            module_key = (key or "").strip().lower()
            if not module_key:
                continue
            visibility = bool(value.get("visibility", True)) if isinstance(value, dict) else bool(value)
            assignment = bool(value.get("assignment", True)) if isinstance(value, dict) else True
            cleaned[module_key] = {
                "visibility": visibility,
                "assignment": assignment,
            }
        self.module_permissions_json = json.dumps(cleaned)
        self._module_permissions_data = cleaned

    def can_view_module(self, module_key):
        if self.is_admin:
            return True
        module_key = (module_key or "").strip().lower()
        if not module_key:
            return True
        settings = self.get_module_permission_settings(module_key)
        return bool(settings.get("visibility", True))

    def can_be_assigned_module(self, module_key):
        if self.is_admin:
            return True
        module_key = (module_key or "").strip().lower()
        if not module_key:
            return True
        settings = self.get_module_permission_settings(module_key)
        return bool(settings.get("assignment", True))

    def issue_session_token(self):
        self.session_token = str(uuid.uuid4())
        return self.session_token


def _module_visibility_required(module_key):
    if not current_user.can_view_module(module_key):
        abort(403)


def get_assignable_users_for_module(module_key, order_by="name"):
    module_key = (module_key or "").strip().lower()
    query = User.query
    if order_by == "username":
        query = query.order_by(User.username.asc())
    else:
        query = query.order_by(User.first_name.asc(), User.last_name.asc(), User.username.asc())
    users = query.all()
    return [user for user in users if user.can_be_assigned_module(module_key)]


@app.context_processor
def inject_workspace_modules():
    return {
        "workspace_modules": WORKSPACE_MODULES,
        "workspace_module_map": WORKSPACE_MODULE_MAP,
    }


@app.context_processor
def inject_service_form_options():
    return {
        "SERVICE_BRANCH_OPTIONS": SERVICE_BRANCH_OPTIONS,
        "SERVICE_BRANCH_OPTION_SET": SERVICE_BRANCH_OPTION_SET,
        "LIFT_TYPE_OPTIONS": LIFT_TYPE_OPTIONS,
        "LIFT_CAPACITY_PERSON_OPTIONS": LIFT_CAPACITY_PERSON_OPTIONS,
        "MACHINE_TYPE_OPTIONS": MACHINE_TYPE_OPTIONS,
        "DOOR_TYPE_OPTIONS": DOOR_TYPE_OPTIONS,
        "DOOR_FINISH_OPTIONS": DOOR_FINISH_OPTIONS,
        "POWER_SUPPLY_OPTIONS": POWER_SUPPLY_OPTIONS,
        "AMC_STATUS_OPTIONS": AMC_STATUS_OPTIONS,
        "LIFT_STATUS_OPTIONS": LIFT_STATUS_OPTIONS,
    }


@app.context_processor
def inject_switchable_users():
    users = []
    if current_user.is_authenticated:
        try:
            users = (
                User.query.order_by(
                    User.first_name.asc(),
                    User.last_name.asc(),
                    User.username.asc(),
                ).all()
            )
        except Exception:
            users = []
    return {"switchable_users": users}


class Project(db.Model):
    __tablename__ = "project"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    site_name = db.Column(db.String(200), nullable=True)
    site_address = db.Column(db.Text, nullable=True)
    customer_name = db.Column(db.String(200), nullable=True)
    lift_type = db.Column(db.String(40), nullable=True)
    floors = db.Column(db.String(40), nullable=True)
    stops = db.Column(db.Integer, nullable=True)
    opening_type = db.Column(db.String(40), nullable=True)
    location = db.Column(db.String(40), nullable=True)
    structure_type = db.Column(db.String(40), nullable=True)
    cladding_type = db.Column(db.String(40), nullable=True)
    cabin_finish = db.Column(db.String(40), nullable=True)
    door_operation_type = db.Column(db.String(40), nullable=True)
    door_finish = db.Column(db.String(40), nullable=True)
    handover_date = db.Column(db.Date, nullable=True)
    priority = db.Column(db.String(20), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    comments = db.relationship(
        "ProjectComment",
        back_populates="project",
        cascade="all, delete-orphan",
    )


class ProjectComment(db.Model):
    __tablename__ = "project_comment"

    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=False)
    author_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    body = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    project = db.relationship("Project", back_populates="comments")
    author = db.relationship("User")


class FormSchema(db.Model):
    __tablename__ = "form_schema"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    schema_json = db.Column(db.Text, nullable=False, default="[]")
    min_photos_if_all_good = db.Column(db.Integer, default=0)

    # NEW
    stage = db.Column(db.String(40), nullable=True)      # e.g., "Stage 1", "Completion", etc.
    lift_type = db.Column(db.String(40), nullable=True)  # e.g., "MRL", "Hydraulic", etc.


class Submission(db.Model):
    __tablename__ = "submission"
    id = db.Column(db.Integer, primary_key=True)
    form_id = db.Column(db.Integer, db.ForeignKey("form_schema.id"), nullable=False)
    submitted_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    data_json = db.Column(db.Text, nullable=False, default="{}")
    photos_json = db.Column(db.Text, nullable=False, default="[]")
    videos_json = db.Column(db.Text, nullable=False, default="[]")

    # Optional link to a QC work item (created in /qc tab). Safe-migrated.
    work_id = db.Column(db.Integer, nullable=True)

    form = db.relationship("FormSchema")
    user = db.relationship("User")


class ProjectTemplate(db.Model):
    __tablename__ = "project_template"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    creator = db.relationship("User")


class ProjectTemplateTask(db.Model):
    __tablename__ = "project_template_task"
    id = db.Column(db.Integer, primary_key=True)
    template_id = db.Column(db.Integer, db.ForeignKey("project_template.id"), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=True)
    order_index = db.Column(db.Integer, default=0)
    depends_on_id = db.Column(db.Integer, db.ForeignKey("project_template_task.id"), nullable=True)
    default_assignee_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    form_template_id = db.Column(db.Integer, db.ForeignKey("form_schema.id"), nullable=True)
    start_mode = db.Column(db.String(20), default="immediate")
    planned_start_date = db.Column(db.Date, nullable=True)
    duration_days = db.Column(db.Integer, nullable=True)
    milestone = db.Column(db.String(120), nullable=True)

    template = db.relationship("ProjectTemplate", backref=db.backref("tasks", cascade="all, delete-orphan", order_by="ProjectTemplateTask.order_index"))
    depends_on = db.relationship("ProjectTemplateTask", remote_side=[id], backref=db.backref("primary_dependents", cascade="all"))
    default_assignee = db.relationship("User", foreign_keys=[default_assignee_id])
    form_template = db.relationship("FormSchema")
    dependency_links = db.relationship(
        "ProjectTemplateTaskDependency",
        foreign_keys="ProjectTemplateTaskDependency.task_id",
        cascade="all, delete-orphan",
        back_populates="task"
    )

    @property
    def planned_due_date(self):
        if self.planned_start_date and self.duration_days:
            return self.planned_start_date + datetime.timedelta(days=self.duration_days)
        return None

    @property
    def dependency_ids(self):
        ids = []
        if self.depends_on_id:
            ids.append(self.depends_on_id)
        for link in getattr(self, "dependency_links", []):
            if link.depends_on_id and link.depends_on_id not in ids:
                ids.append(link.depends_on_id)
        return ids

    @property
    def dependencies(self):
        seen = set()
        ordered = []
        if self.depends_on and self.depends_on.id not in seen:
            ordered.append(self.depends_on)
            seen.add(self.depends_on.id)
        for link in getattr(self, "dependency_links", []):
            if link.dependency and link.dependency.id not in seen:
                ordered.append(link.dependency)
                seen.add(link.dependency.id)
        return ordered


class ProjectTemplateTaskDependency(db.Model):
    __tablename__ = "project_template_task_dependency"
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey("project_template_task.id"), nullable=False)
    depends_on_id = db.Column(db.Integer, db.ForeignKey("project_template_task.id"), nullable=False)

    __table_args__ = (
        db.UniqueConstraint("task_id", "depends_on_id", name="uq_template_task_dependency"),
    )

    task = db.relationship(
        "ProjectTemplateTask",
        foreign_keys=[task_id],
        back_populates="dependency_links"
    )
    dependency = db.relationship(
        "ProjectTemplateTask",
        foreign_keys=[depends_on_id],
        backref=db.backref("dependent_links", cascade="all, delete-orphan")
    )


class TaskTemplate(db.Model):
    __tablename__ = "task_template"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False, unique=True)
    description = db.Column(db.Text, nullable=True)
    blueprint_json = db.Column(db.Text, nullable=False)
    created_from_template_id = db.Column(db.Integer, db.ForeignKey("project_template.id"), nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    creator = db.relationship("User")
    source_template = db.relationship("ProjectTemplate")

    @property
    def task_count(self):
        try:
            data = json.loads(self.blueprint_json or "[]")
        except json.JSONDecodeError:
            return 0
        if isinstance(data, list):
            return len(data)
        return 0


class SalesClient(db.Model):
    __tablename__ = "sales_client"

    id = db.Column(db.Integer, primary_key=True)
    display_name = db.Column(db.String(150), nullable=False)
    company_name = db.Column(db.String(200), nullable=True)
    email = db.Column(db.String(200), nullable=True)
    phone = db.Column(db.String(50), nullable=True)
    tag = db.Column(db.String(60), nullable=True)
    category = db.Column(db.String(60), default="Individual")
    owner_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    lifecycle_stage = db.Column(db.String(120), nullable=True)
    description = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(
        db.DateTime,
        default=datetime.datetime.utcnow,
        onupdate=datetime.datetime.utcnow,
    )

    owner = db.relationship("User")
    opportunities = db.relationship(
        "SalesOpportunity",
        back_populates="client",
        cascade="all, delete-orphan",
    )

    @property
    def open_opportunity_count(self):
        return sum(1 for opp in self.opportunities if not opp.is_closed)


class SalesOpportunity(db.Model):
    __tablename__ = "sales_opportunity"

    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    pipeline = db.Column(db.String(40), nullable=False, default="lift")
    stage = db.Column(db.String(120), nullable=False)
    status = db.Column(db.String(40), default="Open")
    temperature = db.Column(db.String(20), nullable=True)
    amount = db.Column(db.Float, nullable=True)
    currency = db.Column(db.String(8), default="₹")
    expected_close_date = db.Column(db.Date, nullable=True)
    probability = db.Column(db.Integer, nullable=True)
    owner_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    client_id = db.Column(db.Integer, db.ForeignKey("sales_client.id"), nullable=True)
    related_project = db.Column(db.String(200), nullable=True)
    description = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(
        db.DateTime,
        default=datetime.datetime.utcnow,
        onupdate=datetime.datetime.utcnow,
    )

    owner = db.relationship("User")
    client = db.relationship("SalesClient", back_populates="opportunities")
    comments = db.relationship(
        "SalesOpportunityComment",
        back_populates="opportunity",
        cascade="all, delete-orphan",
    )
    files = db.relationship(
        "SalesOpportunityFile",
        back_populates="opportunity",
        cascade="all, delete-orphan",
    )
    engagements = db.relationship(
        "SalesOpportunityEngagement",
        back_populates="opportunity",
        cascade="all, delete-orphan",
    )
    items = db.relationship(
        "SalesOpportunityItem",
        back_populates="opportunity",
        cascade="all, delete-orphan",
    )

    @property
    def display_amount(self):
        if self.amount is None:
            return "—"
        return f"{self.currency or '₹'}{self.amount:,.2f}"

    @property
    def is_closed(self):
        status = (self.status or "").strip().lower()
        if status == "closed":
            return True

        stage = (self.stage or "").strip().lower()
        return stage.startswith("closed") if stage else False

    @property
    def badge_variant(self):
        mapping = {
            "hot": ("Hot", "bg-rose-500/20 text-rose-300 border border-rose-500/40"),
            "warm": ("Warm", "bg-amber-500/20 text-amber-200 border border-amber-500/40"),
            "cold": ("Cold", "bg-sky-500/20 text-sky-200 border border-sky-500/40"),
        }
        key = (self.temperature or "").strip().lower()
        return mapping.get(key, (None, "bg-slate-800/60 text-slate-300 border border-slate-700/60"))


class SalesActivity(db.Model):
    __tablename__ = "sales_activity"

    id = db.Column(db.Integer, primary_key=True)
    parent_type = db.Column(db.String(30), nullable=False)
    parent_id = db.Column(db.Integer, nullable=False)
    actor_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    title = db.Column(db.String(200), nullable=False)
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    actor = db.relationship("User")


class SalesOpportunityComment(db.Model):
    __tablename__ = "sales_opportunity_comment"

    id = db.Column(db.Integer, primary_key=True)
    opportunity_id = db.Column(db.Integer, db.ForeignKey("sales_opportunity.id"), nullable=False)
    author_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    body = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    opportunity = db.relationship("SalesOpportunity", back_populates="comments")
    author = db.relationship("User")


class SalesOpportunityFile(db.Model):
    __tablename__ = "sales_opportunity_file"

    id = db.Column(db.Integer, primary_key=True)
    opportunity_id = db.Column(db.Integer, db.ForeignKey("sales_opportunity.id"), nullable=False)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    original_filename = db.Column(db.String(255), nullable=False)
    stored_path = db.Column(db.String(400), nullable=False)
    content_type = db.Column(db.String(120), nullable=True)
    file_size = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    opportunity = db.relationship("SalesOpportunity", back_populates="files")
    uploaded_by = db.relationship("User")

    @property
    def display_size(self):
        return format_file_size(self.file_size or 0)


class SalesOpportunityEngagement(db.Model):
    __tablename__ = "sales_opportunity_engagement"

    id = db.Column(db.Integer, primary_key=True)
    opportunity_id = db.Column(db.Integer, db.ForeignKey("sales_opportunity.id"), nullable=False)
    activity_type = db.Column(db.String(40), default="meeting")
    subject = db.Column(db.String(200), nullable=True)
    scheduled_for = db.Column(db.DateTime, nullable=True)
    reminder_option = db.Column(db.String(20), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    created_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    opportunity = db.relationship("SalesOpportunity", back_populates="engagements")
    created_by = db.relationship("User")

    @property
    def display_activity_type(self):
        return OPPORTUNITY_ACTIVITY_LABELS.get(self.activity_type, (self.activity_type or "").title() or "Activity")

    @property
    def display_schedule(self):
        if not self.scheduled_for:
            return "Date not set"
        return self.scheduled_for.strftime("%d %b %Y, %I:%M %p")

    @property
    def display_reminder(self):
        key = (self.reminder_option or "").strip()
        return REMINDER_OPTION_LABELS.get(key, "No reminder")


class SalesOpportunityItem(db.Model):
    __tablename__ = "sales_opportunity_item"

    id = db.Column(db.Integer, primary_key=True)
    opportunity_id = db.Column(db.Integer, db.ForeignKey("sales_opportunity.id"), nullable=False)
    details = db.Column(db.Text, nullable=True)
    lift_type = db.Column(db.String(80), nullable=True)
    quantity = db.Column(db.Integer, nullable=True)
    floors = db.Column(db.String(80), nullable=True)
    cabin_finish = db.Column(db.String(120), nullable=True)
    door_type = db.Column(db.String(120), nullable=True)
    structure_required = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    opportunity = db.relationship("SalesOpportunity", back_populates="items")

    @property
    def structure_label(self):
        return "Yes" if self.structure_required else "No"


# NEW: QC Work table (simple tracker for “create work for new site QC”)
class QCWork(db.Model):
    __tablename__ = "qc_work"
    id = db.Column(db.Integer, primary_key=True)
    site_name = db.Column(db.String(200), nullable=False)
    client_name = db.Column(db.String(200), nullable=True)
    address = db.Column(db.Text, nullable=True)

    name = db.Column(db.String(200), nullable=True)
    description = db.Column(db.Text, nullable=True)

    template_id = db.Column(db.Integer, db.ForeignKey("form_schema.id"), nullable=False)
    stage = db.Column(db.String(40), nullable=True)
    lift_type = db.Column(db.String(40), nullable=True)

    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=True)

    template_task_id = db.Column(db.Integer, db.ForeignKey("project_template_task.id"), nullable=True)
    depends_on_id = db.Column(db.Integer, db.ForeignKey("qc_work.id"), nullable=True)

    status = db.Column(db.String(40), default="Open")  # Open / In Progress / Closed
    due_date = db.Column(db.DateTime, nullable=True)
    planned_start_date = db.Column(db.Date, nullable=True)
    planned_duration_days = db.Column(db.Integer, nullable=True)
    milestone = db.Column(db.String(120), nullable=True)

    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    assigned_to = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    template = db.relationship("FormSchema")
    creator = db.relationship("User", foreign_keys=[created_by])
    assignee = db.relationship("User", foreign_keys=[assigned_to])
    project = db.relationship("Project", backref=db.backref("tasks", lazy="dynamic"))
    template_task = db.relationship("ProjectTemplateTask", backref=db.backref("project_tasks", lazy="dynamic"))
    primary_dependency = db.relationship(
        "QCWork",
        remote_side=[id],
        backref=db.backref("primary_dependents", cascade="all"),
        foreign_keys=[depends_on_id]
    )
    dependency_links = db.relationship(
        "QCWorkDependency",
        foreign_keys="QCWorkDependency.task_id",
        cascade="all, delete-orphan",
        back_populates="task"
    )

    @property
    def display_title(self):
        if self.name:
            return self.name
        if self.site_name:
            return self.site_name
        return f"Task #{self.id}"

    @property
    def dependency_satisfied(self):
        for dependency in self.dependencies:
            if (dependency.status or "").lower() != "closed":
                return False
        return True

    @property
    def is_blocked(self):
        return not self.dependency_satisfied

    @property
    def dependencies(self):
        seen = set()
        ordered = []
        if self.primary_dependency and self.primary_dependency.id not in seen:
            ordered.append(self.primary_dependency)
            seen.add(self.primary_dependency.id)
        for link in getattr(self, "dependency_links", []):
            if link.dependency and link.dependency.id not in seen:
                ordered.append(link.dependency)
                seen.add(link.dependency.id)
        return ordered

    @property
    def dependency(self):
        deps = self.dependencies
        return deps[0] if deps else None

    @property
    def dependency_ids(self):
        return [dep.id for dep in self.dependencies]

    @property
    def all_dependents(self):
        dependents = []
        seen = set()
        for dependent in getattr(self, "primary_dependents", []):
            if dependent.id not in seen:
                dependents.append(dependent)
                seen.add(dependent.id)
        for link in getattr(self, "dependent_links", []):
            if link.task and link.task.id not in seen:
                dependents.append(link.task)
                seen.add(link.task.id)
        return dependents

    @property
    def planned_due_date(self):
        if self.planned_start_date and self.planned_duration_days:
            return self.planned_start_date + datetime.timedelta(days=self.planned_duration_days)
        return None


class QCWorkDependency(db.Model):
    __tablename__ = "qc_work_dependency"
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey("qc_work.id"), nullable=False)
    depends_on_id = db.Column(db.Integer, db.ForeignKey("qc_work.id"), nullable=False)

    __table_args__ = (
        db.UniqueConstraint("task_id", "depends_on_id", name="uq_qc_work_dependency"),
    )

    task = db.relationship(
        "QCWork",
        foreign_keys=[task_id],
        back_populates="dependency_links"
    )
    dependency = db.relationship(
        "QCWork",
        foreign_keys=[depends_on_id],
        backref=db.backref("dependent_links", cascade="all, delete-orphan")
    )


class QCWorkComment(db.Model):
    __tablename__ = "qc_work_comment"
    id = db.Column(db.Integer, primary_key=True)
    work_id = db.Column(db.Integer, db.ForeignKey("qc_work.id"), nullable=False)
    author_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    body = db.Column(db.Text, nullable=False)
    attachments_json = db.Column(db.Text, default="[]")
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    work = db.relationship("QCWork", backref=db.backref("comments", cascade="all, delete-orphan"))
    author = db.relationship("User")


class QCWorkLog(db.Model):
    __tablename__ = "qc_work_log"
    id = db.Column(db.Integer, primary_key=True)
    work_id = db.Column(db.Integer, db.ForeignKey("qc_work.id"), nullable=False)
    actor_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    action = db.Column(db.String(120), nullable=False)
    from_status = db.Column(db.String(40), nullable=True)
    to_status = db.Column(db.String(40), nullable=True)
    details_json = db.Column(db.Text, default="{}")
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    work = db.relationship("QCWork", backref=db.backref("logs", cascade="all, delete-orphan"))
    actor = db.relationship("User")


class ServiceRoute(db.Model):
    __tablename__ = "service_route"

    id = db.Column(db.Integer, primary_key=True)
    state = db.Column(db.String(120), unique=True, nullable=False)
    branch = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    @property
    def display_name(self):
        if self.branch:
            return f"{self.state} · {self.branch}"
        return self.state

    @property
    def route_name(self):
        return self.state


class Customer(db.Model):
    __tablename__ = "customer"

    id = db.Column(db.Integer, primary_key=True)
    customer_code = db.Column(db.String(32), unique=True, nullable=False, index=True)
    external_customer_id = db.Column(
        db.String(100), unique=True, nullable=True, index=True
    )
    company_name = db.Column(db.String(255), nullable=False)
    contact_person = db.Column(db.String(255), nullable=True)
    phone = db.Column(db.String(50), nullable=True)
    mobile = db.Column(db.String(50), nullable=True)
    email = db.Column(db.String(255), nullable=True)
    gst_no = db.Column(db.String(40), nullable=True)
    billing_address_line1 = db.Column(db.String(255), nullable=True)
    billing_address_line2 = db.Column(db.String(255), nullable=True)
    city = db.Column(db.String(120), nullable=True)
    state = db.Column(db.String(120), nullable=True)
    pincode = db.Column(db.String(20), nullable=True)
    country = db.Column(db.String(120), nullable=True)
    route = db.Column(db.String(120), nullable=True)
    sector = db.Column(db.String(60), nullable=True)
    branch = db.Column(db.String(120), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    office_address_line1 = db.Column(db.String(255), nullable=True)
    office_address_line2 = db.Column(db.String(255), nullable=True)
    office_city = db.Column(db.String(120), nullable=True)
    office_state = db.Column(db.String(120), nullable=True)
    office_pincode = db.Column(db.String(20), nullable=True)
    office_country = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(
        db.DateTime,
        default=datetime.datetime.utcnow,
        onupdate=datetime.datetime.utcnow,
    )

    lifts = db.relationship(
        "Lift",
        back_populates="customer",
        foreign_keys="Lift.customer_code",
        primaryjoin="Customer.customer_code==Lift.customer_code",
    )

    comments = db.relationship(
        "CustomerComment",
        back_populates="customer",
        cascade="all, delete-orphan",
    )

    def display_name(self):
        return f"{self.customer_code} – {self.company_name}" if self.company_name else self.customer_code


class CustomerComment(db.Model):
    __tablename__ = "customer_comment"

    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey("customer.id"), nullable=False)
    author_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    body = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    customer = db.relationship("Customer", back_populates="comments")
    author = db.relationship("User")


class Lift(db.Model):
    __tablename__ = "lift"

    id = db.Column(db.Integer, primary_key=True)
    lift_code = db.Column(db.String(50), unique=True, nullable=False, index=True)
    external_lift_id = db.Column(db.String(100), nullable=True)
    customer_code = db.Column(db.String(32), db.ForeignKey("customer.customer_code"), nullable=True, index=True)
    site_address_line1 = db.Column(db.String(255), nullable=True)
    site_address_line2 = db.Column(db.String(255), nullable=True)
    city = db.Column(db.String(120), nullable=True)
    state = db.Column(db.String(120), nullable=True)
    pincode = db.Column(db.String(20), nullable=True)
    geo_location = db.Column(db.String(255), nullable=True)
    country = db.Column(db.String(120), nullable=True)
    building_villa_number = db.Column(db.String(120), nullable=True)
    route = db.Column(db.String(20), nullable=True)
    building_floors = db.Column(db.String(40), nullable=True)
    lift_type = db.Column(db.String(40), nullable=True)
    lift_brand = db.Column(db.String(120), nullable=True)
    capacity_persons = db.Column(db.Integer, nullable=True)
    capacity_kg = db.Column(db.Integer, nullable=True)
    speed_mps = db.Column(db.Float, nullable=True)
    machine_type = db.Column(db.String(40), nullable=True)
    machine_brand = db.Column(db.String(120), nullable=True)
    controller_brand = db.Column(db.String(120), nullable=True)
    door_type = db.Column(db.String(120), nullable=True)
    door_brand = db.Column(db.String(120), nullable=True)
    cabin_finish = db.Column(db.String(120), nullable=True)
    power_supply = db.Column(db.String(40), nullable=True)
    install_date = db.Column(db.Date, nullable=True)
    warranty_expiry = db.Column(db.Date, nullable=True)
    amc_status = db.Column(db.String(40), nullable=True)
    amc_start = db.Column(db.Date, nullable=True)
    amc_end = db.Column(db.Date, nullable=True)
    amc_duration_key = db.Column(db.String(40), nullable=True)
    amc_contract_id = db.Column(db.String(60), nullable=True)
    qr_code_url = db.Column(db.String(255), nullable=True)
    status = db.Column(db.String(40), nullable=True)
    last_service_date = db.Column(db.Date, nullable=True)
    next_service_due = db.Column(db.Date, nullable=True)
    notes = db.Column(db.Text, nullable=True)
    remarks = db.Column(db.Text, nullable=True)
    preferred_service_day = db.Column(db.String(20), nullable=True)
    preferred_service_date = db.Column(db.Date, nullable=True)
    preferred_service_time = db.Column(db.Time, nullable=True)
    preferred_service_days_json = db.Column(db.Text, nullable=True)
    lifetime_metrics_json = db.Column(db.Text, nullable=True)
    amc_contacts_json = db.Column(db.Text, nullable=True)
    timeline_entries_json = db.Column(db.Text, nullable=True)
    service_schedule_json = db.Column(db.Text, nullable=True)
    last_updated_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(
        db.DateTime,
        default=datetime.datetime.utcnow,
        onupdate=datetime.datetime.utcnow,
    )

    capacity_display = db.Column(db.String(120), nullable=True)

    customer = db.relationship(
        "Customer",
        back_populates="lifts",
        foreign_keys=[customer_code],
    )
    attachments = db.relationship(
        "LiftFile",
        back_populates="lift",
        cascade="all, delete-orphan",
    )
    comments = db.relationship(
        "LiftComment",
        back_populates="lift",
        cascade="all, delete-orphan",
    )

    def set_capacity_display(self):
        if self.capacity_persons and self.capacity_kg:
            self.capacity_display = f"{self.capacity_persons} persons / {self.capacity_kg} kg"
        elif self.capacity_persons:
            self.capacity_display = f"{self.capacity_persons} persons"
        elif self.capacity_kg:
            self.capacity_display = f"{self.capacity_kg} kg"
        else:
            self.capacity_display = None

    @property
    def door_finish(self):
        return self.door_brand

    @door_finish.setter
    def door_finish(self, value):
        self.door_brand = value

    @property
    def preferred_service_days(self):
        if not self.preferred_service_days_json:
            fallback = clean_str(self.preferred_service_day)
            return [fallback.lower()] if fallback else []
        try:
            data = json.loads(self.preferred_service_days_json)
        except (TypeError, ValueError):
            fallback = clean_str(self.preferred_service_day)
            return [fallback.lower()] if fallback else []
        if not isinstance(data, list):
            fallback = clean_str(self.preferred_service_day)
            return [fallback.lower()] if fallback else []
        cleaned = []
        for item in data:
            if not isinstance(item, str):
                continue
            cleaned_item = item.strip().lower()
            if cleaned_item:
                cleaned.append(cleaned_item)
        if not cleaned:
            fallback = clean_str(self.preferred_service_day)
            if fallback:
                return [fallback.lower()]
        return cleaned

    @preferred_service_days.setter
    def preferred_service_days(self, values):
        if not values:
            self.preferred_service_days_json = None
            self.preferred_service_day = None
            return
        unique_values = []
        for value in values:
            if not isinstance(value, str):
                continue
            cleaned = value.strip().lower()
            if cleaned and cleaned not in unique_values:
                unique_values.append(cleaned)
        if unique_values:
            self.preferred_service_day = unique_values[0]
            self.preferred_service_days_json = json.dumps(unique_values)
        else:
            self.preferred_service_day = None
            self.preferred_service_days_json = None

    @property
    def lifetime_metrics(self):
        if not self.lifetime_metrics_json:
            return []
        try:
            data = json.loads(self.lifetime_metrics_json)
        except (TypeError, ValueError):
            return []
        if not isinstance(data, list):
            return []
        normalized = []
        for item in data:
            if not isinstance(item, dict):
                continue
            label = clean_str(item.get("label")) or "Metric"
            display = item.get("display")
            if display is None:
                display = clean_str(item.get("value")) or "—"
            normalized.append({"label": label, "display": display})
        return normalized

    @lifetime_metrics.setter
    def lifetime_metrics(self, values):
        if not values:
            self.lifetime_metrics_json = None
            return
        cleaned = []
        for item in values:
            if not isinstance(item, dict):
                continue
            label = clean_str(item.get("label"))
            display = item.get("display")
            if display is None:
                display = clean_str(item.get("value"))
            if not label and not display:
                continue
            cleaned.append(
                {
                    "label": label or "Metric",
                    "display": display or "—",
                }
            )
        self.lifetime_metrics_json = json.dumps(cleaned, ensure_ascii=False)

    @property
    def amc_contacts(self):
        if not self.amc_contacts_json:
            return []
        try:
            data = json.loads(self.amc_contacts_json)
        except (TypeError, ValueError):
            return []
        if not isinstance(data, list):
            return []
        contacts = []
        for item in data:
            if not isinstance(item, dict):
                continue
            contacts.append(
                {
                    "name": clean_str(item.get("name")) or "—",
                    "designation": clean_str(item.get("designation")) or "—",
                    "phone": clean_str(item.get("phone")) or "—",
                    "email": clean_str(item.get("email")) or "—",
                }
            )
        return contacts

    @amc_contacts.setter
    def amc_contacts(self, values):
        if not values:
            self.amc_contacts_json = None
            return
        cleaned = []
        for item in values:
            if not isinstance(item, dict):
                continue
            name = clean_str(item.get("name"))
            designation = clean_str(item.get("designation"))
            phone = clean_str(item.get("phone"))
            email = clean_str(item.get("email"))
            if not any([name, designation, phone, email]):
                continue
            cleaned.append(
                {
                    "name": name or "—",
                    "designation": designation or "—",
                    "phone": phone or "—",
                    "email": email or "—",
                }
            )
        self.amc_contacts_json = json.dumps(cleaned, ensure_ascii=False)

    @property
    def service_schedule(self):
        if not self.service_schedule_json:
            return []
        try:
            data = json.loads(self.service_schedule_json)
        except (TypeError, ValueError):
            return []
        if not isinstance(data, list):
            return []
        schedule = []
        for item in data:
            if not isinstance(item, dict):
                continue
            raw_date = item.get("date")
            visit_date = None
            if isinstance(raw_date, datetime.datetime):
                visit_date = raw_date.date()
            elif isinstance(raw_date, datetime.date):
                visit_date = raw_date
            elif isinstance(raw_date, str):
                try:
                    visit_date = datetime.datetime.strptime(raw_date, "%Y-%m-%d").date()
                except ValueError:
                    visit_date = None
            technician = clean_str(item.get("technician"))
            status_value = clean_str(item.get("status"))
            status_key = (
                status_value.lower()
                if status_value and status_value.lower() in SERVICE_VISIT_STATUS_LABELS
                else None
            )
            slip_url = clean_str(item.get("slip_url") or item.get("slip"))
            slip_label = clean_str(item.get("slip_label") or item.get("label"))
            schedule.append(
                {
                    "date": visit_date,
                    "technician": technician,
                    "status": status_key,
                    "slip_url": slip_url,
                    "slip_label": slip_label,
                }
            )
        return schedule

    @service_schedule.setter
    def service_schedule(self, values):
        if not values:
            self.service_schedule_json = None
            return
        cleaned = []
        for item in values:
            if not isinstance(item, dict):
                continue
            raw_date = item.get("date")
            iso_date = None
            if isinstance(raw_date, datetime.datetime):
                iso_date = raw_date.date().isoformat()
            elif isinstance(raw_date, datetime.date):
                iso_date = raw_date.isoformat()
            else:
                date_str = clean_str(raw_date)
                if date_str:
                    try:
                        iso_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date().isoformat()
                    except ValueError:
                        continue
            if not iso_date:
                continue
            status_value = clean_str(item.get("status"))
            status_key = (
                status_value.lower()
                if status_value and status_value.lower() in SERVICE_VISIT_STATUS_LABELS
                else "scheduled"
            )
            technician = clean_str(item.get("technician"))
            slip_url = clean_str(item.get("slip_url"))
            slip_label = clean_str(item.get("slip_label"))
            cleaned.append(
                {
                    "date": iso_date,
                    "technician": technician,
                    "status": status_key,
                    "slip_url": slip_url,
                    "slip_label": slip_label,
                }
            )
        self.service_schedule_json = (
            json.dumps(cleaned, ensure_ascii=False) if cleaned else None
        )

    @property
    def timeline_entries(self):
        if not self.timeline_entries_json:
            return []
        try:
            data = json.loads(self.timeline_entries_json)
        except (TypeError, ValueError):
            return []
        if not isinstance(data, list):
            return []
        entries = []
        for item in data:
            if not isinstance(item, dict):
                continue
            raw_date = item.get("date")
            parsed_date = None
            if isinstance(raw_date, datetime.date):
                parsed_date = raw_date
            elif isinstance(raw_date, str) and raw_date:
                try:
                    parsed_date = datetime.datetime.strptime(raw_date, "%Y-%m-%d").date()
                except ValueError:
                    parsed_date = None
            actor_info = apply_actor_context(item)
            entries.append(
                {
                    "date": parsed_date,
                    "title": item.get("title") or "—",
                    "detail": item.get("detail") or "",
                    "category": item.get("category") or "Update",
                    "actor": actor_info.get("actor"),
                    "actor_role": actor_info.get("actor_role"),
                    "actor_label": actor_info.get("actor_label"),
                }
            )
        return entries

    @timeline_entries.setter
    def timeline_entries(self, values):
        if not values:
            self.timeline_entries_json = None
            return
        cleaned = []
        for item in values:
            if not isinstance(item, dict):
                continue
            title = clean_str(item.get("title")) or "—"
            detail = clean_str(item.get("detail"))
            category = clean_str(item.get("category")) or "Update"
            raw_date = item.get("date")
            iso_date = None
            if isinstance(raw_date, datetime.date):
                iso_date = raw_date.isoformat()
            elif isinstance(raw_date, datetime.datetime):
                iso_date = raw_date.date().isoformat()
            else:
                date_str = clean_str(raw_date)
                if date_str:
                    try:
                        iso_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date().isoformat()
                    except ValueError:
                        iso_date = None
            actor_info = apply_actor_context(item)
            cleaned.append(
                {
                    "date": iso_date,
                    "title": title,
                    "detail": detail or "",
                    "category": category,
                    "actor": actor_info.get("actor"),
                    "actor_role": actor_info.get("actor_role"),
                    "actor_label": actor_info.get("actor_label"),
                }
            )
        self.timeline_entries_json = json.dumps(cleaned, ensure_ascii=False)


class LiftFile(db.Model):
    __tablename__ = "lift_file"

    id = db.Column(db.Integer, primary_key=True)
    lift_id = db.Column(db.Integer, db.ForeignKey("lift.id"), nullable=False, index=True)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    label = db.Column(db.String(150), nullable=True)
    description = db.Column(db.Text, nullable=True)
    category = db.Column(db.String(20), nullable=True, default="other")
    original_filename = db.Column(db.String(255), nullable=False)
    stored_path = db.Column(db.String(400), nullable=False)
    content_type = db.Column(db.String(120), nullable=True)
    file_size = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    lift = db.relationship("Lift", back_populates="attachments")
    uploaded_by = db.relationship("User")

    @property
    def display_label(self):
        return self.label or self.original_filename

    @property
    def display_size(self):
        return format_file_size(self.file_size or 0)

    @property
    def uploaded_display(self):
        if not self.created_at:
            return "—"
        return self.created_at.strftime("%d %b %Y, %I:%M %p")


class LiftComment(db.Model):
    __tablename__ = "lift_comment"

    id = db.Column(db.Integer, primary_key=True)
    lift_id = db.Column(db.Integer, db.ForeignKey("lift.id"), nullable=False, index=True)
    author_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    body = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    lift = db.relationship("Lift", back_populates="comments")
    author = db.relationship("User")

    @property
    def author_name(self):
        if self.author:
            return self.author.display_name
        return "System"

    @property
    def created_display(self):
        if not self.created_at:
            return "—"
        return self.created_at.strftime("%d %b %Y, %I:%M %p")


class DropdownOption(db.Model):
    __tablename__ = "dropdown_option"

    id = db.Column(db.Integer, primary_key=True)
    field_key = db.Column(db.String(50), nullable=False, index=True)
    value = db.Column(db.String(120), nullable=True)
    label = db.Column(db.String(150), nullable=False)
    order_index = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint("field_key", "label", name="uq_dropdown_option_field_label"),
    )

    def as_choice(self):
        option_value = self.value if self.value is not None else self.label
        return {"id": self.id, "value": option_value, "label": self.label}

@login_manager.user_loader
def load_user(user_id):
    # Ensure that any pending bootstrap/migration tasks run before we try to
    # query the database. When the application starts for the very first time
    # Flask-Login may attempt to load the user (via `current_user`) before our
    # `@app.before_request` hook that calls `ensure_bootstrap()` executes. In
    # that scenario the legacy SQLite schema might still be missing newer
    # columns such as `user.active`, resulting in an OperationalError during the
    # initial SELECT. Proactively invoking `ensure_bootstrap()` here guarantees
    # the schema has been patched before any queries are issued.
    ensure_bootstrap()
    try:
        user_obj = db.session.get(User, int(user_id))
    except (TypeError, ValueError):
        return None
    if user_obj and not user_obj.is_active:
        return None
    return user_obj


ALLOWED_PHOTO = {"png", "jpg", "jpeg", "webp"}
ALLOWED_VIDEO = {"mp4", "mov", "avi", "mkv"}
ALLOWED_ATTACHMENTS = ALLOWED_PHOTO.union(ALLOWED_VIDEO).union({"pdf", "doc", "docx", "xls", "xlsx"})

def allowed_file(filename, kind="photo"):
    if "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    if kind == "photo":
        return ext in ALLOWED_PHOTO
    if kind == "video":
        return ext in ALLOWED_VIDEO
    return ext in ALLOWED_ATTACHMENTS


@app.before_request
def enforce_user_session():
    if not current_user.is_authenticated:
        return

    if not current_user.is_active:
        session.pop("session_token", None)
        logout_user()
        flash("Your account has been deactivated.", "error")
        return

    if not current_user.session_token:
        current_user.issue_session_token()
        db.session.commit()

    token = session.get("session_token")
    if token and token == current_user.session_token:
        return

    session.pop("session_token", None)
    logout_user()
    flash("You have been signed out. Please log in again.", "info")


def log_work_event(work_id, action, actor_id=None, from_status=None, to_status=None, details=None):
    entry = QCWorkLog(
        work_id=work_id,
        actor_id=actor_id,
        action=action,
        from_status=from_status,
        to_status=to_status,
        details_json=json.dumps(details or {}, ensure_ascii=False)
    )
    db.session.add(entry)


def get_or_create_default_task_form():
    form = FormSchema.query.filter_by(name=DEFAULT_TASK_FORM_NAME).first()
    if form:
        return form
    fallback_schema = [
        {"label": "Summary", "type": "textarea", "required": False, "allow_remark": False},
        {"label": "Status", "type": "select", "required": True, "options": ["Not Started", "In Progress", "Completed"]}
    ]
    form = FormSchema(
        name=DEFAULT_TASK_FORM_NAME,
        schema_json=json.dumps(fallback_schema, ensure_ascii=False),
        min_photos_if_all_good=0
    )
    db.session.add(form)
    db.session.flush()
    return form


def release_dependent_tasks(work, actor_id=None):
    for dependent in work.all_dependents:
        if dependent.status == "Blocked" and dependent.dependency_satisfied:
            dependent.status = "Open"
            db.session.flush()
            log_work_event(
                dependent.id,
                "dependency_released",
                actor_id=actor_id,
                details={"dependency": work.id}
            )


def block_child_tasks(work, actor_id=None):
    for dependent in work.all_dependents:
        if dependent.status != "Closed" and work.id in dependent.dependency_ids:
            dependent.status = "Blocked"
            db.session.flush()
            log_work_event(
                dependent.id,
                "dependency_reinstated",
                actor_id=actor_id,
                details={"dependency": work.id}
            )


# ---------------------- SAFE DB REPAIR / MIGRATIONS ----------------------
def ensure_qc_columns():
    """Check and auto-add 'stage' and 'lift_type' in form_schema, and 'work_id' in submission."""
    db_path = os.path.join("instance", "eleva.db")
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # form_schema
    cur.execute("PRAGMA table_info(form_schema)")
    fs_cols = [r[1] for r in cur.fetchall()]
    added_fs = []
    if "stage" not in fs_cols:
        cur.execute("ALTER TABLE form_schema ADD COLUMN stage TEXT;")
        added_fs.append("stage")
    if "lift_type" not in fs_cols:
        cur.execute("ALTER TABLE form_schema ADD COLUMN lift_type TEXT;")
        added_fs.append("lift_type")

    # submission
    cur.execute("PRAGMA table_info(submission)")
    sub_cols = [r[1] for r in cur.fetchall()]
    added_sub = []
    if "work_id" not in sub_cols:
        cur.execute("ALTER TABLE submission ADD COLUMN work_id INTEGER;")
        added_sub.append("work_id")

    # user profile additions
    cur.execute("PRAGMA table_info(user)")
    user_cols = [r[1] for r in cur.fetchall()]
    added_user = []
    user_column_defs = {
        "first_name": "TEXT",
        "last_name": "TEXT",
        "department": "TEXT",
        "role": "TEXT",
        "mobile_number": "TEXT",
        "email": "TEXT",
        "display_picture": "TEXT",
        "active": "INTEGER DEFAULT 1",
        "session_token": "TEXT",
        "position_id": "INTEGER",
        "module_permissions_json": "TEXT"
    }
    for col, col_type in user_column_defs.items():
        if col not in user_cols:
            cur.execute(f"ALTER TABLE user ADD COLUMN {col} {col_type};")
            added_user.append(col)

    # department additions
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='department'")
    department_exists = cur.fetchone() is not None
    added_department_cols = []
    if department_exists:
        cur.execute("PRAGMA table_info(department)")
        department_cols = [r[1] for r in cur.fetchall()]
        if "branch" not in department_cols:
            cur.execute("ALTER TABLE department ADD COLUMN branch TEXT DEFAULT 'Goa';")
            cur.execute("UPDATE department SET branch = COALESCE(branch, 'Goa');")
            added_department_cols.append("branch")

    # qc_work (only attempt to alter when table exists)
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='qc_work'")
    qc_exists = cur.fetchone() is not None
    added_qc = []
    if qc_exists:
        cur.execute("PRAGMA table_info(qc_work)")
        qc_cols = [r[1] for r in cur.fetchall()]
        if "assigned_to" not in qc_cols:
            cur.execute("ALTER TABLE qc_work ADD COLUMN assigned_to INTEGER;")
            added_qc.append("assigned_to")
        if "project_id" not in qc_cols:
            cur.execute("ALTER TABLE qc_work ADD COLUMN project_id INTEGER;")
            added_qc.append("project_id")
        if "name" not in qc_cols:
            cur.execute("ALTER TABLE qc_work ADD COLUMN name TEXT;")
            added_qc.append("name")
        if "description" not in qc_cols:
            cur.execute("ALTER TABLE qc_work ADD COLUMN description TEXT;")
            added_qc.append("description")
        if "template_task_id" not in qc_cols:
            cur.execute("ALTER TABLE qc_work ADD COLUMN template_task_id INTEGER;")
            added_qc.append("template_task_id")
        if "depends_on_id" not in qc_cols:
            cur.execute("ALTER TABLE qc_work ADD COLUMN depends_on_id INTEGER;")
            added_qc.append("depends_on_id")
        if "planned_start_date" not in qc_cols:
            cur.execute("ALTER TABLE qc_work ADD COLUMN planned_start_date TEXT;")
            added_qc.append("planned_start_date")
        if "planned_duration_days" not in qc_cols:
            cur.execute("ALTER TABLE qc_work ADD COLUMN planned_duration_days INTEGER;")
            added_qc.append("planned_duration_days")
        if "milestone" not in qc_cols:
            cur.execute("ALTER TABLE qc_work ADD COLUMN milestone TEXT;")
            added_qc.append("milestone")

    # project_template_task additions
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='project_template_task'")
    template_task_exists = cur.fetchone() is not None
    added_template_cols = []
    if template_task_exists:
        cur.execute("PRAGMA table_info(project_template_task)")
        template_cols = [r[1] for r in cur.fetchall()]
        if "start_mode" not in template_cols:
            cur.execute("ALTER TABLE project_template_task ADD COLUMN start_mode TEXT DEFAULT 'immediate';")
            added_template_cols.append("start_mode")
        if "planned_start_date" not in template_cols:
            cur.execute("ALTER TABLE project_template_task ADD COLUMN planned_start_date TEXT;")
            added_template_cols.append("planned_start_date")
        if "duration_days" not in template_cols:
            cur.execute("ALTER TABLE project_template_task ADD COLUMN duration_days INTEGER;")
            added_template_cols.append("duration_days")
        if "milestone" not in template_cols:
            cur.execute("ALTER TABLE project_template_task ADD COLUMN milestone TEXT;")
            added_template_cols.append("milestone")

    # project additions
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='project'")
    project_exists = cur.fetchone() is not None
    added_project_cols = []
    if project_exists:
        cur.execute("PRAGMA table_info(project)")
        project_cols = {row[1]: (row[2] or "").upper() for row in cur.execute("PRAGMA table_info(project)")}

        if "floors" in project_cols and project_cols["floors"] not in {"TEXT", "VARCHAR", "NVARCHAR"}:
            cur.execute("ALTER TABLE project RENAME COLUMN floors TO floors_numeric;")
            project_cols = {row[1]: (row[2] or "").upper() for row in cur.execute("PRAGMA table_info(project)")}

        project_column_defs = {
            "floors": "TEXT",
            "stops": "INTEGER",
            "opening_type": "TEXT",
            "location": "TEXT",
            "handover_date": "TEXT",
            "priority": "TEXT",
            "structure_type": "TEXT",
            "cladding_type": "TEXT",
            "cabin_finish": "TEXT",
            "door_operation_type": "TEXT",
            "door_finish": "TEXT",
        }
        for col, col_type in project_column_defs.items():
            if col not in project_cols:
                cur.execute(f"ALTER TABLE project ADD COLUMN {col} {col_type};")
                added_project_cols.append(col)

        if "floors_numeric" in project_cols and "floors" in project_column_defs:
            cur.execute(
                "UPDATE project SET floors = CASE WHEN floors IS NULL OR floors = '' THEN CAST(floors_numeric AS TEXT) ELSE floors END WHERE floors_numeric IS NOT NULL;"
            )

    conn.commit()
    conn.close()

    if added_fs:
        print(f"✅ Auto-added in form_schema: {', '.join(added_fs)}")
    else:
        print("✔️ form_schema OK")

    if added_sub:
        print(f"✅ Auto-added in submission: {', '.join(added_sub)}")
    else:
        print("✔️ submission OK")

    if added_user:
        print(f"✅ Auto-added in user: {', '.join(added_user)}")
    else:
        print("✔️ user OK")

    if qc_exists:
        if added_qc:
            print(f"✅ Auto-added in qc_work: {', '.join(added_qc)}")
        else:
            print("✔️ qc_work OK")
    else:
        print("ℹ️ qc_work table did not exist prior to ensure_qc_columns")

    if department_exists:
        if added_department_cols:
            print(f"✅ Auto-added in department: {', '.join(added_department_cols)}")
        else:
            print("✔️ department OK")

    if template_task_exists:
        if added_template_cols:
            print(f"✅ Auto-added in project_template_task: {', '.join(added_template_cols)}")
        else:
            print("✔️ project_template_task OK")

    if added_project_cols:
        print(f"✅ Auto-added in project: {', '.join(added_project_cols)}")
    else:
        print("✔️ project OK")


def ensure_lift_columns():
    db_path = os.path.join("instance", "eleva.db")
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(lift)")
    lift_cols = [row[1] for row in cur.fetchall()]
    added_cols = []

    if "remarks" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN remarks TEXT;")
        added_cols.append("remarks")

    if "preferred_service_date" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN preferred_service_date DATE;")
        added_cols.append("preferred_service_date")

    if "preferred_service_time" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN preferred_service_time TEXT;")
        added_cols.append("preferred_service_time")

    if "preferred_service_day" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN preferred_service_day TEXT;")
        added_cols.append("preferred_service_day")

    if "lift_brand" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN lift_brand TEXT;")
        added_cols.append("lift_brand")

    if "amc_contract_id" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN amc_contract_id TEXT;")
        added_cols.append("amc_contract_id")

    if "building_villa_number" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN building_villa_number TEXT;")
        added_cols.append("building_villa_number")

    if "geo_location" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN geo_location TEXT;")
        added_cols.append("geo_location")

    if "amc_duration_key" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN amc_duration_key TEXT;")
        added_cols.append("amc_duration_key")

    if "preferred_service_days_json" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN preferred_service_days_json TEXT;")
        added_cols.append("preferred_service_days_json")

    if "lifetime_metrics_json" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN lifetime_metrics_json TEXT;")
        added_cols.append("lifetime_metrics_json")

    if "amc_contacts_json" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN amc_contacts_json TEXT;")
        added_cols.append("amc_contacts_json")

    if "timeline_entries_json" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN timeline_entries_json TEXT;")
        added_cols.append("timeline_entries_json")

    if "service_schedule_json" not in lift_cols:
        cur.execute("ALTER TABLE lift ADD COLUMN service_schedule_json TEXT;")
        added_cols.append("service_schedule_json")

    conn.commit()
    conn.close()

    if added_cols:
        print(f"✅ Auto-added in lift: {', '.join(added_cols)}")
    else:
        print("✔️ lift OK")


def ensure_service_route_columns():
    db_path = os.path.join("instance", "eleva.db")
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(service_route)")
    route_cols = {row[1] for row in cur.fetchall()}
    added_cols = []

    if "branch" not in route_cols:
        cur.execute("ALTER TABLE service_route ADD COLUMN branch TEXT;")
        added_cols.append("branch")

    before_update = conn.total_changes
    cur.execute(
        """
        UPDATE service_route
        SET branch = CASE
            WHEN lower(state) = 'goa' THEN 'Goa'
            WHEN lower(state) = 'maharashtra' THEN 'Mumbai'
            WHEN lower(state) = 'karnataka' THEN 'Mumbai'
            ELSE branch
        END
        WHERE branch IS NULL
        """
    )
    branch_backfill_changes = conn.total_changes - before_update

    conn.commit()
    conn.close()

    if added_cols:
        print(f"✅ Auto-added in service_route: {', '.join(added_cols)}")
    elif branch_backfill_changes:
        print("♻️ service_route branches backfilled")
    else:
        print("✔️ service_route OK")


def ensure_customer_columns():
    db_path = os.path.join("instance", "eleva.db")
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(customer)")
    customer_cols = {row[1] for row in cur.fetchall()}
    added_cols = []

    column_defs = [
        ("office_address_line1", "TEXT"),
        ("office_address_line2", "TEXT"),
        ("office_city", "TEXT"),
        ("office_state", "TEXT"),
        ("office_pincode", "TEXT"),
        ("office_country", "TEXT"),
        ("external_customer_id", "TEXT"),
    ]

    for column_name, column_type in column_defs:
        if column_name not in customer_cols:
            cur.execute(f"ALTER TABLE customer ADD COLUMN {column_name} {column_type};")
            added_cols.append(column_name)

    conn.commit()
    conn.close()

    if added_cols:
        print(f"✅ Auto-added in customer: {', '.join(added_cols)}")
    else:
        print("✔️ customer OK")


def ensure_tables():
    """Ensure all known tables exist. Creates them if missing."""
    created_tables = []
    inspector = inspect(db.engine)
    try:
        existing_tables = set(inspector.get_table_names())
    except OperationalError:
        # Database file might be missing – create all tables fresh.
        db.create_all()
        existing_tables = set(inspect(db.engine).get_table_names())

    models = [
        Department.__table__,
        Position.__table__,
        User.__table__,
        Project.__table__,
        ProjectComment.__table__,
        FormSchema.__table__,
        Submission.__table__,
        ProjectTemplate.__table__,
        ProjectTemplateTask.__table__,
        ProjectTemplateTaskDependency.__table__,
        TaskTemplate.__table__,
        SalesClient.__table__,
        SalesOpportunity.__table__,
        SalesActivity.__table__,
        SalesOpportunityComment.__table__,
        SalesOpportunityFile.__table__,
        SalesOpportunityEngagement.__table__,
        SalesOpportunityItem.__table__,
        ServiceRoute.__table__,
        Customer.__table__,
        CustomerComment.__table__,
        Lift.__table__,
        LiftFile.__table__,
        LiftComment.__table__,
        DropdownOption.__table__,
        QCWork.__table__,
        QCWorkDependency.__table__,
        QCWorkComment.__table__,
        QCWorkLog.__table__,
    ]

    for table in models:
        if table.name not in existing_tables:
            table.create(bind=db.engine, checkfirst=True)
            created_tables.append(table.name)

    if created_tables:
        print(f"✅ Created missing tables: {', '.join(created_tables)}")


def ensure_project_comment_table():
    """Backfill the project_comment table for legacy databases."""
    try:
        inspector = inspect(db.engine)
        table_names = {name.lower() for name in inspector.get_table_names()}
    except OperationalError:
        table_names = set()

    if "project_comment" in table_names:
        return

    # Use SQLAlchemy metadata to create the table if it is missing. checkfirst
    # guards against race conditions with ensure_tables.
    ProjectComment.__table__.create(bind=db.engine, checkfirst=True)
    print("✅ Backfilled missing table: project_comment")


def bootstrap_db():
    ensure_tables()
    ensure_project_comment_table()
    ensure_qc_columns()    # adds missing columns safely
    ensure_lift_columns()
    ensure_service_route_columns()
    ensure_customer_columns()
    ensure_dropdown_options_seed()
    purge_legacy_demo_records()

    default_users = [("user1", "pass"), ("user2", "pass"), ("admin", "admin")]
    for u, p in default_users:
        if not User.query.filter_by(username=u).first():
            new_user = User(username=u, password=p)
            new_user.issue_session_token()
            db.session.add(new_user)

    admin_user = User.query.filter_by(username="admin").first()
    if admin_user and not admin_user.role:
        admin_user.role = "Admin"

    # Ensure legacy accounts have activation flag and session tokens
    for user in User.query.filter(or_(User.session_token.is_(None), User.session_token == "")).all():
        user.issue_session_token()
    for user in User.query.filter(User.active.is_(None)).all():
        user.active = True
    for user in User.query.filter(or_(User.module_permissions_json.is_(None), User.module_permissions_json == "")).all():
        user.module_permissions_json = "{}"

    get_or_create_default_task_form()

    if not FormSchema.query.filter_by(name="QC - New Installation").first():
        sample_schema = [
            {"label": "Lift Cabin Condition", "type": "select", "required": True, "options": ["OK", "Not OK", "Need Client Input"], "photo_required_if_ng": True},
            {"label": "Machine Room Cleanliness", "type": "select", "required": True, "options": ["OK", "Not OK", "Need Client Input"], "photo_required_if_ng": True},
            {"label": "Lift Shaft Obstruction", "type": "select", "required": True, "options": ["OK", "Not OK", "Need Client Input"], "photo_required_if_ng": True},
            {"label": "General Remarks", "type": "textarea", "required": False},
        ]
        fs = FormSchema(
            name="QC - New Installation",
            schema_json=json.dumps(sample_schema, ensure_ascii=False),
            min_photos_if_all_good=0,
            stage="Template QC",
            lift_type="MRL"
        )
        db.session.add(fs)

    if not ProjectTemplate.query.filter_by(name="NI Project").first():
        ni_template = ProjectTemplate(
            name="NI Project",
            description="Baseline new installation delivery with sequential QC checks.",
            created_by=admin_user.id if admin_user else None
        )
        db.session.add(ni_template)

    if SalesClient.query.count() == 0:
        pass

    if ServiceRoute.query.count() == 0:
        default_routes = [
            ("Goa", "Goa"),
            ("Maharashtra", "Mumbai"),
            ("Karnataka", "Mumbai"),
        ]
        for state_name, branch_name in default_routes:
            db.session.add(ServiceRoute(state=state_name, branch=branch_name))
        db.session.flush()

    if Customer.query.count() == 0:
        pass

    if Lift.query.count() == 0:
        pass

    db.session.commit()
    synchronize_dependency_links()
# -----------------------------------------------------------------------


@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    return render_template("index.html", category_label=None)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        user = User.query.filter_by(username=username).first()
        if user and user.password == password:
            if not user.is_active:
                flash("Your account is deactivated. Please contact an administrator.", "error")
            else:
                user.issue_session_token()
                db.session.commit()
                login_user(user)
                session["session_token"] = user.session_token
                flash("Welcome back!", "success")
                return redirect(url_for("dashboard"))
        else:
            flash("Invalid credentials", "error")
    return render_template("login.html", category_label=None)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    session.pop("session_token", None)
    flash("Logged out", "info")
    return redirect(url_for("index"))


@app.route("/switch-user", methods=["POST"])
@login_required
def switch_user():
    target_id = request.form.get("user_id")
    redirect_to = request.form.get("next") or request.referrer or url_for("dashboard")

    if not target_id:
        flash("Select a user to switch.", "error")
        return redirect(redirect_to)

    user = User.query.filter_by(id=target_id).first()
    if not user:
        flash("Unable to find the selected user.", "error")
        return redirect(redirect_to)

    if not user.is_active:
        flash("The selected user is deactivated.", "error")
        return redirect(redirect_to)

    user.issue_session_token()
    db.session.commit()

    login_user(user)
    session["session_token"] = user.session_token
    flash(f"Switched to {user.display_name}", "info")
    return redirect(redirect_to)


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    if request.method == "POST":
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()
        remove_avatar = request.form.get("remove_avatar") == "1"

        current_user.first_name = first_name or None
        current_user.last_name = last_name or None

        file = request.files.get("display_picture")
        if remove_avatar:
            current_user.display_picture = None
        elif file and file.filename:
            if not allowed_file(file.filename, kind="photo"):
                flash("Please upload a PNG, JPG, JPEG or WEBP image for the display picture.", "error")
                return redirect(url_for("profile"))
            fname = secure_filename(file.filename)
            dest_name = f"avatar_{current_user.id}_{int(datetime.datetime.utcnow().timestamp())}_{fname}"
            dest_path = os.path.join(app.config["UPLOAD_FOLDER"], dest_name)
            file.save(dest_path)
            rel_path = os.path.relpath(dest_path, "static") if dest_path.startswith("static") else os.path.join("uploads", dest_name)
            current_user.display_picture = rel_path.replace("\\", "/")

        db.session.commit()
        flash("Profile updated.", "success")
        return redirect(url_for("profile"))

    return render_template("profile.html")


@app.route("/settings")
@login_required
def settings():
    tab = (request.args.get("tab") or "admin").lower()
    allowed_tabs = {"admin", "account", "display", "modules"}
    active_tab = tab if tab in allowed_tabs else "admin"

    users = []
    departments = []
    positions = []
    department_options = []
    position_options = []

    service_routes = ServiceRoute.query.order_by(
        func.lower(ServiceRoute.state), func.lower(ServiceRoute.branch)
    ).all()
    dropdown_options = get_dropdown_options_map()

    if current_user.is_admin:
        departments = sorted(
            Department.query.order_by(Department.name.asc()).all(),
            key=lambda d: (d.full_name or "").lower(),
        )
        positions = sorted(
            Position.query.order_by(Position.title.asc()).all(),
            key=lambda p: (p.display_label or "").lower(),
        )
        department_options = departments
        position_options = positions
        users = User.query.order_by(User.username.asc()).all()

    return render_template(
        "settings.html",
        active_tab=active_tab,
        allowed_tabs=sorted(allowed_tabs),
        users=users,
        departments=departments,
        department_options=department_options,
        department_branches=DEPARTMENT_BRANCHES,
        positions=positions,
        position_options=position_options,
        support_categories=CUSTOMER_SUPPORT_CATEGORIES,
        support_channels=CUSTOMER_SUPPORT_CHANNELS,
        support_sla_presets=CUSTOMER_SUPPORT_SLA_PRESETS,
        service_routes=service_routes,
        dropdown_options=dropdown_options,
        dropdown_meta=DROPDOWN_FIELD_DEFINITIONS,
    )


@app.route("/settings/service/routes/create", methods=["POST"])
@login_required
def settings_service_route_create():
    if not current_user.is_admin:
        abort(403)

    route_name = clean_str(request.form.get("route_name") or request.form.get("state"))
    if not route_name:
        flash("Route name is required to add a route.", "error")
        return redirect(url_for("settings", tab="modules"))

    branch_value, error = validate_branch(request.form.get("branch"), required=True)
    if error:
        flash(error, "error")
        return redirect(url_for("settings", tab="modules"))

    existing = ServiceRoute.query.filter(func.lower(ServiceRoute.state) == route_name.lower()).first()
    if existing:
        flash("A route with that name already exists.", "error")
        return redirect(url_for("settings", tab="modules"))

    db.session.add(ServiceRoute(state=route_name, branch=branch_value))
    db.session.commit()
    flash(f"Route '{route_name}' for branch {branch_value} added.", "success")
    return redirect(url_for("settings", tab="modules"))


@app.route("/settings/service/routes/<int:route_id>/update", methods=["POST"])
@login_required
def settings_service_route_update(route_id):
    if not current_user.is_admin:
        abort(403)

    route = db.session.get(ServiceRoute, route_id)
    if not route:
        flash("Route not found.", "error")
        return redirect(url_for("settings", tab="modules"))

    route_name = clean_str(request.form.get("route_name") or request.form.get("state"))
    if not route_name:
        flash("Route name cannot be empty.", "error")
        return redirect(url_for("settings", tab="modules"))

    branch_value, error = validate_branch(request.form.get("branch"), required=True)
    if error:
        flash(error, "error")
        return redirect(url_for("settings", tab="modules"))

    duplicate = (
        ServiceRoute.query.filter(func.lower(ServiceRoute.state) == route_name.lower(), ServiceRoute.id != route.id)
        .first()
    )
    if duplicate:
        flash("Another route already uses that name.", "error")
        return redirect(url_for("settings", tab="modules"))

    route.state = route_name
    route.branch = branch_value
    db.session.commit()

    flash("Route updated.", "success")
    return redirect(url_for("settings", tab="modules"))


@app.route("/settings/service/routes/<int:route_id>/delete", methods=["POST"])
@login_required
def settings_service_route_delete(route_id):
    if not current_user.is_admin:
        abort(403)

    route = db.session.get(ServiceRoute, route_id)
    if not route:
        flash("Route not found.", "error")
        return redirect(url_for("settings", tab="modules"))

    db.session.delete(route)
    db.session.commit()
    flash(f"Route '{route.display_name}' removed.", "success")
    return redirect(url_for("settings", tab="modules"))


@app.route("/admin/reset-workspace", methods=["POST"])
@login_required
def admin_reset_workspace():
    if not current_user.is_admin:
        abort(403)

    redirect_target = url_for("settings", tab="admin")
    password = (request.form.get("confirm_password") or "").strip()

    if not password:
        flash("Enter your password to confirm the workspace reset.", "error")
        return redirect(redirect_target)

    if password != current_user.password:
        flash("Password verification failed. Reset cancelled.", "error")
        return redirect(redirect_target)

    try:
        summary = reset_workspace_data()
        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception("Workspace reset failed")
        flash("An unexpected error occurred while resetting the workspace.", "error")
        return redirect(redirect_target)

    removed_parts = []
    labels = [
        ("customers", "customer"),
        ("lifts", "lift"),
        ("opportunities", "opportunity"),
        ("sales_clients", "sales client"),
        ("projects", "project"),
        ("submissions", "submission"),
        ("qc_tasks", "QC task"),
        ("service_complaints", "complaint"),
        ("service_contracts", "contract"),
        ("service_parts_entries", "parts/material entry"),
        ("srt_sites", "SRT site"),
    ]

    for key, label in labels:
        count = summary.get(key, 0)
        if count:
            plural = "s" if count != 1 else ""
            removed_parts.append(f"{count} {label}{plural}")

    details = ", ".join(removed_parts) if removed_parts else "no records"
    flash(f"Workspace reset complete. Removed {details}.", "success")
    return redirect(redirect_target)


def _get_dropdown_definition_or_404(field_key):
    definition = DROPDOWN_FIELD_DEFINITIONS.get(field_key)
    if not definition:
        abort(404)
    return definition


@app.route("/settings/dropdowns/<field_key>/options", methods=["POST"])
@login_required
def settings_dropdown_option_create(field_key):
    if not current_user.is_admin:
        abort(403)
    definition = _get_dropdown_definition_or_404(field_key)
    label = clean_str(request.form.get("label"))
    value = clean_str(request.form.get("value")) if definition.get("value_editable") else None
    if not label:
        flash("Option label cannot be empty.", "error")
        return redirect(url_for("settings", tab="modules"))
    if not definition.get("value_editable"):
        value = label
    elif not value:
        value = label
    existing = (
        DropdownOption.query.filter(
            DropdownOption.field_key == field_key,
            func.lower(DropdownOption.label) == label.lower(),
        )
        .first()
    )
    if existing:
        flash("An option with that label already exists.", "error")
        return redirect(url_for("settings", tab="modules"))
    max_order = (
        db.session.query(func.coalesce(func.max(DropdownOption.order_index), -1))
        .filter(DropdownOption.field_key == field_key)
        .scalar()
    )
    option = DropdownOption(
        field_key=field_key,
        label=label,
        value=value,
        order_index=max_order + 1,
    )
    db.session.add(option)
    db.session.commit()
    flash("Dropdown option added.", "success")
    return redirect(url_for("settings", tab="modules"))


@app.route("/settings/dropdowns/<field_key>/options/<int:option_id>", methods=["POST"])
@login_required
def settings_dropdown_option_update(field_key, option_id):
    if not current_user.is_admin:
        abort(403)
    definition = _get_dropdown_definition_or_404(field_key)
    option = DropdownOption.query.filter_by(field_key=field_key, id=option_id).first()
    if not option:
        flash("Option not found.", "error")
        return redirect(url_for("settings", tab="modules"))
    label = clean_str(request.form.get("label"))
    value = clean_str(request.form.get("value")) if definition.get("value_editable") else option.label
    if not label:
        flash("Option label cannot be empty.", "error")
        return redirect(url_for("settings", tab="modules"))
    if not definition.get("value_editable") or not value:
        value = label
    duplicate = (
        DropdownOption.query.filter(
            DropdownOption.field_key == field_key,
            func.lower(DropdownOption.label) == label.lower(),
            DropdownOption.id != option.id,
        )
        .first()
    )
    if duplicate:
        flash("Another option already uses that label.", "error")
        return redirect(url_for("settings", tab="modules"))
    option.label = label
    option.value = value
    db.session.commit()
    flash("Option updated.", "success")
    return redirect(url_for("settings", tab="modules"))


@app.route("/settings/dropdowns/<field_key>/options/<int:option_id>/delete", methods=["POST"])
@login_required
def settings_dropdown_option_delete(field_key, option_id):
    if not current_user.is_admin:
        abort(403)
    _get_dropdown_definition_or_404(field_key)
    option = DropdownOption.query.filter_by(field_key=field_key, id=option_id).first()
    if not option:
        flash("Option not found.", "error")
        return redirect(url_for("settings", tab="modules"))
    db.session.delete(option)
    db.session.commit()
    flash("Option removed.", "success")
    return redirect(url_for("settings", tab="modules"))


@app.route("/settings/dropdowns/<field_key>/reorder", methods=["POST"])
@login_required
def settings_dropdown_option_reorder(field_key):
    if not current_user.is_admin:
        abort(403)
    _get_dropdown_definition_or_404(field_key)
    payload = request.get_json(silent=True) or {}
    order_ids = payload.get("order")
    if not isinstance(order_ids, list):
        return jsonify({"status": "error", "message": "Invalid payload."}), 400
    options = {
        option.id: option
        for option in DropdownOption.query.filter_by(field_key=field_key).all()
    }
    for index, option_id in enumerate(order_ids):
        try:
            option_id = int(option_id)
        except (TypeError, ValueError):
            continue
        option = options.get(option_id)
        if option:
            option.order_index = index
    db.session.commit()
    return jsonify({"status": "ok"})


@app.route("/sales")
@login_required
def sales_home():
    _module_visibility_required("sales")
    today = datetime.date.today()
    month_start = today.replace(day=1)
    next_month_start = (month_start + datetime.timedelta(days=32)).replace(day=1)

    duration = request.args.get("duration", "month")
    period_label = {
        "month": "This Month",
        "quarter": "This Quarter",
        "ytd": "Year to Date",
    }.get(duration, "This Month")
    period_descriptions = {
        "month": "Current month snapshot",
        "quarter": "Performance for the current quarter",
        "ytd": "Year-to-date performance overview",
    }

    if duration == "quarter":
        quarter_index = (today.month - 1) // 3
        start_month = quarter_index * 3 + 1
        period_start = today.replace(month=start_month, day=1)
        next_quarter_month = start_month + 3
        if next_quarter_month > 12:
            period_end = today.replace(year=today.year + 1, month=1, day=1)
        else:
            period_end = today.replace(month=next_quarter_month, day=1)
    elif duration == "ytd":
        period_start = today.replace(month=1, day=1)
        period_end = today.replace(year=today.year + 1, month=1, day=1)
    else:
        duration = "month"
        period_start = month_start
        period_end = next_month_start
    period_description = period_descriptions.get(duration, period_descriptions["month"])

    closed_won_clause = func.lower(SalesOpportunity.stage).like("closed won%")
    period_filters = [
        closed_won_clause,
        SalesOpportunity.updated_at >= period_start,
        SalesOpportunity.updated_at < period_end,
    ]

    won_count = SalesOpportunity.query.filter(*period_filters).count()
    won_value = (
        db.session.query(func.coalesce(func.sum(SalesOpportunity.amount), 0.0))
        .filter(*period_filters)
        .scalar()
        or 0.0
    )

    closed_lost_clause = func.lower(SalesOpportunity.stage).like("closed lost%")
    period_closed_total = (
        SalesOpportunity.query
        .filter(
            or_(closed_won_clause, closed_lost_clause),
            SalesOpportunity.updated_at >= period_start,
            SalesOpportunity.updated_at < period_end,
        )
        .count()
    )
    win_rate = (won_count / period_closed_total * 100) if period_closed_total else 0.0
    average_deal_value = (won_value / won_count) if won_count else 0.0

    open_pipeline_clause = ~func.lower(SalesOpportunity.stage).like("closed%")
    open_deals_count = SalesOpportunity.query.filter(open_pipeline_clause).count()
    open_pipeline_value = (
        db.session.query(func.coalesce(func.sum(SalesOpportunity.amount), 0.0))
        .filter(open_pipeline_clause)
        .scalar()
        or 0.0
    )

    stage_rows = (
        db.session.query(
            SalesOpportunity.stage,
            func.count(SalesOpportunity.id),
            func.coalesce(func.sum(SalesOpportunity.amount), 0.0),
        )
        .group_by(SalesOpportunity.stage)
        .order_by(func.count(SalesOpportunity.id).desc())
        .all()
    )
    stage_distribution = [
        {
            "stage": stage or "(No Stage)",
            "count": count,
            "value": float(total or 0.0),
        }
        for stage, count, total in stage_rows
    ]

    previous_months = []
    cursor = month_start
    for _ in range(3):
        cursor = (cursor - datetime.timedelta(days=1)).replace(day=1)
        next_cursor = (cursor + datetime.timedelta(days=32)).replace(day=1)
        total_value = (
            db.session.query(func.coalesce(func.sum(SalesOpportunity.amount), 0.0))
            .filter(
                closed_won_clause,
                SalesOpportunity.updated_at >= cursor,
                SalesOpportunity.updated_at < next_cursor,
            )
            .scalar()
            or 0.0
        )
        previous_months.append(
            {
                "label": cursor.strftime("%b %Y"),
                "total": float(total_value or 0.0),
            }
        )
    previous_months.reverse()

    team_rows = (
        db.session.query(
            SalesOpportunity.owner_id,
            func.count(SalesOpportunity.id),
            func.coalesce(func.sum(SalesOpportunity.amount), 0.0),
        )
        .filter(*period_filters)
        .group_by(SalesOpportunity.owner_id)
        .all()
    )

    team_breakdown = []
    owner_cache = {}
    for owner_id, deal_count, total_value in team_rows:
        if owner_id:
            owner = owner_cache.get(owner_id)
            if owner is None:
                owner = db.session.get(User, owner_id)
                owner_cache[owner_id] = owner
            owner_name = owner.display_name if owner else "Unknown"
        else:
            owner_name = "Unassigned"
        team_breakdown.append(
            {
                "owner": owner_name,
                "deals": deal_count,
                "value": float(total_value or 0.0),
            }
        )
    team_breakdown.sort(key=lambda row: row["value"], reverse=True)

    now = datetime.datetime.utcnow()
    due_activities = (
        SalesOpportunityEngagement.query
        .filter(
            SalesOpportunityEngagement.scheduled_for.isnot(None),
            SalesOpportunityEngagement.scheduled_for <= now,
        )
        .order_by(SalesOpportunityEngagement.scheduled_for.asc())
        .limit(10)
        .all()
    )

    return render_template(
        "sales/dashboard.html",
        won_count=won_count,
        won_value=won_value,
        win_rate=win_rate,
        average_deal_value=average_deal_value,
        open_deals_count=open_deals_count,
        open_pipeline_value=open_pipeline_value,
        stage_distribution=stage_distribution,
        previous_months=previous_months,
        team_breakdown=team_breakdown,
        due_activities=due_activities,
        format_currency=format_currency,
        selected_duration=duration,
        period_label=period_label,
        period_description=period_description,
    )


@app.route("/sales/clients")
@login_required
def sales_clients():
    _module_visibility_required("sales")
    clients = (
        SalesClient.query
        .order_by(SalesClient.display_name.asc())
        .all()
    )
    return render_template(
        "sales/clients_list.html",
        clients=clients,
        pipeline_map=SALES_PIPELINES,
        temperature_choices=SALES_TEMPERATURES,
        lifecycle_options=SALES_CLIENT_LIFECYCLE_STAGES,
    )


@app.route("/sales/clients/export")
@login_required
def sales_clients_export():
    _module_visibility_required("sales")

    clients = SalesClient.query.order_by(SalesClient.display_name.asc()).all()
    timestamp = datetime.datetime.utcnow().strftime("%Y%m%d")
    filename = f"sales_clients_{timestamp}.xlsx"

    if OPENPYXL_AVAILABLE:
        workbook = build_sales_client_export_workbook(clients)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    csv_rows = [_sales_client_upload_row(client) for client in clients]
    csv_output = _build_csv_output(SALES_CLIENT_UPLOAD_HEADERS, csv_rows)
    return send_file(
        csv_output,
        as_attachment=True,
        download_name=filename.replace(".xlsx", ".csv"),
        mimetype="text/csv",
    )


@app.route("/sales/clients/upload-template")
@login_required
def sales_clients_upload_template():
    _module_visibility_required("sales")

    filename = f"sales_clients_template_{datetime.datetime.utcnow():%Y%m%d}.xlsx"

    if OPENPYXL_AVAILABLE:
        workbook = build_sales_client_upload_workbook()
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    csv_output = _build_csv_output(
        SALES_CLIENT_UPLOAD_HEADERS,
        [["" for _ in SALES_CLIENT_UPLOAD_HEADERS]],
    )
    return send_file(
        csv_output,
        as_attachment=True,
        download_name=filename.replace(".xlsx", ".csv"),
        mimetype="text/csv",
    )


@app.route("/sales/clients/upload", methods=["POST"])
@login_required
def sales_clients_upload():
    _module_visibility_required("sales")

    upload = request.files.get("client_upload_file")
    if not upload or not upload.filename:
        flash("Select an Excel or CSV file to upload.", "error")
        return redirect(url_for("sales_clients"))

    filename = (upload.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv")):
        flash("Upload a .xlsx or .csv file exported from the clients template.", "error")
        return redirect(url_for("sales_clients"))

    try:
        header_cells, data_rows = _extract_tabular_upload(
            upload, sheet_name=SALES_CLIENT_TEMPLATE_SHEET_NAME
        )
    except MissingDependencyError:
        flash(OPENPYXL_MISSING_MESSAGE, "error")
        return redirect(url_for("sales_clients"))
    except ValueError:
        flash("Upload a .xlsx or .csv file exported from the clients template.", "error")
        return redirect(url_for("sales_clients"))
    except Exception:
        flash("Could not read the uploaded file. Ensure it is a valid spreadsheet.", "error")
        return redirect(url_for("sales_clients"))

    header_map = {}
    for idx, header in enumerate(header_cells or []):
        header_label = stringify_cell(header)
        if header_label:
            header_map[header_label] = idx

    if "Display Name" not in header_map:
        flash("The uploaded sheet must include the 'Display Name' column.", "error")
        return redirect(url_for("sales_clients"))

    users = User.query.all()
    user_lookup = {}
    for user in users:
        email_value = clean_str(getattr(user, "email", None))
        if email_value:
            user_lookup[email_value.lower()] = user

    existing_clients = SalesClient.query.all()
    client_lookup = {
        (client.display_name or "").strip().lower(): client
        for client in existing_clients
        if client.display_name
    }

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for row_index, row in enumerate(data_rows, start=2):
        row_data = {}
        for header, column_index in header_map.items():
            if column_index < len(row):
                row_data[header] = stringify_cell(row[column_index])
            else:
                row_data[header] = None

        name = clean_str(row_data.get("Display Name"))
        if not name:
            skipped += 1
            continue

        email = clean_str(row_data.get("Email"))
        phone = clean_str(row_data.get("Phone"))
        tag = clean_str(row_data.get("Tag"))
        category = clean_str(row_data.get("Category")) or "Individual"
        lifecycle = clean_str(row_data.get("Lifecycle Stage"))
        description = clean_str(row_data.get("Description"))
        owner_email = clean_str(row_data.get("Owner Email"))

        row_errors = []
        owner_user = None
        if owner_email:
            candidate = user_lookup.get(owner_email.lower())
            if not candidate or not candidate.can_be_assigned_module("sales"):
                row_errors.append(
                    f"Row {row_index}: Owner email '{owner_email}' is not linked to a sales user."
                )
            else:
                owner_user = candidate

        if lifecycle and lifecycle not in SALES_CLIENT_LIFECYCLE_STAGES:
            row_errors.append(
                f"Row {row_index}: Lifecycle Stage '{lifecycle}' is not valid."
            )

        if row_errors:
            errors.extend(row_errors)
            skipped += 1
            continue

        client = client_lookup.get(name.lower())
        is_new = client is None
        if is_new:
            client = SalesClient(display_name=name)
            client.lifecycle_stage = lifecycle or (
                SALES_CLIENT_LIFECYCLE_STAGES[0]
                if SALES_CLIENT_LIFECYCLE_STAGES
                else None
            )
            client.owner = owner_user or current_user
            db.session.add(client)
            db.session.flush()
            log_sales_activity(
                "client",
                client.id,
                "Client created via bulk upload",
                actor=current_user,
            )
            client_lookup[name.lower()] = client
            created += 1
        else:
            updated += 1

        client.company_name = clean_str(row_data.get("Company Name"))
        client.email = email
        client.phone = phone
        client.tag = tag
        client.category = category
        if owner_user:
            client.owner = owner_user
        if lifecycle:
            client.lifecycle_stage = lifecycle
        client.description = description

    if created or updated:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            flash("Could not save the uploaded clients due to a database error.", "error")
            return redirect(url_for("sales_clients"))

    summary_bits = []
    if created:
        summary_bits.append(f"{created} created")
    if updated:
        summary_bits.append(f"{updated} updated")
    if skipped:
        summary_bits.append(f"{skipped} skipped")

    if summary_bits:
        flash("Client import complete: " + ", ".join(summary_bits) + ".", "success")
    else:
        flash("No client rows were imported.", "warning")

    if errors:
        preview = errors[:5]
        more = len(errors) - len(preview)
        message = "\n".join(preview)
        if more > 0:
            message += f"\n…and {more} more issue(s)."
        flash(message, "warning")

    return redirect(url_for("sales_clients"))


@app.route("/sales/clients/create", methods=["POST"])
@login_required
def sales_clients_create():
    _module_visibility_required("sales")
    name = (request.form.get("display_name") or "").strip()
    if not name:
        flash("Client name is required.", "error")
        return redirect(url_for("sales_clients"))

    lifecycle_value = normalize_lifecycle_stage(request.form.get("lifecycle_stage"))
    if lifecycle_value is None and SALES_CLIENT_LIFECYCLE_STAGES:
        lifecycle_value = SALES_CLIENT_LIFECYCLE_STAGES[0]

    client = SalesClient(
        display_name=name,
        company_name=(request.form.get("company_name") or "").strip() or None,
        email=(request.form.get("email") or "").strip() or None,
        phone=(request.form.get("phone") or "").strip() or None,
        category=(request.form.get("category") or "Individual").strip() or "Individual",
        description=(request.form.get("description") or "").strip() or None,
    )

    client.lifecycle_stage = lifecycle_value
    client.owner = current_user

    db.session.add(client)
    db.session.flush()
    log_sales_activity("client", client.id, "Client created")
    db.session.commit()
    flash(f"Client '{client.display_name}' created.", "success")
    return redirect(url_for("sales_client_detail", client_id=client.id))


@app.route("/sales/clients/<int:client_id>", methods=["GET", "POST"])
@login_required
def sales_client_detail(client_id):
    _module_visibility_required("sales")
    client = db.session.get(SalesClient, client_id)
    if not client:
        flash("Client not found.", "error")
        return redirect(url_for("sales_clients"))

    if request.method == "POST":
        action = request.form.get("form_action") or "update"
        if action == "update":
            client.display_name = (request.form.get("display_name") or "").strip() or client.display_name
            client.company_name = (request.form.get("company_name") or "").strip() or None
            client.email = (request.form.get("email") or "").strip() or None
            client.phone = (request.form.get("phone") or "").strip() or None
            client.category = (request.form.get("category") or "").strip() or "Individual"
            lifecycle_value = normalize_lifecycle_stage(request.form.get("lifecycle_stage"))
            client.lifecycle_stage = lifecycle_value
            client.description = (request.form.get("description") or "").strip() or None

            owner_id_raw = request.form.get("owner_id")
            if owner_id_raw:
                try:
                    owner_candidate = db.session.get(User, int(owner_id_raw))
                except (TypeError, ValueError):
                    owner_candidate = None
                if owner_candidate and not owner_candidate.can_be_assigned_module("sales"):
                    flash("The selected owner cannot be assigned to Sales records.", "error")
                    return redirect(url_for("sales_client_detail", client_id=client.id))
                client.owner = owner_candidate
            else:
                client.owner = None

            log_sales_activity("client", client.id, "Client updated", actor=current_user)
            db.session.commit()
            flash("Client details updated.", "success")
            return redirect(url_for("sales_client_detail", client_id=client.id))

        elif action == "add_note":
            note_title = (request.form.get("note_title") or "").strip() or "Timeline update"
            note_body = (request.form.get("note_body") or "").strip() or None
            log_sales_activity("client", client.id, note_title, notes=note_body)
            db.session.commit()
            flash("Timeline note added.", "success")
            return redirect(url_for("sales_client_detail", client_id=client.id))

    activities = (
        SalesActivity.query
        .filter_by(parent_type="client", parent_id=client.id)
        .order_by(SalesActivity.created_at.desc())
        .all()
    )
    owners = get_assignable_users_for_module("sales", order_by="name")
    open_opportunities = [opp for opp in client.opportunities if not opp.is_closed]
    all_clients = SalesClient.query.order_by(SalesClient.display_name.asc()).all()
    return render_template(
        "sales/client_detail.html",
        client=client,
        owners=owners,
        activities=activities,
        open_opportunities=open_opportunities,
        pipeline_map=SALES_PIPELINES,
        opportunity_clients=all_clients,
        temperature_choices=SALES_TEMPERATURES,
        lifecycle_options=SALES_CLIENT_LIFECYCLE_STAGES,
    )


@app.route("/sales/opportunities/<pipeline_key>")
@login_required
def sales_opportunities_pipeline(pipeline_key):
    _module_visibility_required("sales")
    pipeline_key = (pipeline_key or "lift").lower()
    if pipeline_key not in SALES_PIPELINES:
        pipeline_key = "lift"
    config = get_pipeline_config(pipeline_key)
    stages = list(config["stages"])
    opportunities = (
        SalesOpportunity.query
        .filter(SalesOpportunity.pipeline == pipeline_key)
        .order_by(SalesOpportunity.stage.asc(), SalesOpportunity.updated_at.desc())
        .all()
    )

    grouped = {stage: [] for stage in stages}
    for opp in opportunities:
        grouped.setdefault(opp.stage, []).append(opp)

    for stage_list in grouped.values():
        stage_list.sort(key=lambda o: o.updated_at or o.created_at, reverse=True)

    owners = get_assignable_users_for_module("sales", order_by="name")
    clients = SalesClient.query.order_by(SalesClient.display_name.asc()).all()
    temperature_choices = SALES_TEMPERATURES
    total_opportunities = sum(len(items) for items in grouped.values())

    stage_totals_raw = {}
    stage_currencies = {}
    stage_totals_display = {}
    for stage in stages:
        opportunities_in_stage = grouped.get(stage, [])
        total_amount = sum(opp.amount for opp in opportunities_in_stage if opp.amount is not None)
        currency = next((opp.currency for opp in opportunities_in_stage if opp.currency), "₹")
        stage_totals_raw[stage] = total_amount
        stage_currencies[stage] = currency
        stage_totals_display[stage] = format_currency(total_amount if opportunities_in_stage else 0, currency)

    return render_template(
        "sales/opportunity_board.html",
        pipeline_key=pipeline_key,
        pipeline_config=config,
        stages=stages,
        grouped=grouped,
        owners=owners,
        clients=clients,
        temperature_choices=temperature_choices,
        pipeline_map=SALES_PIPELINES,
        total_opportunities=total_opportunities,
        stage_totals=stage_totals_display,
        stage_totals_raw=stage_totals_raw,
        stage_currencies=stage_currencies,
    )


@app.route("/sales/opportunities/<pipeline_key>/export")
@login_required
def sales_opportunities_export(pipeline_key):
    _module_visibility_required("sales")

    pipeline_key = (pipeline_key or "lift").lower()
    if pipeline_key not in SALES_PIPELINES:
        pipeline_key = "lift"

    opportunities = (
        SalesOpportunity.query
        .filter(SalesOpportunity.pipeline == pipeline_key)
        .order_by(SalesOpportunity.stage.asc(), SalesOpportunity.updated_at.desc())
        .all()
    )

    timestamp = datetime.datetime.utcnow().strftime("%Y%m%d")
    filename = f"sales_opportunities_{pipeline_key}_{timestamp}.xlsx"

    if OPENPYXL_AVAILABLE:
        workbook = build_sales_opportunity_export_workbook(opportunities)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    csv_rows = [_sales_opportunity_upload_row(opp) for opp in opportunities]
    csv_output = _build_csv_output(SALES_OPPORTUNITY_UPLOAD_HEADERS, csv_rows)
    return send_file(
        csv_output,
        as_attachment=True,
        download_name=filename.replace(".xlsx", ".csv"),
        mimetype="text/csv",
    )


@app.route("/sales/opportunities/<pipeline_key>/upload-template")
@login_required
def sales_opportunities_upload_template(pipeline_key):
    _module_visibility_required("sales")

    pipeline_key = (pipeline_key or "lift").lower()
    if pipeline_key not in SALES_PIPELINES:
        pipeline_key = "lift"

    filename = f"sales_opportunities_template_{pipeline_key}_{datetime.datetime.utcnow():%Y%m%d}.xlsx"

    if OPENPYXL_AVAILABLE:
        workbook = build_sales_opportunity_upload_workbook(pipeline_key)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    csv_output = _build_csv_output(
        SALES_OPPORTUNITY_UPLOAD_HEADERS,
        [["" for _ in SALES_OPPORTUNITY_UPLOAD_HEADERS]],
    )
    return send_file(
        csv_output,
        as_attachment=True,
        download_name=filename.replace(".xlsx", ".csv"),
        mimetype="text/csv",
    )


@app.route("/sales/opportunities/upload", methods=["POST"])
@login_required
def sales_opportunities_upload():
    _module_visibility_required("sales")

    default_pipeline = (request.form.get("pipeline") or "lift").lower()
    if default_pipeline not in SALES_PIPELINES:
        default_pipeline = "lift"

    redirect_url = url_for("sales_opportunities_pipeline", pipeline_key=default_pipeline)

    upload = request.files.get("opportunity_upload_file")
    if not upload or not upload.filename:
        flash("Select an Excel or CSV file to upload.", "error")
        return redirect(redirect_url)

    filename = (upload.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv")):
        flash("Upload a .xlsx or .csv file exported from the opportunity template.", "error")
        return redirect(redirect_url)

    try:
        header_cells, data_rows = _extract_tabular_upload(
            upload, sheet_name=SALES_OPPORTUNITY_TEMPLATE_SHEET_NAME
        )
    except MissingDependencyError:
        flash(OPENPYXL_MISSING_MESSAGE, "error")
        return redirect(redirect_url)
    except ValueError:
        flash("Upload a .xlsx or .csv file exported from the opportunity template.", "error")
        return redirect(redirect_url)
    except Exception:
        flash("Could not read the uploaded file. Ensure it is a valid spreadsheet.", "error")
        return redirect(redirect_url)

    header_map = {}
    for idx, header in enumerate(header_cells or []):
        header_label = stringify_cell(header)
        if header_label:
            header_map[header_label] = idx

    if "Title" not in header_map:
        flash("The uploaded sheet must include the 'Title' column.", "error")
        return redirect(redirect_url)

    allowed_temperatures = {value for value, _ in SALES_TEMPERATURES}

    users = User.query.all()
    user_lookup = {}
    for user in users:
        email_value = clean_str(getattr(user, "email", None))
        if email_value:
            user_lookup[email_value.lower()] = user

    existing_clients = SalesClient.query.all()
    client_lookup = {
        (client.display_name or "").strip().lower(): client
        for client in existing_clients
        if client.display_name
    }

    created = 0
    updated = 0
    skipped = 0
    created_clients = 0
    errors = []

    for row_index, row in enumerate(data_rows, start=2):
        row_data = {}
        for header, column_index in header_map.items():
            if column_index < len(row):
                row_data[header] = stringify_cell(row[column_index])
            else:
                row_data[header] = None

        title = clean_str(row_data.get("Title"))
        if not title:
            skipped += 1
            continue

        pipeline_value = clean_str(row_data.get("Pipeline")) or default_pipeline
        pipeline_key = (pipeline_value or "lift").lower()
        if pipeline_key not in SALES_PIPELINES:
            pipeline_key = default_pipeline
        config = get_pipeline_config(pipeline_key)
        stages = config.get("stages", [])

        stage_value = clean_str(row_data.get("Stage")) or (stages[0] if stages else None)
        if stage_value and stage_value not in stages:
            errors.append(
                f"Row {row_index}: Stage '{stage_value}' is not valid for the {config['label']} pipeline."
            )
            skipped += 1
            continue

        status_value = clean_str(row_data.get("Status")) or "Open"
        temperature_raw = clean_str(row_data.get("Temperature"))
        temperature_value = None
        if temperature_raw:
            lowered = temperature_raw.lower()
            if lowered not in allowed_temperatures:
                errors.append(
                    f"Row {row_index}: Temperature '{temperature_raw}' is not valid."
                )
                skipped += 1
                continue
            temperature_value = lowered

        amount_value, amount_error = parse_float_field(row_data.get("Amount"), "Amount")
        if amount_error:
            errors.append(f"Row {row_index}: {amount_error}")
            skipped += 1
            continue

        currency_value = clean_str(row_data.get("Currency")) or "₹"

        expected_close_value, date_error = parse_date_field(
            row_data.get("Expected Close Date"), "Expected Close Date"
        )
        if date_error:
            errors.append(f"Row {row_index}: {date_error}")
            skipped += 1
            continue

        probability_value, probability_error = parse_int_field(
            row_data.get("Probability"), "Probability"
        )
        if probability_error:
            errors.append(f"Row {row_index}: {probability_error}")
            skipped += 1
            continue
        if probability_value is not None and not (0 <= probability_value <= 100):
            errors.append(
                f"Row {row_index}: Probability '{probability_value}' must be between 0 and 100."
            )
            skipped += 1
            continue

        owner_email = clean_str(row_data.get("Owner Email"))
        owner_user = None
        if owner_email:
            candidate = user_lookup.get(owner_email.lower())
            if not candidate or not candidate.can_be_assigned_module("sales"):
                errors.append(
                    f"Row {row_index}: Owner email '{owner_email}' is not linked to a sales user."
                )
                skipped += 1
                continue
            owner_user = candidate

        client_name = clean_str(row_data.get("Client Name"))
        related_project = clean_str(row_data.get("Related Project"))
        description = clean_str(row_data.get("Description"))

        client = None
        if client_name:
            client = client_lookup.get(client_name.lower())
            if not client:
                client = SalesClient(
                    display_name=client_name,
                    owner=owner_user or current_user,
                    lifecycle_stage=(
                        SALES_CLIENT_LIFECYCLE_STAGES[0]
                        if SALES_CLIENT_LIFECYCLE_STAGES
                        else None
                    ),
                )
                db.session.add(client)
                db.session.flush()
                log_sales_activity(
                    "client",
                    client.id,
                    "Client created via opportunity upload",
                    actor=current_user,
                )
                client_lookup[client_name.lower()] = client
                created_clients += 1

        query = (
            SalesOpportunity.query
            .filter(func.lower(SalesOpportunity.title) == title.lower())
            .filter(SalesOpportunity.pipeline == pipeline_key)
        )
        if client:
            query = query.filter(SalesOpportunity.client_id == client.id)
        else:
            query = query.filter(SalesOpportunity.client_id.is_(None))

        opportunity = query.first()
        is_new = opportunity is None
        if is_new:
            opportunity = SalesOpportunity(title=title, pipeline=pipeline_key)
            db.session.add(opportunity)
            created += 1
        else:
            updated += 1

        opportunity.stage = stage_value or opportunity.stage
        opportunity.status = status_value
        opportunity.temperature = temperature_value
        opportunity.amount = amount_value
        opportunity.currency = currency_value
        opportunity.expected_close_date = expected_close_value
        opportunity.probability = probability_value
        if owner_user:
            opportunity.owner = owner_user
        elif is_new and not opportunity.owner:
            opportunity.owner = current_user
        if client:
            opportunity.client = client
        opportunity.related_project = related_project
        opportunity.description = description

        if is_new:
            db.session.flush()
            log_sales_activity(
                "opportunity",
                opportunity.id,
                "Opportunity imported via bulk upload",
                actor=current_user,
            )

    if created or updated or created_clients:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            flash("Could not save the uploaded opportunities due to a database error.", "error")
            return redirect(redirect_url)

    summary_bits = []
    if created:
        summary_bits.append(f"{created} created")
    if updated:
        summary_bits.append(f"{updated} updated")
    if created_clients:
        summary_bits.append(f"{created_clients} client(s) created")
    if skipped:
        summary_bits.append(f"{skipped} skipped")

    if summary_bits:
        flash("Opportunity import complete: " + ", ".join(summary_bits) + ".", "success")
    else:
        flash("No opportunity rows were imported.", "warning")

    if errors:
        preview = errors[:5]
        more = len(errors) - len(preview)
        message = "\n".join(preview)
        if more > 0:
            message += f"\n…and {more} more issue(s)."
        flash(message, "warning")

    return redirect(redirect_url)


@app.route("/sales/opportunities/create", methods=["POST"])
@login_required
def sales_opportunities_create():
    _module_visibility_required("sales")
    title = (request.form.get("title") or "").strip()
    pipeline_key = (request.form.get("pipeline") or "lift").lower()
    config = get_pipeline_config(pipeline_key)

    if not title:
        flash("Opportunity title is required.", "error")
        return redirect(url_for("sales_opportunities_pipeline", pipeline_key=pipeline_key))

    stage_value = (request.form.get("stage") or config["stages"][0]).strip()
    if stage_value not in config["stages"]:
        stage_value = config["stages"][0]

    amount_raw = (request.form.get("amount") or "").strip()
    amount_value = None
    if amount_raw:
        try:
            amount_value = float(amount_raw)
        except ValueError:
            flash("Amount must be a valid number.", "error")
            return redirect(url_for("sales_opportunities_pipeline", pipeline_key=pipeline_key))

    opportunity = SalesOpportunity(
        title=title,
        pipeline=pipeline_key,
        stage=stage_value,
        temperature=(request.form.get("temperature") or "").strip() or None,
        amount=amount_value,
        description=(request.form.get("description") or "").strip() or None,
    )

    opportunity.owner = current_user

    client_id_raw = request.form.get("client_id")
    if client_id_raw:
        try:
            opportunity.client = db.session.get(SalesClient, int(client_id_raw))
        except (TypeError, ValueError):
            opportunity.client = None

    db.session.add(opportunity)
    db.session.flush()
    log_sales_activity("opportunity", opportunity.id, "Opportunity created")
    db.session.commit()
    flash("Opportunity created.", "success")
    return redirect(url_for("sales_opportunities_pipeline", pipeline_key=pipeline_key))


@app.route("/sales/opportunities/<int:opportunity_id>", methods=["GET", "POST"])
@login_required
def sales_opportunity_detail(opportunity_id):
    _module_visibility_required("sales")
    opportunity = db.session.get(SalesOpportunity, opportunity_id)
    if not opportunity:
        flash("Opportunity not found.", "error")
        return redirect(url_for("sales_opportunities_pipeline", pipeline_key="lift"))

    pipeline_key = opportunity.pipeline
    stages = get_pipeline_stages(pipeline_key)
    pipeline_config = get_pipeline_config(pipeline_key)
    current_stage_index = stages.index(opportunity.stage) if opportunity.stage in stages else 0

    if request.method == "POST":
        action = request.form.get("form_action") or "update"
        if action == "update":
            opportunity.title = (request.form.get("title") or "").strip() or opportunity.title
            stage_value = (request.form.get("stage") or stages[0]).strip()
            if stage_value not in stages:
                stage_value = stages[0]
            opportunity.stage = stage_value
            opportunity.temperature = (request.form.get("temperature") or "").strip() or None
            amount_raw = (request.form.get("amount") or "").strip()
            if amount_raw:
                try:
                    opportunity.amount = float(amount_raw)
                except ValueError:
                    flash("Amount must be a valid number.", "error")
                    return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))
            else:
                opportunity.amount = None
            probability_raw = (request.form.get("probability") or "").strip()
            if probability_raw:
                try:
                    opportunity.probability = int(probability_raw)
                except ValueError:
                    flash("Probability must be a whole number.", "error")
                    return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))
            else:
                opportunity.probability = None
            expected_close_raw = (request.form.get("expected_close_date") or "").strip()
            if expected_close_raw:
                try:
                    opportunity.expected_close_date = datetime.datetime.strptime(expected_close_raw, "%Y-%m-%d").date()
                except ValueError:
                    flash("Expected close date must be YYYY-MM-DD.", "error")
                    return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))
            else:
                opportunity.expected_close_date = None
            opportunity.related_project = (request.form.get("related_project") or "").strip() or None
            opportunity.description = (request.form.get("description") or "").strip() or None

            client_id_raw = request.form.get("client_id")
            if client_id_raw:
                try:
                    opportunity.client = db.session.get(SalesClient, int(client_id_raw))
                except (TypeError, ValueError):
                    opportunity.client = None
            else:
                opportunity.client = None

            owner_id_raw = request.form.get("owner_id")
            if owner_id_raw:
                try:
                    owner_candidate = db.session.get(User, int(owner_id_raw))
                except (TypeError, ValueError):
                    owner_candidate = None
                if owner_candidate and not owner_candidate.can_be_assigned_module("sales"):
                    flash("The selected owner cannot be assigned to Sales opportunities.", "error")
                    return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))
                opportunity.owner = owner_candidate
            else:
                opportunity.owner = None

            log_sales_activity("opportunity", opportunity.id, "Opportunity updated", actor=current_user)
            db.session.commit()
            flash("Opportunity updated.", "success")
            return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

        elif action == "add_note":
            note_title = (request.form.get("note_title") or "").strip() or "Timeline update"
            note_body = (request.form.get("note_body") or "").strip() or None
            log_sales_activity("opportunity", opportunity.id, note_title, notes=note_body)
            db.session.commit()
            flash("Timeline note added.", "success")
            return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

        elif action == "add_comment":
            body = (request.form.get("comment_body") or "").strip()
            if not body:
                flash("Comment cannot be empty.", "error")
                return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

            comment = SalesOpportunityComment(
                opportunity=opportunity,
                body=body,
                author=current_user if current_user.is_authenticated else None,
            )
            db.session.add(comment)

            snippet = body if len(body) <= 120 else f"{body[:117]}..."
            log_sales_activity(
                "opportunity",
                opportunity.id,
                "Comment added",
                notes=snippet,
                actor=current_user,
            )
            db.session.commit()
            flash("Comment saved.", "success")
            return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

        elif action == "upload_file":
            uploaded_file = request.files.get("attachment")
            if not uploaded_file or not uploaded_file.filename:
                flash("Select a file to upload.", "error")
                return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

            safe_name = secure_filename(uploaded_file.filename)
            if not safe_name:
                flash("The selected file name is not valid.", "error")
                return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

            upload_root = os.path.join(app.config["UPLOAD_FOLDER"], "opportunities", str(opportunity.id))
            os.makedirs(upload_root, exist_ok=True)

            unique_name = f"{uuid.uuid4().hex}_{safe_name}"
            destination_path = os.path.join(upload_root, unique_name)
            uploaded_file.save(destination_path)

            static_root = os.path.join(BASE_DIR, "static")
            stored_relative = os.path.relpath(destination_path, static_root).replace(os.sep, "/")

            record = SalesOpportunityFile(
                opportunity=opportunity,
                original_filename=uploaded_file.filename,
                stored_path=stored_relative,
                content_type=uploaded_file.mimetype,
                file_size=os.path.getsize(destination_path),
                uploaded_by=current_user if current_user.is_authenticated else None,
            )
            db.session.add(record)

            log_sales_activity(
                "opportunity",
                opportunity.id,
                "File uploaded",
                notes=uploaded_file.filename,
                actor=current_user,
            )
            db.session.commit()
            flash("File uploaded.", "success")
            return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

        elif action == "add_item":
            details = (request.form.get("item_details") or "").strip()
            lift_type = (request.form.get("item_lift_type") or "").strip()
            if lift_type and lift_type not in LIFT_TYPES:
                lift_type = None

            quantity_raw = (request.form.get("item_quantity") or "").strip()
            quantity_value = None
            if quantity_raw:
                try:
                    quantity_value = int(quantity_raw)
                    if quantity_value < 0:
                        raise ValueError
                except ValueError:
                    flash("Quantity must be a positive whole number.", "error")
                    return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

            floors_value = (request.form.get("item_floors") or "").strip() or None
            cabin_finish = (request.form.get("item_cabin_finish") or "").strip() or None
            door_type = (request.form.get("item_door_type") or "").strip() or None
            structure_value = (request.form.get("item_structure") or "no").strip().lower()
            structure_required = structure_value in {"yes", "true", "1", "on"}

            if not any([details, lift_type, quantity_value, floors_value, cabin_finish, door_type]):
                flash("Provide at least one detail for the opportunity item.", "error")
                return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

            item = SalesOpportunityItem(
                opportunity=opportunity,
                details=details or None,
                lift_type=lift_type or None,
                quantity=quantity_value,
                floors=floors_value,
                cabin_finish=cabin_finish,
                door_type=door_type,
                structure_required=structure_required,
            )
            db.session.add(item)

            summary_bits = []
            if lift_type:
                summary_bits.append(f"Lift type: {lift_type}")
            if quantity_value is not None:
                summary_bits.append(f"Qty: {quantity_value}")
            if floors_value:
                summary_bits.append(f"Floors: {floors_value}")

            log_sales_activity(
                "opportunity",
                opportunity.id,
                "Item added",
                notes="; ".join(summary_bits) or None,
                actor=current_user,
            )
            db.session.commit()
            flash("Opportunity item added.", "success")
            return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

        elif action == "schedule_activity":
            activity_type = (request.form.get("activity_type") or "meeting").strip().lower()
            if activity_type not in {"meeting", "call", "email"}:
                activity_type = "meeting"

            subject = (request.form.get("activity_subject") or "").strip()
            activity_date_raw = (request.form.get("activity_date") or "").strip()
            activity_time_raw = (request.form.get("activity_time") or "").strip()
            reminder_option = (request.form.get("reminder_option") or "").strip()
            if reminder_option not in {value for value, _ in OPPORTUNITY_REMINDER_OPTIONS}:
                reminder_option = ""
            additional_notes = (request.form.get("activity_notes") or "").strip()

            scheduled_parts = []
            if activity_date_raw:
                try:
                    scheduled_date = datetime.datetime.strptime(activity_date_raw, "%Y-%m-%d").date()
                    scheduled_parts.append(scheduled_date.strftime("%d %b %Y"))
                except ValueError:
                    flash("Provide a valid date for the activity (YYYY-MM-DD).", "error")
                    return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))
            else:
                scheduled_date = None

            if activity_time_raw:
                try:
                    scheduled_time = datetime.datetime.strptime(activity_time_raw, "%H:%M").time()
                    scheduled_parts.append(scheduled_time.strftime("%I:%M %p"))
                except ValueError:
                    flash("Provide a valid time for the activity (HH:MM).", "error")
                    return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))
            else:
                scheduled_time = None

            scheduled_for = None
            if scheduled_date and scheduled_time:
                scheduled_for = datetime.datetime.combine(scheduled_date, scheduled_time)
            elif scheduled_date:
                scheduled_for = datetime.datetime.combine(scheduled_date, datetime.time())
            elif scheduled_time:
                scheduled_for = datetime.datetime.combine(datetime.date.today(), scheduled_time)

            engagement = SalesOpportunityEngagement(
                opportunity=opportunity,
                activity_type=activity_type,
                subject=subject or None,
                scheduled_for=scheduled_for,
                reminder_option=reminder_option or None,
                notes=additional_notes or None,
                created_by=current_user if current_user.is_authenticated else None,
            )
            db.session.add(engagement)

            details = []
            if scheduled_parts:
                details.append(f"Scheduled for {' '.join(scheduled_parts)}.")
            elif not subject:
                details.append("Scheduled without a specific date. Update when confirmed.")

            reminder_label = REMINDER_OPTION_LABELS.get(reminder_option or "", "No reminder")
            if reminder_option:
                details.append(f"Reminder set: {reminder_label}.")

            if additional_notes:
                details.append(additional_notes)

            activity_label = OPPORTUNITY_ACTIVITY_LABELS.get(activity_type, activity_type.title())
            title = subject or f"Scheduled {activity_label}"
            log_sales_activity(
                "opportunity",
                opportunity.id,
                title,
                notes="\n\n".join(details) if details else None,
                actor=current_user,
            )
            db.session.commit()
            flash("Activity scheduled.", "success")
            return redirect(url_for("sales_opportunity_detail", opportunity_id=opportunity.id))

    activities = (
        SalesActivity.query
        .filter_by(parent_type="opportunity", parent_id=opportunity.id)
        .order_by(SalesActivity.created_at.desc())
        .all()
    )
    comments = (
        SalesOpportunityComment.query
        .filter_by(opportunity_id=opportunity.id)
        .order_by(SalesOpportunityComment.created_at.desc())
        .all()
    )
    files = (
        SalesOpportunityFile.query
        .filter_by(opportunity_id=opportunity.id)
        .order_by(SalesOpportunityFile.created_at.desc())
        .all()
    )
    scheduled_activities = (
        SalesOpportunityEngagement.query
        .filter_by(opportunity_id=opportunity.id)
        .order_by(SalesOpportunityEngagement.created_at.desc())
        .all()
    )
    items = (
        SalesOpportunityItem.query
        .filter_by(opportunity_id=opportunity.id)
        .order_by(SalesOpportunityItem.created_at.desc())
        .all()
    )
    owners = get_assignable_users_for_module("sales", order_by="name")
    clients = SalesClient.query.order_by(SalesClient.display_name.asc()).all()

    return render_template(
        "sales/opportunity_detail.html",
        opportunity=opportunity,
        pipeline_key=pipeline_key,
        pipeline_config=pipeline_config,
        stages=stages,
        current_stage_index=current_stage_index,
        owners=owners,
        clients=clients,
        activities=activities,
        comments=comments,
        files=files,
        scheduled_activities=scheduled_activities,
        items=items,
        reminder_options=OPPORTUNITY_REMINDER_OPTIONS,
        lift_types=LIFT_TYPES,
        temperature_choices=SALES_TEMPERATURES,
        pipeline_map=SALES_PIPELINES,
    )


@app.route("/sales/opportunities/<int:opportunity_id>/stage", methods=["POST"])
@login_required
def sales_opportunity_stage(opportunity_id):
    _module_visibility_required("sales")
    opportunity = db.session.get(SalesOpportunity, opportunity_id)
    accepts = request.accept_mimetypes.best
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or accepts == "application/json"
    if not opportunity:
        if wants_json:
            return jsonify({"success": False, "message": "Opportunity not found."}), 404
        flash("Opportunity not found.", "error")
        return redirect(url_for("sales_opportunities_pipeline", pipeline_key="lift"))

    pipeline_key = opportunity.pipeline
    stage = (request.form.get("stage") or "").strip()
    stages = get_pipeline_stages(pipeline_key)
    if stage not in stages:
        if wants_json:
            return jsonify({"success": False, "message": "Invalid stage selected."}), 400
        flash("Invalid stage selected.", "error")
        return redirect(url_for("sales_opportunities_pipeline", pipeline_key=pipeline_key))

    opportunity.stage = stage
    log_sales_activity("opportunity", opportunity.id, f"Stage moved to {stage}", actor=current_user)
    db.session.commit()

    if wants_json:
        return jsonify({"success": True, "stage": stage, "pipeline": pipeline_key})

    flash("Opportunity stage updated.", "success")
    return redirect(url_for("sales_opportunities_pipeline", pipeline_key=pipeline_key))

def _require_admin():
    if not current_user.is_admin:
        abort(403)


def _form_truthy(value):
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "on", "yes"}


def _department_cycle(department, candidate_parent):
    current = candidate_parent
    while current is not None:
        if current.id == department.id:
            return True
        current = current.parent
    return False


def _position_cycle(position, candidate_manager):
    current = candidate_manager
    while current is not None:
        if current.id == position.id:
            return True
        current = current.reports_to
    return False


# ---------------------- ADMINISTRATION ----------------------
def _admin_users_context(create_defaults=None, show_create=False):
    departments = sorted(
        Department.query.order_by(Department.name.asc()).all(),
        key=lambda d: (d.full_name or "").lower(),
    )
    positions = sorted(
        Position.query.order_by(Position.title.asc()).all(),
        key=lambda p: (p.display_label or "").lower(),
    )
    defaults = dict(create_defaults or {})
    defaults.setdefault("active", "1")

    return dict(
        users=User.query.order_by(User.username.asc()).all(),
        departments=departments,
        department_options=departments,
        department_branches=DEPARTMENT_BRANCHES,
        positions=positions,
        position_options=positions,
        category_label="Admin",
        category_url=url_for("admin_users"),
        create_user_defaults=defaults,
        show_create_user_form=show_create,
    )


@app.route("/admin/users")
@login_required
def admin_users():
    _require_admin()

    return render_template("admin_users.html", **_admin_users_context())


@app.route("/admin/users/create", methods=["POST"])
@login_required
def admin_users_create():
    _require_admin()

    username = (request.form.get("username") or "").strip()
    password = (request.form.get("password") or "").strip()
    first_name = (request.form.get("first_name") or "").strip()
    last_name = (request.form.get("last_name") or "").strip()
    email = (request.form.get("email") or "").strip()
    mobile_number = (request.form.get("mobile_number") or "").strip()
    department_id_raw = request.form.get("department_id")
    position_id_raw = request.form.get("position_id")
    active_flag = _form_truthy(request.form.get("active", "1"))

    if not username or not password:
        flash("Username and password are required to create a user.", "error")
        create_defaults = request.form.to_dict(flat=True)
        if "active" not in create_defaults:
            create_defaults["active"] = ""
        return render_template(
            "admin_users.html",
            **_admin_users_context(
                create_defaults=create_defaults,
                show_create=True,
            ),
        )

    existing = (
        User.query.filter(func.lower(User.username) == username.lower()).first()
        if username
        else None
    )
    if existing:
        flash("A user with that username already exists.", "error")
        create_defaults = request.form.to_dict(flat=True)
        if "active" not in create_defaults:
            create_defaults["active"] = ""
        return render_template(
            "admin_users.html",
            **_admin_users_context(
                create_defaults=create_defaults,
                show_create=True,
            ),
        )

    department = None
    if department_id_raw:
        try:
            department = db.session.get(Department, int(department_id_raw))
        except (TypeError, ValueError):
            department = None

    position = None
    if position_id_raw:
        try:
            position = db.session.get(Position, int(position_id_raw))
        except (TypeError, ValueError):
            position = None

    user = User(
        username=username,
        password=password,
        first_name=first_name or None,
        last_name=last_name or None,
        email=email or None,
        mobile_number=mobile_number or None,
        department=department.name if department else None,
        active=active_flag,
    )
    user.set_module_permissions(
        {
            module["key"]: {
                "visibility": False,
                "assignment": False,
            }
            for module in WORKSPACE_MODULES
        }
    )
    if position:
        user.position = position
        if not user.department and position.department:
            user.department = position.department.name

    user.issue_session_token()
    db.session.add(user)
    db.session.commit()

    flash(f"User '{username}' created successfully.", "success")
    return redirect(url_for("admin_users") + f"#user-{user.id}")


@app.route("/admin/users/<int:user_id>/update", methods=["POST"])
@login_required
def admin_users_update(user_id):
    _require_admin()

    user = db.session.get(User, user_id)
    if not user:
        flash("User not found.", "error")
        return redirect(url_for("admin_users"))

    action = request.form.get("action") or "update"
    if action == "reset_sessions":
        user.issue_session_token()
        db.session.commit()
        flash(f"All active sessions for '{user.username}' have been revoked.", "success")
        if user.id == current_user.id:
            session.pop("session_token", None)
            logout_user()
            flash("You signed yourself out of all sessions.", "info")
            return redirect(url_for("login"))
        return redirect(url_for("admin_users") + f"#user-{user.id}")

    username = (request.form.get("username") or "").strip()
    first_name = (request.form.get("first_name") or "").strip()
    last_name = (request.form.get("last_name") or "").strip()
    email = (request.form.get("email") or "").strip()
    mobile_number = (request.form.get("mobile_number") or "").strip()
    department_id_raw = request.form.get("department_id")
    position_id_raw = request.form.get("position_id")
    active_flag = _form_truthy(request.form.get("active"))
    password = (request.form.get("password") or "").strip()

    if not username:
        flash("Username is required.", "error")
        return redirect(url_for("admin_users") + f"#user-{user.id}")

    if username.lower() != user.username.lower():
        existing = (
            User.query.filter(func.lower(User.username) == username.lower(), User.id != user.id).first()
        )
        if existing:
            flash("Another user already uses that username.", "error")
            return redirect(url_for("admin_users") + f"#user-{user.id}")

    department = None
    if department_id_raw:
        try:
            department = db.session.get(Department, int(department_id_raw))
        except (TypeError, ValueError):
            department = None

    position = None
    if position_id_raw:
        try:
            position = db.session.get(Position, int(position_id_raw))
        except (TypeError, ValueError):
            position = None

    if user.id == current_user.id and not active_flag:
        flash("You cannot deactivate your own account while logged in.", "error")
        return redirect(url_for("admin_users") + f"#user-{user.id}")

    user.username = username
    user.first_name = first_name or None
    user.last_name = last_name or None
    user.email = email or None
    user.mobile_number = mobile_number or None
    user.department = department.name if department else None
    user.active = active_flag
    user.position = position
    if not user.department and position and position.department:
        user.department = position.department.name

    if password:
        user.password = password

    permissions_payload = {}
    for module in WORKSPACE_MODULES:
        module_key = module["key"]
        visibility_flag = _form_truthy(request.form.get(f"module_{module_key}_visibility"))
        assignment_flag = _form_truthy(request.form.get(f"module_{module_key}_assignment"))
        permissions_payload[module_key] = {
            "visibility": visibility_flag,
            "assignment": assignment_flag,
        }
    user.set_module_permissions(permissions_payload)

    db.session.commit()
    flash(f"User '{user.username}' updated.", "success")
    return redirect(url_for("admin_users") + "#users")


@app.route("/admin/departments/create", methods=["POST"])
@login_required
def admin_departments_create():
    _require_admin()

    name = (request.form.get("name") or "").strip()
    branch = (request.form.get("branch") or "").strip()
    description = (request.form.get("description") or "").strip()
    parent_id_raw = request.form.get("parent_id")
    active_flag = _form_truthy(request.form.get("active", "1"))

    if not name:
        flash("Department name is required.", "error")
        return redirect(url_for("admin_users") + "#departments")

    if branch not in DEPARTMENT_BRANCHES:
        branch = DEPARTMENT_BRANCHES[0]

    existing = Department.query.filter(func.lower(Department.name) == name.lower()).first()
    if existing:
        flash("A department with that name already exists.", "error")
        return redirect(url_for("admin_users") + "#departments")

    parent = None
    if parent_id_raw:
        try:
            parent = db.session.get(Department, int(parent_id_raw))
        except (TypeError, ValueError):
            parent = None

    department = Department(
        name=name,
        branch=branch,
        description=description or None,
        active=active_flag,
    )
    if parent:
        department.parent = parent

    db.session.add(department)
    db.session.commit()

    flash(f"Department '{department.full_name}' created.", "success")
    return redirect(url_for("admin_users") + f"#department-{department.id}")


@app.route("/admin/departments/<int:department_id>/update", methods=["POST"])
@login_required
def admin_departments_update(department_id):
    _require_admin()

    department = db.session.get(Department, department_id)
    if not department:
        flash("Department not found.", "error")
        return redirect(url_for("admin_users") + "#departments")

    name = (request.form.get("name") or "").strip()
    branch = (request.form.get("branch") or "").strip()
    description = (request.form.get("description") or "").strip()
    parent_id_raw = request.form.get("parent_id")
    active_flag = _form_truthy(request.form.get("active"))

    if not name:
        flash("Department name is required.", "error")
        return redirect(url_for("admin_users") + f"#department-{department.id}")

    if branch not in DEPARTMENT_BRANCHES:
        branch = DEPARTMENT_BRANCHES[0]

    if name.lower() != department.name.lower():
        existing = Department.query.filter(
            func.lower(Department.name) == name.lower(),
            Department.id != department.id,
        ).first()
        if existing:
            flash("Another department already uses that name.", "error")
            return redirect(url_for("admin_users") + f"#department-{department.id}")

    parent = None
    if parent_id_raw:
        try:
            parent = db.session.get(Department, int(parent_id_raw))
        except (TypeError, ValueError):
            parent = None
    if parent and _department_cycle(department, parent):
        flash("Cannot assign a department to one of its descendants.", "error")
        return redirect(url_for("admin_users") + f"#department-{department.id}")

    department.name = name
    department.branch = branch
    department.description = description or None
    department.active = active_flag
    department.parent = parent

    db.session.commit()
    flash(f"Department '{department.full_name}' updated.", "success")
    return redirect(url_for("admin_users") + f"#department-{department.id}")


@app.route("/admin/departments/<int:department_id>/delete", methods=["POST"])
@login_required
def admin_departments_delete(department_id):
    _require_admin()

    department = db.session.get(Department, department_id)
    if not department:
        flash("Department not found.", "error")
        return redirect(url_for("admin_users") + "#departments")

    for child in list(department.children or []):
        child.parent = None
    for position in list(department.positions or []):
        position.department = None

    db.session.delete(department)
    db.session.commit()

    flash(f"Department '{department.name}' deleted.", "success")
    return redirect(url_for("admin_users") + "#departments")


@app.route("/admin/positions/create", methods=["POST"])
@login_required
def admin_positions_create():
    _require_admin()

    title = (request.form.get("title") or "").strip()
    department_id_raw = request.form.get("department_id")
    reports_to_id_raw = request.form.get("reports_to_id")
    active_flag = _form_truthy(request.form.get("active", "1"))

    if not title:
        flash("Position title is required.", "error")
        return redirect(url_for("admin_users") + "#positions")

    department = None
    if department_id_raw:
        try:
            department = db.session.get(Department, int(department_id_raw))
        except (TypeError, ValueError):
            department = None

    manager = None
    if reports_to_id_raw:
        try:
            manager = db.session.get(Position, int(reports_to_id_raw))
        except (TypeError, ValueError):
            manager = None

    position = Position(
        title=title,
        department=department,
        reports_to=manager,
        active=active_flag,
    )
    db.session.add(position)
    db.session.commit()

    flash(f"Position '{position.display_label}' created.", "success")
    return redirect(url_for("admin_users") + f"#position-{position.id}")


@app.route("/admin/positions/<int:position_id>/update", methods=["POST"])
@login_required
def admin_positions_update(position_id):
    _require_admin()

    position = db.session.get(Position, position_id)
    if not position:
        flash("Position not found.", "error")
        return redirect(url_for("admin_users") + "#positions")

    title = (request.form.get("title") or "").strip()
    department_id_raw = request.form.get("department_id")
    reports_to_id_raw = request.form.get("reports_to_id")
    active_flag = _form_truthy(request.form.get("active"))

    if not title:
        flash("Position title is required.", "error")
        return redirect(url_for("admin_users") + f"#position-{position.id}")

    department = None
    if department_id_raw:
        try:
            department = db.session.get(Department, int(department_id_raw))
        except (TypeError, ValueError):
            department = None

    manager = None
    if reports_to_id_raw:
        try:
            manager = db.session.get(Position, int(reports_to_id_raw))
        except (TypeError, ValueError):
            manager = None

    if manager and _position_cycle(position, manager):
        flash("Cannot assign a position to report to itself or its descendants.", "error")
        return redirect(url_for("admin_users") + f"#position-{position.id}")

    position.title = title
    position.department = department
    position.reports_to = manager
    position.active = active_flag

    db.session.commit()
    flash(f"Position '{position.display_label}' updated.", "success")
    return redirect(url_for("admin_users") + f"#position-{position.id}")


@app.route("/admin/positions/<int:position_id>/delete", methods=["POST"])
@login_required
def admin_positions_delete(position_id):
    _require_admin()

    position = db.session.get(Position, position_id)
    if not position:
        flash("Position not found.", "error")
        return redirect(url_for("admin_users") + "#positions")

    for report in list(position.direct_reports or []):
        report.reports_to = None

    db.session.delete(position)
    db.session.commit()

    flash(f"Position '{position.title}' deleted.", "success")
    return redirect(url_for("admin_users") + "#positions")


def _build_task_overview(viewing_user: "User"):
    def _describe_due_date(target_date, now):
        if not target_date:
            return None, None, "none"
        if isinstance(target_date, datetime.date) and not isinstance(target_date, datetime.datetime):
            target_dt = datetime.datetime.combine(target_date, datetime.time.min)
        else:
            target_dt = target_date
        delta_days = (target_dt.date() - now.date()).days
        display = target_dt.strftime("%d %b %Y")
        if delta_days < 0:
            days = abs(delta_days)
            label = f"Overdue by {days} day{'s' if days != 1 else ''}"
            variant = "overdue"
        elif delta_days == 0:
            label = "Due today"
            variant = "today"
        elif delta_days == 1:
            label = "Due tomorrow"
            variant = "upcoming"
        else:
            label = f"Due in {delta_days} days"
            variant = "upcoming"
        return label, display, variant

    def _status_badge_class(key):
        mapping = {
            "open": "bg-amber-500/20 text-amber-100 border border-amber-500/40",
            "in_progress": "bg-sky-500/20 text-sky-100 border border-sky-500/40",
            "blocked": "bg-rose-500/20 text-rose-100 border border-rose-500/40",
            "scheduled": "bg-sky-500/20 text-sky-100 border border-sky-500/40",
            "overdue": "bg-rose-500/20 text-rose-100 border border-rose-500/40",
        }
        return mapping.get(key, "bg-slate-800/60 text-slate-200 border border-slate-700/60")

    def _due_badge_class(variant):
        mapping = {
            "overdue": "bg-rose-500/20 text-rose-100 border border-rose-500/40",
            "today": "bg-amber-500/20 text-amber-100 border border-amber-500/40",
            "upcoming": "bg-slate-800/60 text-slate-200 border border-slate-700/60",
        }
        return mapping.get(variant, "bg-slate-800/60 text-slate-200 border border-slate-700/60")

    def _ensure_module(modules_map, order, label, empty_message, description=None):
        module = modules_map.get(label)
        if module is None:
            module = {
                "module": label,
                "items": [],
                "empty_message": empty_message,
                "description": description,
            }
            modules_map[label] = module
            order.append(label)
        else:
            if description and not module.get("description"):
                module["description"] = description
        return module

    def _status_key(status):
        key = (status or "").strip().lower()
        if key == "in progress":
            return "in_progress"
        if key == "blocked":
            return "blocked"
        if key == "closed":
            return "closed"
        if key == "overdue":
            return "overdue"
        if key == "scheduled":
            return "scheduled"
        return "open"

    def _build_pending_modules(
        viewing_user,
        open_tasks,
        now,
        assignee_lookup=None,
        sales_user_ids=None,
    ):
        modules_map = OrderedDict()
        module_order = []

        show_projects = viewing_user.can_view_module("operations") if viewing_user else True
        show_qc = viewing_user.can_view_module("qc") if viewing_user else True
        show_sales = viewing_user.can_view_module("sales") if viewing_user else True

        if show_projects:
            _ensure_module(
                modules_map,
                module_order,
                "Projects",
                "No pending project tasks.",
                "Tasks from active projects assigned to you.",
            )
        if show_sales:
            _ensure_module(
                modules_map,
                module_order,
                "Sales",
                "No pending sales activities.",
                "Upcoming and overdue sales engagements on your opportunities.",
            )

        for task in open_tasks:
            module_label = "Projects" if task.project else "Quality Control"
            if module_label == "Projects" and not show_projects:
                continue
            if module_label == "Quality Control" and not show_qc:
                continue
            module_description = (
                "Project execution tasks awaiting your action."
                if module_label == "Projects"
                else "Quality inspections and tasks that still need attention."
            )
            module = _ensure_module(
                modules_map,
                module_order,
                module_label,
                "No pending QC tasks." if module_label == "Quality Control" else "No pending project tasks.",
                module_description,
            )

            due_label, due_display, due_variant = _describe_due_date(task.due_date, now)
            metadata = []
            if task.stage:
                metadata.append(task.stage)
            if task.lift_type:
                metadata.append(task.lift_type)
            if task.template and task.template.name:
                metadata.append(f"Form: {task.template.name}")
            if task.project and task.project.name:
                metadata.append(f"Project: {task.project.name}")

            assignee_user = None
            if assignee_lookup:
                assignee_user = assignee_lookup.get(getattr(task, "assigned_to", None))

            module["items"].append(
                {
                    "title": task.display_title,
                    "subtitle": task.client_name or task.site_name,
                    "description": task.description,
                    "identifier": f"#{task.id}",
                    "status": task.status or "Open",
                    "status_class": _status_badge_class(_status_key(task.status)),
                    "due_description": due_label,
                    "due_display": due_display,
                    "due_class": _due_badge_class(due_variant),
                    "url": url_for("qc_work_detail", work_id=task.id),
                    "secondary_url": url_for("forms_fill", form_id=task.template_id, work_id=task.id)
                    if task.template_id
                    else None,
                    "secondary_label": "New Submission" if task.template_id else None,
                    "metadata": metadata
                    + (
                        []
                        if not assignee_user
                        or (viewing_user and assignee_user.id == viewing_user.id)
                        else [
                            f"Owner: {assignee_user.display_name}"
                            if assignee_user.display_name
                            else f"Owner: {assignee_user.username}"
                        ]
                    ),
                }
            )

        if show_sales:
            sales_module = modules_map.get("Sales")
            if sales_module is None:
                sales_module = _ensure_module(
                    modules_map,
                    module_order,
                    "Sales",
                    "No pending sales activities.",
                    "Upcoming and overdue sales engagements on your opportunities.",
                )
            sales_filters = []
            if sales_user_ids is None:
                if viewing_user:
                    sales_filters = [
                        SalesOpportunity.owner_id == viewing_user.id,
                        SalesOpportunityEngagement.created_by_id == viewing_user.id,
                    ]
            else:
                allowed_ids = [uid for uid in set(sales_user_ids) if uid]
                if allowed_ids:
                    sales_filters = [
                        SalesOpportunity.owner_id.in_(allowed_ids),
                        SalesOpportunityEngagement.created_by_id.in_(allowed_ids),
                    ]
            sales_items = []
            if sales_filters:
                sales_items = (
                    SalesOpportunityEngagement.query
                    .join(SalesOpportunity, SalesOpportunity.id == SalesOpportunityEngagement.opportunity_id)
                    .filter(SalesOpportunityEngagement.scheduled_for.isnot(None))
                    .filter(or_(*sales_filters))
                    .filter(func.lower(SalesOpportunity.status) != "closed")
                    .order_by(
                        SalesOpportunityEngagement.scheduled_for.asc(),
                        SalesOpportunityEngagement.id.asc(),
                    )
                    .limit(50)
                    .all()
                )

            for activity in sales_items:
                opportunity = activity.opportunity
                due_label, due_display, due_variant = _describe_due_date(activity.scheduled_for, now)
                status_key = "overdue" if due_variant == "overdue" else "scheduled"
                metadata = [activity.display_activity_type]
                subtitle = None
                if opportunity:
                    subtitle = opportunity.title
                    if opportunity.stage:
                        metadata.append(opportunity.stage)
                    if opportunity.client and opportunity.client.display_name:
                        metadata.append(f"Client: {opportunity.client.display_name}")

                if assignee_lookup:
                    owner_user = None
                    if opportunity and opportunity.owner_id:
                        owner_user = assignee_lookup.get(opportunity.owner_id)
                    if owner_user and owner_user.display_name:
                        owner_label = f"Owner: {owner_user.display_name}"
                        if owner_label not in metadata:
                            metadata.append(owner_label)
                    planner_user = None
                    if activity.created_by_id:
                        planner_user = assignee_lookup.get(activity.created_by_id)
                    if planner_user and planner_user.display_name:
                        planner_label = f"Planner: {planner_user.display_name}"
                        if planner_label not in metadata:
                            metadata.append(planner_label)

                sales_module["items"].append(
                    {
                        "title": activity.subject or activity.display_activity_type,
                        "subtitle": subtitle,
                        "description": activity.notes,
                        "identifier": f"Activity #{activity.id}",
                        "status": "Overdue" if status_key == "overdue" else "Scheduled",
                        "status_class": _status_badge_class(status_key),
                        "due_description": due_label,
                        "due_display": due_display,
                        "due_class": _due_badge_class(due_variant),
                        "url": url_for("sales_opportunity_detail", opportunity_id=opportunity.id)
                        if opportunity
                        else None,
                        "secondary_url": None,
                        "secondary_label": None,
                        "metadata": metadata,
                    }
                )

        pending_modules = [modules_map[label] for label in module_order]
        pending_total = sum(len(module["items"]) for module in pending_modules)
        return pending_modules, pending_total

    def _team_members_for(user):
        if not user:
            return []

        members = []
        seen_positions = set()
        stack = []

        position = getattr(user, "position", None)
        if position and getattr(position, "direct_reports", None):
            stack.extend(list(position.direct_reports))

        if not stack and user.is_admin:
            query = User.query.filter(User.id != user.id)
            query = query.filter(User.active.is_(True))
            ordered = query.order_by(
                User.first_name.asc(),
                User.last_name.asc(),
                User.username.asc(),
            ).all()
            return ordered

        while stack:
            pos = stack.pop()
            if not pos:
                continue
            pos_id = getattr(pos, "id", None)
            if pos_id in seen_positions:
                continue
            seen_positions.add(pos_id)
            direct_reports = list(getattr(pos, "direct_reports", []) or [])
            if direct_reports:
                stack.extend(direct_reports)
            if not getattr(pos, "active", True):
                continue
            for member in list(getattr(pos, "users", []) or []):
                if not member or getattr(member, "id", None) is None:
                    continue
                if member.id == user.id:
                    continue
                if not member.is_active:
                    continue
                members.append(member)

        unique_members = []
        seen_member_ids = set()
        for member in members:
            if member.id in seen_member_ids:
                continue
            seen_member_ids.add(member.id)
            unique_members.append(member)
        return unique_members

    status_order = case(
        (QCWork.status == "In Progress", 0),
        (QCWork.status == "Open", 1),
        (QCWork.status == "Blocked", 2),
        (QCWork.status == "Closed", 3),
        else_=4
    )
    now = datetime.datetime.utcnow()
    tasks = (
        QCWork.query
        .filter(QCWork.assigned_to == viewing_user.id)
        .order_by(status_order, QCWork.due_date.asc().nullslast(), QCWork.created_at.desc())
        .all()
    )
    actionable_tasks = [task for task in tasks if task.dependency_satisfied]
    blocked_tasks = [task for task in tasks if not task.dependency_satisfied and task.status != "Closed"]
    open_tasks = [task for task in actionable_tasks if task.status != "Closed"]
    closed_tasks = [task for task in tasks if task.status == "Closed"]

    open_count = sum(1 for task in open_tasks if (task.status or "").lower() in {"open", "blocked"})
    in_progress_count = sum(1 for task in open_tasks if (task.status or "").lower() == "in progress")
    overdue_count = sum(
        1
        for task in open_tasks
        if task.due_date and task.due_date < now and (task.status or "").lower() != "closed"
    )

    team_load = []
    if current_user.is_admin:
        users = User.query.order_by(User.username.asc()).all()
        assignments = {user.id: [] for user in users}
        for task in QCWork.query.filter(QCWork.assigned_to.isnot(None)).all():
            assignments.setdefault(task.assigned_to, []).append(task)

        for member in users:
            member_tasks = assignments.get(member.id, [])
            actionable = [task for task in member_tasks if task.dependency_satisfied]
            open_items = [task for task in actionable if task.status != "Closed"]
            team_load.append({
                "user": member,
                "total": len(actionable),
                "open": sum(1 for task in open_items if (task.status or "").lower() in {"open", "blocked"}),
                "in_progress": sum(1 for task in open_items if (task.status or "").lower() == "in progress"),
                "overdue": sum(
                    1 for task in open_items if task.due_date and task.due_date < now
                )
            })

    pending_modules, pending_total = _build_pending_modules(viewing_user, open_tasks, now)

    team_members = _team_members_for(viewing_user)
    team_user_ids = sorted(
        {member.id for member in team_members if getattr(member, "id", None)}
    )
    team_pending_modules = []
    team_pending_total = 0
    if team_user_ids:
        team_tasks = (
            QCWork.query
            .filter(QCWork.assigned_to.in_(team_user_ids))
            .order_by(status_order, QCWork.due_date.asc().nullslast(), QCWork.created_at.desc())
            .all()
        )
        team_actionable_tasks = [task for task in team_tasks if task.dependency_satisfied]
        team_open_tasks = [task for task in team_actionable_tasks if task.status != "Closed"]
        assignment_lookup = {member.id: member for member in team_members}
        team_pending_modules, team_pending_total = _build_pending_modules(
            viewing_user,
            team_open_tasks,
            now,
            assignee_lookup=assignment_lookup,
            sales_user_ids=team_user_ids,
        )

    return {
        "open_tasks": open_tasks,
        "closed_tasks": closed_tasks,
        "open_count": open_count,
        "in_progress_count": in_progress_count,
        "overdue_count": overdue_count,
        "blocked_tasks": blocked_tasks,
        "team_load": team_load,
        "pending_modules": pending_modules,
        "pending_total": pending_total,
        "team_pending_modules": team_pending_modules,
        "team_pending_total": team_pending_total,
        "team_members": team_members,
    }


@app.route("/dashboard")
@login_required
def dashboard():
    viewing_user = current_user
    selected_user_id = request.args.get("user_id", type=int)
    if selected_user_id and current_user.is_admin:
        candidate = db.session.get(User, selected_user_id)
        if candidate:
            viewing_user = candidate

    context = _build_task_overview(viewing_user)
    context.update({
        "viewing_user": viewing_user,
        "category_label": None,
        "page_mode": "dashboard",
        "switch_user_endpoint": "dashboard",
    })

    return render_template("dashboard.html", **context)


@app.route("/projects/pending")
@login_required
def projects_pending():
    _module_visibility_required("operations")
    viewing_user = current_user
    selected_user_id = request.args.get("user_id", type=int)
    if selected_user_id and current_user.is_admin:
        candidate = db.session.get(User, selected_user_id)
        if candidate:
            viewing_user = candidate

    context = _build_task_overview(viewing_user)
    context.update({
        "viewing_user": viewing_user,
        "category_label": "Projects",
        "category_url": url_for("projects_pending"),
        "page_mode": "projects",
        "switch_user_endpoint": "projects_pending",
    })

    project_modules = [
        module
        for module in context.get("pending_modules", [])
        if module.get("module") in {"Projects", "Quality Control"}
    ]
    context["pending_modules"] = project_modules
    context["pending_total"] = sum(len(module.get("items", [])) for module in project_modules)

    team_project_modules = [
        module
        for module in context.get("team_pending_modules", [])
        if module.get("module") in {"Projects", "Quality Control"}
    ]
    context["team_pending_modules"] = team_project_modules
    context["team_pending_total"] = sum(
        len(module.get("items", [])) for module in team_project_modules
    )

    return render_template("dashboard.html", **context)


# ---------------------- FORMS (TEMPLATES) ----------------------
@app.route("/forms")
@login_required
def forms_list():
    _module_visibility_required("qc")
    forms = FormSchema.query.order_by(FormSchema.name.asc()).all()
    return render_template("forms_list.html", forms=forms, category_label="Forms", category_url=url_for('forms_list'))


@app.route("/forms/new", methods=["GET", "POST"])
@login_required
def forms_new():
    _module_visibility_required("qc")
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        stage = (request.form.get("stage") or "").strip()
        lift_type = (request.form.get("lift_type") or "").strip()
        try:
            schema = json.loads(request.form.get("schema_json", "[]"))
        except Exception:
            flash("Invalid JSON schema", "error")
            return render_template(
                "forms_edit.html",
                item=None,
                STAGES=STAGES,
                LIFT_TYPES=LIFT_TYPES,
                initial_schema=[],
                category_label="Forms",
                category_url=url_for('forms_list')
            )

        if not name:
            flash("Name is required", "error")
            return render_template(
                "forms_edit.html",
                item=None,
                STAGES=STAGES,
                LIFT_TYPES=LIFT_TYPES,
                initial_schema=schema,
                category_label="Forms",
                category_url=url_for('forms_list')
            )

        if stage and stage not in STAGES:
            flash("Select a valid Stage.", "error")
            return render_template(
                "forms_edit.html",
                item=None,
                STAGES=STAGES,
                LIFT_TYPES=LIFT_TYPES,
                initial_schema=schema,
                category_label="Forms",
                category_url=url_for('forms_list')
            )
        if lift_type and lift_type not in LIFT_TYPES:
            flash("Select a valid Lift Type.", "error")
            return render_template(
                "forms_edit.html",
                item=None,
                STAGES=STAGES,
                LIFT_TYPES=LIFT_TYPES,
                initial_schema=schema,
                category_label="Forms",
                category_url=url_for('forms_list')
            )

        item = FormSchema(
            name=name,
            schema_json=json.dumps(schema, ensure_ascii=False),
            min_photos_if_all_good=0,
            stage=stage or None,
            lift_type=lift_type or None
        )
        db.session.add(item)
        db.session.commit()
        flash("Form created", "success")
        return redirect(url_for("forms_list"))

    return render_template(
        "forms_edit.html",
        item=None,
        STAGES=STAGES,
        LIFT_TYPES=LIFT_TYPES,
        initial_schema=[],
        category_label="Forms",
        category_url=url_for('forms_list')
    )


@app.route("/forms/<int:form_id>/edit", methods=["GET", "POST"])
@login_required
def forms_edit(form_id):
    _module_visibility_required("qc")
    item = db.session.get(FormSchema, form_id)
    if not item:
        flash("Form not found", "error")
        return redirect(url_for("forms_list"))

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        stage = (request.form.get("stage") or "").strip()
        lift_type = (request.form.get("lift_type") or "").strip()
        try:
            schema = json.loads(request.form.get("schema_json", "[]"))
        except Exception:
            flash("Invalid JSON schema", "error")
            return render_template(
                "forms_edit.html",
                item=item,
                STAGES=STAGES,
                LIFT_TYPES=LIFT_TYPES,
                initial_schema=[],
                category_label="Forms",
                category_url=url_for('forms_list')
            )

        if name:
            item.name = name
        item.schema_json = json.dumps(schema, ensure_ascii=False)

        if stage and stage not in STAGES:
            flash("Select a valid Stage.", "error")
            return render_template(
                "forms_edit.html",
                item=item,
                STAGES=STAGES,
                LIFT_TYPES=LIFT_TYPES,
                initial_schema=schema,
                category_label="Forms",
                category_url=url_for('forms_list')
            )
        if lift_type and lift_type not in LIFT_TYPES:
            flash("Select a valid Lift Type.", "error")
            return render_template(
                "forms_edit.html",
                item=item,
                STAGES=STAGES,
                LIFT_TYPES=LIFT_TYPES,
                initial_schema=schema,
                category_label="Forms",
                category_url=url_for('forms_list')
            )

        item.stage = stage or None
        item.lift_type = lift_type or None

        db.session.commit()
        flash("Form updated", "success")
        return redirect(url_for("forms_list"))

    return render_template(
        "forms_edit.html",
        item=item,
        STAGES=STAGES,
        LIFT_TYPES=LIFT_TYPES,
        initial_schema=json.loads(item.schema_json or "[]"),
        category_label="Forms",
        category_url=url_for('forms_list')
    )


@app.route("/forms/<int:form_id>/delete", methods=["POST"])
@login_required
def forms_delete(form_id):
    _module_visibility_required("qc")
    item = db.session.get(FormSchema, form_id)
    if item:
        db.session.delete(item)
        db.session.commit()
        flash("Form deleted", "info")
    return redirect(url_for("forms_list"))


@app.route("/forms/<int:form_id>/fill", methods=["GET", "POST"])
@login_required
def forms_fill(form_id):
    _module_visibility_required("qc")
    fs = db.session.get(FormSchema, form_id)
    if not fs:
        flash("Form not found", "error")
        return redirect(url_for("forms_list"))

    schema_raw = json.loads(fs.schema_json)
    sections, is_sectioned = _normalize_form_schema(schema_raw)
    if request.method == "POST":
        values = {} if is_sectioned else {}
        any_ng = False
        saved_photos = []
        per_item_photo_count = 0
        for s_idx, section in enumerate(sections):
            section_label = section.get("section") or f"Section {s_idx + 1}"
            section_items = section.get("items") or []
            section_entries = []
            if is_sectioned:
                values[section_label] = section_entries
            for f_idx, field in enumerate(section_items):
                label = field.get("label") or f"Item {f_idx + 1}"
                ftype = field.get("type")
                required = field.get("required", False)
                field_name = f"field__{s_idx}__{f_idx}"
                remark_name = f"remark__{s_idx}__{f_idx}"
                photo_name = f"photo__{s_idx}__{f_idx}"
                options = field.get("options") or []

                normalized_val = ""
                if ftype == "table":
                    rows = field.get("rows") or []
                    columns = field.get("columns") or []
                    table_values = []
                    missing_required = False
                    for r_idx, _ in enumerate(rows):
                        row_entries = []
                        for c_idx, _ in enumerate(columns):
                            cell_name = f"table__{s_idx}__{f_idx}__{r_idx}__{c_idx}"
                            cell_val = request.form.get(cell_name, "").strip()
                            if required and not cell_val:
                                missing_required = True
                            row_entries.append(cell_val)
                        table_values.append(row_entries)
                    if required and missing_required:
                        flash(f"'{label}' requires a value in every cell.", "error")
                        return render_template(
                            "form_render.html",
                            fs=fs,
                            sections=sections,
                            is_sectioned=is_sectioned,
                            category_label="Forms",
                            category_url=url_for('forms_list'),
                            subcategory_label=fs.name
                        )
                    table_payload = {
                        "type": "table",
                        "rows": rows,
                        "columns": columns,
                        "values": table_values,
                        "reference_image": field.get("reference_image") or ""
                    }
                    if is_sectioned:
                        section_entries.append({
                            "label": label,
                            "type": "table",
                            "table": table_payload
                        })
                    else:
                        values[label] = table_payload
                    continue

                if ftype in ["text", "textarea"]:
                    val = request.form.get(field_name, "").strip()
                    if required and not val:
                        flash(f"'{label}' is required", "error")
                        return render_template(
                            "form_render.html",
                            fs=fs,
                            sections=sections,
                            is_sectioned=is_sectioned,
                            category_label="Forms",
                            category_url=url_for('forms_list'),
                            subcategory_label=fs.name
                        )
                elif ftype == "select":
                    val = request.form.get(field_name, "")
                    normalized_val = val.strip().lower() if isinstance(val, str) else ""
                    if required and not val:
                        flash(f"'{label}' is required", "error")
                        return render_template(
                            "form_render.html",
                            fs=fs,
                            sections=sections,
                            is_sectioned=is_sectioned,
                            category_label="Forms",
                            category_url=url_for('forms_list'),
                            subcategory_label=fs.name
                        )
                    if options and val and val not in options:
                        flash(f"'{label}' has an invalid selection", "error")
                        return render_template(
                            "form_render.html",
                            fs=fs,
                            sections=sections,
                            is_sectioned=is_sectioned,
                            category_label="Forms",
                            category_url=url_for('forms_list'),
                            subcategory_label=fs.name
                        )
                    if normalized_val in {"ng", "not ok"}:
                        any_ng = True
                else:
                    val = request.form.get(field_name, "").strip()

                remark_val = ""
                if field.get("allow_remark"):
                    remark_val = request.form.get(remark_name, "").strip()

                item_saved_photos = []
                valid_item_files = []
                if field.get("allow_photo"):
                    upload_list = request.files.getlist(photo_name)
                    for f in upload_list:
                        if f and f.filename:
                            ext = f.filename.rsplit(".", 1)[1].lower() if "." in f.filename else ""
                            if ext in {"png", "jpg", "jpeg", "webp"}:
                                valid_item_files.append(f)
                    if (
                        field.get("photo_required_if_ng")
                        and normalized_val in {"ng", "not ok"}
                        and not valid_item_files
                    ):
                        flash(f"Photo evidence is required for '{label}' when marked Not OK.", "error")
                        return render_template(
                            "form_render.html",
                            fs=fs,
                            sections=sections,
                            is_sectioned=is_sectioned,
                            category_label="Forms",
                            category_url=url_for('forms_list'),
                            subcategory_label=fs.name
                        )
                    for f in valid_item_files:
                        fname = secure_filename(f.filename)
                        dest = os.path.join(app.config["UPLOAD_FOLDER"], f"{datetime.datetime.utcnow().timestamp()}_{fname}")
                        f.save(dest)
                        saved_path = dest
                        item_saved_photos.append(saved_path)
                        saved_photos.append(saved_path)
                        per_item_photo_count += 1

                if is_sectioned:
                    section_entries.append({
                        "label": label,
                        "type": ftype,
                        "value": val,
                        "remark": remark_val or None,
                        "photos": item_saved_photos
                    })
                else:
                    values[label] = val
                    if field.get("allow_remark") and remark_val:
                        values[f"{label} - Remark"] = remark_val

        photo_files = request.files.getlist("photos")
        video_files = request.files.getlist("videos")
        saved_videos = []

        valid_photo_files = [
            p for p in photo_files
            if p and p.filename and "." in p.filename and p.filename.rsplit(".", 1)[1].lower() in {"png", "jpg", "jpeg", "webp"}
        ]

        if any_ng and per_item_photo_count + len(valid_photo_files) == 0:
            flash("At least one photo is required when any item is marked Not OK.", "error")
            return render_template(
                "form_render.html",
                fs=fs,
                sections=sections,
                is_sectioned=is_sectioned,
                category_label="Forms",
                category_url=url_for('forms_list'),
                subcategory_label=fs.name
            )

        for f in valid_photo_files:
            fname = secure_filename(f.filename)
            dest = os.path.join(app.config["UPLOAD_FOLDER"], f"{datetime.datetime.utcnow().timestamp()}_{fname}")
            f.save(dest)
            saved_photos.append(dest)

        for f in video_files:
            if f and f.filename:
                ext = f.filename.rsplit(".",1)[1].lower() if "." in f.filename else ""
                if ext in {"mp4","mov","avi","mkv"}:
                    fname = secure_filename(f.filename)
                    dest = os.path.join(app.config["UPLOAD_FOLDER"], f"{datetime.datetime.utcnow().timestamp()}_{fname}")
                    f.save(dest)
                    saved_videos.append(dest)

        linked_work_id = request.args.get("work_id", type=int)
        sub = Submission(
            form_id=fs.id,
            submitted_by=current_user.id,
            data_json=json.dumps(values, ensure_ascii=False),
            photos_json=json.dumps(saved_photos, ensure_ascii=False),
            videos_json=json.dumps(saved_videos, ensure_ascii=False),
            work_id=linked_work_id
        )
        db.session.add(sub)
        db.session.commit()
        if linked_work_id:
            log_work_event(
                linked_work_id,
                "submission_created",
                actor_id=current_user.id,
                details={"submission_id": sub.id}
            )
            db.session.commit()
        flash("Submitted successfully!", "success")
        return redirect(url_for("dashboard"))

    return render_template(
        "form_render.html",
        fs=fs,
        sections=sections,
        is_sectioned=is_sectioned,
        category_label="Forms",
        category_url=url_for('forms_list'),
        subcategory_label=fs.name
    )


@app.route("/submissions/<int:sub_id>")
@login_required
def submission_view(sub_id):
    sub = db.session.get(Submission, sub_id)
    if not sub:
        flash("Submission not found", "error")
        return redirect(url_for("dashboard"))
    data = json.loads(sub.data_json or "{}")
    photos = json.loads(sub.photos_json or "[]")
    videos = json.loads(sub.videos_json or "[]")
    data_sectioned = any(
        isinstance(entries, list) and entries and isinstance(entries[0], dict) and "label" in entries[0]
        for entries in (data.values() if isinstance(data, dict) else [])
    )
    return render_template("submission_view.html", sub=sub, data=data, photos=photos, videos=videos,
                           data_sectioned=data_sectioned,
                           category_label="Dashboard", category_url=url_for('dashboard'),
                           subcategory_label=f"Submission #{sub.id}", subcategory_url=None)


# ---------------------- PROJECTS ----------------------
@app.route("/projects", methods=["GET", "POST"])
@login_required
def projects_list():
    _module_visibility_required("operations")
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        site_name = (request.form.get("site_name") or "").strip()
        site_address = (request.form.get("site_address") or "").strip()
        customer_name = (request.form.get("customer_name") or "").strip()
        lift_type = (request.form.get("lift_type") or "").strip()
        floors_raw = (request.form.get("floors") or "").strip()
        stops_raw = (request.form.get("stops") or "").strip()
        opening_type = (request.form.get("opening_type") or "").strip()
        location = (request.form.get("location") or "").strip()
        handover_raw = (request.form.get("handover_date") or "").strip()
        priority = (request.form.get("priority") or "").strip()
        structure_type = (request.form.get("structure_type") or "").strip()
        cladding_type = (request.form.get("cladding_type") or "").strip()
        cabin_finish = (request.form.get("cabin_finish") or "").strip()
        door_operation_type = (request.form.get("door_operation_type") or "").strip()
        door_finish = (request.form.get("door_finish") or "").strip()

        if not name:
            flash("Project name is required.", "error")
            return redirect(url_for("projects_list"))

        if lift_type and lift_type not in LIFT_TYPES:
            flash("Select a valid lift type.", "error")
            return redirect(url_for("projects_list"))

        floors = normalize_floor_label(floors_raw)

        stops = None
        if stops_raw:
            try:
                stops = int(stops_raw)
                if stops < 0:
                    raise ValueError
            except ValueError:
                flash("Number of stops must be a positive whole number.", "error")
                return redirect(url_for("projects_list"))

        if opening_type and opening_type not in PROJECT_OPENING_TYPES:
            flash("Choose a valid opening type.", "error")
            return redirect(url_for("projects_list"))

        if location and location not in PROJECT_LOCATIONS:
            flash("Choose a valid location.", "error")
            return redirect(url_for("projects_list"))

        if structure_type and structure_type not in PROJECT_STRUCTURE_TYPES:
            flash("Choose a valid structure type.", "error")
            return redirect(url_for("projects_list"))

        if cladding_type and cladding_type not in PROJECT_CLADDING_TYPES:
            flash("Choose a valid cladding type.", "error")
            return redirect(url_for("projects_list"))

        if cabin_finish and cabin_finish not in PROJECT_CABIN_FINISHES:
            flash("Choose a valid cabin finish.", "error")
            return redirect(url_for("projects_list"))

        if door_operation_type and door_operation_type not in PROJECT_DOOR_OPERATION_TYPES:
            flash("Choose a valid door operation type.", "error")
            return redirect(url_for("projects_list"))

        if door_finish and door_finish not in PROJECT_DOOR_FINISHES:
            flash("Choose a valid door finish.", "error")
            return redirect(url_for("projects_list"))

        if priority and priority not in PROJECT_PRIORITIES:
            flash("Choose a valid project priority.", "error")
            return redirect(url_for("projects_list"))

        handover_date = None
        if handover_raw:
            try:
                handover_date = datetime.datetime.strptime(handover_raw, "%Y-%m-%d").date()
            except ValueError:
                flash("Provide a valid handover date (YYYY-MM-DD).", "error")
                return redirect(url_for("projects_list"))

        project = Project(
            name=name,
            site_name=site_name or None,
            site_address=site_address or None,
            customer_name=customer_name or None,
            lift_type=lift_type or None,
            floors=floors,
            stops=stops,
            opening_type=opening_type or None,
            location=location or None,
            handover_date=handover_date,
            priority=priority or None,
            structure_type=structure_type or None,
            cladding_type=cladding_type or None,
            cabin_finish=cabin_finish or None,
            door_operation_type=door_operation_type or None,
            door_finish=door_finish or None,
        )
        db.session.add(project)
        db.session.commit()
        flash("Project created.", "success")
        return redirect(url_for("project_detail", project_id=project.id))

    projects = Project.query.order_by(Project.created_at.desc()).all()
    stats_rows = (
        db.session.query(
            QCWork.project_id,
            func.count(QCWork.id).label("total"),
            func.coalesce(func.sum(case((QCWork.status == "Open", 1), else_=0)), 0).label("open"),
            func.coalesce(func.sum(case((QCWork.status == "In Progress", 1), else_=0)), 0).label("in_progress"),
            func.coalesce(func.sum(case((QCWork.status == "Closed", 1), else_=0)), 0).label("closed")
        )
        .filter(QCWork.project_id.isnot(None))
        .group_by(QCWork.project_id)
        .all()
    )
    stats_map = {
        row.project_id: {
            "total": int(row.total or 0),
            "open": int(row.open or 0),
            "in_progress": int(row.in_progress or 0),
            "closed": int(row.closed or 0)
        }
        for row in stats_rows
    }
    return render_template(
        "projects.html",
        projects=projects,
        stats_map=stats_map,
        LIFT_TYPES=LIFT_TYPES,
        PROJECT_PRIORITIES=PROJECT_PRIORITIES,
        PROJECT_OPENING_TYPES=PROJECT_OPENING_TYPES,
        PROJECT_LOCATIONS=PROJECT_LOCATIONS,
        PROJECT_STRUCTURE_TYPES=PROJECT_STRUCTURE_TYPES,
        PROJECT_CLADDING_TYPES=PROJECT_CLADDING_TYPES,
        PROJECT_CABIN_FINISHES=PROJECT_CABIN_FINISHES,
        PROJECT_DOOR_OPERATION_TYPES=PROJECT_DOOR_OPERATION_TYPES,
        PROJECT_DOOR_FINISHES=PROJECT_DOOR_FINISHES,
    )


@app.route("/projects/<int:project_id>")
@login_required
def project_detail(project_id):
    _module_visibility_required("operations")
    project = Project.query.get_or_404(project_id)
    project_tasks = (
        QCWork.query
        .filter(QCWork.project_id == project.id)
        .order_by(QCWork.created_at.asc())
        .all()
    )
    active_tasks = [task for task in project_tasks if task.status != "Closed" and task.dependency_satisfied]
    waiting_tasks = [task for task in project_tasks if task.status != "Closed" and not task.dependency_satisfied]
    closed_tasks = [task for task in project_tasks if task.status == "Closed"]

    templates = ProjectTemplate.query.order_by(ProjectTemplate.name.asc()).all()
    form_templates = FormSchema.query.order_by(FormSchema.name.asc()).all()
    users = get_assignable_users_for_module("operations", order_by="username")
    comments = (
        ProjectComment.query.filter_by(project_id=project.id)
        .order_by(ProjectComment.created_at.desc())
        .all()
    )

    return render_template(
        "project_detail.html",
        project=project,
        active_tasks=active_tasks,
        waiting_tasks=waiting_tasks,
        closed_tasks=closed_tasks,
        all_tasks=project_tasks,
        templates=templates,
        form_templates=form_templates,
        users=users,
        comments=comments,
        LIFT_TYPES=LIFT_TYPES,
        DEFAULT_TASK_FORM_NAME=DEFAULT_TASK_FORM_NAME,
        STAGES=STAGES,
        TASK_MILESTONES=TASK_MILESTONES,
        PROJECT_PRIORITIES=PROJECT_PRIORITIES,
        PROJECT_OPENING_TYPES=PROJECT_OPENING_TYPES,
        PROJECT_LOCATIONS=PROJECT_LOCATIONS,
        PROJECT_STRUCTURE_TYPES=PROJECT_STRUCTURE_TYPES,
        PROJECT_CLADDING_TYPES=PROJECT_CLADDING_TYPES,
        PROJECT_CABIN_FINISHES=PROJECT_CABIN_FINISHES,
        PROJECT_DOOR_OPERATION_TYPES=PROJECT_DOOR_OPERATION_TYPES,
        PROJECT_DOOR_FINISHES=PROJECT_DOOR_FINISHES,
    )


@app.route("/projects/<int:project_id>/edit", methods=["POST"])
@login_required
def project_edit(project_id):
    _module_visibility_required("operations")
    project = Project.query.get_or_404(project_id)

    name = (request.form.get("name") or "").strip()
    site_name = (request.form.get("site_name") or "").strip()
    site_address = (request.form.get("site_address") or "").strip()
    customer_name = (request.form.get("customer_name") or "").strip()
    lift_type = (request.form.get("lift_type") or "").strip()
    floors_raw = (request.form.get("floors") or "").strip()
    stops_raw = (request.form.get("stops") or "").strip()
    opening_type = (request.form.get("opening_type") or "").strip()
    location = (request.form.get("location") or "").strip()
    handover_raw = (request.form.get("handover_date") or "").strip()
    priority = (request.form.get("priority") or "").strip()
    structure_type = (request.form.get("structure_type") or "").strip()
    cladding_type = (request.form.get("cladding_type") or "").strip()
    cabin_finish = (request.form.get("cabin_finish") or "").strip()
    door_operation_type = (request.form.get("door_operation_type") or "").strip()
    door_finish = (request.form.get("door_finish") or "").strip()

    if not name:
        flash("Project name is required.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    if lift_type and lift_type not in LIFT_TYPES:
        flash("Select a valid lift type.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    floors = normalize_floor_label(floors_raw)

    stops = None
    if stops_raw:
        try:
            stops = int(stops_raw)
            if stops < 0:
                raise ValueError
        except ValueError:
            flash("Number of stops must be a positive whole number.", "error")
            return redirect(url_for("project_detail", project_id=project.id))

    if opening_type and opening_type not in PROJECT_OPENING_TYPES:
        flash("Choose a valid opening type.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    if location and location not in PROJECT_LOCATIONS:
        flash("Choose a valid location.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    if structure_type and structure_type not in PROJECT_STRUCTURE_TYPES:
        flash("Choose a valid structure type.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    if cladding_type and cladding_type not in PROJECT_CLADDING_TYPES:
        flash("Choose a valid cladding type.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    if cabin_finish and cabin_finish not in PROJECT_CABIN_FINISHES:
        flash("Choose a valid cabin finish.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    if door_operation_type and door_operation_type not in PROJECT_DOOR_OPERATION_TYPES:
        flash("Choose a valid door operation type.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    if door_finish and door_finish not in PROJECT_DOOR_FINISHES:
        flash("Choose a valid door finish.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    if priority and priority not in PROJECT_PRIORITIES:
        flash("Choose a valid project priority.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    handover_date = None
    if handover_raw:
        try:
            handover_date = datetime.datetime.strptime(handover_raw, "%Y-%m-%d").date()
        except ValueError:
            flash("Provide a valid handover date (YYYY-MM-DD).", "error")
            return redirect(url_for("project_detail", project_id=project.id))

    project.name = name
    project.site_name = site_name or None
    project.site_address = site_address or None
    project.customer_name = customer_name or None
    project.lift_type = lift_type or None
    project.floors = floors
    project.stops = stops
    project.opening_type = opening_type or None
    project.location = location or None
    project.handover_date = handover_date
    project.priority = priority or None
    project.structure_type = structure_type or None
    project.cladding_type = cladding_type or None
    project.cabin_finish = cabin_finish or None
    project.door_operation_type = door_operation_type or None
    project.door_finish = door_finish or None

    db.session.commit()
    flash("Project updated.", "success")
    return redirect(url_for("project_detail", project_id=project.id))


@app.route("/projects/<int:project_id>/comments", methods=["POST"])
@login_required
def project_add_comment(project_id):
    _module_visibility_required("operations")

    project = Project.query.get_or_404(project_id)
    body = (request.form.get("body") or "").strip()
    if not body:
        flash("Comment cannot be empty.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    comment = ProjectComment(
        project=project,
        body=body,
        author=current_user if current_user.is_authenticated else None,
    )
    db.session.add(comment)
    db.session.commit()

    flash("Comment added.", "success")
    return redirect(url_for("project_detail", project_id=project.id))


@app.route("/projects/<int:project_id>/apply-template", methods=["POST"])
@login_required
def project_apply_template(project_id):
    _module_visibility_required("operations")
    project = Project.query.get_or_404(project_id)
    template_id = request.form.get("template_id", type=int)
    if not template_id:
        flash("Select a template to apply.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    template = ProjectTemplate.query.get(template_id)
    if not template:
        flash("Template not found.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    fallback_form = get_or_create_default_task_form()
    existing_template_task_ids = {
        task.template_task_id
        for task in QCWork.query.filter(
            QCWork.project_id == project.id,
            QCWork.template_task_id.isnot(None)
        ).all()
        if task.template_task_id is not None
    }

    created = []
    task_lookup = {}
    ordered_template_tasks = sorted(template.tasks, key=lambda t: ((t.order_index or 0), t.id))

    for template_task in ordered_template_tasks:
        if template_task.id in existing_template_task_ids:
            continue

        dependency_tasks = []
        for dep_id in template_task.dependency_ids:
            dependency = task_lookup.get(dep_id)
            if not dependency:
                dependency = QCWork.query.filter_by(
                    project_id=project.id,
                    template_task_id=dep_id
                ).first()
            if dependency:
                dependency_tasks.append(dependency)

        form_template = template_task.form_template or fallback_form
        if not form_template:
            continue

        status = "Open"
        if any((dep.status or "").lower() != "closed" for dep in dependency_tasks):
            status = "Blocked"

        planned_start = template_task.planned_start_date
        planned_duration = template_task.duration_days
        due_date = None
        if planned_start and planned_duration:
            due_date = datetime.datetime.combine(planned_start, datetime.time.min) + datetime.timedelta(days=planned_duration)
        elif planned_duration and (template_task.start_mode or "immediate") != "after_previous":
            due_date = datetime.datetime.utcnow() + datetime.timedelta(days=planned_duration)
        elif template_task.planned_due_date:
            due_date = datetime.datetime.combine(template_task.planned_due_date, datetime.time.min)

        milestone_value = template_task.milestone or None

        assigned_user = None
        default_assignee_id = template_task.default_assignee_id
        if default_assignee_id:
            assigned_user = db.session.get(User, default_assignee_id)
            if not assigned_user or not assigned_user.can_be_assigned_module("operations"):
                assigned_user = None
                default_assignee_id = None

        new_task = QCWork(
            site_name=project.site_name or project.name,
            client_name=project.customer_name,
            address=project.site_address,
            template_id=form_template.id,
            stage=template_task.template.name,
            lift_type=project.lift_type or form_template.lift_type,
            project_id=project.id,
            created_by=current_user.id,
            assigned_to=default_assignee_id,
            name=template_task.name,
            description=template_task.description,
            template_task_id=template_task.id,
            status=status,
            due_date=due_date,
            planned_start_date=planned_start,
            planned_duration_days=planned_duration,
            milestone=milestone_value
        )
        db.session.add(new_task)
        db.session.flush()
        set_qc_work_dependencies(new_task, [dep.id for dep in dependency_tasks])

        task_lookup[template_task.id] = new_task
        created.append(new_task)

        log_work_event(
            new_task.id,
            "created_from_project_template",
            actor_id=current_user.id,
            details={
                "project_id": project.id,
                "template": template.name,
                "template_task": template_task.name,
                "planned_start_date": planned_start.strftime("%Y-%m-%d") if planned_start else None,
                "planned_duration_days": planned_duration,
                "milestone": milestone_value,
                "due_date": due_date.strftime("%Y-%m-%d") if due_date else None
            }
        )
        if assigned_user:
            log_work_event(
                new_task.id,
                "assigned",
                actor_id=current_user.id,
                details={"assigned_to": assigned_user.id}
            )
        if status == "Blocked" and dependency_tasks:
            log_work_event(
                new_task.id,
                "waiting_on_dependency",
                actor_id=current_user.id,
                details={"depends_on": [dep.id for dep in dependency_tasks]}
            )

    if not created:
        flash("No new tasks were created – they may already exist for this project.", "info")
        return redirect(url_for("project_detail", project_id=project.id))

    db.session.commit()
    flash(f"Added {len(created)} tasks from template {template.name}.", "success")
    return redirect(url_for("project_detail", project_id=project.id))


@app.route("/projects/<int:project_id>/tasks/create", methods=["POST"])
@login_required
def project_task_create(project_id):
    _module_visibility_required("operations")
    project = Project.query.get_or_404(project_id)
    name = (request.form.get("name") or "").strip()
    description = (request.form.get("description") or "").strip()
    form_template_id = request.form.get("form_template_id", type=int)
    assigned_to = request.form.get("assigned_to", type=int)
    due = (request.form.get("due_date") or "").strip()
    depends_on_ids = []
    for raw in request.form.getlist("depends_on_ids"):
        try:
            dep_id = int(raw)
        except (TypeError, ValueError):
            continue
        if dep_id not in depends_on_ids:
            depends_on_ids.append(dep_id)
    stage = (request.form.get("stage") or "").strip()
    planned_start = (request.form.get("planned_start_date") or "").strip()
    duration_raw = (request.form.get("planned_duration_days") or "").strip()
    milestone_value = (request.form.get("milestone") or "").strip()

    if not name:
        flash("Provide a task name.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    form_template = db.session.get(FormSchema, form_template_id) if form_template_id else None
    if not form_template:
        form_template = get_or_create_default_task_form()
    if not form_template:
        flash("Set up a form template before creating tasks.", "error")
        return redirect(url_for("project_detail", project_id=project.id))

    due_dt = None
    if due:
        try:
            due_dt = datetime.datetime.strptime(due, "%Y-%m-%d")
        except Exception:
            flash("Invalid due date format.", "error")
            return redirect(url_for("project_detail", project_id=project.id))

    planned_start_date = None
    if planned_start:
        try:
            planned_start_date = datetime.datetime.strptime(planned_start, "%Y-%m-%d").date()
        except ValueError:
            flash("Invalid planned start date format.", "error")
            return redirect(url_for("project_detail", project_id=project.id))

    duration_days = None
    if duration_raw:
        try:
            duration_days = int(duration_raw)
        except ValueError:
            flash("Duration must be a whole number of days.", "error")
            return redirect(url_for("project_detail", project_id=project.id))
        if duration_days < 0:
            flash("Duration must be zero or positive.", "error")
            return redirect(url_for("project_detail", project_id=project.id))

    if due_dt is None:
        if planned_start_date and duration_days is not None:
            due_dt = datetime.datetime.combine(planned_start_date, datetime.time.min) + datetime.timedelta(days=duration_days)
        elif duration_days is not None:
            due_dt = datetime.datetime.utcnow() + datetime.timedelta(days=duration_days)

    dependency_tasks = []
    assignee_user = None
    if assigned_to:
        assignee_user = db.session.get(User, assigned_to)
        if not assignee_user or not assignee_user.can_be_assigned_module("operations"):
            flash("Choose an assignee who is available for Operations tasks.", "error")
            return redirect(url_for("project_detail", project_id=project.id))
    if depends_on_ids:
        dependencies = (
            QCWork.query
            .filter(QCWork.project_id == project.id, QCWork.id.in_(depends_on_ids))
            .all()
        )
        found_ids = {dep.id for dep in dependencies}
        missing = [dep_id for dep_id in depends_on_ids if dep_id not in found_ids]
        if missing:
            flash("Choose dependencies from the same project.", "error")
            return redirect(url_for("project_detail", project_id=project.id))
        id_map = {dep.id: dep for dep in dependencies}
        dependency_tasks = [id_map[dep_id] for dep_id in depends_on_ids if dep_id in id_map]

    status = "Open"
    if any((dep.status or "").lower() != "closed" for dep in dependency_tasks):
        status = "Blocked"

    work = QCWork(
        site_name=project.site_name or project.name,
        client_name=project.customer_name,
        address=project.site_address,
        template_id=form_template.id,
        stage=stage or None,
        lift_type=project.lift_type or form_template.lift_type,
        project_id=project.id,
        due_date=due_dt,
        created_by=current_user.id,
        assigned_to=assigned_to if assignee_user else None,
        name=name,
        description=description or None,
        status=status,
        planned_start_date=planned_start_date,
        planned_duration_days=duration_days,
        milestone=milestone_value or None
    )
    db.session.add(work)
    db.session.flush()
    set_qc_work_dependencies(work, [dep.id for dep in dependency_tasks])
    log_work_event(
        work.id,
        "created_from_project",
        actor_id=current_user.id,
        details={
            "project_id": project.id,
            "stage": stage or None,
            "assigned_to": assigned_to,
            "due_date": work.due_date.strftime("%Y-%m-%d") if work.due_date else None,
            "dependencies": [dep.id for dep in dependency_tasks],
            "planned_start_date": planned_start_date.strftime("%Y-%m-%d") if planned_start_date else None,
            "planned_duration_days": duration_days,
            "milestone": work.milestone
        }
    )
    if assignee_user:
        log_work_event(
            work.id,
            "assigned",
            actor_id=current_user.id,
            details={"assigned_to": assigned_to}
        )
    if status == "Blocked" and dependency_tasks:
        log_work_event(
            work.id,
            "waiting_on_dependency",
            actor_id=current_user.id,
            details={"depends_on": [dep.id for dep in dependency_tasks]}
        )
    db.session.commit()
    flash("Project task created.", "success")
    return redirect(url_for("qc_work_detail", work_id=work.id))


@app.route("/project-templates", methods=["GET", "POST"])
@login_required
def project_templates():
    _module_visibility_required("operations")
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        description = (request.form.get("description") or "").strip()
        if not name:
            flash("Template name is required.", "error")
        else:
            template = ProjectTemplate(
                name=name,
                description=description or None,
                created_by=current_user.id
            )
            db.session.add(template)
            db.session.commit()
            flash("Project template created.", "success")
            return redirect(url_for("project_template_detail", template_id=template.id))

    templates = ProjectTemplate.query.order_by(ProjectTemplate.name.asc()).all()
    template_counts = {
        tpl.id: len(tpl.tasks)
        for tpl in templates
    }
    task_templates = TaskTemplate.query.order_by(TaskTemplate.created_at.desc()).all()
    return render_template(
        "project_templates.html",
        templates=templates,
        template_counts=template_counts,
        task_templates=task_templates,
        DEFAULT_TASK_FORM_NAME=DEFAULT_TASK_FORM_NAME
    )


@app.route("/project-templates/<int:template_id>", methods=["GET", "POST"])
@login_required
def project_template_detail(template_id):
    _module_visibility_required("operations")
    template = ProjectTemplate.query.get_or_404(template_id)
    users = get_assignable_users_for_module("operations", order_by="username")
    forms = FormSchema.query.order_by(FormSchema.name.asc()).all()

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        description = (request.form.get("description") or "").strip()
        requested_order = request.form.get("order_index", type=int)
        depends_on_ids = []
        for raw in request.form.getlist("depends_on_ids"):
            try:
                dep_id = int(raw)
            except (TypeError, ValueError):
                continue
            if dep_id not in depends_on_ids:
                depends_on_ids.append(dep_id)
        default_assignee_id = request.form.get("default_assignee_id", type=int)
        form_template_id = request.form.get("form_template_id", type=int)

        if not name:
            flash("Task name is required.", "error")
            return redirect(url_for("project_template_detail", template_id=template.id))

        if depends_on_ids:
            dependencies = ProjectTemplateTask.query.filter(
                ProjectTemplateTask.template_id == template.id,
                ProjectTemplateTask.id.in_(depends_on_ids)
            ).all()
            if len(dependencies) != len(depends_on_ids):
                flash("Select dependencies from the same template.", "error")
                return redirect(url_for("project_template_detail", template_id=template.id))

        if default_assignee_id:
            assignee_candidate = db.session.get(User, default_assignee_id)
            if not assignee_candidate or not assignee_candidate.can_be_assigned_module("operations"):
                flash("Choose a valid default assignee with Operations access.", "error")
                return redirect(url_for("project_template_detail", template_id=template.id))

        if form_template_id and not db.session.get(FormSchema, form_template_id):
            flash("Choose a valid form template.", "error")
            return redirect(url_for("project_template_detail", template_id=template.id))

        start_mode, planned_start_date, duration_value, milestone_value, timing_error = _extract_task_timing(request.form)
        if timing_error:
            flash(timing_error, "error")
            return redirect(url_for("project_template_detail", template_id=template.id))

        existing_tasks = ProjectTemplateTask.query.filter_by(template_id=template.id).order_by(
            ProjectTemplateTask.order_index.asc(),
            ProjectTemplateTask.id.asc()
        ).all()
        max_position = len(existing_tasks) + 1
        if not requested_order or requested_order < 1:
            requested_order = max_position
        else:
            requested_order = min(requested_order, max_position)

        shifts = False
        for idx, existing in enumerate(existing_tasks, start=1):
            new_index = idx if idx < requested_order else idx + 1
            if existing.order_index != new_index:
                shifts = True
            existing.order_index = new_index

        task = ProjectTemplateTask(
            template_id=template.id,
            name=name,
            description=description or None,
            order_index=requested_order,
            default_assignee_id=default_assignee_id,
            form_template_id=form_template_id,
            start_mode=start_mode,
            planned_start_date=planned_start_date,
            duration_days=duration_value,
            milestone=milestone_value
        )
        db.session.add(task)
        db.session.flush()
        set_template_task_dependencies(task, depends_on_ids)
        normalize_template_task_order(template.id)
        db.session.commit()
        message = "Template task added."
        if shifts:
            message = "Template task added. Existing tasks were re-ordered automatically."
        flash(message, "success")
        return redirect(url_for("project_template_detail", template_id=template.id))

    tasks = ProjectTemplateTask.query.filter_by(template_id=template.id).order_by(
        ProjectTemplateTask.order_index.asc(),
        ProjectTemplateTask.id.asc()
    ).all()
    return render_template(
        "project_template_detail.html",
        template=template,
        tasks=tasks,
        users=users,
        forms=forms,
        DEFAULT_TASK_FORM_NAME=DEFAULT_TASK_FORM_NAME,
        TASK_MILESTONES=TASK_MILESTONES
    )


@app.route("/project-templates/<int:template_id>/update", methods=["POST"])
@login_required
def project_template_update(template_id):
    _module_visibility_required("operations")
    template = ProjectTemplate.query.get_or_404(template_id)
    name = (request.form.get("template_name") or request.form.get("name") or "").strip()
    description = (request.form.get("template_description") or request.form.get("description") or "").strip()
    if not name:
        flash("Template name is required.", "error")
        return redirect(url_for("project_template_detail", template_id=template.id))

    template.name = name
    template.description = description or None
    db.session.commit()
    flash("Template details saved.", "success")
    return redirect(url_for("project_template_detail", template_id=template.id))


@app.route("/project-templates/<int:template_id>/save-as-task-template", methods=["POST"])
@login_required
def project_template_save_as_task_template(template_id):
    _module_visibility_required("operations")
    template = ProjectTemplate.query.get_or_404(template_id)
    name = (request.form.get("task_template_name") or "").strip()
    description = (request.form.get("task_template_description") or "").strip()
    if not name:
        flash("Provide a name for the task template.", "error")
        return redirect(url_for("project_template_detail", template_id=template.id))

    existing = TaskTemplate.query.filter(func.lower(TaskTemplate.name) == name.lower()).first()
    if existing:
        flash("A task template with that name already exists.", "error")
        return redirect(url_for("project_template_detail", template_id=template.id))

    blueprint = build_task_template_blueprint(template)
    task_template = TaskTemplate(
        name=name,
        description=description or None,
        blueprint_json=json.dumps(blueprint, ensure_ascii=False),
        created_from_template_id=template.id,
        created_by=current_user.id
    )
    db.session.add(task_template)
    db.session.commit()
    flash("Task template saved for future reuse.", "success")
    return redirect(url_for("project_template_detail", template_id=template.id))


@app.route("/task-templates/<int:task_template_id>/create-project-template", methods=["POST"])
@login_required
def create_project_template_from_task_template(task_template_id):
    task_template = TaskTemplate.query.get_or_404(task_template_id)
    name = (request.form.get("name") or task_template.name or "").strip()
    description = (request.form.get("description") or task_template.description or "").strip()
    if not name:
        flash("Template name is required.", "error")
        return redirect(url_for("project_templates"))

    new_template = ProjectTemplate(
        name=name,
        description=description or None,
        created_by=current_user.id
    )
    db.session.add(new_template)
    db.session.flush()

    try:
        blueprint = json.loads(task_template.blueprint_json or "[]")
    except json.JSONDecodeError:
        blueprint = []

    apply_blueprint_to_template(new_template, blueprint)
    db.session.commit()
    flash(f"Project template '{name}' created from task template.", "success")
    return redirect(url_for("project_template_detail", template_id=new_template.id))


@app.route("/project-templates/<int:template_id>/delete", methods=["POST"])
@login_required
def project_template_delete(template_id):
    _module_visibility_required("operations")
    template = ProjectTemplate.query.get_or_404(template_id)
    if not (current_user.is_admin or template.created_by == current_user.id):
        flash("You do not have permission to delete this template.", "error")
        return redirect(url_for("project_templates"))

    name = template.name or f'Template {template.id}'
    db.session.delete(template)
    db.session.commit()
    flash(f"Project template '{name}' deleted.", "success")
    return redirect(url_for("project_templates"))



# ---- Template task management helpers ----
@app.route("/project-templates/<int:template_id>/tasks/<int:task_id>/edit", methods=["POST"])
@login_required
def project_template_task_edit(template_id, task_id):
    _module_visibility_required("operations")
    template = ProjectTemplate.query.get_or_404(template_id)
    task = ProjectTemplateTask.query.filter_by(id=task_id, template_id=template.id).first()
    if not task:
        flash("Task not found for this template.", "error")
        return redirect(url_for("project_template_detail", template_id=template.id))

    name = (request.form.get("name") or "").strip()
    description = (request.form.get("description") or "").strip()
    requested_order = request.form.get("order_index", type=int)
    depends_on_ids = []
    for raw in request.form.getlist("depends_on_ids"):
        try:
            dep_id = int(raw)
        except (TypeError, ValueError):
            continue
        if dep_id not in depends_on_ids:
            depends_on_ids.append(dep_id)
    default_assignee_id = request.form.get("default_assignee_id", type=int)
    form_template_id = request.form.get("form_template_id", type=int)

    if not name:
        flash("Task name is required.", "error")
        return redirect(url_for("project_template_detail", template_id=template.id))

    if depends_on_ids:
        if task.id in depends_on_ids:
            flash("A task cannot depend on itself.", "error")
            return redirect(url_for("project_template_detail", template_id=template.id))
        dependencies = ProjectTemplateTask.query.filter(
            ProjectTemplateTask.template_id == template.id,
            ProjectTemplateTask.id.in_(depends_on_ids)
        ).all()
        if len(dependencies) != len(depends_on_ids):
            flash("Select dependencies from the same template.", "error")
            return redirect(url_for("project_template_detail", template_id=template.id))

    if default_assignee_id:
        assignee_candidate = db.session.get(User, default_assignee_id)
        if not assignee_candidate or not assignee_candidate.can_be_assigned_module("operations"):
            flash("Choose a valid default assignee with Operations access.", "error")
            return redirect(url_for("project_template_detail", template_id=template.id))

    if form_template_id and not db.session.get(FormSchema, form_template_id):
        flash("Choose a valid form template.", "error")
        return redirect(url_for("project_template_detail", template_id=template.id))

    start_mode, planned_start_date, duration_value, milestone_value, timing_error = _extract_task_timing(request.form)
    if timing_error:
        flash(timing_error, "error")
        return redirect(url_for("project_template_detail", template_id=template.id))

    existing_tasks = ProjectTemplateTask.query.filter_by(template_id=template.id).order_by(
        ProjectTemplateTask.order_index.asc(),
        ProjectTemplateTask.id.asc()
    ).all()
    others = [t for t in existing_tasks if t.id != task.id]
    max_position = len(others) + 1
    if not requested_order or requested_order < 1:
        requested_order = max_position if task.order_index is None else min(task.order_index, max_position)
    else:
        requested_order = min(requested_order, max_position)

    for idx, existing in enumerate(others, start=1):
        new_index = idx if idx < requested_order else idx + 1
        existing.order_index = new_index

    task.name = name
    task.description = description or None
    task.order_index = requested_order
    task.default_assignee_id = default_assignee_id
    task.form_template_id = form_template_id
    task.start_mode = start_mode
    task.planned_start_date = planned_start_date
    task.duration_days = duration_value
    task.milestone = milestone_value

    set_template_task_dependencies(task, depends_on_ids)

    normalize_template_task_order(template.id)
    db.session.commit()
    flash("Template task updated.", "success")
    return redirect(url_for("project_template_detail", template_id=template.id))


@app.route("/project-templates/<int:template_id>/tasks/reorder", methods=["POST"])
@login_required
def project_template_task_reorder(template_id):
    _module_visibility_required("operations")
    template = ProjectTemplate.query.get_or_404(template_id)
    payload = request.get_json(silent=True) or {}
    ordered_ids = payload.get("ordered_ids")
    if not isinstance(ordered_ids, list) or not ordered_ids:
        return jsonify({"ok": False, "message": "No ordering received."}), 400

    try:
        ordered_ids = [int(task_id) for task_id in ordered_ids]
    except (TypeError, ValueError):
        return jsonify({"ok": False, "message": "Invalid task identifiers."}), 400

    template_task_ids = {task.id for task in template.tasks}
    if set(ordered_ids) != template_task_ids:
        return jsonify({"ok": False, "message": "Ordering must include all template tasks."}), 400

    tasks = {
        task.id: task for task in ProjectTemplateTask.query.filter(
            ProjectTemplateTask.template_id == template.id,
            ProjectTemplateTask.id.in_(ordered_ids)
        ).all()
    }

    if len(tasks) != len(ordered_ids):
        return jsonify({"ok": False, "message": "One or more tasks were not found."}), 400

    for idx, task_id in enumerate(ordered_ids, start=1):
        tasks[task_id].order_index = idx

    db.session.commit()
    return jsonify({"ok": True})


# ---------------------- CUSTOMER SUPPORT MODULE ----------------------
@app.route("/customer-support")
@login_required
def customer_support_home():
    _module_visibility_required("customer_support")
    return redirect(url_for("customer_support_overview"))


@app.route("/customer-support/overview")
@login_required
def customer_support_overview():
    _module_visibility_required("customer_support")
    summary = _customer_support_summary()
    counts = summary["counts"]
    recent_tickets = []
    for ticket in summary["recent"][:5]:
        recent_tickets.append(
            {
                **ticket,
                "created_display": ticket["created_at"].strftime("%d %b %Y · %I:%M %p"),
                "updated_display": (ticket.get("updated_at") or ticket["created_at"]).strftime("%d %b %Y · %I:%M %p"),
            }
        )

    kpis = [
        {"label": "Open", "value": counts.get("Open", 0), "tone": "rose"},
        {"label": "In Progress", "value": counts.get("In Progress", 0), "tone": "amber"},
        {"label": "Resolved", "value": counts.get("Resolved", 0), "tone": "emerald"},
        {"label": "Closed", "value": counts.get("Closed", 0), "tone": "sky"},
    ]

    return render_template(
        "customer_support_overview.html",
        kpis=kpis,
        total_tickets=summary["total"],
        recent_tickets=recent_tickets,
        categories=CUSTOMER_SUPPORT_CATEGORIES,
        channels=CUSTOMER_SUPPORT_CHANNELS,
    )


@app.route("/customer-support/tasks", methods=["GET", "POST"])
@login_required
def customer_support_tasks():
    _module_visibility_required("customer_support")
    if request.method == "POST":
        response = _handle_customer_support_ticket_creation()
        if response is not None:
            return response
    now = datetime.datetime.utcnow()
    tickets = []
    for ticket in CUSTOMER_SUPPORT_TICKETS:
        _resolve_ticket_assignee_user(ticket)
        sla_due_at = _calculate_ticket_sla_due(ticket)
        ticket["_sla_due_at"] = sla_due_at
        ticket["_sla_due_iso"] = sla_due_at.isoformat() if sla_due_at else ""
        ticket["_sla_due_display"] = sla_due_at.strftime("%d %b %Y %H:%M") if sla_due_at else "No SLA"
        ticket["_sla_seconds_remaining"] = (sla_due_at - now).total_seconds() if sla_due_at else None
        tickets.append(ticket)

    tickets.sort(key=lambda ticket: ticket.get("_sla_due_at") or datetime.datetime.max)
    ticket_open_task_map = {
        ticket.get("id"): _ticket_has_open_linked_tasks(ticket) for ticket in tickets
    }
    ticket_id = request.args.get("ticket")
    selected_ticket = _get_customer_support_ticket(ticket_id)

    timeline = []
    attachments = []
    linked_tasks = []
    open_linked_task_modal = request.args.get("open_linked_task") == "1"
    ticket_first_update = None
    has_open_linked_tasks = False
    if selected_ticket:
        timeline = sorted(
            (
                apply_actor_context(event)
                for event in selected_ticket.get("timeline", [])
            ),
            key=lambda event: event.get("timestamp"),
            reverse=True,
        )
        attachments = selected_ticket.get("attachments", [])
        linked_tasks = selected_ticket.get("linked_tasks", [])
        if linked_tasks:
            has_open_linked_tasks = _ticket_has_open_linked_tasks(selected_ticket)

        timeline_chronological = sorted(
            (
                apply_actor_context(event)
                for event in selected_ticket.get("timeline", [])
            ),
            key=lambda event: event.get("timestamp"),
        )
        for entry in timeline_chronological:
            if entry.get("comment"):
                ticket_first_update = entry
                break
        if ticket_first_update is None and timeline_chronological:
            ticket_first_update = timeline_chronological[0]

    active_support_users = [
        user for user in get_assignable_users_for_module("customer_support") if user.is_active
    ]

    sales_clients = (
        SalesClient.query.order_by(func.lower(func.coalesce(SalesClient.display_name, ""))).all()
    )
    installation_projects = Project.query.order_by(func.lower(Project.name)).all()
    amc_customers = (
        Customer.query.order_by(func.lower(func.coalesce(Customer.company_name, ""))).all()
    )

    return render_template(
        "customer_support_tasks.html",
        tickets=tickets,
        selected_ticket=selected_ticket,
        timeline=timeline,
        attachments=attachments,
        linked_tasks=linked_tasks,
        ticket_first_update=ticket_first_update,
        support_categories=CUSTOMER_SUPPORT_CATEGORIES,
        channels=CUSTOMER_SUPPORT_CHANNELS,
        sla_presets=CUSTOMER_SUPPORT_SLA_PRESETS,
        status_options=["Open", "In Progress", "Resolved", "Closed"],
        priority_options=["Low", "Medium", "High", "Critical"],
        open_ticket_modal=bool(selected_ticket),
        open_linked_task_modal=open_linked_task_modal,
        has_open_linked_tasks=has_open_linked_tasks,
        ticket_open_task_map=ticket_open_task_map,
        active_support_users=active_support_users,
        amc_lifts=_customer_support_amc_site_options(),
        sales_clients=sales_clients,
        installation_projects=installation_projects,
        amc_customers=amc_customers,
    )


@app.route("/customer-support/linked-tasks", methods=["POST"])
@login_required
def customer_support_create_linked_task():
    _module_visibility_required("customer_support")
    ticket_id = (request.form.get("ticket_id") or "").strip()
    if not ticket_id:
        flash("Select a ticket before creating a linked task.", "error")
        return redirect(url_for("customer_support_tasks"))

    ticket = _get_customer_support_ticket(ticket_id)
    if not ticket:
        flash("The referenced ticket could not be found.", "error")
        return redirect(url_for("customer_support_tasks"))

    title = (request.form.get("title") or "").strip()
    details = (request.form.get("details") or "").strip()
    assignee_id_raw = (request.form.get("assignee") or "").strip()
    due_date_raw = (request.form.get("due_date") or "").strip()
    category = (request.form.get("category") or "").strip()
    priority = (request.form.get("priority") or "").strip()

    errors = []
    if not title:
        errors.append("Provide a title for the linked task.")

    due_date = None
    if due_date_raw:
        try:
            due_date = datetime.datetime.strptime(due_date_raw, "%Y-%m-%d").date()
        except ValueError:
            errors.append("Enter the due date in YYYY-MM-DD format.")

    assignee_user = None
    if assignee_id_raw:
        try:
            assignee_user = db.session.get(User, int(assignee_id_raw))
        except (TypeError, ValueError):
            assignee_user = None

        if not assignee_user:
            errors.append("Select a valid assignee for the linked task.")
        elif not assignee_user.is_active:
            errors.append("The selected assignee is not active on the portal.")
        elif not assignee_user.can_be_assigned_module("customer_support"):
            errors.append("The selected user cannot be assigned to customer support tasks.")

    if errors:
        for message in errors:
            flash(message, "error")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id, open_linked_task="1"))

    category_label = None
    if category:
        category_label = next((item.get("label") for item in CUSTOMER_SUPPORT_CATEGORIES if item.get("id") == category), None)

    assignee_label = assignee_user.display_name if assignee_user else "Unassigned"

    new_task = {
        "id": generate_linked_task_id(),
        "title": title,
        "assignee": assignee_label,
        "assignee_id": assignee_user.id if assignee_user else None,
        "status": "Open",
        "due_date": due_date,
        "details": details or None,
        "category": category_label or category or None,
        "priority": priority or "Medium",
        "created_at": datetime.datetime.utcnow(),
    }

    ticket.setdefault("linked_tasks", []).append(new_task)
    flash("Linked task created successfully.", "success")
    return redirect(url_for("customer_support_tasks", ticket=ticket_id))


@app.route("/customer-support/tickets/<ticket_id>/update", methods=["POST"])
@login_required
def customer_support_update_ticket(ticket_id):
    _module_visibility_required("customer_support")
    ticket = _get_customer_support_ticket(ticket_id)
    if not ticket:
        flash("The requested ticket could not be found.", "error")
        return redirect(url_for("customer_support_tasks"))

    status = (request.form.get("status") or ticket.get("status") or "Open").strip()
    priority = (request.form.get("priority") or ticket.get("priority") or "Medium").strip()
    assignee_value = (request.form.get("assignee") or "").strip()
    closing_comment = (request.form.get("closing_comment") or "").strip()

    allowed_status = {"Open", "In Progress", "Resolved", "Closed"}
    allowed_priority = {"Low", "Medium", "High", "Critical"}
    errors = []
    if status not in allowed_status:
        errors.append("Choose a valid status for the ticket.")
    if priority not in allowed_priority:
        errors.append("Choose a valid priority for the ticket.")

    assignee_user = None
    new_assignee_label = "Unassigned"
    new_assignee_id = None
    current_assignee_label = ticket.get("assignee") or "Unassigned"
    current_assignee_id = ticket.get("assignee_user_id")

    if assignee_value:
        try:
            assignee_user_id = int(assignee_value)
        except (TypeError, ValueError):
            assignee_user_id = None
        if assignee_user_id is None:
            if assignee_value == current_assignee_label:
                new_assignee_label = current_assignee_label
                new_assignee_id = current_assignee_id
            else:
                errors.append("Select an assignee from the available team members.")
        else:
            assignee_user = User.query.get(assignee_user_id)
            if (
                not assignee_user
                or not assignee_user.is_active
                or not assignee_user.can_be_assigned_module("customer_support")
            ):
                errors.append("Select an assignee from the available team members.")
            else:
                new_assignee_label = assignee_user.display_name
                new_assignee_id = assignee_user.id
    else:
        new_assignee_label = "Unassigned"
        new_assignee_id = None

    if errors:
        for message in errors:
            flash(message, "error")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id))

    current_status = (ticket.get("status") or "").strip().lower()
    new_status = status.lower()
    requires_closing_comment = new_status in {"resolved", "closed"} and new_status != current_status

    if requires_closing_comment and _ticket_has_open_linked_tasks(ticket):
        flash("Resolve or close all linked tasks before marking the ticket resolved or closed.", "error")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id))

    if requires_closing_comment and not closing_comment:
        flash("Add closing remarks before completing the ticket.", "error")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id))

    changes = []
    if status != ticket.get("status"):
        changes.append(f"Status updated to {status}")
        ticket["status"] = status
    if priority != ticket.get("priority"):
        changes.append(f"Priority updated to {priority}")
        ticket["priority"] = priority
    if (
        new_assignee_label != current_assignee_label
        or (new_assignee_id or None) != (current_assignee_id or None)
    ):
        changes.append(f"Assigned to {new_assignee_label}")
        ticket["assignee"] = new_assignee_label
        ticket["assignee_user_id"] = new_assignee_id

    if not changes:
        flash("No changes detected to update.", "info")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id))

    ticket["updated_at"] = datetime.datetime.utcnow()
    actor_info = timeline_actor_context()
    ticket.setdefault("timeline", []).append(
        {
            "timestamp": datetime.datetime.utcnow(),
            "type": "status",
            "label": "Ticket details updated",
            "visibility": "internal",
            "detail": "; ".join(changes),
            **actor_info,
        }
    )

    if closing_comment:
        ticket.setdefault("timeline", []).append(
            {
                "timestamp": datetime.datetime.utcnow(),
                "type": "comment",
                "label": "Closing remarks",
                "visibility": "internal",
                "comment": closing_comment,
                **actor_info,
            }
        )

    flash("Ticket details updated successfully.", "success")
    return redirect(url_for("customer_support_tasks", ticket=ticket_id))


@app.route("/customer-support/tickets/<ticket_id>/resolve", methods=["POST"])
@login_required
def customer_support_mark_ticket_resolved(ticket_id):
    _module_visibility_required("customer_support")
    ticket = _get_customer_support_ticket(ticket_id)
    if not ticket:
        flash("The requested ticket could not be found.", "error")
        return redirect(url_for("customer_support_tasks"))

    closing_comment = (request.form.get("closing_comment") or "").strip()
    if not closing_comment:
        flash("Add closing remarks before completing the ticket.", "error")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id))

    if _ticket_has_open_linked_tasks(ticket):
        flash("Resolve linked tasks before marking this ticket resolved.", "error")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id))

    current_status = (ticket.get("status") or "").strip()
    now = datetime.datetime.utcnow()
    actor_info = timeline_actor_context()

    if current_status.lower() not in {"resolved", "closed"}:
        ticket["status"] = "Resolved"
        ticket["updated_at"] = now
        previous_label = current_status or "Open"
        ticket.setdefault("timeline", []).append(
            {
                "timestamp": now,
                "type": "status",
                "label": "Ticket marked resolved",
                "visibility": "internal",
                "detail": f"Status changed from {previous_label} to Resolved.",
                **actor_info,
            }
        )
    else:
        ticket["updated_at"] = now

    ticket.setdefault("timeline", []).append(
        {
            "timestamp": datetime.datetime.utcnow(),
            "type": "comment",
            "label": "Closing remarks",
            "visibility": "internal",
            "comment": closing_comment,
            **actor_info,
        }
    )

    flash(f"Ticket {ticket_id} marked as resolved.", "success")
    return redirect(url_for("customer_support_tasks"))


@app.route("/customer-support/tickets/<ticket_id>/delete", methods=["POST"])
@login_required
def customer_support_delete_ticket(ticket_id):
    _module_visibility_required("customer_support")
    if not current_user.is_admin:
        flash("You do not have permission to delete tickets.", "error")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id))

    ticket = _get_customer_support_ticket(ticket_id)
    if not ticket:
        flash("The requested ticket could not be found.", "error")
        return redirect(url_for("customer_support_tasks"))

    CUSTOMER_SUPPORT_TICKETS[:] = [item for item in CUSTOMER_SUPPORT_TICKETS if item.get("id") != ticket_id]
    flash(f"Ticket {ticket_id} deleted successfully.", "success")
    return redirect(url_for("customer_support_tasks"))


@app.route("/customer-support/tickets/<ticket_id>/comment", methods=["POST"])
@login_required
def customer_support_post_update(ticket_id):
    _module_visibility_required("customer_support")
    ticket = _get_customer_support_ticket(ticket_id)
    if not ticket:
        flash("The requested ticket could not be found.", "error")
        return redirect(url_for("customer_support_tasks"))

    comment = (request.form.get("comment") or "").strip()
    is_external = request.form.get("is_external") == "1"
    uploaded_files = request.files.getlist("attachments") or []
    valid_files = [file for file in uploaded_files if file and file.filename]

    if not comment and not valid_files:
        flash("Add a comment or attach at least one file before posting the update.", "error")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id))

    attachments_added = _save_customer_support_attachments(valid_files)

    if not comment and not attachments_added:
        flash("No valid attachments were uploaded.", "error")
        return redirect(url_for("customer_support_tasks", ticket=ticket_id))

    if attachments_added:
        ticket.setdefault("attachments", []).extend(attachments_added)

    actor_info = timeline_actor_context()
    visibility_label = "External update" if is_external else "Internal note"
    timeline_entry = {
        "timestamp": datetime.datetime.utcnow(),
        "type": "comment" if comment else "attachment",
        "label": visibility_label,
        "visibility": "external" if is_external else "internal",
    }
    timeline_entry.update(actor_info)

    if comment:
        timeline_entry["comment"] = comment

    if attachments_added:
        names = ", ".join(item["label"] for item in attachments_added)
        timeline_entry["detail"] = f"Uploaded: {names}"

    ticket.setdefault("timeline", []).append(timeline_entry)
    ticket["updated_at"] = datetime.datetime.utcnow()

    flash("Ticket update posted successfully.", "success")
    return redirect(url_for("customer_support_tasks", ticket=ticket_id))


@app.route("/customer-support/calls")
@login_required
def customer_support_calls():
    _module_visibility_required("customer_support")
    status_filter = request.args.get("status") or ""
    category_filter = request.args.get("category") or ""
    search_term = request.args.get("q") or ""

    call_logs = _customer_support_filter_calls(
        category=category_filter or None,
        status=status_filter or None,
        search=search_term or None,
    )

    return render_template(
        "customer_support_calls.html",
        call_logs=call_logs,
        status_filter=status_filter,
        category_filter=category_filter,
        search_term=search_term,
        status_options=["Open", "In Progress", "Resolved", "Closed"],
        categories=CUSTOMER_SUPPORT_CATEGORIES,
    )


# ---------------------- SERVICE MODULE ----------------------


def _coerce_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in (
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M:%S.%f",
        ):
            try:
                return datetime.datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _coerce_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"yes", "true", "1"}:
            return True
        if lowered in {"no", "false", "0"}:
            return False
    return None


def _coerce_float(value):
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _coerce_minutes(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().lower()
        if not cleaned:
            return None
        time_match = re.match(r"^(-?\d+):(\d{1,2})$", cleaned)
        if time_match:
            hours_part = float(time_match.group(1))
            minutes_part = float(time_match.group(2))
            return hours_part * 60 + minutes_part
        hour_match = re.search(r"(-?\d+(?:\.\d+)?)\s*h", cleaned)
        minute_match = re.search(r"(-?\d+(?:\.\d+)?)\s*m", cleaned)
        total_minutes = 0.0
        matched = False
        if hour_match:
            total_minutes += float(hour_match.group(1)) * 60
            matched = True
        if minute_match:
            total_minutes += float(minute_match.group(1))
            matched = True
        if matched:
            return total_minutes
        for suffix in ("minutes", "minute", "mins", "min", "m"):
            if cleaned.endswith(suffix):
                cleaned = cleaned[: -len(suffix)].strip()
                break
        if cleaned.endswith("h"):
            try:
                return float(cleaned[:-1].strip()) * 60
            except ValueError:
                return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _format_minutes_display(value):
    numeric = _coerce_float(value)
    if numeric is None:
        return "—"
    total_minutes = int(round(numeric))
    sign = "-" if total_minutes < 0 else ""
    minutes_abs = abs(total_minutes)
    hours, minutes = divmod(minutes_abs, 60)
    parts = []
    if hours:
        parts.append(f"{hours}h")
    if minutes or not parts:
        parts.append(f"{minutes}m")
    return sign + " ".join(parts)


def _format_percentage(numerator, denominator):
    if not denominator:
        return None
    try:
        ratio = (numerator / denominator) * 100
    except ZeroDivisionError:
        return None
    if ratio.is_integer():
        return f"{int(ratio)}%"
    return f"{ratio:.1f}%"


def _format_delta_label(delta_days):
    if delta_days == 0:
        return "Today"
    if delta_days == 1:
        return "Tomorrow"
    if delta_days > 1:
        return f"In {delta_days} days"
    if delta_days == -1:
        return "Yesterday"
    return f"{abs(delta_days)} days ago"


def get_service_schedule_snapshot():
    today = datetime.date.today()
    lifts = (
        Lift.query.options(joinedload(Lift.customer))
        .order_by(func.lower(Lift.lift_code))
        .all()
    )
    entries = []
    lifts_with_schedule = 0
    branches = set()

    for lift in lifts:
        branch_value = clean_str(lift.route)
        if branch_value:
            branches.add(branch_value)
        schedule = lift.service_schedule or []
        if schedule:
            lifts_with_schedule += 1
        for raw_entry in schedule:
            visit_date = _coerce_date(raw_entry.get("date"))
            status_value = clean_str(raw_entry.get("status")) or "scheduled"
            status_key = status_value.lower()
            if status_key not in SERVICE_VISIT_STATUS_LABELS:
                status_key = "scheduled"
            technician = clean_str(raw_entry.get("technician"))
            entries.append(
                {
                    "lift": lift,
                    "date": visit_date,
                    "status": status_key,
                    "technician": technician,
                    "first_time_fix": _coerce_bool(raw_entry.get("first_time_fix")),
                    "on_time": _coerce_bool(
                        raw_entry.get("on_time")
                        or raw_entry.get("on_time_completion")
                    ),
                    "travel_minutes": _coerce_minutes(
                        raw_entry.get("travel_minutes")
                        or raw_entry.get("travel_time")
                    ),
                    "repair_minutes": _coerce_minutes(
                        raw_entry.get("repair_minutes")
                        or raw_entry.get("duration_minutes")
                    ),
                    "rating": _coerce_float(raw_entry.get("rating")),
                    "checklist": clean_str(raw_entry.get("checklist")),
                }
            )

    return {
        "today": today,
        "lifts": lifts,
        "entries": entries,
        "branches": branches,
        "lifts_with_schedule": lifts_with_schedule,
    }


def build_service_overview_payload():
    snapshot = get_service_schedule_snapshot()
    entries = snapshot["entries"]
    lifts = snapshot["lifts"]
    today = snapshot["today"]

    status_counts = Counter()
    technicians_assigned = set()
    branches = snapshot.get("branches", set())

    def _initial_tech_stats():
        return {
            "name": "Unassigned",
            "open": 0,
            "completed": 0,
            "overdue": 0,
            "first_time_fix_total": 0,
            "first_time_fix_success": 0,
            "on_time_total": 0,
            "on_time_success": 0,
            "travel_minutes_total": 0.0,
            "travel_entries": 0,
            "repair_minutes_total": 0.0,
            "repair_entries": 0,
            "ratings": [],
        }

    technician_stats = defaultdict(_initial_tech_stats)

    overall_completed_this_month = 0
    overall_first_time_fix_total = 0
    overall_first_time_fix_success = 0
    overall_on_time_total = 0
    overall_on_time_success = 0

    for entry in entries:
        status_counts[entry["status"]] += 1
        if entry["technician"]:
            technicians_assigned.add(entry["technician"])
        entry["is_overdue"] = (
            entry["status"] == "overdue"
            or (
                entry["status"] != "completed"
                and isinstance(entry.get("date"), datetime.date)
                and entry["date"] < today
            )
        )
        if (
            entry["status"] == "completed"
            and isinstance(entry.get("date"), datetime.date)
            and entry["date"].year == today.year
            and entry["date"].month == today.month
        ):
            overall_completed_this_month += 1
        if entry["status"] == "completed" and entry["first_time_fix"] is not None:
            overall_first_time_fix_total += 1
            if entry["first_time_fix"]:
                overall_first_time_fix_success += 1
        if entry["status"] == "completed" and entry["on_time"] is not None:
            overall_on_time_total += 1
            if entry["on_time"]:
                overall_on_time_success += 1

        stat_key = entry["technician"] or "__unassigned__"
        stat = technician_stats[stat_key]
        stat["name"] = entry["technician"] or "Unassigned"
        if entry["status"] == "completed":
            stat["completed"] += 1
        else:
            stat["open"] += 1
        if entry["is_overdue"]:
            stat["overdue"] += 1
        if entry["status"] == "completed" and entry["first_time_fix"] is not None:
            stat["first_time_fix_total"] += 1
            if entry["first_time_fix"]:
                stat["first_time_fix_success"] += 1
        if entry["status"] == "completed" and entry["on_time"] is not None:
            stat["on_time_total"] += 1
            if entry["on_time"]:
                stat["on_time_success"] += 1
        if entry["travel_minutes"] is not None:
            stat["travel_minutes_total"] += entry["travel_minutes"]
            stat["travel_entries"] += 1
        if entry["repair_minutes"] is not None:
            stat["repair_minutes_total"] += entry["repair_minutes"]
            stat["repair_entries"] += 1
        if entry["rating"] is not None:
            stat["ratings"].append(entry["rating"])

    open_entries = [entry for entry in entries if entry["status"] != "completed"]
    open_count = len(open_entries)
    overdue_count = sum(1 for entry in entries if entry.get("is_overdue"))
    unassigned_stats = technician_stats.get("__unassigned__")
    unassigned_open = unassigned_stats["open"] if unassigned_stats else 0

    lifts_without_schedule = len(lifts) - snapshot.get("lifts_with_schedule", 0)

    amc_due_within_30 = 0
    amc_dates_recorded = 0
    for lift in lifts:
        if isinstance(lift.amc_end, datetime.date):
            amc_dates_recorded += 1
            delta = (lift.amc_end - today).days
            if 0 <= delta <= 30:
                amc_due_within_30 += 1

    first_time_fix_display = _format_percentage(
        overall_first_time_fix_success,
        overall_first_time_fix_total,
    )
    first_time_fix_rate_value = None
    if first_time_fix_display:
        first_time_fix_rate_value = (
            (overall_first_time_fix_success / overall_first_time_fix_total) * 100
            if overall_first_time_fix_total
            else None
        )

    on_time_display = _format_percentage(
        overall_on_time_success, overall_on_time_total
    )
    on_time_rate_value = None
    if on_time_display:
        on_time_rate_value = (
            (overall_on_time_success / overall_on_time_total) * 100
            if overall_on_time_total
            else None
        )

    kpis = []
    kpis.append(
        {
            "label": "Open service visits",
            "value": str(open_count),
            "descriptor": (
                "Visits awaiting completion"
                if entries
                else "Add service schedule entries on lifts to start tracking workload."
            ),
            "tone": "emerald" if open_count == 0 else "amber",
        }
    )
    kpis.append(
        {
            "label": "Overdue visits",
            "value": str(overdue_count),
            "descriptor": (
                "Past-due or marked overdue visits"
                if entries
                else "Overdue metrics appear once visits are scheduled."
            ),
            "tone": "emerald" if overdue_count == 0 else "rose",
        }
    )
    kpis.append(
        {
            "label": "Visits completed this month",
            "value": str(overall_completed_this_month),
            "descriptor": (
                f"{today.strftime('%b %Y')} closures"
                if overall_completed_this_month
                else "Complete a visit to track monthly closures."
            ),
            "tone": "emerald" if overall_completed_this_month else "slate",
        }
    )
    kpis.append(
        {
            "label": "AMC renewals due (30 days)",
            "value": str(amc_due_within_30),
            "descriptor": (
                f"Across {amc_dates_recorded} lifts with AMC end dates"
                if amc_dates_recorded
                else "Capture AMC end dates on lifts to monitor renewals."
            ),
            "tone": (
                "slate"
                if not amc_dates_recorded
                else ("amber" if amc_due_within_30 else "emerald")
            ),
        }
    )
    kpis.append(
        {
            "label": "Lifts without schedule",
            "value": str(max(lifts_without_schedule, 0)),
            "descriptor": (
                "Lifts without any recorded service visits"
                if lifts
                else "Add lifts to begin tracking service schedules."
            ),
            "tone": (
                "slate"
                if not lifts
                else ("rose" if lifts_without_schedule else "emerald")
            ),
        }
    )
    kpis.append(
        {
            "label": "First-time-fix rate",
            "value": first_time_fix_display or "—",
            "descriptor": (
                f"Based on {overall_first_time_fix_total} completed visits"
                if first_time_fix_display
                else "Record first-time-fix outcome on completed visits to populate this metric."
            ),
            "tone": (
                "slate"
                if not first_time_fix_display
                else ("emerald" if first_time_fix_rate_value and first_time_fix_rate_value >= 90 else "amber")
            ),
        }
    )
    kpis.append(
        {
            "label": "On-time completion",
            "value": on_time_display or "—",
            "descriptor": (
                f"Across {overall_on_time_total} completed visits"
                if on_time_display
                else "Capture on-time status when closing visits to see this KPI."
            ),
            "tone": (
                "slate"
                if not on_time_display
                else ("emerald" if on_time_rate_value and on_time_rate_value >= 90 else "amber")
            ),
        }
    )
    if unassigned_open:
        kpis.append(
            {
                "label": "Unassigned visits",
                "value": str(unassigned_open),
                "descriptor": "Assign technicians to clear these visits from the queue.",
                "tone": "rose",
            }
        )

    calendar_candidates = [
        entry
        for entry in entries
        if entry["status"] != "completed" and isinstance(entry.get("date"), datetime.date)
    ]
    calendar_candidates.sort(key=lambda item: item["date"])
    calendar_items = []
    for entry in calendar_candidates:
        lift = entry["lift"]
        label_bits = []
        if lift.customer and lift.customer.company_name:
            label_bits.append(lift.customer.company_name)
        if lift.lift_code:
            label_bits.append(lift.lift_code)
        label = " · ".join(label_bits) or (lift.lift_code or f"Lift #{lift.id}")
        delta_days = (entry["date"] - today).days
        calendar_items.append(
            {
                "label": label,
                "type": "Overdue" if entry.get("is_overdue") else "Scheduled",
                "date": entry["date"].strftime("%d %b %Y"),
                "delta": _format_delta_label(delta_days),
            }
        )

    if entries and not calendar_items:
        calendar_empty_message = "All tracked visits are completed."
    elif not entries:
        calendar_empty_message = "Add service visits to lifts to build the preventive calendar."
    else:
        calendar_empty_message = "No upcoming visits logged."

    filters = []
    for branch in sorted(branches):
        filters.append(f"Branch · {branch}")
    for technician in sorted(technicians_assigned):
        filters.append(f"Technician · {technician}")
    if not filters:
        filters.append("Add branches and technician assignments to unlock filters.")
    filters = filters[:8]

    chart_sets = {
        "complaints_by_category": [],
        "tasks_by_status": [],
        "parts_consumption": [],
        "technician_workload": [],
    }
    chart_notes = {
        "complaints_by_category": "Log support tickets with categories to populate this chart.",
        "tasks_by_status": "Add service visits to lifts to see task status distribution.",
        "parts_consumption": "Capture parts usage on service tasks to view monthly consumption.",
        "technician_workload": "Assign technicians to service visits to see workload split.",
    }

    category_counts = Counter()
    for ticket in CUSTOMER_SUPPORT_TICKETS:
        category_value = clean_str(ticket.get("category")) or "Uncategorised"
        category_counts[category_value] += 1
    if category_counts:
        chart_sets["complaints_by_category"] = [
            {"label": label, "value": count}
            for label, count in category_counts.most_common()
        ]
        chart_notes["complaints_by_category"] = ""

    status_items = []
    for status_key in ("scheduled", "completed", "overdue"):
        count = status_counts.get(status_key, 0)
        if count:
            status_label = SERVICE_VISIT_STATUS_LABELS.get(status_key, status_key.title())
            status_items.append({"label": status_label, "value": count})
    if status_items:
        chart_sets["tasks_by_status"] = status_items
        chart_notes["tasks_by_status"] = ""

    workload_items = []
    for stat in technician_stats.values():
        if stat["open"] > 0:
            workload_items.append({"label": stat["name"], "value": stat["open"]})
    if workload_items:
        chart_sets["technician_workload"] = sorted(
            workload_items, key=lambda item: item["value"], reverse=True
        )
        chart_notes["technician_workload"] = ""

    technician_cards = []
    for key, stat in technician_stats.items():
        total_handled = stat["completed"] + stat["open"]
        if key == "__unassigned__" or total_handled == 0:
            continue
        first_time_fix_stat = _format_percentage(
            stat["first_time_fix_success"], stat["first_time_fix_total"]
        )
        on_time_stat = _format_percentage(
            stat["on_time_success"], stat["on_time_total"]
        )
        travel_average = (
            stat["travel_minutes_total"] / stat["travel_entries"]
            if stat["travel_entries"]
            else None
        )
        repair_average = (
            stat["repair_minutes_total"] / stat["repair_entries"]
            if stat["repair_entries"]
            else None
        )
        if stat["ratings"]:
            rating_value = sum(stat["ratings"]) / len(stat["ratings"])
            rating_display = f"{rating_value:.1f}".rstrip("0").rstrip(".")
        else:
            rating_value = None
            rating_display = "—"
        notes = []
        if stat["completed"] == 0:
            notes.append(
                "Complete visits assigned to this technician to track closure metrics."
            )
        if stat["first_time_fix_total"] == 0:
            notes.append("Record first-time-fix outcome on completed visits.")
        if stat["on_time_total"] == 0:
            notes.append("Capture on-time status when closing visits.")
        if stat["travel_entries"] == 0:
            notes.append("Log travel time against visits to compute averages.")
        if stat["repair_entries"] == 0:
            notes.append("Log repair duration to compute averages.")
        if not stat["ratings"]:
            notes.append("Collect customer feedback ratings to track satisfaction.")
        technician_cards.append(
            {
                "name": stat["name"],
                "tasks_closed": stat["completed"],
                "first_time_fix_rate": first_time_fix_stat or "—",
                "on_time": on_time_stat or "—",
                "travel_time": _format_minutes_display(travel_average),
                "repair_time": _format_minutes_display(repair_average),
                "rating_display": rating_display,
                "rating_value": rating_value,
                "rating_stars": int(round(rating_value)) if rating_value is not None else 0,
                "rating_note": None if rating_value is not None else "No rating captured yet.",
                "notes": " • ".join(notes),
            }
        )

    technician_cards.sort(key=lambda item: item["tasks_closed"], reverse=True)
    technician_empty_message = (
        "Assign technicians to service visits to populate performance metrics."
        if not technician_cards
        else ""
    )

    return {
        "kpis": kpis,
        "chart_sets": chart_sets,
        "chart_notes": chart_notes,
        "calendar_items": calendar_items,
        "calendar_empty_message": calendar_empty_message,
        "filters": filters,
        "technicians": technician_cards,
        "technician_empty_message": technician_empty_message,
    }


@app.route("/service")
@login_required
def service_home():
    _module_visibility_required("service")
    return redirect(url_for("service_overview"))


@app.route("/service/overview")
@login_required
def service_overview():
    _module_visibility_required("service")
    payload = build_service_overview_payload()
    return render_template("service/overview.html", **payload)


@app.route("/service/tasks")
@login_required
def service_tasks():
    _module_visibility_required("service")
    tasks = []
    for task in SERVICE_TASKS:
        tasks.append(
            {
                **task,
                "requires_media_label": "Photos mandatory" if task.get("requires_media") else "Flexible",
                "technician_display": ", ".join(task.get("technicians") or []),
            }
        )
    for task in _service_complaint_tasks_from_support():
        tasks.append(
            {
                **task,
                "requires_media_label": "Photos mandatory" if task.get("requires_media") else "Flexible",
                "technician_display": ", ".join(task.get("technicians") or []),
            }
        )
    return render_template("service/tasks.html", tasks=tasks)


@app.route("/service/customers")
@login_required
def service_customers():
    _module_visibility_required("service")
    search_query = (request.args.get("q") or "").strip()

    query = _customer_query_for_export(search_query)
    query = query.options(joinedload(Customer.lifts))

    customers = query.all()
    for customer in customers:
        open_lifts = [lift for lift in customer.lifts if is_lift_open(lift)]
        customer.open_lifts = open_lifts
    next_customer_code = generate_next_customer_code()
    return render_template(
        "service/customers.html",
        customers=customers,
        search_query=search_query,
        next_customer_code=next_customer_code,
    )


@app.route("/service/customers/export")
@login_required
def service_customers_export():
    _module_visibility_required("service")

    search_query = (request.args.get("q") or "").strip()
    customers = _customer_query_for_export(search_query).all()

    timestamp = datetime.datetime.utcnow().strftime("%Y%m%d")
    filename = f"customers_export_{timestamp}.xlsx"

    if OPENPYXL_AVAILABLE:
        workbook = build_customer_export_workbook(customers)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    csv_rows = [_customer_upload_row(customer) for customer in customers]
    csv_output = _build_csv_output(CUSTOMER_UPLOAD_TEMPLATE_HEADERS, csv_rows)
    return send_file(
        csv_output,
        as_attachment=True,
        download_name=filename.replace(".xlsx", ".csv"),
        mimetype="text/csv",
    )


@app.route("/service/customers/create", methods=["POST"])
@login_required
def service_customers_create():
    _module_visibility_required("service")

    redirect_url = request.form.get("next") or url_for("service_customers")

    customer_code = generate_next_customer_code()
    company_name = clean_str(request.form.get("company_name"))
    category_value = clean_str(request.form.get("category"))
    branch_value, branch_error = validate_branch(request.form.get("branch"))
    if branch_error:
        flash(branch_error, "error")
        return redirect(redirect_url)

    if not company_name:
        flash("Company name is required.", "error")
        return redirect(redirect_url)

    existing = Customer.query.filter(func.lower(Customer.customer_code) == customer_code.lower()).first()
    if existing:
        flash("Another customer already uses that customer code. Please try again.", "error")
        return redirect(redirect_url)

    customer = Customer(
        customer_code=customer_code,
        company_name=company_name,
        contact_person=clean_str(request.form.get("contact_person")),
        phone=clean_str(request.form.get("phone")),
        mobile=clean_str(request.form.get("mobile")),
        email=clean_str(request.form.get("email")),
        gst_no=clean_str(request.form.get("gst_no")),
        billing_address_line1=clean_str(request.form.get("billing_address_line1")),
        billing_address_line2=clean_str(request.form.get("billing_address_line2")),
        city=clean_str(request.form.get("city")),
        state=clean_str(request.form.get("state")),
        pincode=clean_str(request.form.get("pincode")),
        country="India",
        sector=category_value,
        branch=branch_value,
        notes=clean_str(request.form.get("notes")),
        office_address_line1=clean_str(request.form.get("office_address_line1")),
        office_address_line2=clean_str(request.form.get("office_address_line2")),
        office_city=clean_str(request.form.get("office_city")),
        office_state=clean_str(request.form.get("office_state")),
        office_pincode=clean_str(request.form.get("office_pincode")),
        office_country="India",
    )

    db.session.add(customer)
    db.session.commit()

    flash(f"Customer {customer.customer_code} created.", "success")
    return redirect(redirect_url)


@app.route("/service/customers/<int:customer_id>")
@login_required
def service_customer_detail(customer_id):
    _module_visibility_required("service")

    customer = db.session.get(Customer, customer_id)
    if not customer:
        flash("Customer not found.", "error")
        return redirect(url_for("service_customers"))

    lifts = (
        Lift.query.filter_by(customer_code=customer.customer_code)
        .order_by(func.lower(Lift.lift_code))
        .options(joinedload(Lift.customer))
        .all()
    )

    customer.open_lifts = [lift for lift in lifts if is_lift_open(lift)]
    comments = (
        CustomerComment.query.filter_by(customer_id=customer.id)
        .order_by(CustomerComment.created_at.desc())
        .all()
    )
    return render_template(
        "service/customer_detail.html",
        customer=customer,
        lifts=lifts,
        comments=comments,
    )


@app.route("/service/customers/<int:customer_id>/update", methods=["POST"])
@login_required
def service_customer_update(customer_id):
    _module_visibility_required("service")

    customer = db.session.get(Customer, customer_id)
    if not customer:
        flash("Customer not found.", "error")
        return redirect(url_for("service_customers"))

    company_name = clean_str(request.form.get("company_name"))
    category_value = clean_str(request.form.get("category"))
    branch_value, branch_error = validate_branch(request.form.get("branch"))
    if branch_error:
        flash(branch_error, "error")
        return redirect(url_for("service_customer_detail", customer_id=customer.id))

    if not company_name:
        flash("Company name is required.", "error")
        return redirect(url_for("service_customer_detail", customer_id=customer.id))

    customer.company_name = company_name
    customer.contact_person = clean_str(request.form.get("contact_person"))
    customer.phone = clean_str(request.form.get("phone"))
    customer.mobile = clean_str(request.form.get("mobile"))
    customer.email = clean_str(request.form.get("email"))
    customer.gst_no = clean_str(request.form.get("gst_no"))
    customer.sector = category_value
    customer.branch = branch_value
    customer.notes = clean_str(request.form.get("notes"))

    db.session.commit()

    flash("Customer updated.", "success")
    return redirect(url_for("service_customer_detail", customer_id=customer.id))


@app.route("/service/customers/<int:customer_id>/address", methods=["POST"])
@login_required
def service_customer_update_address(customer_id):
    _module_visibility_required("service")

    customer = db.session.get(Customer, customer_id)
    if not customer:
        flash("Customer not found.", "error")
        return redirect(url_for("service_customers"))

    address_type = (request.form.get("address_type") or "").strip().lower()
    redirect_url = url_for("service_customer_detail", customer_id=customer.id)

    if address_type == "billing":
        customer.billing_address_line1 = clean_str(request.form.get("billing_address_line1"))
        customer.billing_address_line2 = clean_str(request.form.get("billing_address_line2"))
        customer.city = clean_str(request.form.get("city"))
        customer.state = clean_str(request.form.get("state"))
        customer.pincode = clean_str(request.form.get("pincode"))
        customer.country = "India"
        success_message = "Billing address updated."
    elif address_type == "office":
        customer.office_address_line1 = clean_str(request.form.get("office_address_line1"))
        customer.office_address_line2 = clean_str(request.form.get("office_address_line2"))
        customer.office_city = clean_str(request.form.get("office_city"))
        customer.office_state = clean_str(request.form.get("office_state"))
        customer.office_pincode = clean_str(request.form.get("office_pincode"))
        customer.office_country = "India"
        success_message = "Office address updated."
    else:
        flash("Unknown address type.", "error")
        return redirect(redirect_url)

    db.session.commit()
    flash(success_message, "success")
    return redirect(redirect_url)


@app.route("/service/customers/<int:customer_id>/delete", methods=["POST"])
@login_required
def service_customer_delete(customer_id):
    _module_visibility_required("service")

    if not current_user.is_admin:
        abort(403)

    customer = db.session.get(Customer, customer_id)
    if not customer:
        flash("Customer not found.", "error")
        return redirect(url_for("service_customers"))

    redirect_url = request.form.get("next") or url_for("service_customers")
    if not redirect_url.startswith("/"):
        redirect_url = url_for("service_customers")

    lifts = (
        Lift.query.options(subqueryload(Lift.attachments))
        .filter_by(customer_code=customer.customer_code)
        .all()
    )
    for lift in lifts:
        delete_lift_record(lift)

    db.session.delete(customer)
    db.session.commit()

    flash(f"Customer {customer.customer_code} deleted.", "success")
    return redirect(redirect_url)


@app.route("/service/customers/<int:customer_id>/comments", methods=["POST"])
@login_required
def service_customer_add_comment(customer_id):
    _module_visibility_required("service")

    customer = db.session.get(Customer, customer_id)
    if not customer:
        flash("Customer not found.", "error")
        return redirect(url_for("service_customers"))

    body = (request.form.get("body") or "").strip()
    if not body:
        flash("Comment cannot be empty.", "error")
        return redirect(url_for("service_customer_detail", customer_id=customer.id))

    comment = CustomerComment(
        customer=customer,
        body=body,
        author=current_user if current_user.is_authenticated else None,
    )
    db.session.add(comment)
    db.session.commit()

    flash("Comment added.", "success")
    return redirect(url_for("service_customer_detail", customer_id=customer.id))


AMC_LIFT_TEMPLATE_FALLBACK_B64 = """
UEsDBBQAAAAIANV+bFtGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9S
dreZih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHk
e64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+
ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIANV+bFuI4JN86gAAAMsBAAARAAAAZG9j
UHJvcHMvY29yZS54bWylkcFOwzAMhl9l6r11064coi4Xpp1AQmISiFvkeFtE00aJUbu3py1bB4Ib
x/j//NlWavQSu0BPofMU2FJcDa5po0S/SU7MXgJEPJHTMRuJdgwPXXCax2c4gtf4ro8ERZ7fgSPW
RrOGSZj6xZhclAYXpf8IzSwwCNSQo5YjiEzAjWUKLv7ZMCcLOUS7UH3fZ305c+NGAl4fH57n5VPb
RtYtUqJqgxIDae6Cmi7y56Gp4Vuxvsz+KpBZjRMknz1tkmvyUt5v97tEFXlRpUKkotiLSlZruS7f
JteP/pvQdcYe7D+MV4Gq4de/qU9QSwMEFAAAAAgA1X5sW5lcnCMQBgAAnCcAABMAAAB4bC90aGVt
ZS90aGVtZTEueG1s7Vpbc9o4FH7vr9B4Z/ZtC8Y2gba0E3Npdtu0mYTtTh+FEViNbHlkkYR/v0c2
EMuWDe2STbqbPAQs6fvORUfn6Dh58+4uYuiGiJTyeGDZL9vWu7cv3uBXMiQRQTAZp6/wwAqlTF61
WmkAwzh9yRMSw9yCiwhLeBTL1lzgWxovI9bqtNvdVoRpbKEYR2RgfV4saEDQVFFab18gtOUfM/gV
y1SNZaMBE1dBJrmItPL5bMX82t4+Zc/pOh0ygW4wG1ggf85vp+ROWojhVMLEwGpnP1Zrx9HSSICC
yX2UBbpJ9qPTFQgyDTs6nVjOdnz2xO2fjMradDRtGuDj8Xg4tsvSi3AcBOBRu57CnfRsv6RBCbSj
adBk2PbarpGmqo1TT9P3fd/rm2icCo1bT9Nrd93TjonGrdB4Db7xT4fDronGq9B062kmJ/2ua6Tp
FmhCRuPrehIVteVA0yAAWHB21szSA5ZeKfp1lBrZHbvdQVzwWO45iRH+xsUE1mnSGZY0RnKdkAUO
ADfE0UxQfK9BtorgwpLSXJDWzym1UBoImsiB9UeCIcXcr/31l7vJpDN6nX06zmuUf2mrAaftu5vP
k/xz6OSfp5PXTULOcLwsCfH7I1thhyduOxNyOhxnQnzP9vaRpSUyz+/5CutOPGcfVpawXc/P5J6M
ciO73fZYffZPR24j16nAsyLXlEYkRZ/ILbrkETi1SQ0yEz8InYaYalAcAqQJMZahhvi0xqwR4BN9
t74IyN+NiPerb5o9V6FYSdqE+BBGGuKcc+Zz0Wz7B6VG0fZVvNyjl1gVAZcY3zSqNSzF1niVwPGt
nDwdExLNlAsGQYaXJCYSqTl+TUgT/iul2v6c00DwlC8k+kqRj2mzI6d0Js3oMxrBRq8bdYdo0jx6
/gX5nDUKHJEbHQJnG7NGIYRpu/AerySOmq3CEStCPmIZNhpytRaBtnGphGBaEsbReE7StBH8Waw1
kz5gyOzNkXXO1pEOEZJeN0I+Ys6LkBG/HoY4SprtonFYBP2eXsNJweiCy2b9uH6G1TNsLI73R9QX
SuQPJqc/6TI0B6OaWQm9hFZqn6qHND6oHjIKBfG5Hj7lengKN5bGvFCugnsB/9HaN8Kr+ILAOX8u
fc+l77n0PaHStzcjfWfB04tb3kZuW8T7rjHa1zQuKGNXcs3Ix1SvkynYOZ/A7P1oPp7x7frZJISv
mlktIxaQS4GzQSS4/IvK8CrECehkWyUJy1TTZTeKEp5CG27pU/VKldflr7kouDxb5OmvoXQ+LM/5
PF/ntM0LM0O3ckvqtpS+tSY4SvSxzHBOHssMO2c8kh22d6AdNfv2XXbkI6UwU5dDuBpCvgNtup3c
OjiemJG5CtNSkG/D+enFeBriOdkEuX2YV23n2NHR++fBUbCj7zyWHceI8qIh7qGGmM/DQ4d5e1+Y
Z5XGUDQUbWysJCxGt2C41/EsFOBkYC2gB4OvUQLyUlVgMVvGAyuQonxMjEXocOeXXF/j0ZLj26Zl
tW6vKXcZbSJSOcJpmBNnq8reZbHBVR3PVVvysL5qPbQVTs/+Wa3InwwRThYLEkhjlBemSqLzGVO+
5ytJxFU4v0UzthKXGLzj5sdxTlO4Ena2DwIyubs5qXplMWem8t8tDAksW4hZEuJNXe3V55ucrnoi
dvqXd8Fg8v1wyUcP5TvnX/RdQ65+9t3j+m6TO0hMnHnFEQF0RQIjlRwGFhcy5FDukpAGEwHNlMlE
8AKCZKYcgJj6C73yDLkpFc6tPjl/RSyDhk5e0iUSFIqwDAUhF3Lj7++TaneM1/osgW2EVDJk1RfK
Q4nBPTNyQ9hUJfOu2iYLhdviVM27Gr4mYEvDem6dLSf/217UPbQXPUbzo5ngHrOHc5t6uMJFrP9Y
1h75Mt85cNs63gNe5hMsQ6R+wX2KioARq2K+uq9P+SWcO7R78YEgm/zW26T23eAMfNSrWqVkKxE/
Swd8H5IGY4xb9DRfjxRiraaxrcbaMQx5gFjzDKFmON+HRZoaM9WLrDmNCm9B1UDlP9vUDWj2DTQc
kQVeMZm2NqPkTgo83P7vDbDCxI7h7Yu/AVBLAwQUAAAACADVfmxb7y5T53ECAABaBQAAGAAAAHhs
L3dvcmtzaGVldHMvc2hlZXQxLnhtbIVUbW/aMBD+K6d82iRKoBRaoRCJwqpVGhJq1Vb7aOILseqX
1Haa8u93NiVttcGQUM7n8/M8d75z1hr77CpED29KajdLKu/raZq6okLFXN/UqGmnNFYxT0u7TV1t
kfF4SMn0fDCYpIoJneRZ9K1tnpnGS6FxbcE1SjG7u0Zp2lkyTA6OO7GtfHSkeVazLd6jf6jpAC3T
DocLhdoJo8FiOUvmw+n8Kp6IEY8CW/fJhpDMxpjnsLjls2QQNKHEwgcIRp9XXKCUAYmUvLyDJh+k
4eRn+wB/E/MneRvmcGHkk+C+miUkhmPJGunvTPsT33Maf0hcMs/yzJoWbEg2z4pgRPaYPIULHUp1
7y3tCuLz+Xy1AClKD00tDePgUdWSecxST9pCTFrQn1A76FEHPToCeiOkBF8hBPRfhO5gf++t8BUY
jXvKGi0QZv8E10XHdXGEa23Nq+AY6YrGeaMIFd88Ws0k3C6htEZ93X1P1div7sJwTDVTeErQuBM0
PiLokUnBY+bOM984dFOYx3bowY+3WljkPbhDjS3pW6PmQm97sGBUsmvmhDtBPunIJ/8l541loRWJ
fQg7ZLYH5/HrejA6GBcHY3ww9rGwEcpoX8ld53lpmKWiBs8E4uYppZed0ssjSqlb0YGi4sMGQWj4
Tb+z1epsSVcWJ6APaxoatFQwcGhfRYHA6RQNV4E19dRgeDYanLqsq07F1dHu+Zth56BgOqgqDL0f
tFMzqiXFfMP+tg8roymo94RcoyPr+z8lpJ9mMrw5K2a3QjuQWJKSQf+S+sfuh3i/8KaOY7oxnvox
mhW9fWhDAO2XxvhuER6R7jnN/wBQSwMEFAAAAAgA1X5sW5J2+ZcYBAAATxIAABgAAAB4bC93b3Jr
c2hlZXRzL3NoZWV0Mi54bWyVmG1v4jgQx7+KxauedDSxgT4JkFrobrtbemyhT69OLhnAamJnHVPK
ffpzQsJWZZykVVUS4t9/xjPDjEt3rfRrsgQw5D0KZdJrLI2JzzwvmS0h4smhikHaJ3OlI27srV54
SayBBxkUhR7z/SMv4kI2+t3svbHud9XKhELCWJNkFUVcby4gVOtegzaKN+7EYmmyN7x+N+YLmIC5
jy1gb72dTiAikIlQkmiY9xrn9OyZZUS24kHAOvlwTdLNvCj1mt5cB72Gn/oEIcxMKsHtyxsMIAxT
JevJ71y08cdoSn68LuS/Zfu37r3wBAYqfBSBWfYaJw0SwJyvQnOn1leQ76mTCc5UmGR/yXq7mFnX
Z6vEqCinreFIyO0rfy+C8YGgJw6C5QSrTbRyolWbaOdEuzbRyYnOZ8K586OcONojqIM4zonj2sRJ
TpzU3sdpTpzWJqhfpNCvz+zSXj/vtEg8rZ95WqSe7uWe+S6mSD6tn31apJ/u5d/NFAVA9yrAzRQl
QPdroO1iiiKg9auAFmVA9+rAWc+sqAO2VwdO31hRB2yvDtipi9k1gL06aDl9K+qA7dVBy/XZYUUd
sPp1wIo6YNs68LaNMGujQ254v6vVmugUsILpRdaMs2Zmu6eQ6eSYGG2fCsuZ/iCzAppcvhvQkofk
etj1jFVOn3uzXOWiXOVGzA0ZqAAQdFCO7uxmGqjxYc0tOBy4rInf8gjDv5XjFysRBkIuiEceRBhy
cqsOEZXv5SoTYYCcB4GGJLGRkEAoInL1VRGGiFxXhEOYDUL9qDBtuMGC97McGws5w5N2Uw7e2UMQ
ho1q1Ol0E2PobUVceMxnNjbkIAadKJn8hWj8U1fjdYHh44ogxwABOYg81Pavcvh8NCBpllYJwt7V
YrUhB8/2pzkaNYdDzIdJtc5wpXl6YkToaTV9KYMqH+4rSs6eS0FrG8gJ6DcxAzLkGywmD1/XsR+/
A9ypx6+KTe3hnBxcXZ2NRpjeU7neLbybP36toCpmzxVyysCnEHl2zuyGDdsNG2fl30+mzcunadP3
scZ28ZnL5mE2PZxzw4pR1mpj86LECx934NI9nngUh0CuZWDnhBaf47AdES6asnNsFlQZG4f8P471
f+fGQBptZ+j5G8gV1tyuXeR3hRn68bXlP13L234Lj/fN1wyMXMtHdzdYK3ctP8F6tmtxp40V19iZ
60Mf68rObpb9w4z1YhfBfNZu+vYXC+jE6da/G+Aa67YldjpNv9VsYXbunZlQMuCbvx8hkJDYK6yj
On3sYC3Tudo/87FQP5UH7hgP3LPziAES1vYzFdj2KST5RbEG6H04eadftIy4XgiZkBDmVtQ/PLbH
dL395mJ7Y1ScneZflLGnzuxyCTwAnS6wz+fK9triJj3g775D6v8PUEsDBBQAAAAIANV+bFsX/Mrx
iwIAAHgLAAANAAAAeGwvc3R5bGVzLnhtbN1W246bMBD9FcQHlCS0KFQJUhspUqW2Wmnz0FcTDLHk
CzVmS/r19dgOJLsZtq3Ul4ISxnN85u6QTWfOnD6eKDXRILjstvHJmPZ9knTHExWke6NaKi1SKy2I
sUvdJF2rKak6IAmerBaLLBGEybjYyF7shemio+ql2caLOEqKTa3kpEpjr7B7iaDRE+HbeEc4KzXz
m4lg/Oz1K6c5Kq50ZGw0dBsvnar76TcswxJCDbYEk0o7beLd+O8yEG7pb1/ZGBD36OwOxvmYyCr2
imLTEmOolnu78CSnfYkF+XBubSKNJufl6l18xXAP66ZUuqJ6dLSML6piw2ltgKFZc3KCUS08SmWM
EiBVjDRKEh/JhRYEa/tIOX+Eln+rbxwMdeR796lybYOML6KNKojeTFiAg2tz3viV3fTv7LbsSZmP
vU1IuvX3Xhn6oGnNBrce6ikAzPzy35pf4eZJ2/LzB84aKagv7m97LDbkwot+aNIe6GDGSRzqUO6x
0q7uN00ctRGcrW38Fc4sv0qp7Bk3TIbViVUVlS97ae0bUtpfhRsHdldFa9JzcxjBbTzJX2jFepGP
ux4g67Brkj/D4C6z6WBbZ0xWdKDVLix1UzoxsoJ1Gy7HeA7t3YVAKMuDCAQg6gsNA2V5Hurrf8xr
jeflQTTC9X1ojbPWOMvz7kI7d6O+EFZuLyTlPE/TLEPLu9vdD2OH1jDL4IMYRCMEDuoLvP1p5WcG
YGZsXpkNtMuzY4OmPDOiaMozlQcIqSFw8hwZANQXcNCmoBMFQSC+YNQQVppCn9EI0WM+A+U5CsGQ
ItObZVihMriRfqGHKE3zHIEARMJIUxSCAzsDoWFAICiUpv5F+ux9llzec8n0X7v4BVBLAwQUAAAA
CADVfmxbt0frisAAAAAWAgAACwAAAF9yZWxzLy5yZWxznZJLbgIxDECvEmVfTKnEAjGs2LBDiAu4
ieejmcSRY8T09o3YwCBoEUv/np4trw80oHYcc9ulbMYwxFzZVjWtALJrKWCecaJYKjVLQC2hNJDQ
9dgQLObzJcgtw27Wt0xz/En0CpHrunO0ZXcKFPUB+K7DmiNKQ1rZcYAzS//N3M8K1Jqdr6zs/Kc1
8KbM8/UgkKJHRXAs9JGkTIt2lK8+nt2+pPOlY2K0eN/o//PQqBQ9+b+dMKWJ0tdFCSZvsPkFUEsD
BBQAAAAIANV+bFt9A0CpSAEAAHECAAAPAAAAeGwvd29ya2Jvb2sueG1sjZFvS8MwEMa/SskHsN3Q
gWMdyIY6mH9wsvdpe12PJbmSXDfdpzdJqRYE8VVyzx2/PM9lcSZ7LIiOyYdWxs1tLhrmdp6mrmxA
S3dFLRjfq8lqyb60h5TqGktYU9lpMJxOs2yWWlCSkYxrsHWip/2H5VoLsnINAGvVo7REI5aLwdmr
TdJxRQxleCmoQdkjnN3PQCiTEzosUCF/5iLeFYhEo0GNF6hykYnENXR+JIsXMizVrrSkVC4mfWMP
lrH8Je+CzXdZuKiwLN5C5lzMMg+s0TqOE5EvvckT+OG+6pjuUTHYtWR4sNS1aA4R42OkoxxxFcOZ
GKkhFxvj2HYxtQtOfGdT9a7Y40YZ7Rx9w26qHjyG3D2tki3WPCZM/yBMe2uDnwpqNFA9e5YLDb+d
0n9NOKKX6fXN5NZvoVNq5bUXsyVZfQccfmf5BVBLAwQUAAAACADVfmxbq15yLrQAAACNAgAAGgAA
AHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzxZJNCoMwEEavEnIAR23poqirbtwWLxB0/MHEhMyU
6u0rulChi26kq/BNyPseTJInasWdHajtHInR6IFS2TK7OwCVLRpFgXU4zDe19UbxHH0DTpW9ahDi
MLyB3zNkluyZopgc/kK0dd2V+LDly+DAX8Dwtr6nFpGlKJRvkFMJo97GBMsRBTNZirxKpc+rSAr4
t1F8MIrPNCKeNNKms+ZD/+XMfp7f4la/xHV4XMt1kYDD78s+UEsDBBQAAAAIANV+bFul4RtYHwEA
AGAEAAATAAAAW0NvbnRlbnRfVHlwZXNdLnhtbMVUy07DMBD8lcjXKnbpgQNqeqFcoQd+wCSbxopf
8m5L+vdsEloJVFqqIHGJFe/szHjH8vL1EAGzzlmPhWiI4oNSWDbgNMoQwXOlDslp4t+0VVGXrd6C
Wszn96oMnsBTTj2HWC3XUOudpeyp4200wRcigUWRPY7AXqsQOkZrSk1cV3tffVPJPxUkdw4YbEzE
GQNEps5KDKUfFY6NL3tIyVSQbXSiZ+0YpjqrkA4WUF7mOOMy1LUpoQrlznGLxJhAV9gAkLNyJJ1d
kSYeMozfu8kGBpqLigzdpBCRU0twu94xlr47j0wEicyVQ54kmXvyCaFPvILqt+I84feQ2iETVMMy
fcxfcz7x32pk8Z9G3kJo//rC96t02viTATU8LKsPUEsBAhQDFAAAAAgA1X5sW0bHTUiVAAAAzQAA
ABAAAAAAAAAAAAAAAIABAAAAAGRvY1Byb3BzL2FwcC54bWxQSwECFAMUAAAACADVfmxbiOCTfOoA
AADLAQAAEQAAAAAAAAAAAAAAgAHDAAAAZG9jUHJvcHMvY29yZS54bWxQSwECFAMUAAAACADVfmxb
mVycIxAGAACcJwAAEwAAAAAAAAAAAAAAgAHcAQAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQIUAxQA
AAAIANV+bFvvLlPncQIAAFoFAAAYAAAAAAAAAAAAAACAgR0IAAB4bC93b3Jrc2hlZXRzL3NoZWV0
MS54bWxQSwECFAMUAAAACADVfmxbknb5lxgEAABPEgAAGAAAAAAAAAAAAAAAgIHECgAAeGwvd29y
a3NoZWV0cy9zaGVldDIueG1sUEsBAhQDFAAAAAgA1X5sWxf8yvGLAgAAeAsAAA0AAAAAAAAAAAAA
AIABEg8AAHhsL3N0eWxlcy54bWxQSwECFAMUAAAACADVfmxbt0frisAAAAAWAgAACwAAAAAAAAAA
AAAAgAHIEQAAX3JlbHMvLnJlbHNQSwECFAMUAAAACADVfmxbfQNAqUgBAABxAgAADwAAAAAAAAAA
AAAAgAGxEgAAeGwvd29ya2Jvb2sueG1sUEsBAhQDFAAAAAgA1X5sW6teci60AAAAjQIAABoAAAAA
AAAAAAAAAIABJhQAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAhQDFAAAAAgA1X5sW6Xh
G1gfAQAAYAQAABMAAAAAAAAAAAAAAIABEhUAAFtDb250ZW50X1R5cGVzXS54bWxQSwUGAAAAAAoA
CgCEAgAAYhYAAAAA
"""
CUSTOMER_TEMPLATE_FALLBACK_B64 = """
UEsDBBQAAAAIANV+bFtGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9S
dreZih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHk
e64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+
ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIANV+bFuI4JN86gAAAMsBAAARAAAAZG9j
UHJvcHMvY29yZS54bWylkcFOwzAMhl9l6r11064coi4Xpp1AQmISiFvkeFtE00aJUbu3py1bB4Ib
x/j//NlWavQSu0BPofMU2FJcDa5po0S/SU7MXgJEPJHTMRuJdgwPXXCax2c4gtf4ro8ERZ7fgSPW
RrOGSZj6xZhclAYXpf8IzSwwCNSQo5YjiEzAjWUKLv7ZMCcLOUS7UH3fZ305c+NGAl4fH57n5VPb
RtYtUqJqgxIDae6Cmi7y56Gp4Vuxvsz+KpBZjRMknz1tkmvyUt5v97tEFXlRpUKkotiLSlZruS7f
JteP/pvQdcYe7D+MV4Gq4de/qU9QSwMEFAAAAAgA1X5sW5lcnCMQBgAAnCcAABMAAAB4bC90aGVt
ZS90aGVtZTEueG1s7Vpbc9o4FH7vr9B4Z/ZtC8Y2gba0E3Npdtu0mYTtTh+FEViNbHlkkYR/v0c2
EMuWDe2STbqbPAQs6fvORUfn6Dh58+4uYuiGiJTyeGDZL9vWu7cv3uBXMiQRQTAZp6/wwAqlTF61
WmkAwzh9yRMSw9yCiwhLeBTL1lzgWxovI9bqtNvdVoRpbKEYR2RgfV4saEDQVFFab18gtOUfM/gV
y1SNZaMBE1dBJrmItPL5bMX82t4+Zc/pOh0ygW4wG1ggf85vp+ROWojhVMLEwGpnP1Zrx9HSSICC
yX2UBbpJ9qPTFQgyDTs6nVjOdnz2xO2fjMradDRtGuDj8Xg4tsvSi3AcBOBRu57CnfRsv6RBCbSj
adBk2PbarpGmqo1TT9P3fd/rm2icCo1bT9Nrd93TjonGrdB4Db7xT4fDronGq9B062kmJ/2ua6Tp
FmhCRuPrehIVteVA0yAAWHB21szSA5ZeKfp1lBrZHbvdQVzwWO45iRH+xsUE1mnSGZY0RnKdkAUO
ADfE0UxQfK9BtorgwpLSXJDWzym1UBoImsiB9UeCIcXcr/31l7vJpDN6nX06zmuUf2mrAaftu5vP
k/xz6OSfp5PXTULOcLwsCfH7I1thhyduOxNyOhxnQnzP9vaRpSUyz+/5CutOPGcfVpawXc/P5J6M
ciO73fZYffZPR24j16nAsyLXlEYkRZ/ILbrkETi1SQ0yEz8InYaYalAcAqQJMZahhvi0xqwR4BN9
t74IyN+NiPerb5o9V6FYSdqE+BBGGuKcc+Zz0Wz7B6VG0fZVvNyjl1gVAZcY3zSqNSzF1niVwPGt
nDwdExLNlAsGQYaXJCYSqTl+TUgT/iul2v6c00DwlC8k+kqRj2mzI6d0Js3oMxrBRq8bdYdo0jx6
/gX5nDUKHJEbHQJnG7NGIYRpu/AerySOmq3CEStCPmIZNhpytRaBtnGphGBaEsbReE7StBH8Waw1
kz5gyOzNkXXO1pEOEZJeN0I+Ys6LkBG/HoY4SprtonFYBP2eXsNJweiCy2b9uH6G1TNsLI73R9QX
SuQPJqc/6TI0B6OaWQm9hFZqn6qHND6oHjIKBfG5Hj7lengKN5bGvFCugnsB/9HaN8Kr+ILAOX8u
fc+l77n0PaHStzcjfWfB04tb3kZuW8T7rjHa1zQuKGNXcs3Ix1SvkynYOZ/A7P1oPp7x7frZJISv
mlktIxaQS4GzQSS4/IvK8CrECehkWyUJy1TTZTeKEp5CG27pU/VKldflr7kouDxb5OmvoXQ+LM/5
PF/ntM0LM0O3ckvqtpS+tSY4SvSxzHBOHssMO2c8kh22d6AdNfv2XXbkI6UwU5dDuBpCvgNtup3c
OjiemJG5CtNSkG/D+enFeBriOdkEuX2YV23n2NHR++fBUbCj7zyWHceI8qIh7qGGmM/DQ4d5e1+Y
Z5XGUDQUbWysJCxGt2C41/EsFOBkYC2gB4OvUQLyUlVgMVvGAyuQonxMjEXocOeXXF/j0ZLj26Zl
tW6vKXcZbSJSOcJpmBNnq8reZbHBVR3PVVvysL5qPbQVTs/+Wa3InwwRThYLEkhjlBemSqLzGVO+
5ytJxFU4v0UzthKXGLzj5sdxTlO4Ena2DwIyubs5qXplMWem8t8tDAksW4hZEuJNXe3V55ucrnoi
dvqXd8Fg8v1wyUcP5TvnX/RdQ65+9t3j+m6TO0hMnHnFEQF0RQIjlRwGFhcy5FDukpAGEwHNlMlE
8AKCZKYcgJj6C73yDLkpFc6tPjl/RSyDhk5e0iUSFIqwDAUhF3Lj7++TaneM1/osgW2EVDJk1RfK
Q4nBPTNyQ9hUJfOu2iYLhdviVM27Gr4mYEvDem6dLSf/217UPbQXPUbzo5ngHrOHc5t6uMJFrP9Y
1h75Mt85cNs63gNe5hMsQ6R+wX2KioARq2K+uq9P+SWcO7R78YEgm/zW26T23eAMfNSrWqVkKxE/
Swd8H5IGY4xb9DRfjxRiraaxrcbaMQx5gFjzDKFmON+HRZoaM9WLrDmNCm9B1UDlP9vUDWj2DTQc
kQVeMZm2NqPkTgo83P7vDbDCxI7h7Yu/AVBLAwQUAAAACADVfmxbx0021wkCAAAwBAAAGAAAAHhs
L3dvcmtzaGVldHMvc2hlZXQxLnhtbIVUwW7bMAz9FcLnLU7XJhsKx0CarlsxBAhSbDsrNh0LkURX
ouPu70c5iZvDkh0MiyL13uOz6Kwjvws1IsObNS7Mkpq5uU/TUNRoVRhRg04yFXmrWEK/TUPjUZX9
IWvST+PxNLVKuyTP+r2VzzNq2WiHKw+htVb5Pw9oqJslN8lpY623NfcbaZ41aosvyD8bOSBhOuCU
2qILmhx4rGbJ/OZ+Pu1P9BW/NHbhbA2xmQ3RLgbP5SwZR01osOAIoeS1xwUaE5FEyesRNHknjSfP
1yf4p75/kbdRARdkfuuS61nyJYESK9UaXlP3HY89Td4lPipWeeapAx+bzbMiLnr2vnkp1y5a9cJe
slr4OF+0gcmih7YxpEpgtI1RjFnKoi3WpIU8gjpA3w7QtxdAn7QxwDXCCT3A4bt3mmsgh1CcaBt5
BHd0he9u4Lu7wLfytNclgnLw9Y3RO2UGanh+BCawqgGjKw5Qtl67LcyXi2PP4Rr5ZCCf/M/BgkoM
oLzoaJk+btGhFydL0BUYrBg2RrndCH4gNtEdC63Try3GfHPooLwmZTpImV6QspZZQLGaWlNKx1zU
0ZLDXYSAfq8LFLelaAQPXjnJW1EPG+w/ClXwjdQHWLZ2o/Q/paRnNy1O0lL5rXah709mYPRZXPKH
q3kImJr+8m2IxaV+WctEo48Fkq+IeAjiaAw/ifwvUEsDBBQAAAAIANV+bFvWeAQ/ugMAAOUQAAAY
AAAAeGwvd29ya3NoZWV0cy9zaGVldDIueG1snZhrc9o4FIb/ioavncSWuKYDzHJJWroNoUCSTr8p
tgBNZInKIoR/v7Kx2cz2yIj9AtblOZJ8Xl5JdPdKv6Ybxgx6T4RMe7WNMdvPQZBGG5bQ9FptmbQt
K6UTamxRr4N0qxmNcygRAQnDVpBQLmv9bl430/2u2hnBJZtplO6ShOrDkAm179VwrayY8/XG5BVB
v7ula7Zg5nFrAVsMTnFinjCZciWRZqtebYA/P5OcyHs8cbZPPzyjbDEvSr1mhUncq4XZnJhgkclC
UPv1xkZMiCySncnvImjt30Ez8uNzGf4uX7+d3gtN2UiJZx6bTa/WqaGYrehOmLnaf2XFmpp5wEiJ
NP9E+2NnYqce7VKjkoK2AydcHr/pe/kyPhC44yBIQRBvol4QdW+iURANb6JZEE1volUQLW+iXRBt
b6JTEJ3/EqThIG4K4sabwGGZwtB7XviUdv+84zLx2D/zuEw99s89LpOP/bOPy/Rj//zjUgDYXwG4
lAD21wAuRYD/VEHdxZQywH/qwMWQUgfEXwek1AG54Pd/MgB/HZBSB+Sog+BoULm9jamh/a5We6Qz
wAbMHnKTzE3GuhqXmaMvjLat3HKmf/tumJZUoFE+HNNoMu4GxkbO2oOoiDKsjnKCRypmAD46g6tk
S+UBTWkC0eNztDR2W0AzplMlAf62mp9tlISGvavG7tULFxD35cwbt/usALCv1diXxRJNFcBNqrkh
F7ZqjQZxrFmaou+2HWEgzrf/EYcAcf4+ky1uDgD1vZpaGGqgd31/JrVcRrAip+c0tZNGQxN9qAbn
9tQEjTc7sz57vFEa4H6cyYqmMtoA3LyamyrDUgBbVGMPqxWP2HktLS8PA0np0SuMQ1FPXrBLWM9e
NKivwFrwyYfJyYedUntcLK9ufy6vwhB6lcMKLoSRkdvsabIVDE1kbA1bc1ACYxc9EOwdjai2GwZk
si7s0w2+6tTJFSb1RrPVhoy2Cr3ptFvNRp3gELJaF6m26V/suNrrSCWQ3bpQ0h4MR+PbbLp3+FcT
clwXiskAMlZX94mU6o3m15oZ1a+Qlzp3A0UhE72s+72reyOsw8qaupcSc2iIh8tmNKs4IdjjRcQp
tHP+uGyQuXO3sFdJe4ZAv3e5yMXBXljfsitpRIUA3dJp5ztuGKqHkKktndDrIatCS7W3s4Ds8LJ1
Pl3W/fkSNRw9Lvhw7syu//dUr7lMkWArGym8bttLhD7ep48Fo7b5WfZFGXtezB83jMZMZx1s+0rZ
faksZMfb0z8b/X8AUEsDBBQAAAAIANV+bFsX/MrxiwIAAHgLAAANAAAAeGwvc3R5bGVzLnhtbN1W
246bMBD9FcQHlCS0KFQJUhspUqW2Wmnz0FcTDLHkCzVmS/r19dgOJLsZtq3Ul4ISxnN85u6QTWfO
nD6eKDXRILjstvHJmPZ9knTHExWke6NaKi1SKy2IsUvdJF2rKak6IAmerBaLLBGEybjYyF7shemi
o+ql2caLOEqKTa3kpEpjr7B7iaDRE+HbeEc4KzXzm4lg/Oz1K6c5Kq50ZGw0dBsvnar76TcswxJC
DbYEk0o7beLd+O8yEG7pb1/ZGBD36OwOxvmYyCr2imLTEmOolnu78CSnfYkF+XBubSKNJufl6l18
xXAP66ZUuqJ6dLSML6piw2ltgKFZc3KCUS08SmWMEiBVjDRKEh/JhRYEa/tIOX+Eln+rbxwMdeR7
96lybYOML6KNKojeTFiAg2tz3viV3fTv7LbsSZmPvU1IuvX3Xhn6oGnNBrce6ikAzPzy35pf4eZJ
2/LzB84aKagv7m97LDbkwot+aNIe6GDGSRzqUO6x0q7uN00ctRGcrW38Fc4sv0qp7Bk3TIbViVUV
lS97ae0bUtpfhRsHdldFa9JzcxjBbTzJX2jFepGPux4g67Brkj/D4C6z6WBbZ0xWdKDVLix1Uzox
soJ1Gy7HeA7t3YVAKMuDCAQg6gsNA2V5Hurrf8xrjeflQTTC9X1ojbPWOMvz7kI7d6O+EFZuLyTl
PE/TLEPLu9vdD2OH1jDL4IMYRCMEDuoLvP1p5WcGYGZsXpkNtMuzY4OmPDOiaMozlQcIqSFw8hwZ
ANQXcNCmoBMFQSC+YNQQVppCn9EI0WM+A+U5CsGQItObZVihMriRfqGHKE3zHIEARMJIUxSCAzsD
oWFAICiUpv5F+ux9llzec8n0X7v4BVBLAwQUAAAACADVfmxbt0frisAAAAAWAgAACwAAAF9yZWxz
Ly5yZWxznZJLbgIxDECvEmVfTKnEAjGs2LBDiAu4ieejmcSRY8T09o3YwCBoEUv/np4trw80oHYc
c9ulbMYwxFzZVjWtALJrKWCecaJYKjVLQC2hNJDQ9dgQLObzJcgtw27Wt0xz/En0CpHrunO0ZXcK
FPUB+K7DmiNKQ1rZcYAzS//N3M8K1Jqdr6zs/Kc18KbM8/UgkKJHRXAs9JGkTIt2lK8+nt2+pPOl
Y2K0eN/o//PQqBQ9+b+dMKWJ0tdFCSZvsPkFUEsDBBQAAAAIANV+bFtGlf0ARQEAAHECAAAPAAAA
eGwvd29ya2Jvb2sueG1sjZFhS8MwEIb/SskPsF3RgWPdF4c6EBUn+5611/VYkiuX66b79SYt1YIg
fkruvePJ+16WZ+LjnuiYfFjj/IIL1Yi0izT1ZQNW+ytqwYVeTWy1hJIPKdU1lrCmsrPgJM2zbJ4y
GC1IzjfYejXQ/sPyLYOufAMg1gwoq9Gp1XJ09spJOq1IoIwvRTUqO4Sz/xmIZXJCj3s0KJ+F6u8G
VGLRocULVIXKVOIbOj8S44WcaLMtmYwp1Gxo7IAFy1/yNtp813vfK6L3bzFzoeZZANbIXvqJnq+D
yROE4aHqhO7RCPBaCzwwdS26Q48JMdJJjn4V45k4baFQG+eFuz61j05CZ1MNriTgJhl5gaHBm2oA
TyF3nReywFNC/gchH6yNfiqo0UH1HFg+NsJ2yvA18ei95Nc3s9uwhc6Yu6C9uCfS1XfA8XdWX1BL
AwQUAAAACADVfmxbq15yLrQAAACNAgAAGgAAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzxZJN
CoMwEEavEnIAR23poqirbtwWLxB0/MHEhMyU6u0rulChi26kq/BNyPseTJInasWdHajtHInR6IFS
2TK7OwCVLRpFgXU4zDe19UbxHH0DTpW9ahDiMLyB3zNkluyZopgc/kK0dd2V+LDly+DAX8Dwtr6n
FpGlKJRvkFMJo97GBMsRBTNZirxKpc+rSAr4t1F8MIrPNCKeNNKms+ZD/+XMfp7f4la/xHV4XMt1
kYDD78s+UEsDBBQAAAAIANV+bFul4RtYHwEAAGAEAAATAAAAW0NvbnRlbnRfVHlwZXNdLnhtbMVU
y07DMBD8lcjXKnbpgQNqeqFcoQd+wCSbxopf8m5L+vdsEloJVFqqIHGJFe/szHjH8vL1EAGzzlmP
hWiI4oNSWDbgNMoQwXOlDslp4t+0VVGXrd6CWszn96oMnsBTTj2HWC3XUOudpeyp4200wRcigUWR
PY7AXqsQOkZrSk1cV3tffVPJPxUkdw4YbEzEGQNEps5KDKUfFY6NL3tIyVSQbXSiZ+0YpjqrkA4W
UF7mOOMy1LUpoQrlznGLxJhAV9gAkLNyJJ1dkSYeMozfu8kGBpqLigzdpBCRU0twu94xlr47j0wE
icyVQ54kmXvyCaFPvILqt+I84feQ2iETVMMyfcxfcz7x32pk8Z9G3kJo//rC96t02viTATU8LKsP
UEsBAhQDFAAAAAgA1X5sW0bHTUiVAAAAzQAAABAAAAAAAAAAAAAAAIABAAAAAGRvY1Byb3BzL2Fw
cC54bWxQSwECFAMUAAAACADVfmxbiOCTfOoAAADLAQAAEQAAAAAAAAAAAAAAgAHDAAAAZG9jUHJv
cHMvY29yZS54bWxQSwECFAMUAAAACADVfmxbmVycIxAGAACcJwAAEwAAAAAAAAAAAAAAgAHcAQAA
eGwvdGhlbWUvdGhlbWUxLnhtbFBLAQIUAxQAAAAIANV+bFvHTTbXCQIAADAEAAAYAAAAAAAAAAAA
AACAgR0IAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECFAMUAAAACADVfmxb1ngEP7oDAADl
EAAAGAAAAAAAAAAAAAAAgIFcCgAAeGwvd29ya3NoZWV0cy9zaGVldDIueG1sUEsBAhQDFAAAAAgA
1X5sWxf8yvGLAgAAeAsAAA0AAAAAAAAAAAAAAIABTA4AAHhsL3N0eWxlcy54bWxQSwECFAMUAAAA
CADVfmxbt0frisAAAAAWAgAACwAAAAAAAAAAAAAAgAECEQAAX3JlbHMvLnJlbHNQSwECFAMUAAAA
CADVfmxbRpX9AEUBAABxAgAADwAAAAAAAAAAAAAAgAHrEQAAeGwvd29ya2Jvb2sueG1sUEsBAhQD
FAAAAAgA1X5sW6teci60AAAAjQIAABoAAAAAAAAAAAAAAIABXRMAAHhsL19yZWxzL3dvcmtib29r
LnhtbC5yZWxzUEsBAhQDFAAAAAgA1X5sW6XhG1gfAQAAYAQAABMAAAAAAAAAAAAAAIABSRQAAFtD
b250ZW50X1R5cGVzXS54bWxQSwUGAAAAAAoACgCEAgAAmRUAAAAA
"""


def _serve_base64_excel_template(encoded_value, filename):
    if not encoded_value:
        return None
    output = BytesIO(base64.b64decode(encoded_value))
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/service/lifts/upload-template")
@login_required
def service_lifts_upload_template():
    _module_visibility_required("service")

    filename = f"amc_lifts_template_{datetime.datetime.utcnow():%Y%m%d}.xlsx"

    if OPENPYXL_AVAILABLE:
        workbook = build_amc_lift_upload_workbook()
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    csv_output = _build_csv_output(
        AMC_LIFT_TEMPLATE_HEADERS,
        [["" for _ in AMC_LIFT_TEMPLATE_HEADERS]],
    )
    return send_file(
        csv_output,
        as_attachment=True,
        download_name=filename.replace(".xlsx", ".csv"),
        mimetype="text/csv",
    )


@app.route("/service/customers/upload-template")
@login_required
def service_customers_upload_template():
    _module_visibility_required("service")

    filename = f"customers_template_{datetime.datetime.utcnow():%Y%m%d}.xlsx"

    if OPENPYXL_AVAILABLE:
        workbook = build_customer_upload_workbook()
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    csv_output = _build_csv_output(
        CUSTOMER_UPLOAD_TEMPLATE_HEADERS,
        [["" for _ in CUSTOMER_UPLOAD_TEMPLATE_HEADERS]],
    )
    return send_file(
        csv_output,
        as_attachment=True,
        download_name=filename.replace(".xlsx", ".csv"),
        mimetype="text/csv",
    )


@app.route("/service/customers/upload", methods=["POST"])
@login_required
def service_customers_upload():
    _module_visibility_required("service")

    pending_token = request.form.get("pending_token")
    action = request.form.get("action")

    if pending_token:
        pending_uploads = session.get("pending_uploads", {})
        pending = pending_uploads.get(pending_token)
        if not pending or pending.get("type") != "service_customers":
            flash("The upload preview has expired. Please re-upload the file.", "error")
            return redirect(url_for("service_customers"))

        file_path = pending.get("path")
        if not file_path or not os.path.exists(file_path):
            pending_uploads.pop(pending_token, None)
            session["pending_uploads"] = pending_uploads
            session.modified = True
            flash("The staged upload file was not found. Please try uploading again.", "error")
            return redirect(url_for("service_customers"))

        if action == "discard":
            pending_uploads.pop(pending_token, None)
            session["pending_uploads"] = pending_uploads
            session.modified = True
            try:
                os.remove(file_path)
            except OSError:
                pass
            flash("Customer upload discarded.", "info")
            return redirect(url_for("service_customers"))

        if action != "merge":
            flash("Select a valid action for the upload.", "error")
            return redirect(url_for("service_customers"))

        outcome = process_customer_upload_file(
            file_path,
            apply_changes=True,
        )

        pending_uploads.pop(pending_token, None)
        session["pending_uploads"] = pending_uploads
        session.modified = True

        try:
            os.remove(file_path)
        except OSError:
            pass

        if outcome.created_count or outcome.updated_count:
            flash(
                f"Customer upload complete: {outcome.created_count} added, {outcome.updated_count} updated.",
                "success",
            )
        elif outcome.processed_rows and not outcome.row_errors:
            flash("No changes were detected in the uploaded workbook.", "info")
        else:
            flash("No rows were imported from the uploaded workbook.", "info")

        for message in outcome.row_errors[:25]:
            flash(message, "error")
        if len(outcome.row_errors) > 25:
            flash(
                f"Additional {len(outcome.row_errors) - 25} errors were omitted from the alert. Check the file and retry.",
                "warning",
            )

        return redirect(url_for("service_customers"))

    upload = request.files.get("customer_upload_file")
    if not upload or not upload.filename:
        flash("Select an Excel workbook to upload.", "error")
        return redirect(url_for("service_customers"))

    filename = (upload.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv")):
        flash(
            "Upload the .xlsx or .csv template exported from the customer upload modal.",
            "error",
        )
        return redirect(url_for("service_customers"))

    saved_path = None
    try:
        saved_path = save_pending_upload_file(upload)
        outcome = process_customer_upload_file(
            saved_path,
            apply_changes=False,
        )
    except MissingDependencyError:
        if saved_path:
            try:
                os.remove(saved_path)
            except OSError:
                pass
        flash(OPENPYXL_MISSING_MESSAGE, "error")
        return redirect(url_for("service_customers"))
    except ValueError:
        if saved_path:
            try:
                os.remove(saved_path)
            except OSError:
                pass
        flash(
            "Upload the .xlsx or .csv template exported from the customer upload modal.",
            "error",
        )
        return redirect(url_for("service_customers"))
    except Exception:
        if saved_path:
            try:
                os.remove(saved_path)
            except OSError:
                pass
        flash(
            "Could not read the uploaded workbook. Ensure it is a valid .xlsx or .csv file.",
            "error",
        )
        return redirect(url_for("service_customers"))

    if not outcome.header_map:
        try:
            os.remove(saved_path)
        except OSError:
            pass
        flash("The uploaded workbook is missing header labels.", "error")
        return redirect(url_for("service_customers"))

    missing_headers = []
    if "Company Name" not in outcome.header_map:
        missing_headers.append("Company Name")
    if "Customer Code" not in outcome.header_map and "External Customer ID" not in outcome.header_map:
        missing_headers.append("Customer Code / External Customer ID")
    if missing_headers:
        try:
            os.remove(saved_path)
        except OSError:
            pass
        flash(
            "The uploaded workbook is missing required columns: "
            + ", ".join(missing_headers)
            + ".",
            "error",
        )
        return redirect(url_for("service_customers"))

    pending_token = uuid.uuid4().hex
    pending_uploads = session.get("pending_uploads", {})
    pending_uploads[pending_token] = {
        "type": "service_customers",
        "path": saved_path,
        "original_filename": upload.filename,
    }
    session["pending_uploads"] = pending_uploads
    session.modified = True

    return redirect(
        url_for(
            "service_upload_review",
            upload_type="customers",
            pending_token=pending_token,
        )
    )

@app.route("/service/lifts/upload", methods=["POST"])
@login_required
def service_lifts_upload():
    _module_visibility_required("service")

    pending_token = request.form.get("pending_token")
    action = request.form.get("action")

    if pending_token:
        pending_uploads = session.get("pending_uploads", {})
        pending = pending_uploads.get(pending_token)
        if not pending or pending.get("type") != "service_lifts":
            flash("The upload preview has expired. Please re-upload the file.", "error")
            return redirect(url_for("service_lifts"))

        file_path = pending.get("path")
        if not file_path or not os.path.exists(file_path):
            pending_uploads.pop(pending_token, None)
            session["pending_uploads"] = pending_uploads
            session.modified = True
            flash("The staged upload file was not found. Please try uploading again.", "error")
            return redirect(url_for("service_lifts"))

        if action == "discard":
            pending_uploads.pop(pending_token, None)
            session["pending_uploads"] = pending_uploads
            session.modified = True
            try:
                os.remove(file_path)
            except OSError:
                pass
            flash("Lift upload discarded.", "info")
            return redirect(url_for("service_lifts"))

        if action != "merge":
            flash("Select a valid action for the upload.", "error")
            return redirect(url_for("service_lifts"))

        outcome = process_lift_upload_file(
            file_path,
            apply_changes=True,
        )

        pending_uploads.pop(pending_token, None)
        session["pending_uploads"] = pending_uploads
        session.modified = True

        try:
            os.remove(file_path)
        except OSError:
            pass

        if outcome.created_count or outcome.updated_count:
            flash(
                f"AMC lift upload complete: {outcome.created_count} added, {outcome.updated_count} updated.",
                "success",
            )
        elif outcome.processed_rows and not outcome.row_errors:
            flash("No changes were detected in the uploaded workbook.", "info")
        else:
            flash("No rows were imported from the uploaded workbook.", "info")

        for message in outcome.row_errors[:25]:
            flash(message, "error")
        if len(outcome.row_errors) > 25:
            flash(
                f"Additional {len(outcome.row_errors) - 25} errors were omitted from the alert. Check the file and retry.",
                "warning",
            )

        return redirect(url_for("service_lifts"))

    upload = request.files.get("amc_lift_file")
    if not upload or not upload.filename:
        flash("Select an Excel workbook to upload.", "error")
        return redirect(url_for("service_lifts"))

    filename = (upload.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv")):
        flash(
            "Upload the .xlsx or .csv template exported from the AMC lifts modal.",
            "error",
        )
        return redirect(url_for("service_lifts"))

    saved_path = None
    try:
        saved_path = save_pending_upload_file(upload)
        outcome = process_lift_upload_file(
            saved_path,
            apply_changes=False,
        )
    except MissingDependencyError:
        if saved_path:
            try:
                os.remove(saved_path)
            except OSError:
                pass
        flash(OPENPYXL_MISSING_MESSAGE, "error")
        return redirect(url_for("service_lifts"))
    except ValueError:
        if saved_path:
            try:
                os.remove(saved_path)
            except OSError:
                pass
        flash(
            "Upload the .xlsx or .csv template exported from the AMC lifts modal.",
            "error",
        )
        return redirect(url_for("service_lifts"))
    except Exception:
        if saved_path:
            try:
                os.remove(saved_path)
            except OSError:
                pass
        flash(
            "Could not read the uploaded workbook. Ensure it is a valid .xlsx or .csv file.",
            "error",
        )
        return redirect(url_for("service_lifts"))

    header_map = outcome.header_map
    if not header_map:
        try:
            os.remove(saved_path)
        except OSError:
            pass
        flash("The uploaded workbook is missing header labels.", "error")
        return redirect(url_for("service_lifts"))

    missing_headers = [
        column
        for column in ["AMC Status", "AMC Duration"]
        if column not in header_map
    ]
    if (
        "Customer External ID" not in header_map
        and "Customer Code" not in header_map
        and "Customer Name" not in header_map
    ):
        missing_headers.append("Customer External ID / Customer Code / Customer Name")
    if missing_headers:
        try:
            os.remove(saved_path)
        except OSError:
            pass
        flash(
            "The uploaded workbook is missing required columns: "
            + ", ".join(missing_headers)
            + ".",
            "error",
        )
        return redirect(url_for("service_lifts"))

    pending_token = uuid.uuid4().hex
    pending_uploads = session.get("pending_uploads", {})
    pending_uploads[pending_token] = {
        "type": "service_lifts",
        "path": saved_path,
        "original_filename": upload.filename,
    }
    session["pending_uploads"] = pending_uploads
    session.modified = True

    return redirect(
        url_for(
            "service_upload_review",
            upload_type="lifts",
            pending_token=pending_token,
        )
    )


@app.route("/service/upload-review/<upload_type>/<pending_token>")
@login_required
def service_upload_review(upload_type, pending_token):
    _module_visibility_required("service")

    upload_key = (upload_type or "").strip().lower()
    config_map = {
        "customers": {
            "session_type": "service_customers",
            "processor": process_customer_upload_file,
            "redirect_endpoint": "service_customers",
            "confirm_endpoint": "service_customers_upload",
            "review_title": "Review customer upload",
            "invalid_template_message": (
                "Upload the .xlsx or .csv template exported from the customer upload modal."
            ),
            "read_error_message": (
                "Could not read the staged workbook. Ensure it is a valid .xlsx or .csv file."
            ),
        },
        "lifts": {
            "session_type": "service_lifts",
            "processor": process_lift_upload_file,
            "redirect_endpoint": "service_lifts",
            "confirm_endpoint": "service_lifts_upload",
            "review_title": "Review AMC lift upload",
            "invalid_template_message": (
                "Upload the .xlsx or .csv template exported from the AMC lifts modal."
            ),
            "read_error_message": (
                "Could not read the staged workbook. Ensure it is a valid .xlsx or .csv file."
            ),
        },
    }

    config = config_map.get(upload_key)
    if not config:
        abort(404)

    pending_uploads = session.get("pending_uploads", {})
    pending = pending_uploads.get(pending_token)
    if not pending or pending.get("type") != config["session_type"]:
        flash("The upload preview has expired. Please re-upload the file.", "error")
        return redirect(url_for(config["redirect_endpoint"]))

    file_path = pending.get("path")
    if not file_path or not os.path.exists(file_path):
        _clear_pending_upload(pending_token, remove_file=True)
        flash("The staged upload file was not found. Please try uploading again.", "error")
        return redirect(url_for(config["redirect_endpoint"]))

    try:
        outcome = config["processor"](file_path, apply_changes=False)
    except MissingDependencyError:
        _clear_pending_upload(pending_token, remove_file=True)
        flash(OPENPYXL_MISSING_MESSAGE, "error")
        return redirect(url_for(config["redirect_endpoint"]))
    except ValueError:
        _clear_pending_upload(pending_token, remove_file=True)
        flash(config["invalid_template_message"], "error")
        return redirect(url_for(config["redirect_endpoint"]))
    except Exception:
        _clear_pending_upload(pending_token, remove_file=True)
        flash(config["read_error_message"], "error")
        return redirect(url_for(config["redirect_endpoint"]))

    return render_template(
        "service/upload_review.html",
        review_title=config["review_title"],
        original_filename=pending.get("original_filename"),
        confirm_url=url_for(config["confirm_endpoint"]),
        pending_token=pending_token,
        created_items=outcome.created_items,
        updated_items=outcome.updated_items,
        created_count=outcome.created_count,
        updated_count=outcome.updated_count,
        row_errors=outcome.row_errors,
        processed_rows=outcome.processed_rows,
        upload_type=upload_key,
    )


@app.route("/service/lifts")
@login_required
def service_lifts():
    _module_visibility_required("service")

    search_query = (request.args.get("q") or "").strip()
    query = Lift.query.options(
        joinedload(Lift.customer),
        subqueryload(Lift.attachments),
        subqueryload(Lift.comments),
    )

    if search_query:
        like = f"%{search_query.lower()}%"
        query = query.filter(
            or_(
                func.lower(Lift.lift_code).like(like),
                func.lower(Lift.customer_code).like(like),
                func.lower(Lift.city).like(like),
                func.lower(Lift.state).like(like),
                func.lower(Lift.route).like(like),
                func.lower(Lift.lift_type).like(like),
                func.lower(Lift.lift_brand).like(like),
                func.lower(Lift.status).like(like),
            )
        )

    lifts = query.order_by(func.lower(Lift.lift_code)).all()
    customers = Customer.query.order_by(func.lower(Customer.company_name)).all()
    service_routes = ServiceRoute.query.order_by(
        func.lower(ServiceRoute.state), func.lower(ServiceRoute.branch)
    ).all()
    next_lift_code = generate_next_lift_code()
    next_customer_code = generate_next_customer_code()
    dropdown_options = get_dropdown_options_map()

    return render_template(
        "service/lifts.html",
        lifts=lifts,
        customers=customers,
        service_routes=service_routes,
        search_query=search_query,
        next_lift_code=next_lift_code,
        next_customer_code=next_customer_code,
        service_contracts=SERVICE_CONTRACTS,
        service_day_options=SERVICE_PREFERRED_DAY_OPTIONS,
        dropdown_options=dropdown_options,
        dropdown_meta=DROPDOWN_FIELD_DEFINITIONS,
        amc_duration_choices=AMC_DURATION_CHOICES,
        amc_duration_months=AMC_DURATION_MONTHS,
        amc_status_options=AMC_STATUS_OPTIONS,
    )


@app.route("/service/lifts/create", methods=["POST"])
@login_required
def service_lifts_create():
    _module_visibility_required("service")

    redirect_url = request.form.get("next") or url_for("service_lifts")

    lift_code = generate_next_lift_code()

    customer_name_input = clean_str(request.form.get("customer_name"))
    customer_code = None
    customer = None
    if customer_name_input:
        lowered = customer_name_input.lower()
        customer = Customer.query.filter(func.lower(Customer.company_name) == lowered).first()
        if not customer:
            customer = Customer.query.filter(func.lower(Customer.customer_code) == lowered).first()
        if not customer:
            flash("Select a valid customer from the list or create a new customer.", "error")
            return redirect(redirect_url)
        customer_code = customer.customer_code

    route_value = clean_str(request.form.get("route"))
    if route_value:
        valid_route = ServiceRoute.query.filter(func.lower(ServiceRoute.state) == route_value.lower()).first()
        if not valid_route:
            flash("Select a valid service route from the dropdown.", "error")
            return redirect(redirect_url)
        route_value = valid_route.state

    capacity_persons, error = parse_int_field(request.form.get("capacity_persons"), "Capacity (persons)")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    capacity_kg, error = parse_int_field(request.form.get("capacity_kg"), "Capacity (kg)")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    speed_mps, error = parse_float_field(request.form.get("speed_mps"), "Speed (m/s)")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    install_date, error = parse_date_field(request.form.get("install_date"), "Install date")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    warranty_expiry, error = parse_date_field(request.form.get("warranty_expiry"), "Warranty expiry")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    amc_start, error = parse_date_field(request.form.get("amc_start"), "AMC start")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    amc_duration_value = clean_str(request.form.get("amc_duration_key"))
    amc_duration_key = amc_duration_value.lower() if amc_duration_value else None
    if amc_duration_key and amc_duration_key not in AMC_DURATION_MONTHS:
        flash("Select a valid AMC duration option.", "error")
        return redirect(redirect_url)
    computed_amc_end = calculate_amc_end_date(amc_start, amc_duration_key)

    preferred_date, error = parse_preferred_service_date(
        request.form.get("preferred_service_date")
    )
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    preferred_time, error = parse_time_field(
        request.form.get("preferred_service_time"), "Preferred service time"
    )
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    preferred_days_raw = request.form.getlist("preferred_service_days")
    preferred_days, day_error = parse_preferred_service_days(preferred_days_raw)
    if day_error:
        flash(day_error, "error")
        return redirect(redirect_url)

    contract_input = clean_str(request.form.get("amc_contract_id"))
    amc_contract_id = None
    if contract_input:
        contract_record = get_service_contract_by_id(contract_input)
        if not contract_record:
            flash("Select a valid AMC contract.", "error")
            return redirect(redirect_url)
        amc_contract_id = contract_record.get("id")

    lift = Lift(
        lift_code=lift_code,
        external_lift_id=clean_str(request.form.get("external_lift_id")),
        customer_code=customer_code,
        site_address_line1=clean_str(request.form.get("site_address_line1")),
        site_address_line2=clean_str(request.form.get("site_address_line2")),
        city=clean_str(request.form.get("city")),
        state=clean_str(request.form.get("state")),
        pincode=clean_str(request.form.get("pincode")),
        geo_location=clean_str(request.form.get("geo_location")),
        country="India",
        building_villa_number=clean_str(request.form.get("building_villa_number")),
        route=route_value,
        building_floors=clean_str(request.form.get("building_floors")),
        lift_type=clean_str(request.form.get("lift_type")),
        lift_brand=clean_str(request.form.get("lift_brand")),
        capacity_persons=capacity_persons,
        capacity_kg=capacity_kg,
        speed_mps=speed_mps,
        machine_type=clean_str(request.form.get("machine_type")),
        machine_brand=clean_str(request.form.get("machine_brand")),
        controller_brand=clean_str(request.form.get("controller_brand")),
        door_type=clean_str(request.form.get("door_type")),
        door_finish=clean_str(request.form.get("door_finish")),
        cabin_finish=clean_str(request.form.get("cabin_finish")),
        power_supply=clean_str(request.form.get("power_supply")),
        install_date=install_date,
        warranty_expiry=warranty_expiry,
        amc_status=clean_str(request.form.get("amc_status")),
        amc_start=amc_start,
        amc_end=computed_amc_end,
        amc_duration_key=amc_duration_key,
        amc_contract_id=amc_contract_id,
        qr_code_url=clean_str(request.form.get("qr_code_url")),
        status=clean_str(request.form.get("status")),
        remarks=clean_str(request.form.get("remarks")),
        preferred_service_date=preferred_date,
        preferred_service_time=preferred_time,
        last_updated_by=current_user.id if current_user.is_authenticated else None,
    )
    lift.preferred_service_days = preferred_days
    lift.set_capacity_display()

    db.session.add(lift)
    db.session.commit()

    flash(f"Lift {lift.lift_code} created.", "success")
    return redirect(redirect_url)


@app.route("/service/lifts/<int:lift_id>")
@login_required
def service_lift_detail(lift_id):
    _module_visibility_required("service")

    lift = db.session.get(Lift, lift_id)
    if not lift:
        flash("Lift not found.", "error")
        return redirect(url_for("service_lifts"))

    payload = build_lift_payload(lift)
    attachments = (
        LiftFile.query.filter_by(lift_id=lift.id)
        .order_by(LiftFile.created_at.desc())
        .all()
    )
    comments = (
        LiftComment.query.filter_by(lift_id=lift.id)
        .order_by(LiftComment.created_at.desc())
        .all()
    )

    customers = Customer.query.order_by(func.lower(Customer.company_name)).all()
    service_routes = ServiceRoute.query.order_by(
        func.lower(ServiceRoute.state), func.lower(ServiceRoute.branch)
    ).all()
    dropdown_options = get_dropdown_options_map()

    return render_template(
        "service/lift_detail.html",
        lift=lift,
        payload=payload,
        attachments=attachments,
        comments=comments,
        customers=customers,
        service_routes=service_routes,
        service_contracts=SERVICE_CONTRACTS,
        service_day_options=SERVICE_PREFERRED_DAY_OPTIONS,
        service_visit_status_options=SERVICE_VISIT_STATUS_OPTIONS,
        dropdown_options=dropdown_options,
        dropdown_meta=DROPDOWN_FIELD_DEFINITIONS,
        amc_duration_choices=AMC_DURATION_CHOICES,
        amc_duration_months=AMC_DURATION_MONTHS,
    )


@app.route("/service/lifts/<int:lift_id>/edit")
@login_required
def service_lift_edit(lift_id):
    _module_visibility_required("service")

    lift = db.session.get(Lift, lift_id)
    if not lift:
        flash("Lift not found.", "error")
        return redirect(url_for("service_lifts"))

    customers = Customer.query.order_by(func.lower(Customer.company_name)).all()
    service_routes = ServiceRoute.query.order_by(
        func.lower(ServiceRoute.state), func.lower(ServiceRoute.branch)
    ).all()
    attachments = (
        LiftFile.query.filter_by(lift_id=lift.id)
        .order_by(LiftFile.created_at.desc())
        .all()
    )
    comments = (
        LiftComment.query.filter_by(lift_id=lift.id)
        .order_by(LiftComment.created_at.desc())
        .all()
    )
    dropdown_options = get_dropdown_options_map()
    return render_template(
        "service/lift_edit.html",
        lift=lift,
        customers=customers,
        service_routes=service_routes,
        attachments=attachments,
        comments=comments,
        service_contracts=SERVICE_CONTRACTS,
        service_day_options=SERVICE_PREFERRED_DAY_OPTIONS,
        dropdown_options=dropdown_options,
        dropdown_meta=DROPDOWN_FIELD_DEFINITIONS,
        amc_duration_choices=AMC_DURATION_CHOICES,
        amc_duration_months=AMC_DURATION_MONTHS,
    )


@app.route("/service/lifts/<int:lift_id>/update", methods=["POST"])
@login_required
def service_lift_update(lift_id):
    _module_visibility_required("service")

    lift = db.session.get(Lift, lift_id)
    if not lift:
        flash("Lift not found.", "error")
        return redirect(url_for("service_lifts"))

    redirect_url = request.form.get("next") or url_for("service_lift_detail", lift_id=lift.id)
    form_section = (request.form.get("form_section") or "full").strip().lower()

    if form_section == "summary":
        customer_code_input = clean_str(request.form.get("customer_code"))
        if customer_code_input:
            new_customer_code = customer_code_input.upper()
            customer = Customer.query.filter(
                func.lower(Customer.customer_code) == new_customer_code.lower()
            ).first()
            if not customer:
                flash("Select a valid customer from the list or create a new customer.", "error")
                return redirect(redirect_url)
            lift.customer_code = new_customer_code
        else:
            lift.customer_code = None

        install_date, error = parse_date_field(request.form.get("install_date"), "Completion date")
        if error:
            flash(error, "error")
            return redirect(redirect_url)

        warranty_expiry, error = parse_date_field(request.form.get("warranty_expiry"), "Warranty expiry")
        if error:
            flash(error, "error")
            return redirect(redirect_url)

        lift.external_lift_id = clean_str(request.form.get("external_lift_id"))
        lift.site_address_line1 = clean_str(request.form.get("site_address_line1"))
        lift.site_address_line2 = clean_str(request.form.get("site_address_line2"))
        lift.building_villa_number = clean_str(request.form.get("building_villa_number"))
        lift.city = clean_str(request.form.get("city"))
        lift.state = clean_str(request.form.get("state"))
        lift.pincode = clean_str(request.form.get("pincode"))
        lift.geo_location = clean_str(request.form.get("geo_location"))
        lift.country = "India"
        lift.install_date = install_date
        lift.warranty_expiry = warranty_expiry
        lift.qr_code_url = clean_str(request.form.get("qr_code_url"))
        lift.last_updated_by = current_user.id if current_user.is_authenticated else None

        db.session.commit()
        flash("Lift summary updated.", "success")
        return redirect(redirect_url)

    if form_section == "specifications":
        capacity_persons, error = parse_int_field(
            request.form.get("capacity_persons"), "Capacity (persons)"
        )
        if error:
            flash(error, "error")
            return redirect(redirect_url)

        capacity_kg, error = parse_int_field(request.form.get("capacity_kg"), "Capacity (kg)")
        if error:
            flash(error, "error")
            return redirect(redirect_url)

        speed_mps, error = parse_float_field(request.form.get("speed_mps"), "Speed (m/s)")
        if error:
            flash(error, "error")
            return redirect(redirect_url)

        lift.lift_type = clean_str(request.form.get("lift_type"))
        lift.lift_brand = clean_str(request.form.get("lift_brand"))
        lift.building_floors = clean_str(request.form.get("building_floors"))
        lift.capacity_persons = capacity_persons
        lift.capacity_kg = capacity_kg
        lift.speed_mps = speed_mps
        lift.machine_type = clean_str(request.form.get("machine_type"))
        lift.machine_brand = clean_str(request.form.get("machine_brand"))
        lift.controller_brand = clean_str(request.form.get("controller_brand"))
        lift.door_type = clean_str(request.form.get("door_type"))
        lift.door_finish = clean_str(request.form.get("door_finish"))
        lift.cabin_finish = clean_str(request.form.get("cabin_finish"))
        lift.power_supply = clean_str(request.form.get("power_supply"))
        lift.remarks = clean_str(request.form.get("remarks"))
        lift.set_capacity_display()
        lift.last_updated_by = current_user.id if current_user.is_authenticated else None

        db.session.commit()
        flash("Lift specifications updated.", "success")
        return redirect(redirect_url)

    if form_section == "amc_details":
        route_value = clean_str(request.form.get("route"))
        if route_value:
            valid_route = ServiceRoute.query.filter(
                func.lower(ServiceRoute.state) == route_value.lower()
            ).first()
            if not valid_route:
                flash("Select a valid service route from the dropdown.", "error")
                return redirect(redirect_url)
            route_value = valid_route.state
        else:
            route_value = None

        amc_start, error = parse_date_field(request.form.get("amc_start"), "AMC start date")
        if error:
            flash(error, "error")
            return redirect(redirect_url)

        amc_duration_value = clean_str(request.form.get("amc_duration_key"))
        amc_duration_key = amc_duration_value.lower() if amc_duration_value else None
        if amc_duration_key and amc_duration_key not in AMC_DURATION_MONTHS:
            flash("Select a valid AMC duration option.", "error")
            return redirect(redirect_url)
        computed_amc_end = calculate_amc_end_date(amc_start, amc_duration_key)

        preferred_date, error = parse_preferred_service_date(
            request.form.get("preferred_service_date")
        )
        if error:
            flash(error, "error")
            return redirect(redirect_url)

        preferred_time, error = parse_time_field(
            request.form.get("preferred_service_time"), "Preferred PM time"
        )
        if error:
            flash(error, "error")
            return redirect(redirect_url)

        preferred_days_raw = request.form.getlist("preferred_service_days")
        preferred_days, day_error = parse_preferred_service_days(preferred_days_raw)
        if day_error:
            flash(day_error, "error")
            return redirect(redirect_url)

        contract_input = clean_str(request.form.get("amc_contract_id"))
        amc_contract_id = None
        if contract_input:
            contract_record = get_service_contract_by_id(contract_input)
            if not contract_record:
                flash("Select a valid AMC contract.", "error")
                return redirect(redirect_url)
            amc_contract_id = contract_record.get("id")

        last_service_date, error = parse_date_field(
            request.form.get("last_service_date"), "Last service date"
        )
        if error:
            flash(error, "error")
            return redirect(redirect_url)

        lift.amc_status = clean_str(request.form.get("amc_status"))
        lift.amc_start = amc_start
        lift.amc_end = computed_amc_end
        lift.amc_duration_key = amc_duration_key
        lift.amc_contract_id = amc_contract_id
        lift.route = route_value
        lift.preferred_service_date = preferred_date
        lift.preferred_service_time = preferred_time
        lift.preferred_service_days = preferred_days
        lift.last_service_date = last_service_date
        lift.notes = clean_str(request.form.get("service_notes"))
        lift.status = clean_str(request.form.get("status"))
        lift.last_updated_by = current_user.id if current_user.is_authenticated else None

        db.session.commit()
        flash("AMC details updated.", "success")
        return redirect(redirect_url)

    if form_section == "service_schedule":
        dates = request.form.getlist("service_date")
        technicians = request.form.getlist("service_technician")
        statuses = request.form.getlist("service_status")
        slip_existing_values = request.form.getlist("service_slip_existing")
        slip_labels = request.form.getlist("service_slip_label")
        slip_files = request.files.getlist("service_slip_file")
        max_len = max(
            len(dates),
            len(technicians),
            len(statuses),
            len(slip_existing_values),
            len(slip_labels),
            len(slip_files),
        )
        schedule_entries = []
        static_root = os.path.join(BASE_DIR, "static")
        for idx in range(max_len):
            date_value = dates[idx] if idx < len(dates) else ""
            technician = clean_str(technicians[idx]) if idx < len(technicians) else None
            status_value = clean_str(statuses[idx]) if idx < len(statuses) else None
            slip_value = (
                clean_str(slip_existing_values[idx])
                if idx < len(slip_existing_values)
                else None
            )
            slip_label = (
                clean_str(slip_labels[idx]) if idx < len(slip_labels) else None
            )
            uploaded_file = slip_files[idx] if idx < len(slip_files) else None
            if uploaded_file and getattr(uploaded_file, "filename", None):
                if not uploaded_file.filename:
                    uploaded_file = None
            if uploaded_file and uploaded_file.filename:
                safe_name = secure_filename(uploaded_file.filename)
                if not safe_name:
                    flash(
                        "The uploaded service slip file name is not valid.",
                        "error",
                    )
                    return redirect(redirect_url)
                if not allowed_file(safe_name, kind="attachment"):
                    flash(
                        "Unsupported file type for service slip. Upload images, videos or documents only.",
                        "error",
                    )
                    return redirect(redirect_url)
                upload_root = os.path.join(
                    app.config["UPLOAD_FOLDER"], "service_slips", str(lift.id)
                )
                os.makedirs(upload_root, exist_ok=True)
                unique_name = f"{uuid.uuid4().hex}_{safe_name}"
                destination_path = os.path.join(upload_root, unique_name)
                uploaded_file.save(destination_path)
                slip_value = (
                    os.path.relpath(destination_path, static_root)
                    .replace(os.sep, "/")
                    .lstrip("/")
                )
                slip_label = uploaded_file.filename
            if slip_value and isinstance(slip_value, str) and not slip_value.lower().startswith(("http://", "https://")):
                normalized_value = slip_value.lstrip("/")
                if normalized_value.startswith("static/"):
                    normalized_value = normalized_value.split("static/", 1)[1]
                slip_value = normalized_value or None
            if not any([date_value, technician, status_value, slip_value, slip_label]):
                continue
            if not date_value:
                flash(
                    f"Service date is required for each scheduled visit (row {idx + 1}).",
                    "error",
                )
                return redirect(redirect_url)
            visit_date, error = parse_date_field(
                date_value, f"Service date (row {idx + 1})"
            )
            if error:
                flash(error, "error")
                return redirect(redirect_url)
            status_key = (
                status_value.lower()
                if status_value and status_value.lower() in SERVICE_VISIT_STATUS_LABELS
                else "scheduled"
            )
            schedule_entries.append(
                {
                    "date": visit_date.isoformat() if visit_date else None,
                    "technician": technician,
                    "status": status_key,
                    "slip_url": slip_value,
                    "slip_label": slip_label,
                }
            )
        lift.service_schedule = schedule_entries
        lift.last_updated_by = current_user.id if current_user.is_authenticated else None
        db.session.commit()
        flash("Service schedule updated.", "success")
        return redirect(redirect_url)

    if form_section == "amc_contacts":
        names = request.form.getlist("contact_name")
        designations = request.form.getlist("contact_designation")
        phones = request.form.getlist("contact_phone")
        emails = request.form.getlist("contact_email")
        max_len = max(len(names), len(designations), len(phones), len(emails))
        contacts = []
        for idx in range(max_len):
            name = clean_str(names[idx]) if idx < len(names) else None
            designation = clean_str(designations[idx]) if idx < len(designations) else None
            phone = clean_str(phones[idx]) if idx < len(phones) else None
            email = clean_str(emails[idx]) if idx < len(emails) else None
            if not any([name, designation, phone, email]):
                continue
            contacts.append(
                {
                    "name": name,
                    "designation": designation,
                    "phone": phone,
                    "email": email,
                }
            )
        lift.amc_contacts = contacts
        lift.last_updated_by = current_user.id if current_user.is_authenticated else None
        db.session.commit()
        flash("AMC contacts updated.", "success")
        return redirect(redirect_url)

    if form_section == "lifetime_metrics":
        labels = request.form.getlist("metric_label")
        displays = request.form.getlist("metric_display")
        max_len = max(len(labels), len(displays))
        metrics = []
        for idx in range(max_len):
            label = clean_str(labels[idx]) if idx < len(labels) else None
            display_raw = displays[idx] if idx < len(displays) else ""
            display = display_raw.strip() if isinstance(display_raw, str) else None
            if not label and not display:
                continue
            metrics.append({"label": label or "Metric", "display": display or "—"})
        lift.lifetime_metrics = metrics
        lift.last_updated_by = current_user.id if current_user.is_authenticated else None
        db.session.commit()
        flash("Lifetime value metrics updated.", "success")
        return redirect(redirect_url)

    if form_section == "timeline":
        timeline_date, error = parse_date_field(request.form.get("timeline_date"), "Timeline date")
        if error:
            flash(error, "error")
            return redirect(redirect_url)
        title = clean_str(request.form.get("timeline_title")) or "Update"
        detail = clean_str(request.form.get("timeline_detail"))
        category = clean_str(request.form.get("timeline_category")) or "Update"
        entries = list(lift.timeline_entries)
        actor_info = timeline_actor_context()
        entries.insert(
            0,
            {
                "date": timeline_date,
                "title": title,
                "detail": detail,
                "category": category,
                **actor_info,
            },
        )
        lift.timeline_entries = entries
        lift.last_updated_by = current_user.id if current_user.is_authenticated else None
        db.session.commit()
        flash("Timeline entry added.", "success")
        return redirect(redirect_url)

    # Fallback to full update
    customer_code_input = clean_str(request.form.get("customer_code"))
    if customer_code_input:
        new_customer_code = customer_code_input.upper()
        customer = Customer.query.filter(func.lower(Customer.customer_code) == new_customer_code.lower()).first()
        if not customer:
            flash("Select a valid customer from the list or create a new customer.", "error")
            return redirect(redirect_url)
        lift.customer_code = new_customer_code
    else:
        lift.customer_code = None

    route_value = clean_str(request.form.get("route"))
    if route_value:
        valid_route = ServiceRoute.query.filter(func.lower(ServiceRoute.state) == route_value.lower()).first()
        if not valid_route:
            flash("Select a valid service route from the dropdown.", "error")
            return redirect(redirect_url)
        route_value = valid_route.state

    capacity_persons, error = parse_int_field(request.form.get("capacity_persons"), "Capacity (persons)")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    capacity_kg, error = parse_int_field(request.form.get("capacity_kg"), "Capacity (kg)")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    speed_mps, error = parse_float_field(request.form.get("speed_mps"), "Speed (m/s)")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    install_date, error = parse_date_field(request.form.get("install_date"), "Install date")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    warranty_expiry, error = parse_date_field(request.form.get("warranty_expiry"), "Warranty expiry")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    amc_start, error = parse_date_field(request.form.get("amc_start"), "AMC start")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    amc_duration_value = clean_str(request.form.get("amc_duration_key"))
    amc_duration_key = amc_duration_value.lower() if amc_duration_value else None
    if amc_duration_key and amc_duration_key not in AMC_DURATION_MONTHS:
        flash("Select a valid AMC duration option.", "error")
        return redirect(redirect_url)
    computed_amc_end = calculate_amc_end_date(amc_start, amc_duration_key)

    preferred_date, error = parse_preferred_service_date(
        request.form.get("preferred_service_date")
    )
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    preferred_time, error = parse_time_field(
        request.form.get("preferred_service_time"), "Preferred service time"
    )
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    preferred_days_raw = request.form.getlist("preferred_service_days")
    preferred_days, day_error = parse_preferred_service_days(preferred_days_raw)
    if day_error:
        flash(day_error, "error")
        return redirect(redirect_url)

    contract_input = clean_str(request.form.get("amc_contract_id"))
    amc_contract_id = None
    if contract_input:
        contract_record = get_service_contract_by_id(contract_input)
        if not contract_record:
            flash("Select a valid AMC contract.", "error")
            return redirect(redirect_url)
        amc_contract_id = contract_record.get("id")

    lift.external_lift_id = clean_str(request.form.get("external_lift_id"))
    lift.site_address_line1 = clean_str(request.form.get("site_address_line1"))
    lift.site_address_line2 = clean_str(request.form.get("site_address_line2"))
    lift.city = clean_str(request.form.get("city"))
    lift.state = clean_str(request.form.get("state"))
    lift.pincode = clean_str(request.form.get("pincode"))
    lift.geo_location = clean_str(request.form.get("geo_location"))
    lift.country = "India"
    lift.building_villa_number = clean_str(request.form.get("building_villa_number"))
    lift.route = route_value
    lift.building_floors = clean_str(request.form.get("building_floors"))
    lift.lift_type = clean_str(request.form.get("lift_type"))
    lift.capacity_persons = capacity_persons
    lift.capacity_kg = capacity_kg
    lift.speed_mps = speed_mps
    lift.machine_type = clean_str(request.form.get("machine_type"))
    lift.machine_brand = clean_str(request.form.get("machine_brand"))
    lift.controller_brand = clean_str(request.form.get("controller_brand"))
    lift.door_type = clean_str(request.form.get("door_type"))
    lift.door_finish = clean_str(request.form.get("door_finish"))
    lift.cabin_finish = clean_str(request.form.get("cabin_finish"))
    lift.power_supply = clean_str(request.form.get("power_supply"))
    lift.install_date = install_date
    lift.warranty_expiry = warranty_expiry
    lift.amc_status = clean_str(request.form.get("amc_status"))
    lift.amc_start = amc_start
    lift.amc_end = computed_amc_end
    lift.amc_duration_key = amc_duration_key
    lift.amc_contract_id = amc_contract_id
    lift.qr_code_url = clean_str(request.form.get("qr_code_url"))
    lift.status = clean_str(request.form.get("status"))
    lift.remarks = clean_str(request.form.get("remarks"))
    service_notes_value = request.form.get("service_notes")
    if service_notes_value is not None:
        lift.notes = clean_str(service_notes_value)
    elif "notes" in request.form:
        lift.notes = clean_str(request.form.get("notes"))
    lift.preferred_service_date = preferred_date
    lift.preferred_service_time = preferred_time
    lift.preferred_service_days = preferred_days
    lift.last_updated_by = current_user.id if current_user.is_authenticated else None
    lift.set_capacity_display()

    db.session.commit()

    flash("Lift details updated.", "success")
    return redirect(redirect_url)


@app.route("/service/lifts/<int:lift_id>/delete", methods=["POST"])
@login_required
def service_lift_delete(lift_id):
    _module_visibility_required("service")

    if not current_user.is_admin:
        abort(403)

    lift = db.session.get(Lift, lift_id)
    if not lift:
        flash("Lift not found.", "error")
        return redirect(url_for("service_lifts"))

    redirect_url = request.form.get("next") or url_for("service_lifts")
    if not redirect_url.startswith("/"):
        redirect_url = url_for("service_lifts")

    delete_lift_record(lift)
    db.session.commit()

    flash(f"Lift {lift.lift_code} deleted.", "success")
    return redirect(redirect_url)


@app.route("/service/lifts/<int:lift_id>/notes", methods=["POST"])
@login_required
def service_lift_update_notes(lift_id):
    _module_visibility_required("service")

    lift = db.session.get(Lift, lift_id)
    if not lift:
        flash("Lift not found.", "error")
        return redirect(url_for("service_lifts"))

    redirect_url = request.form.get("next") or url_for("service_lift_detail", lift_id=lift.id)

    lift.remarks = clean_str(request.form.get("remarks"))

    preferred_date, error = parse_preferred_service_date(
        request.form.get("preferred_service_date")
    )
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    preferred_time_raw = request.form.get("preferred_service_time")
    preferred_time, error = parse_time_field(preferred_time_raw, "Preferred service time")
    if error:
        flash(error, "error")
        return redirect(redirect_url)

    if "preferred_service_days" in request.form:
        preferred_days_raw = request.form.getlist("preferred_service_days")
        preferred_days, day_error = parse_preferred_service_days(preferred_days_raw)
        if day_error:
            flash(day_error, "error")
            return redirect(redirect_url)
        lift.preferred_service_days = preferred_days

    lift.preferred_service_date = preferred_date
    lift.preferred_service_time = preferred_time
    db.session.commit()

    flash("Preferences updated.", "success")
    return redirect(redirect_url)


@app.route("/service/lifts/<int:lift_id>/comments", methods=["POST"])
@login_required
def service_lift_add_comment(lift_id):
    _module_visibility_required("service")

    lift = db.session.get(Lift, lift_id)
    if not lift:
        flash("Lift not found.", "error")
        return redirect(url_for("service_lifts"))

    redirect_url = request.form.get("next") or url_for("service_lift_detail", lift_id=lift.id)
    body = (request.form.get("body") or "").strip()
    if not body:
        flash("Comment cannot be empty.", "error")
        return redirect(redirect_url)

    comment = LiftComment(
        lift=lift,
        body=body,
        author=current_user if current_user.is_authenticated else None,
    )
    db.session.add(comment)
    db.session.commit()

    flash("Comment added.", "success")
    return redirect(redirect_url)


@app.route("/service/lifts/<int:lift_id>/files", methods=["POST"])
@login_required
def service_lift_upload_file(lift_id):
    _module_visibility_required("service")

    lift = db.session.get(Lift, lift_id)
    if not lift:
        flash("Lift not found.", "error")
        return redirect(url_for("service_lifts"))

    redirect_url = request.form.get("next") or url_for("service_lift_detail", lift_id=lift.id)

    uploaded_file = request.files.get("attachment")
    if not uploaded_file or not uploaded_file.filename:
        flash("Select a file to upload.", "error")
        return redirect(redirect_url)

    safe_name = secure_filename(uploaded_file.filename)
    if not safe_name:
        flash("The selected file name is not valid.", "error")
        return redirect(redirect_url)

    if not allowed_file(safe_name, kind="attachment"):
        flash("Unsupported file type. Upload images, videos or documents only.", "error")
        return redirect(redirect_url)

    upload_root = os.path.join(app.config["UPLOAD_FOLDER"], "lifts", str(lift.id))
    os.makedirs(upload_root, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}_{safe_name}"
    destination_path = os.path.join(upload_root, unique_name)
    uploaded_file.save(destination_path)

    static_root = os.path.join(BASE_DIR, "static")
    stored_relative = os.path.relpath(destination_path, static_root).replace(os.sep, "/")

    category = (request.form.get("category") or "other").strip().lower()
    if category not in {"document", "media", "other"}:
        category = "other"

    record = LiftFile(
        lift=lift,
        original_filename=uploaded_file.filename,
        stored_path=stored_relative,
        content_type=uploaded_file.mimetype,
        file_size=os.path.getsize(destination_path),
        label=clean_str(request.form.get("label")),
        description=clean_str(request.form.get("description")),
        category=category,
        uploaded_by=current_user if current_user.is_authenticated else None,
    )
    db.session.add(record)
    db.session.commit()

    flash("File uploaded.", "success")
    return redirect(redirect_url)


@app.route("/service/complaints")
@login_required
def service_complaints():
    _module_visibility_required("service")
    return render_template("service/complaints.html", complaints=SERVICE_COMPLAINTS)


@app.route("/service/contracts")
@login_required
def service_contracts():
    _module_visibility_required("service")
    contracts = []
    for contract in SERVICE_CONTRACTS:
        contracts.append(
            {
                **contract,
                "start_display": contract.get("start").strftime("%d %b %Y") if isinstance(contract.get("start"), datetime.date) else "—",
                "end_display": contract.get("end").strftime("%d %b %Y") if isinstance(contract.get("end"), datetime.date) else "—",
                "renewal_display": contract.get("renewal").strftime("%d %b %Y") if isinstance(contract.get("renewal"), datetime.date) else "—",
            }
        )
    return render_template("service/contracts.html", contracts=contracts)


@app.route("/service/parts-materials")
@login_required
def service_parts_materials():
    _module_visibility_required("service")
    ledger = SERVICE_PARTS_LEDGER
    return render_template(
        "service/parts_materials.html",
        stock_alerts=ledger.get("stock_alerts", []),
        consumption=ledger.get("consumption", []),
        returns=ledger.get("returns", []),
    )


@app.route("/service/preventive-maintenance")
@login_required
def service_preventive_maintenance():
    _module_visibility_required("service")
    snapshot = get_service_schedule_snapshot()
    today = snapshot["today"]

    def preference_warning(lift_obj, visit_date=None):
        if not lift_obj:
            return None
        messages = []
        if lift_obj.preferred_service_date:
            preferred_display = format_preferred_service_date(
                lift_obj.preferred_service_date
            )
            if isinstance(visit_date, datetime.date):
                if not preferred_service_date_matches(
                    lift_obj.preferred_service_date, visit_date
                ):
                    messages.append(f"Prefers {preferred_display}")
            else:
                messages.append(f"Prefers {preferred_display}")
        if lift_obj.preferred_service_time:
            messages.append(
                f"Prefers {lift_obj.preferred_service_time.strftime('%H:%M')}"
            )
        for day_key in lift_obj.preferred_service_days:
            if day_key and day_key in SERVICE_PREFERRED_DAY_LABELS and day_key != "any":
                day_label = SERVICE_PREFERRED_DAY_LABELS.get(day_key)
                if isinstance(visit_date, datetime.date):
                    if visit_date.strftime("%A").lower() != day_key:
                        messages.append(f"Prefers {day_label}")
                else:
                    messages.append(f"Prefers {day_label}")
                break
        if messages:
            return " · ".join(messages)
        return None

    upcoming = []
    overdue = []
    for entry in snapshot["entries"]:
        visit_date = entry.get("date")
        if not isinstance(visit_date, datetime.date):
            continue
        if entry.get("status") == "completed":
            continue
        lift = entry.get("lift")
        site_name = (
            (lift.customer.company_name if lift and lift.customer else None)
            or (lift.customer.customer_code if lift and lift.customer else None)
            or (lift.customer_code if lift and lift.customer_code else None)
            or (lift.city if lift and lift.city else None)
            or "—"
        )
        base_payload = {
            "site": site_name,
            "lift": (lift.lift_code if lift and lift.lift_code else (f"Lift #{lift.id}" if lift else "Lift")),
            "technician": entry.get("technician") or "Unassigned",
            "checklist": entry.get("checklist") or "—",
            "preference_warning": preference_warning(lift, visit_date),
            "_sort_date": visit_date,
        }
        if visit_date >= today:
            upcoming.append(
                {
                    **base_payload,
                    "visit_display": visit_date.strftime("%d %b %Y"),
                }
            )
        else:
            overdue.append(
                {
                    **base_payload,
                    "due_display": visit_date.strftime("%d %b %Y"),
                    "days_overdue": (today - visit_date).days,
                }
            )

    upcoming.sort(key=lambda item: item.get("_sort_date"))
    overdue.sort(key=lambda item: item.get("_sort_date"))
    for collection in (upcoming, overdue):
        for item in collection:
            item.pop("_sort_date", None)

    checklist_counts = Counter(
        entry.get("checklist")
        for entry in snapshot["entries"]
        if entry.get("checklist")
    )
    checklists = [
        {
            "name": name,
            "items": count,
            "photo_rules": "Logged via service schedule",
        }
        for name, count in sorted(checklist_counts.items())
    ]

    return render_template(
        "service/preventive_maintenance.html",
        upcoming=upcoming,
        overdue=overdue,
        checklists=checklists,
    )


@app.route("/service/automations")
@login_required
def service_automations():
    _module_visibility_required("service")
    roles = []
    for role, capabilities in SERVICE_AUTOMATIONS.get("roles", {}).items():
        roles.append({"role": role, "capabilities": capabilities})
    return render_template(
        "service/automations.html",
        flows=SERVICE_AUTOMATIONS.get("flows", []),
        roles=roles,
        config_options=SERVICE_AUTOMATIONS.get("config", []),
    )


# ---------------------- SRT MODULE ----------------------
@app.route("/srt")
@login_required
def srt_overview():
    _module_visibility_required("srt")
    status_filter = request.args.get("status", "all").lower()
    today = datetime.date.today()

    tasks = []
    for task in SRT_SAMPLE_TASKS:
        due_date = task.get("due_date")
        due_in = (due_date - today).days if due_date else None
        due_date_display = due_date.strftime("%d %b %Y") if due_date else ""
        due_date_iso = due_date.isoformat() if due_date else ""
        tasks.append(
            {
                **task,
                "due_in": due_in,
                "due_date_display": due_date_display,
                "due_date_iso": due_date_iso,
            }
        )

    if status_filter in {"pending", "open"}:
        filtered_tasks = [task for task in tasks if task["status"].lower() != "closed"]
    elif status_filter in {"in-progress", "in_progress"}:
        filtered_tasks = [task for task in tasks if task["status"].lower() == "in progress"]
    elif status_filter in {"closed", "completed"}:
        filtered_tasks = [task for task in tasks if task["status"].lower() == "closed"]
    else:
        filtered_tasks = [task for task in tasks if task["status"].lower() != "closed"]

    summary = {
        "total_pending": sum(1 for task in tasks if task["status"].lower() != "closed"),
        "high_priority": sum(1 for task in tasks if task["priority"].lower() == "high" and task["status"].lower() != "closed"),
        "due_this_week": sum(
            1
            for task in tasks
            if task["status"].lower() != "closed"
            and task.get("due_in") is not None
            and 0 <= task["due_in"] <= 7
        ),
        "oldest_age": max((task["age_days"] for task in tasks), default=0),
    }

    site_options = sorted(
        (
            {
                "key": site.get("key"),
                "name": site.get("name"),
            }
            for site in SRT_SITES
        ),
        key=lambda item: (item.get("name") or "").lower(),
    )

    return render_template(
        "srt_overview.html",
        tasks=filtered_tasks,
        status_filter=status_filter,
        summary=summary,
        site_options=site_options,
        team_members=SRT_TEAM_MEMBERS,
    )


@app.route("/srt/form-templates", methods=["GET", "POST"])
@login_required
def srt_form_templates():
    _module_visibility_required("srt")
    global SRT_FORM_TEMPLATES

    if request.method == "POST":
        action = request.form.get("action", "").lower()
        name = (request.form.get("name") or "").strip()
        category = (request.form.get("category") or "General").strip() or "General"
        description = (request.form.get("description") or "").strip()
        usage_count_raw = request.form.get("usage_count")
        schema_json = (request.form.get("schema_json") or "").strip()

        usage_count = 0
        try:
            if usage_count_raw is not None and usage_count_raw != "":
                usage_count = max(0, int(usage_count_raw))
        except ValueError:
            usage_count = 0

        schema_payload = _default_srt_schema()
        if schema_json:
            try:
                loaded = json.loads(schema_json)
            except json.JSONDecodeError:
                loaded = _default_srt_schema()
            schema_payload = loaded

        schema = copy.deepcopy(_normalise_srt_schema(schema_payload))
        today = datetime.date.today()

        if action == "create":
            if not name:
                flash("Template name is required.", "error")
            else:
                template_id = slugify(name)
                if any(template["id"] == template_id for template in SRT_FORM_TEMPLATES):
                    template_id = f"{template_id}-{_random_digits(4)}"

                SRT_FORM_TEMPLATES.append(
                    {
                        "id": template_id,
                        "name": name,
                        "category": category or "General",
                        "description": description,
                        "usage_count": usage_count,
                        "last_updated": today,
                        "schema": schema,
                    }
                )
                _persist_srt_form_templates()
                flash("SRT form template added.", "success")

        elif action == "delete":
            template_id = request.form.get("template_id")
            if not template_id:
                flash("Template not found.", "error")
            else:
                updated_templates = [
                    item for item in SRT_FORM_TEMPLATES if item["id"] != template_id
                ]
                if len(updated_templates) == len(SRT_FORM_TEMPLATES):
                    flash("Template not found.", "error")
                else:
                    SRT_FORM_TEMPLATES = updated_templates
                    _persist_srt_form_templates()
                    flash("Template deleted.", "success")

        elif action == "update":
            template_id = request.form.get("template_id")
            template = next((item for item in SRT_FORM_TEMPLATES if item["id"] == template_id), None)
            if not template:
                flash("Template not found.", "error")
            elif not name:
                flash("Template name is required.", "error")
            else:
                template.update(
                    {
                        "name": name,
                        "category": category or "General",
                        "description": description,
                        "usage_count": usage_count,
                        "last_updated": today,
                        "schema": schema,
                    }
                )
                _persist_srt_form_templates()
                flash("Template updated.", "success")

        return redirect(url_for("srt_form_templates"))

    templates_normalised = []
    for template in SRT_FORM_TEMPLATES:
        template.setdefault("schema", _default_srt_schema())
        normalised_schema = copy.deepcopy(_normalise_srt_schema(template["schema"]))
        template["schema"] = normalised_schema
        templates_normalised.append(
            {
                **template,
                "schema": copy.deepcopy(normalised_schema),
            }
        )

    templates_normalised.sort(key=lambda item: item["name"].lower())

    return render_template("srt_form_templates.html", templates=templates_normalised)


@app.route("/srt/sites")
@login_required
def srt_sites():
    _module_visibility_required("srt")
    selected_key = request.args.get("site")

    sites = []
    site_map = {}
    for site in sorted(SRT_SITES, key=lambda item: item["name"].lower()):
        site_copy = {key: value for key, value in site.items() if key != "additional_contacts"}
        contacts = [dict(contact) for contact in site.get("additional_contacts", [])]
        site_copy["additional_contacts"] = contacts
        site_copy["tasks"] = [
            dict(task)
            for task in SRT_SAMPLE_TASKS
            if task.get("site") == site["name"]
        ]
        site_copy.setdefault("client_contact", {})
        site_copy.setdefault("form_update", {})
        site_copy.setdefault("updates", [])
        site_copy.setdefault("interactions", [])
        site_map[site["key"]] = site_copy
        sites.append(site_copy)

    if not sites:
        selected_site = None
    else:
        if selected_key and selected_key in site_map:
            selected_site = site_map[selected_key]
        else:
            selected_site = sites[0]
            selected_key = selected_site["key"]

    return render_template(
        "srt_sites.html",
        sites=sites,
        selected_site_key=selected_key,
        team_members=SRT_TEAM_MEMBERS,
    )


@app.route("/srt/task", methods=["POST"])
@login_required
def srt_task_create():
    _module_visibility_required("srt")
    site_name = (request.form.get("site_name") or "").strip()
    summary = (request.form.get("summary") or "").strip()
    task_name = (request.form.get("name") or "").strip()
    priority = (request.form.get("priority") or "Normal").strip().title() or "Normal"
    owner = (request.form.get("owner") or "Unassigned").strip() or "Unassigned"
    due_date_raw = request.form.get("due_date")

    due_date = None
    if due_date_raw:
        try:
            due_date = datetime.datetime.strptime(due_date_raw, "%Y-%m-%d").date()
        except ValueError:
            due_date = None

    if not site_name or not task_name or not summary:
        flash("Site, task name and summary are required to create a task.", "error")
        return redirect(url_for("srt_sites"))

    existing_ids = {task["id"] for task in SRT_SAMPLE_TASKS}
    task_id = f"SRT-{_random_digits(4)}"
    while task_id in existing_ids:
        task_id = f"SRT-{_random_digits(4)}"

    SRT_SAMPLE_TASKS.insert(
        0,
        {
            "id": task_id,
            "site": site_name,
            "name": task_name,
            "summary": summary,
            "priority": priority,
            "status": "Pending",
            "due_date": due_date,
            "owner": owner or "Unassigned",
            "age_days": 0,
        },
    )

    actor_name = None
    actor_role = None
    if current_user.is_authenticated:
        actor_name = current_user.display_name
        actor_role = "admin" if current_user.is_admin else "user"
    elif owner and owner != "Unassigned":
        actor_name = owner
        actor_role = "user"

    _log_srt_activity(
        task_id,
        type="status",
        label="Task created",
        detail=f"{task_name} — {summary}" if summary else task_name,
        actor=actor_name or "System",
        actor_role=actor_role,
    )

    flash("SRT task added to the board.", "success")

    redirect_to = request.form.get("redirect_to")
    if redirect_to and redirect_to.startswith("/"):
        return redirect(redirect_to)

    return redirect(url_for("srt_sites", site=slugify(site_name)))


@app.route("/srt/task/<task_id>/data")
@login_required
def srt_task_data(task_id):
    _module_visibility_required("srt")
    task = _get_srt_task(task_id)
    if not task:
        return jsonify({"ok": False, "message": "Task not found."}), 404

    today = datetime.date.today()
    due_date = task.get("due_date")
    due_in = (due_date - today).days if due_date else None

    timeline_entries = []
    for event in SRT_TASK_ACTIVITY.get(task_id, []):
        enriched = apply_actor_context(event)
        raw_timestamp = event.get("timestamp") or datetime.datetime.utcnow()
        if isinstance(raw_timestamp, str):
            try:
                timestamp_value = datetime.datetime.fromisoformat(raw_timestamp)
            except ValueError:
                timestamp_value = datetime.datetime.utcnow()
        else:
            timestamp_value = raw_timestamp
        timeline_entries.append(
            {
                "type": enriched.get("type", event.get("type", "update")),
                "label": enriched.get("label") or event.get("label") or "Update",
                "detail": enriched.get("detail") or event.get("detail"),
                "actor": enriched.get("actor"),
                "actor_role": enriched.get("actor_role"),
                "actor_label": enriched.get("actor_label"),
                "comment": event.get("comment"),
                "attachment_label": event.get("attachment_label"),
                "attachment_url": event.get("attachment_url"),
                "timestamp": timestamp_value.isoformat(),
                "timestamp_display": timestamp_value.strftime("%d %b %Y • %H:%M"),
                "_sort_timestamp": timestamp_value,
            }
        )

    timeline = sorted(
        timeline_entries,
        key=lambda item: item.get("_sort_timestamp") or datetime.datetime.utcnow(),
        reverse=True,
    )
    for entry in timeline:
        entry.pop("_sort_timestamp", None)

    task_payload = {
        "id": task["id"],
        "site": task["site"],
        "name": task.get("name") or "",
        "summary": task["summary"],
        "priority": task["priority"],
        "status": task["status"],
        "owner": task["owner"],
        "due_date": due_date.isoformat() if due_date else None,
        "due_date_display": due_date.strftime("%d %b %Y") if due_date else "",
        "due_in": due_in,
    }

    return jsonify({"ok": True, "task": task_payload, "timeline": timeline})


@app.route("/srt/task/<task_id>/update", methods=["POST"])
@login_required
def srt_task_update(task_id):
    _module_visibility_required("srt")
    task = _get_srt_task(task_id)
    if not task:
        flash("Task not found.", "error")
        return redirect(url_for("srt_overview"))

    redirect_to = request.form.get("redirect_to") or url_for("srt_overview")
    if not redirect_to.startswith("/"):
        redirect_to = url_for("srt_overview")

    status_raw = (request.form.get("status") or task.get("status") or "").strip()
    status_lookup = status_raw.lower()
    if status_lookup in {"in-progress", "in_progress"}:
        status = "In Progress"
    elif status_lookup == "closed":
        status = "Closed"
    elif status_lookup == "pending":
        status = "Pending"
    else:
        status = status_raw.title() or task.get("status") or "Pending"

    owner = (request.form.get("owner") or task.get("owner") or "").strip() or "Unassigned"
    due_date_raw = request.form.get("due_date")
    due_date = None
    if due_date_raw:
        try:
            due_date = datetime.datetime.strptime(due_date_raw, "%Y-%m-%d").date()
        except ValueError:
            due_date = task.get("due_date")
    else:
        due_date = None

    comment = (request.form.get("comment") or "").strip()
    attachment_label = (request.form.get("attachment_label") or "").strip()
    attachment_url = (request.form.get("attachment_url") or "").strip()

    original_status = task.get("status")
    original_owner = task.get("owner")
    original_due = task.get("due_date")

    events = []
    actor_info = timeline_actor_context()

    if status != original_status:
        events.append(
            {
                "type": "status",
                "label": f"Status updated to {status}",
                "detail": f"{original_status or '—'} → {status}",
                "actor": actor_info.get("actor"),
                "actor_role": actor_info.get("actor_role"),
            }
        )
        task["status"] = status

    if owner != original_owner:
        events.append(
            {
                "type": "assignment",
                "label": "Owner reassigned",
                "detail": f"{original_owner or 'Unassigned'} → {owner or 'Unassigned'}",
                "actor": actor_info.get("actor"),
                "actor_role": actor_info.get("actor_role"),
            }
        )
        task["owner"] = owner or "Unassigned"

    if due_date != original_due:
        new_due_display = due_date.strftime("%d %b %Y") if due_date else "No due date"
        old_due_display = original_due.strftime("%d %b %Y") if original_due else "No due date"
        events.append(
            {
                "type": "due_date",
                "label": "Due date updated",
                "detail": f"{old_due_display} → {new_due_display}",
                "actor": actor_info.get("actor"),
                "actor_role": actor_info.get("actor_role"),
            }
        )
        task["due_date"] = due_date

    if comment:
        events.append(
            {
                "type": "comment",
                "label": "Comment added",
                "comment": comment,
                "actor": actor_info.get("actor"),
                "actor_role": actor_info.get("actor_role"),
            }
        )

    if attachment_url:
        events.append(
            {
                "type": "attachment",
                "label": attachment_label or "Attachment uploaded",
                "attachment_label": attachment_label or attachment_url,
                "attachment_url": attachment_url,
                "actor": actor_info.get("actor"),
                "actor_role": actor_info.get("actor_role"),
            }
        )

    if not events:
        flash("No updates were made to the task.", "info")
        return redirect(redirect_to)

    for event in events:
        _log_srt_activity(task_id, **event)

    flash("Task updated successfully.", "success")
    return redirect(redirect_to)


# ---------------------- QC TABS ----------------------
@app.route("/qc")
@login_required
def qc_home():
    _module_visibility_required("qc")
    status = request.args.get("status", "all")
    status_order = case(
        (QCWork.status == "In Progress", 0),
        (QCWork.status == "Open", 1),
        (QCWork.status == "Blocked", 2),
        (QCWork.status == "Closed", 3),
        else_=4
    )
    query = QCWork.query.order_by(
        status_order,
        QCWork.due_date.asc().nullslast(),
        QCWork.created_at.desc()
    )
    if status == "open":
        query = query.filter(QCWork.status != "Closed")
    elif status == "closed":
        query = query.filter(QCWork.status == "Closed")

    work_items = query.all()
    templates = FormSchema.query.order_by(FormSchema.name.asc()).all()
    users = get_assignable_users_for_module("qc", order_by="username")
    projects = Project.query.order_by(Project.name.asc()).all()
    return render_template(
        "qc.html",
        work_items=work_items,
        templates=templates,
        users=users,
        projects=projects,
        STAGES=STAGES,
        LIFT_TYPES=LIFT_TYPES,
        status_filter=status,
        TASK_MILESTONES=TASK_MILESTONES
    )


@app.route("/qc/work/new", methods=["POST"])
@login_required
def qc_work_new():
    _module_visibility_required("qc")
    site_name = (request.form.get("site_name") or "").strip()
    client_name = (request.form.get("client_name") or "").strip()
    address = (request.form.get("address") or "").strip()
    template_id = request.form.get("template_id", type=int)
    stage = (request.form.get("stage") or "").strip()
    lift_type = (request.form.get("lift_type") or "").strip()
    due = (request.form.get("due_date") or "").strip()
    assigned_to = request.form.get("assigned_to", type=int)
    project_id = request.form.get("project_id", type=int)
    planned_start = (request.form.get("planned_start_date") or "").strip()
    duration_raw = (request.form.get("planned_duration_days") or "").strip()
    milestone_value = (request.form.get("milestone") or "").strip()

    project = db.session.get(Project, project_id) if project_id else None
    if project:
        site_name = site_name or project.site_name or project.name
        client_name = client_name or (project.customer_name or "")
        address = address or (project.site_address or "")
        lift_type = lift_type or (project.lift_type or "")

    if not site_name or not template_id:
        flash("Site name and Template are required.", "error")
        return redirect(url_for("qc_home"))

    due_dt = None
    if due:
        try:
            due_dt = datetime.datetime.strptime(due, "%Y-%m-%d")
        except Exception:
            flash("Invalid due date format.", "error")
            return redirect(url_for("qc_home"))

    planned_start_date = None
    if planned_start:
        try:
            planned_start_date = datetime.datetime.strptime(planned_start, "%Y-%m-%d").date()
        except ValueError:
            flash("Invalid planned start date format.", "error")
            return redirect(url_for("qc_home"))

    duration_days = None
    if duration_raw:
        try:
            duration_days = int(duration_raw)
        except ValueError:
            flash("Duration must be a whole number of days.", "error")
            return redirect(url_for("qc_home"))
        if duration_days < 0:
            flash("Duration must be zero or positive.", "error")
            return redirect(url_for("qc_home"))

    if due_dt is None:
        if planned_start_date and duration_days is not None:
            due_dt = datetime.datetime.combine(planned_start_date, datetime.time.min) + datetime.timedelta(days=duration_days)
        elif duration_days is not None:
            due_dt = datetime.datetime.utcnow() + datetime.timedelta(days=duration_days)

    template = db.session.get(FormSchema, template_id)

    assignee_user = None
    if assigned_to:
        assignee_user = db.session.get(User, assigned_to)
        if not assignee_user or not assignee_user.can_be_assigned_module("qc"):
            flash("Choose an assignee who is available for QC tasks.", "error")
            return redirect(url_for("qc_home"))

    work = QCWork(
        site_name=site_name,
        client_name=client_name or None,
        address=address or None,
        template_id=template_id,
        stage=stage or (template.stage if template else None),
        lift_type=lift_type or (template.lift_type if template else None),
        project_id=project.id if project else None,
        due_date=due_dt,
        created_by=current_user.id,
        assigned_to=assigned_to if assignee_user else None,
        name=site_name,
        planned_start_date=planned_start_date,
        planned_duration_days=duration_days,
        milestone=milestone_value or None
    )
    db.session.add(work)
    db.session.flush()
    log_work_event(
        work.id,
        "created",
        actor_id=current_user.id,
        details={
            "site_name": work.site_name,
            "assigned_to": assigned_to,
            "due_date": work.due_date.strftime("%Y-%m-%d") if work.due_date else None,
            "project_id": work.project_id,
            "planned_start_date": planned_start_date.strftime("%Y-%m-%d") if planned_start_date else None,
            "planned_duration_days": duration_days,
            "milestone": work.milestone
        }
    )
    if assignee_user:
        log_work_event(
            work.id,
            "assigned",
            actor_id=current_user.id,
            details={"assigned_to": assigned_to}
        )
    db.session.commit()
    flash("QC work created.", "success")
    return redirect(url_for("qc_work_detail", work_id=work.id))


@app.route("/qc/work/<int:work_id>")
@login_required
def qc_work_detail(work_id):
    _module_visibility_required("qc")
    work = QCWork.query.get_or_404(work_id)
    submissions = Submission.query.filter_by(work_id=work_id).order_by(Submission.created_at.desc()).all()
    for submission in submissions:
        try:
            submission.photo_count = len(json.loads(submission.photos_json or "[]"))
        except Exception:
            submission.photo_count = 0
        try:
            submission.video_count = len(json.loads(submission.videos_json or "[]"))
        except Exception:
            submission.video_count = 0
    users = get_assignable_users_for_module("qc", order_by="username")
    comments = (
        QCWorkComment.query
        .filter_by(work_id=work_id)
        .order_by(QCWorkComment.created_at.desc())
        .all()
    )
    for comment in comments:
        try:
            comment.attachments = json.loads(comment.attachments_json or "[]")
        except Exception:
            comment.attachments = []
        comment.has_attachments = bool(comment.attachments)
    logs = QCWorkLog.query.filter_by(work_id=work_id).order_by(QCWorkLog.created_at.desc()).all()
    for log in logs:
        try:
            log.details = json.loads(log.details_json or "{}")
        except Exception:
            log.details = {}
        if isinstance(log.details, dict):
            if "assigned_to" in log.details:
                user_id = log.details.get("assigned_to")
                if user_id:
                    user_obj = db.session.get(User, user_id)
                    log.details["assigned_to"] = user_obj.username if user_obj else user_id
                else:
                    log.details["assigned_to"] = "Unassigned"
            if "from" in log.details:
                prev_id = log.details.get("from")
                if prev_id:
                    user_obj = db.session.get(User, prev_id)
                    log.details["from"] = user_obj.username if user_obj else prev_id
                elif prev_id is None:
                    log.details["from"] = "Unassigned"
            if "to" in log.details:
                next_id = log.details.get("to")
                if next_id:
                    user_obj = db.session.get(User, next_id)
                    log.details["to"] = user_obj.username if user_obj else next_id
                elif next_id is None:
                    log.details["to"] = "Unassigned"
    attachment_entries = []
    for comment in comments:
        for item in getattr(comment, "attachments", []) or []:
            web_path = item.get("web_path") or (
                item.get("path", "").split("static/", 1)[1]
                if "path" in item and "static/" in item.get("path", "")
                else item.get("path")
            )
            attachment_entries.append({
                "name": item.get("name") or "Attachment",
                "web_path": web_path,
                "comment_id": comment.id,
                "author": comment.author.username if comment.author else "Unknown",
                "created_at": comment.created_at,
                "body": (comment.body[:160] + ("…" if len(comment.body) > 160 else "")) if comment.body else None,
            })

    attachment_entries.sort(key=lambda entry: entry["created_at"], reverse=True)
    return render_template(
        "qc_work_detail.html",
        work=work,
        submissions=submissions,
        users=users,
        comments=comments,
        logs=logs,
        attachments=attachment_entries
    )


@app.route("/qc/work/<int:work_id>/assign", methods=["POST"])
@login_required
def qc_work_assign(work_id):
    _module_visibility_required("qc")
    work = QCWork.query.get_or_404(work_id)
    assigned_to = request.form.get("assigned_to", type=int)
    assignee_user = None
    if assigned_to:
        assignee_user = db.session.get(User, assigned_to)
        if not assignee_user or not assignee_user.can_be_assigned_module("qc"):
            flash("Choose an assignee who is available for QC tasks.", "error")
            return redirect(url_for("qc_work_detail", work_id=work.id))
    previous = work.assigned_to
    work.assigned_to = assigned_to if assignee_user else None
    if previous != work.assigned_to:
        db.session.flush()
        log_work_event(
            work.id,
            "assignment_updated",
            actor_id=current_user.id,
            details={"from": previous, "to": work.assigned_to}
        )
        db.session.commit()
        flash("Assignment updated.", "success")
    else:
        db.session.rollback()
        flash("Assignment unchanged.", "info")
    return redirect(url_for("qc_work_detail", work_id=work.id))


@app.route("/qc/work/<int:work_id>/comment", methods=["POST"])
@login_required
def qc_work_comment(work_id):
    work = QCWork.query.get_or_404(work_id)
    body = (request.form.get("body") or "").strip()
    if not body and not request.files.getlist("attachments"):
        flash("Add a comment or attachment.", "error")
        return redirect(url_for("qc_work_detail", work_id=work.id))

    attachments = []
    for f in request.files.getlist("attachments"):
        if f and f.filename:
            if not allowed_file(f.filename, kind="attachment"):
                flash(f"Unsupported file type for {f.filename}.", "error")
                return redirect(url_for("qc_work_detail", work_id=work.id))
            fname = secure_filename(f.filename)
            dest_name = f"{datetime.datetime.utcnow().timestamp()}_{fname}"
            dest = os.path.join(app.config["UPLOAD_FOLDER"], dest_name)
            f.save(dest)
            rel_path = dest.split("static/", 1)[1] if "static/" in dest else dest
            attachments.append({"path": dest, "name": fname, "web_path": rel_path})

    comment = QCWorkComment(
        work_id=work.id,
        author_id=current_user.id,
        body=body,
        attachments_json=json.dumps(attachments, ensure_ascii=False)
    )
    db.session.add(comment)
    db.session.flush()
    log_details = {"comment_id": comment.id}
    if body:
        snippet = body if len(body) <= 160 else body[:157] + "…"
        log_details["body"] = snippet
    if attachments:
        log_details["attachments"] = [item.get("name") for item in attachments if item.get("name")]
    log_work_event(
        work.id,
        "comment_added",
        actor_id=current_user.id,
        details=log_details
    )
    db.session.commit()
    flash("Comment added.", "success")
    return redirect(url_for("qc_work_detail", work_id=work.id))


@app.route("/qc/work/<int:work_id>/status/<string:action>", methods=["POST"])
@login_required
def qc_work_status(work_id, action):
    _module_visibility_required("qc")
    """Progress status for work: open -> in_progress -> closed."""
    work = QCWork.query.get_or_404(work_id)
    if action not in {"start", "close", "reopen"}:
        flash("Invalid action.", "error")
        return redirect(url_for("qc_work_detail", work_id=work.id))

    from_status = work.status or "Open"
    if action == "start":
        if from_status == "Blocked" and not work.dependency_satisfied:
            flash("This task is waiting for its dependency to complete.", "error")
            return redirect(url_for("qc_work_detail", work_id=work.id))
        if from_status == "Blocked" and work.dependency_satisfied:
            work.status = "Open"
            db.session.flush()
            log_work_event(
                work.id,
                "dependency_released",
                actor_id=current_user.id,
                details={"dependency": work.depends_on_id}
            )
            from_status = work.status
        if from_status != "Open":
            flash(f"Cannot start: current status is {work.status}.", "error")
            return redirect(url_for("qc_work_detail", work_id=work.id))
        new_status = "In Progress"
    elif action == "close":
        if from_status != "In Progress":
            flash(f"Cannot close: current status is {work.status}.", "error")
            return redirect(url_for("qc_work_detail", work_id=work.id))
        new_status = "Closed"
    else:  # reopen
        if from_status != "Closed":
            flash(f"Cannot reopen: current status is {work.status}.", "error")
            return redirect(url_for("qc_work_detail", work_id=work.id))
        new_status = "In Progress"

    work.status = new_status
    db.session.flush()
    log_work_event(
        work.id,
        "status_changed",
        actor_id=current_user.id,
        from_status=from_status,
        to_status=new_status
    )
    if new_status == "Closed":
        release_dependent_tasks(work, actor_id=current_user.id)
    elif action == "reopen":
        block_child_tasks(work, actor_id=current_user.id)
    db.session.commit()
    flash(f"Work status changed to {new_status}.", "success")
    return redirect(url_for("qc_work_detail", work_id=work.id))
# ----------------------------------------------------


@app.route("/qc/recent-submissions")
@login_required
def qc_recent_submissions():
    _module_visibility_required("qc")
    submissions = (
        Submission.query
        .order_by(Submission.created_at.desc())
        .limit(50)
        .all()
    )
    return render_template(
        "qc_recent_submissions.html",
        submissions=submissions
    )


@app.cli.command("initdb")
def initdb():
    """Initialize database and seed sample data"""
    bootstrap_db()
    print("Database initialized with default users and sample form.")


_bootstrap_lock = threading.Lock()
_bootstrapped = False


def ensure_bootstrap():
    global _bootstrapped
    if _bootstrapped:
        return
    with _bootstrap_lock:
        if _bootstrapped:
            return
        try:
            bootstrap_db()
            _bootstrapped = True
        except Exception as exc:
            app.logger.exception("Database bootstrap failed: %s", exc)


@app.before_request
def _ensure_db_ready():
    ensure_bootstrap()


if __name__ == "__main__":
    with app.app_context():
        bootstrap_db()
    app.run(debug=True)
